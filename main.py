"""
Lister AI — FastAPI Web Dashboard
Replaces Streamlit for real-time performance.
"""
import os
import csv
import io
import re
from datetime import datetime
from dotenv import load_dotenv
from fastapi import FastAPI, Request, HTTPException, UploadFile, File, Body, Response, BackgroundTasks
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from supabase import create_client
from pydantic import BaseModel
from typing import Optional, List

load_dotenv()

SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_KEY")
print(f"Connecting to Supabase: {SUPABASE_URL}")
supabase     = create_client(SUPABASE_URL, SUPABASE_KEY)

app = FastAPI(title="Lister AI")

@app.middleware("http")
async def _no_stale_html(request, call_next):
    # Every page (text/html) is served no-store so a deploy is ALWAYS picked up
    # on the next normal refresh — no more hard-refresh/stale-JS confusion where
    # a fix is live on the server but the browser keeps rendering old JS.
    # API/JSON responses are untouched.
    response = await call_next(request)
    if response.headers.get("content-type", "").startswith("text/html"):
        response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        response.headers["Pragma"] = "no-cache"
    return response

import os as _os
if _os.path.isdir("static"):
    app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")

# Railway restarts this process on every deploy, so process-start time is a
# reliable proxy for "last deployed at" -- no dependency on any Railway-specific
# env var that may or may not exist. Registered as a Jinja2 global (not passed
# per-route) so it shows up on every page automatically, including any route
# added later, without needing to remember to wire it in each time. Converted to
# Mountain time for display since that's where this business is based.
_deploy_time = datetime.utcnow()
try:
    from zoneinfo import ZoneInfo as _ZoneInfo
    _deploy_time_mt = _deploy_time.replace(tzinfo=_ZoneInfo("UTC")).astimezone(_ZoneInfo("America/Denver"))
except Exception:
    _deploy_time_mt = _deploy_time  # fall back to UTC if zoneinfo/tzdata isn't available for some reason
templates.env.globals["deploy_time_mt"] = _deploy_time_mt

def guess_brand_from_title(title: str) -> str:
    first_word = (title or "").split()[0].strip(",.;:-") if title else ""
    return first_word if first_word else "Unbranded"

def _category_is_restricted_ok(cat_id: str, path_map: dict) -> bool:
    """True only if cat_id is known locally (synced via sync_ebay_categories) AND its
    path actually falls under Business & Industrial / eBay Motors. An id that's blank,
    "0", or simply not in our synced tree at all does NOT count as verified-good."""
    path = path_map.get(str(cat_id))
    if path is None:
        return False
    return ("Business & Industrial" in path) or ("eBay Motors" in path)

async def auto_fill_worker():
    """Runs continuously in the background. Any listing missing a brand or eBay category
    gets one filled in automatically within seconds of the scan finishing — no manual click
    needed. Also periodically re-validates listings that already have SOME category (e.g.
    mantle-scanner's own initial guess, which is never restricted to Business & Industrial /
    eBay Motors) — the null/0/blank check alone let a wrong category sail through forever
    once mantle-scanner had written anything non-empty into ebay_category_id. Also retries
    anything still sitting at the generic default fallback category — that fallback is
    technically INSIDE the allowed Business & Industrial/eBay Motors root, so it used to be
    invisible to the revalidation sweep and stayed wrong forever with zero automatic retries,
    even after the underlying matching logic was fixed."""
    import asyncio

    def needs_category(row: dict) -> bool:
        cat = row.get("ebay_category_id")
        # mantle-scanner writes "0" as a placeholder instead of leaving this NULL,
        # so treat None / "" / "0" all as "not actually categorized yet".
        return cat is None or str(cat).strip() in ("", "0")

    def fix_row(row: dict):
        title = row.get("title") or ""
        biz_id = row.get("business_id")
        if not title or title == "Scanning..." or not biz_id:
            print(f"auto_fill_worker: skipping {row['id']} (title={title!r}, biz_id={biz_id})")
            return
        updates = {}
        if not row.get("brand"):
            updates["brand"] = guess_brand_from_title(title)
        try:
            row_mode = row.get("category_mode") or "industrial"
            if row_mode not in ("industrial", "motors"):
                row_mode = "industrial"
            suggestion = suggest_ebay_category(title, biz_id, restrict=True, mode=row_mode)
            if suggestion and suggestion.get("category_id"):
                updates["ebay_category_id"] = suggestion["category_id"]
                print(f"auto_fill_worker: {row['id']} -> category {suggestion['category_id']} "
                      f"({suggestion.get('name')}, mode={row_mode})")
            else:
                fallback = _motors_fallback_id(biz_id) if row_mode == "motors" else \
                    (get_ebay_settings(biz_id).get("EBAY_DEFAULT_CATEGORY_ID", "") or "26261")
                updates["ebay_category_id"] = fallback
                print(f"auto_fill_worker: {row['id']} -> no match in {row_mode} lane, locked to {fallback}")
        except Exception as e:
            print(f"auto_fill_worker category error for {row['id']}: {e}")
        if updates:
            supabase.table("listings").update(updates).eq("id", row["id"]).execute()

    cycle = 0
    while True:
        try:
            res = supabase.table("listings").select("id,title,brand,ebay_category_id,business_id,category_mode")\
                .neq("status", "archived")\
                .or_("ebay_category_id.is.null,ebay_category_id.eq.0,ebay_category_id.eq.")\
                .limit(50).execute()
            rows = [r for r in (res.data or []) if needs_category(r)]
            print(f"auto_fill_worker: query returned {len(res.data or [])} row(s), {len(rows)} need a category")
            for row in rows:
                fix_row(row)

            # Every ~10th cycle (~80s), also sweep listings that already have SOME
            # category and re-validate it against Business & Industrial / eBay Motors.
            cycle += 1
            if cycle % 10 == 0:
                all_rows = []
                start = 0
                while True:
                    page = supabase.table("listings").select("id,title,brand,ebay_category_id,business_id,category_mode")\
                        .neq("status", "archived")\
                        .range(start, start + 999).execute().data or []
                    all_rows.extend(page)
                    if len(page) < 1000:
                        break
                    start += 1000
                already_categorized = [r for r in all_rows if not needs_category(r)]
                cat_ids = list({str(r["ebay_category_id"]) for r in already_categorized})
                path_map = {}
                for i in range(0, len(cat_ids), 500):
                    chunk = cat_ids[i:i+500]
                    pres = supabase.table("ebay_categories").select("category_id,path").in_("category_id", chunk).execute()
                    for prow in (pres.data or []):
                        path_map[str(prow["category_id"])] = prow.get("path") or ""

                def is_stuck_at_generic_default(row: dict) -> bool:
                    # The generic fallback (e.g. "Other Business & Industrial") is
                    # technically INSIDE the allowed root, so _category_is_restricted_ok
                    # alone will never flag it — meaning an item mantle-scanner (or
                    # anything else) drops here because no specific match was found
                    # stays here forever, even after the underlying matching logic
                    # gets fixed, since nothing ever asks it again. Explicitly retry
                    # these too, per business, since the fallback ID is configurable.
                    row_mode = row.get("category_mode") or "industrial"
                    if row_mode == "motors":
                        biz_default = _motors_fallback_id(row["business_id"])
                    else:
                        biz_default = get_ebay_settings(row["business_id"]).get("EBAY_DEFAULT_CATEGORY_ID", "") or "26261"
                    return str(row.get("ebay_category_id") or "") == str(biz_default)

                misfiled = [r for r in already_categorized
                            if not _category_is_restricted_ok(r["ebay_category_id"], path_map)
                            or is_stuck_at_generic_default(r)][:50]
                print(f"auto_fill_worker: revalidation sweep found {len(misfiled)} listing(s) categorized outside Business & Industrial/eBay Motors, or still stuck at the generic fallback")
                for row in misfiled:
                    fix_row(row)
        except Exception as e:
            print(f"auto_fill_worker error: {e}")
        await asyncio.sleep(8)

@app.on_event("startup")
async def start_background_jobs():
    import asyncio
    asyncio.create_task(auto_fill_worker())
    asyncio.create_task(order_sync_worker())
    asyncio.create_task(shopify_sync_worker())
    asyncio.create_task(analytics_snapshot_worker())
    asyncio.create_task(analytics_cache_refresh_worker())
    asyncio.create_task(auction_archive_worker())
    asyncio.create_task(active_listings_sync_worker())
    asyncio.create_task(post_publish_sync_worker())
    asyncio.create_task(shopify_shipping_rule_worker())
    asyncio.create_task(inventory_value_sync_worker())
    asyncio.create_task(ebay_best_offers_sync_worker())
    asyncio.create_task(ebay_notification_subscribe_worker())
    asyncio.create_task(ebay_analytics_sync_worker())
    asyncio.create_task(ebay_sync_check_worker())

async def ebay_analytics_sync_worker():
    """Once-a-day sweep of real view counts via the Sell Analytics API,
    confirmed working 8/8 (fresh sell.analytics.readonly consent verified
    against real data). Checks hourly but only actually syncs a business
    once its last run is >= 1 day old, settling into a daily cadence --
    same pattern as active_listings_sync_worker. Deliberately not more
    frequent than that: this API has its own separate, real daily call
    limit (see _sync_ebay_analytics_work notes above), not the Trading
    API's, but still a finite one not worth burning on a tight clock."""
    import asyncio, datetime as _dt
    while True:
        try:
            res = supabase.table("app_settings").select("business_id").eq("key", "EBAY_REFRESH_TOKEN").execute()
            business_ids = list(set(r["business_id"] for r in (res.data or [])))
            for biz_id in business_ids:
                try:
                    settings = get_ebay_settings(biz_id)
                    last_run = settings.get("EBAY_ANALYTICS_SYNC_LAST_RUN_AT")
                    needs_sync = True
                    if last_run:
                        age = _dt.datetime.utcnow() - _dt.datetime.fromisoformat(
                            last_run.replace("Z", "+00:00")).replace(tzinfo=None)
                        needs_sync = age >= _dt.timedelta(days=1)
                    if not needs_sync:
                        continue
                    result = await asyncio.to_thread(_sync_ebay_analytics_work, biz_id)
                    save_ebay_setting(biz_id, "EBAY_ANALYTICS_SYNC_LAST_RUN_AT", _dt.datetime.utcnow().isoformat())
                    print(f"ebay_analytics_sync_worker: business {biz_id}: {result}")
                except Exception as e:
                    print(f"ebay_analytics_sync_worker: business {biz_id} failed: {e}")
        except Exception as e:
            print(f"ebay_analytics_sync_worker error: {e}")
        await asyncio.sleep(3600)  # check hourly; actual sync per business gated to ~daily above

async def auction_archive_worker():
    """Runs once a day (00:20 UTC) and archives any capture session whose lots'
    auction(s) have all closed -- keyed off the LATEST auction_ends_at among that
    session's lots, since a single capture (e.g. one multi-day watchlist scan) can
    span lots from several different closing dates. Archived sessions are hidden
    from the normal picker (list_capture_sessions) but never deleted -- all lots,
    pricing, and notes stay intact in Supabase in case they're needed later."""
    import asyncio, datetime as _dt
    while True:
        now = _dt.datetime.now(_dt.timezone.utc)
        next_run = now.replace(hour=0, minute=20, second=0, microsecond=0)
        if next_run <= now:
            next_run += _dt.timedelta(days=1)
        await asyncio.sleep(max((next_run - now).total_seconds(), 5))
        try:
            sessions = supabase.table("auction_capture_sessions").select("id").eq("archived", False).execute().data
            for sess in sessions:
                try:
                    lots = supabase.table("auction_lots").select("auction_ends_at").eq("session_id", sess["id"]).execute().data
                    end_dates = [l["auction_ends_at"] for l in lots if l.get("auction_ends_at")]
                    if not end_dates:
                        continue  # no known auction date yet for any lot in this session -- leave it alone
                    latest = max(_dt.datetime.fromisoformat(d.replace("Z", "+00:00")) for d in end_dates)
                    if latest < now:
                        supabase.table("auction_capture_sessions").update({
                            "archived": True, "archived_at": now.isoformat(),
                        }).eq("id", sess["id"]).execute()
                        print(f"auction_archive_worker: archived session {sess['id']} (latest auction ended {latest.isoformat()})")
                except Exception as e:
                    print(f"auction_archive_worker: session {sess['id']} failed: {e}")
        except Exception as e:
            print(f"auction_archive_worker error: {e}")

async def shopify_sync_worker():
    """Runs the eBay-live-qty -> Shopify-qty sync automatically once per hour, at
    the top of the hour. Computes 'today' server-side in Mountain time (not UTC,
    not a browser's local clock) so it stays correct regardless of when this
    process happens to be running. Sleeps until the next :00 rather than a flat
    3600s interval so it doesn't drift off the hour over time."""
    import asyncio, datetime as _dt
    while True:
        now = _dt.datetime.now(_dt.timezone.utc)
        next_hour = (now.replace(minute=0, second=0, microsecond=0) + _dt.timedelta(hours=1))
        await asyncio.sleep(max((next_hour - now).total_seconds(), 5))
        try:
            res = supabase.table("app_settings").select("business_id").eq("key", "SHOPIFY_STORE_DOMAIN").execute()
            business_ids = list(set(r["business_id"] for r in (res.data or [])))
            for biz_id in business_ids:
                try:
                    result = await asyncio.to_thread(run_hourly_shopify_sync_for_business, biz_id)
                    print(f"shopify_sync_worker: business {biz_id} -> {result}")
                except Exception as e:
                    print(f"shopify_sync_worker: business {biz_id} failed: {e}")
        except Exception as e:
            print(f"shopify_sync_worker error: {e}")

def run_hourly_shopify_sync_for_business(business_id: str) -> dict:
    """The actual per-business work for one hourly cycle: pull today's eBay sales,
    refresh the eBay/Shopify snapshot for just those titles, then push a Shopify
    qty update only for items whose Shopify qty doesn't already match eBay's live
    qty — never re-pushes an already-correct value."""
    import datetime as _dt
    from zoneinfo import ZoneInfo

    settings = get_ebay_settings(business_id)
    if not (settings.get("SHOPIFY_STORE_DOMAIN") or "").strip():
        return {"skipped": "no_shopify_connected"}

    mt_today = _dt.datetime.now(ZoneInfo("America/Denver")).strftime("%Y-%m-%d")
    sold_by_title = _shopify_sync_sold_by_title(business_id, mt_today, mt_today)
    if not sold_by_title:
        return {"sold_today": 0}

    items = [{"title": v["title"], "sku": v["sku"], "legacy_item_id": v.get("legacy_item_id", "")}
             for v in sold_by_title.values()]
    refresh_result = _shopify_sync_refresh_work(business_id, items)
    if refresh_result.get("error"):
        return {"sold_today": len(sold_by_title), "refresh_error": refresh_result["error"]}

    norm_titles = set(sold_by_title.keys())
    snap_res = supabase.table("shopify_sync_snapshot").select("*").eq("business_id", business_id).execute()
    snapshots = {row["norm_title"]: row for row in (snap_res.data or []) if row["norm_title"] in norm_titles}

    push_items = []
    for norm_title, snap in snapshots.items():
        if not snap.get("shopify_inventory_item_id"):
            continue
        if snap.get("ebay_live_qty") is None:
            continue
        if snap.get("shopify_live_qty") == snap.get("ebay_live_qty"):
            continue  # already correct — don't re-push an unchanged value every hour
        sold_entry = sold_by_title.get(norm_title, {})
        push_items.append({
            "title": snap.get("title"), "sku": snap.get("sku"),
            "shopify_inventory_item_id": snap.get("shopify_inventory_item_id"),
            "ebay_live_qty": snap.get("ebay_live_qty"),
            "qty_sold_today": sold_entry.get("qty_sold_today", 0),
            "order_ids": sold_entry.get("order_ids", []),
        })

    push_result = {"pushed": 0}
    if push_items:
        try:
            push_result = _push_shopify_qty_updates(business_id, push_items)
        except Exception as e:
            push_result = {"error": str(e)}

    return {
        "sold_today": len(sold_by_title), "refreshed": refresh_result.get("checked", 0),
        "ebay_checked": refresh_result.get("ebay_checked", 0), "ebay_reused": refresh_result.get("ebay_reused", 0),
        "skipped_recent_failures": refresh_result.get("skipped_recent_failures", 0),
        "push_candidates": len(push_items), "push_result": push_result,
    }

@app.post("/api/shopify-sync/run-now")
async def shopify_sync_run_now(request: Request):
    """Manual trigger for the exact same hourly job — useful for testing without
    waiting for the next top of the hour."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    import asyncio
    result = await asyncio.to_thread(run_hourly_shopify_sync_for_business, business_id)
    return result

# ── EBAY SYNC (Shopify -> eBay, manual-only for now) ───────────── #
# Deliberately the reverse of the automated direction above, and deliberately
# NOT wired into any worker -- explicit instruction: prove it works via a
# manual button first, automate later once trusted. Matching uses ONLY a
# confirmed HDID (inventory_match.hd_id not null) -- an item with no HDID is
# skipped outright, never fuzzy-matched, by design.

def _find_shopify_sales_with_hdid_match(business_id: str, start_iso: str, end_iso: str) -> dict:
    """Finds Shopify order lines in [start_iso, end_iso), aggregates quantity
    sold per variant_id, and joins against inventory_match on shopify_id --
    but ONLY rows where hd_id is not null (a confirmed, locked pairing).
    Anything sold with no confirmed HDID is counted but not returned as an
    actionable row -- 'if it doesn't have one, it should do nothing.'"""
    order_lines = fetch_shopify_orders(business_id, start_iso, end_iso)
    sold_by_variant = {}
    no_variant_id = 0
    for li in order_lines:
        vid = li.get("variant_id")
        if not vid:
            no_variant_id += 1
            continue
        entry = sold_by_variant.setdefault(vid, {"qty_sold": 0, "title": li.get("title"), "sku": li.get("sku")})
        entry["qty_sold"] += li.get("quantity", 0)

    if not sold_by_variant:
        return {"rows": [], "unmatched_no_hdid": 0, "no_variant_id": no_variant_id}

    variant_ids = list(sold_by_variant.keys())
    matches = (supabase.table("inventory_match").select("ebay_id,shopify_id,title")
               .eq("business_id", business_id).in_("shopify_id", variant_ids)
               .not_.is_("hd_id", "null").execute()).data or []
    matched_by_variant = {m["shopify_id"]: m for m in matches}

    rows = []
    unmatched = 0
    for vid, sold in sold_by_variant.items():
        match = matched_by_variant.get(vid)
        if not match or not match.get("ebay_id"):
            unmatched += 1
            continue
        rows.append({
            "shopify_variant_id": vid, "ebay_item_id": match["ebay_id"],
            "title": match.get("title") or sold["title"], "sku": sold["sku"],
            "qty_sold": sold["qty_sold"],
        })
    return {"rows": rows, "unmatched_no_hdid": unmatched, "no_variant_id": no_variant_id}

def _attach_current_quantities(business_id: str, rows: list) -> list:
    """Enriches each row with the CURRENT eBay/Shopify quantity. eBay comes
    from local cache (ebay_listing_status.quantity_available) -- refreshed at
    discovery time and via the explicit Refresh button, kept cache-based on
    purpose to respect eBay's much stricter Trading API quota. Shopify is a
    LIVE lookup instead -- Shopify's Admin API has no comparable quota
    pressure, and the cached shopify_inventory table proved unreliable here:
    a real example (Murphy switchgage) sat showing a stale cached quantity
    that didn't match Shopify's actual live count, giving a wrong 'set eBay
    qty to N' target. A bare cache read is cheap but was silently wrong;
    a live batched GraphQL call is barely more expensive and can't be stale.
    Self-heals the shopify_inventory cache with whatever's read live, same
    healing pattern as the eBay listing-status fix."""
    if not rows:
        return rows
    ebay_ids = list({r["ebay_item_id"] for r in rows if r.get("ebay_item_id")})
    shopify_ids = list({r["shopify_variant_id"] for r in rows if r.get("shopify_variant_id")})

    ebay_qty_by_id = {}
    if ebay_ids:
        for i in range(0, len(ebay_ids), 200):
            chunk = ebay_ids[i:i + 200]
            res = (supabase.table("ebay_listing_status").select("item_id,quantity_available")
                   .eq("business_id", business_id).in_("item_id", chunk).execute()).data or []
            for r in res:
                ebay_qty_by_id[r["item_id"]] = r.get("quantity_available")

    shopify_qty_by_id = {}
    if shopify_ids:
        try:
            settings = get_ebay_settings(business_id)
            domain = (settings.get("SHOPIFY_STORE_DOMAIN", "") or "").strip().replace("https://", "").replace("http://", "").strip("/")
            if domain:
                shopify_token = get_shopify_access_token(business_id)
                sp_headers = {"X-Shopify-Access-Token": shopify_token, "Content-Type": "application/json"}
                location_id = _get_shopify_primary_location_id(domain, sp_headers)
                if location_id:
                    variant_data = _fetch_shopify_variants_by_ids(domain, shopify_token, location_id, shopify_ids)
                    for vid, data in variant_data.items():
                        if data.get("qty") is not None:
                            shopify_qty_by_id[vid] = data["qty"]
                            try:
                                supabase.table("shopify_inventory").update({"quantity": data["qty"]}) \
                                    .eq("business_id", business_id).eq("id", vid).execute()
                            except Exception as e:
                                print(f"_attach_current_quantities: shopify_inventory cache heal failed for {vid}: {e}")
        except Exception as e:
            print(f"_attach_current_quantities: live Shopify qty fetch failed, falling back to cache: {e}")
        # Fallback to cache for anything the live call didn't cover (fetch failed entirely, or a
        # specific variant came back missing) rather than showing nothing.
        missing = [vid for vid in shopify_ids if vid not in shopify_qty_by_id]
        if missing:
            for i in range(0, len(missing), 200):
                chunk = missing[i:i + 200]
                res = (supabase.table("shopify_inventory").select("id,quantity")
                       .eq("business_id", business_id).in_("id", chunk).execute()).data or []
                for r in res:
                    shopify_qty_by_id[r["id"]] = r.get("quantity")

    for r in rows:
        r["ebay_qty"] = ebay_qty_by_id.get(r.get("ebay_item_id"))
        r["shopify_qty"] = shopify_qty_by_id.get(r.get("shopify_variant_id"))
    return rows

@app.get("/api/ebay-sync/today")
async def ebay_sync_today(request: Request, start: str, end: str):
    """Manual date-range lookup, kept as a secondary tool for checking a
    specific past window on demand. The PRIMARY path is the automatic queue
    below (ebay_sync_check_worker) -- this one still makes a live Shopify
    call, so it's for deliberately re-checking a range, not the everyday
    flow."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    import asyncio
    start_iso = f"{start}T00:00:00-00:00"
    end_iso = f"{end}T23:59:59-00:00"
    result = await asyncio.to_thread(_find_shopify_sales_with_hdid_match, business_id, start_iso, end_iso)
    result["rows"] = await asyncio.to_thread(_attach_current_quantities, business_id, result["rows"])
    return result

def _live_check_ebay_listing_status(business_id: str, ebay_item_id: str) -> str:
    """One live GetItem call for the TRUE current status, bypassing whatever
    the periodic full-listing sync last cached. Exists because that sync only
    ever pulls eBay's ActiveList -- it has no mechanism to notice a listing
    that was manually ended outside this app, so a row can sit indefinitely
    showing stale 'Active' with a stale quantity until something explicitly
    re-checks it. Self-heals the ebay_listing_status cache row when it
    disagrees, so this correction sticks instead of just being read once and
    forgotten."""
    import datetime as _dt
    token = get_ebay_access_token(business_id)
    try:
        resp = _ebay_get_item_status(token, ebay_item_id)
    except Exception as e:
        print(f"_live_check_ebay_listing_status: GetItem failed for {ebay_item_id}: {e}")
        return "Unknown"
    item = resp.get("Item") or {}
    selling_status = item.get("SellingStatus") or {}
    live_status = selling_status.get("ListingStatus") or "Unknown"
    if resp.get("Ack") not in ("Success", "Warning"):
        return "Unknown"
    try:
        cached = (supabase.table("ebay_listing_status").select("listing_status")
                  .eq("business_id", business_id).eq("item_id", ebay_item_id).execute()).data
        if cached and cached[0].get("listing_status") != live_status:
            quantity = _safe_int(item.get("Quantity")) or 0
            quantity_sold = _safe_int(selling_status.get("QuantitySold")) or 0
            supabase.table("ebay_listing_status").update({
                "listing_status": live_status,
                "quantity_available": max(quantity - quantity_sold, 0) if live_status == "Active" else 0,
                "updated_at": _dt.datetime.utcnow().isoformat(),
            }).eq("business_id", business_id).eq("item_id", ebay_item_id).execute()
    except Exception as e:
        print(f"_live_check_ebay_listing_status: cache correction failed for {ebay_item_id}: {e}")
    return live_status

def _check_new_shopify_sales_for_ebay_sync(business_id: str) -> dict:
    """The actual periodic check: Shopify sales since the last successful
    check (48h lookback on the very first run, then just the gap since last
    time), matched via confirmed HDID, upserted into ebay_sync_queue as
    'pending' rows. Live-verifies each NEWLY matched item's eBay status right
    here at discovery time (cheap -- only a handful of rows per run, not a
    full-catalog sweep) so a manually-ended listing never gets queued
    pretending it still has a live quantity to push. Never calls eBay for
    rows already sitting in the queue -- only brand-new discoveries. A
    variant already pending gets its qty_sold accumulated and discovered_at
    refreshed rather than duplicated, so selling again before someone gets
    to the push button doesn't create two rows for the same item."""
    import datetime as _dt

    settings = get_ebay_settings(business_id)
    last_checked = settings.get("EBAY_SYNC_QUEUE_LAST_CHECKED_AT")
    now = _dt.datetime.now(_dt.timezone.utc)
    if last_checked:
        start_dt = _dt.datetime.fromisoformat(last_checked.replace("Z", "+00:00"))
    else:
        start_dt = now - _dt.timedelta(hours=48)  # first-ever run: reasonable safe lookback

    start_iso = start_dt.strftime("%Y-%m-%dT%H:%M:%S-00:00")
    end_iso = now.strftime("%Y-%m-%dT%H:%M:%S-00:00")
    found = _find_shopify_sales_with_hdid_match(business_id, start_iso, end_iso)

    queued = 0
    for row in found["rows"]:
        existing = (supabase.table("ebay_sync_queue").select("id,qty_sold")
                    .eq("business_id", business_id).eq("shopify_variant_id", row["shopify_variant_id"])
                    .eq("status", "pending").execute()).data
        if existing:
            supabase.table("ebay_sync_queue").update({
                "qty_sold": existing[0]["qty_sold"] + row["qty_sold"],
                "discovered_at": now.isoformat(),
            }).eq("id", existing[0]["id"]).execute()
        else:
            live_status = _live_check_ebay_listing_status(business_id, row["ebay_item_id"])
            supabase.table("ebay_sync_queue").insert({
                "business_id": business_id, "shopify_variant_id": row["shopify_variant_id"],
                "ebay_item_id": row["ebay_item_id"], "title": row["title"], "sku": row["sku"],
                "qty_sold": row["qty_sold"], "status": "pending", "live_listing_status": live_status,
            }).execute()
        queued += 1

    save_ebay_setting(business_id, "EBAY_SYNC_QUEUE_LAST_CHECKED_AT", now.isoformat())
    return {"queued": queued, "unmatched_no_hdid": found["unmatched_no_hdid"], "checked_at": now.isoformat()}

async def ebay_sync_check_worker():
    """Automatic discovery, no button needed -- runs every 20 min. Pure
    read/match, zero eBay calls, zero writes anywhere except the local
    queue table. The corresponding WRITE action (pushing a quantity to
    eBay) stays manual-only via /api/ebay-sync/push -- explicit
    instruction: checking can be automatic, pushing can't be yet."""
    import asyncio
    while True:
        try:
            res = supabase.table("app_settings").select("business_id").eq("key", "EBAY_REFRESH_TOKEN").execute()
            business_ids = list(set(r["business_id"] for r in (res.data or [])))
            for biz_id in business_ids:
                try:
                    settings = get_ebay_settings(biz_id)
                    if not (settings.get("SHOPIFY_STORE_DOMAIN") or "").strip():
                        continue
                    result = await asyncio.to_thread(_check_new_shopify_sales_for_ebay_sync, biz_id)
                    if result.get("queued"):
                        print(f"ebay_sync_check_worker: business {biz_id}: {result}")
                except Exception as e:
                    print(f"ebay_sync_check_worker: business {biz_id} failed: {e}")
        except Exception as e:
            print(f"ebay_sync_check_worker error: {e}")
        await asyncio.sleep(1200)  # 20 min

@app.get("/api/ebay-sync/queue")
async def ebay_sync_queue(request: Request):
    """What the page actually loads on open -- the automatically-discovered,
    still-pending queue. No live Shopify or eBay call happens here at all;
    just a read of what the background worker has already found, plus
    current cached quantities so it's never a bare Push button with no
    visible number behind it."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    import asyncio
    rows = (supabase.table("ebay_sync_queue").select("*")
            .eq("business_id", business_id).eq("status", "pending")
            .order("discovered_at", desc=True).execute()).data or []
    rows = await asyncio.to_thread(_attach_current_quantities, business_id, rows)
    return {"rows": rows}

@app.post("/api/ebay-sync/queue/{row_id}/dismiss")
async def ebay_sync_queue_dismiss(row_id: str, request: Request):
    """Manually removes one row from the pending eBay-sync queue without
    calling eBay or Shopify at all. Exists because nothing else ever clears
    a 'pending' row except a successful push through this page -- a row
    resolved outside this flow (e.g. manually ended directly on eBay) would
    otherwise sit in the list forever. Scoped to status='pending' in the
    WHERE clause so this can't silently no-op-overwrite a row that got
    pushed in the meantime."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    res = (supabase.table("ebay_sync_queue").update({"status": "ignored"})
           .eq("id", row_id).eq("business_id", business_id).eq("status", "pending").execute())
    if not res.data:
        raise HTTPException(404, "Row not found, or it was already pushed/resolved")
    return {"ok": True}

@app.post("/api/ebay-sync/refresh-status")
async def ebay_sync_refresh_status(request: Request):
    """Live re-checks eBay's real status for every currently-pending queue
    row and corrects any that have drifted (e.g. manually ended on eBay
    outside this app). Manual/on-demand -- this is the one action in this
    section that DOES call eBay live, so it's tied to the explicit Refresh
    button rather than running on every page load."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    import asyncio
    rows = (supabase.table("ebay_sync_queue").select("id,ebay_item_id,live_listing_status")
            .eq("business_id", business_id).eq("status", "pending").execute()).data or []
    changed = 0
    for row in rows:
        live_status = await asyncio.to_thread(_live_check_ebay_listing_status, business_id, row["ebay_item_id"])
        if live_status != row.get("live_listing_status"):
            supabase.table("ebay_sync_queue").update({"live_listing_status": live_status}).eq("id", row["id"]).execute()
            changed += 1
    return {"checked": len(rows), "changed": changed}

class EbaySyncPush(BaseModel):
    shopify_variant_id: str

@app.post("/api/ebay-sync/push")
async def ebay_sync_push(body: EbaySyncPush, request: Request):
    """Pushes ONE eBay-only quantity update: sets eBay's available quantity to
    match Shopify's CURRENT live quantity for a confirmed-HDID pair. Shopify
    is the source of truth here and is never written to -- this is the
    mirror-image of push_inventory_quantity's eBay leg, minus the Shopify
    write, since Shopify already has the correct number (that's WHY this is
    running). Re-confirms the HDID match and re-checks both platforms live,
    immediately before writing -- never trusts what the table on screen
    showed, since time may have passed since it loaded."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    import asyncio

    match_res = (supabase.table("inventory_match").select("ebay_id,shopify_id,title,hd_id")
                 .eq("business_id", business_id).eq("shopify_id", body.shopify_variant_id).execute())
    match = (match_res.data or [None])[0]
    if not match or not match.get("hd_id") or not match.get("ebay_id"):
        raise HTTPException(400, "No confirmed HDID match for this Shopify variant — refusing to push.")
    ebay_item_id = match["ebay_id"]
    title = match.get("title") or ""

    log_row = {"business_id": business_id, "ebay_item_id": str(ebay_item_id),
               "shopify_variant_id": str(body.shopify_variant_id), "title": title,
               "requested_available_qty": 0}

    # Fresh live Shopify quantity — the source of truth this push is copying from.
    try:
        settings = get_ebay_settings(business_id)
        domain = (settings.get("SHOPIFY_STORE_DOMAIN", "") or "").strip().replace("https://", "").replace("http://", "").strip("/")
        if not domain:
            raise Exception("Shopify not connected")
        shopify_token = await asyncio.to_thread(get_shopify_access_token, business_id)
        sp_headers = {"X-Shopify-Access-Token": shopify_token, "Content-Type": "application/json"}
        location_id = await asyncio.to_thread(_get_shopify_primary_location_id, domain, sp_headers)
        if not location_id:
            raise Exception("Could not determine Shopify location")
        variant_data = await asyncio.to_thread(_fetch_shopify_variants_by_ids, domain, shopify_token, location_id, [str(body.shopify_variant_id)])
        shopify_qty = (variant_data.get(str(body.shopify_variant_id)) or {}).get("qty")
        if shopify_qty is None:
            raise Exception(f"Could not read live Shopify quantity for variant {body.shopify_variant_id}")
    except Exception as e:
        log_row.update({"ebay_status": "error", "ebay_error": f"Shopify live-qty check failed: {e}"})
        supabase.table("ebay_quantity_push_log").insert(log_row).execute()
        raise HTTPException(502, f"Could not read Shopify's live quantity: {e}")

    log_row["requested_available_qty"] = shopify_qty

    # Fresh eBay Active-check, same discipline as push_inventory_quantity.
    ebay_token = await asyncio.to_thread(get_ebay_access_token, business_id)
    try:
        status_resp = await asyncio.to_thread(_ebay_get_item_status, ebay_token, ebay_item_id)
    except Exception as e:
        log_row.update({"ebay_status": "error", "ebay_error": f"GetItem failed: {e}"})
        supabase.table("ebay_quantity_push_log").insert(log_row).execute()
        raise HTTPException(502, f"Could not check eBay's live listing status: {e}")
    item = status_resp.get("Item") or {}
    selling_status = item.get("SellingStatus") or {}
    listing_status = selling_status.get("ListingStatus")
    if status_resp.get("Ack") not in ("Success", "Warning") or listing_status != "Active":
        log_row.update({"ebay_status": "refused", "ebay_error": f"Listing not Active (status={listing_status})"})
        supabase.table("ebay_quantity_push_log").insert(log_row).execute()
        raise HTTPException(409, f"eBay listing {ebay_item_id} is not currently Active "
                                  f"(status: {listing_status or 'unknown'}) — refusing to push.")

    quantity_sold = _safe_int(selling_status.get("QuantitySold")) or 0
    new_total_qty = shopify_qty + quantity_sold
    log_row["live_quantity_sold"] = quantity_sold
    log_row["new_total_qty_sent"] = new_total_qty

    try:
        if shopify_qty <= 0:
            # 0 available can't go through ReviseItem -- see _ebay_end_item.
            # This is genuinely the case in the screenshot Alex reported:
            # sold out on Shopify means nothing physical is left, so the eBay
            # listing needs to end, not report an available quantity of 0.
            await asyncio.to_thread(_ebay_end_item, business_id, ebay_item_id, "NotAvailable")
            # Re-check live rather than assume a status string -- eBay uses
            # different ListingStatus values (e.g. "Completed") depending on
            # why an item ended, and this helper already self-heals
            # ebay_listing_status.listing_status + quantity_available itself.
            final_status = await asyncio.to_thread(_live_check_ebay_listing_status, business_id, ebay_item_id)
            log_row["ebay_status"] = f"success_ended ({final_status})"
        else:
            await asyncio.to_thread(_ebay_revise_quantity_only, business_id, ebay_item_id, new_total_qty)
            log_row["ebay_status"] = "success"
    except Exception as e:
        log_row.update({"ebay_status": "error", "ebay_error": str(e)})
        supabase.table("ebay_quantity_push_log").insert(log_row).execute()
        raise HTTPException(502, f"eBay quantity update failed: {e}")

    if shopify_qty > 0:
        # The ended-item branch above already wrote listing_status +
        # quantity_available via _live_check_ebay_listing_status.
        try:
            supabase.table("ebay_listing_status").update({
                "quantity": new_total_qty, "quantity_available": shopify_qty,
            }).eq("business_id", business_id).eq("item_id", ebay_item_id).execute()
        except Exception as e:
            print(f"ebay_sync_push: local ebay_listing_status update failed (eBay itself succeeded): {e}")

    try:
        import datetime as _dt
        supabase.table("ebay_sync_queue").update({
            "status": "pushed", "pushed_at": _dt.datetime.utcnow().isoformat(), "new_available_qty": shopify_qty,
        }).eq("business_id", business_id).eq("shopify_variant_id", body.shopify_variant_id).eq("status", "pending").execute()
    except Exception as e:
        print(f"ebay_sync_push: queue status update failed (eBay itself succeeded): {e}")

    supabase.table("ebay_quantity_push_log").insert(log_row).execute()
    return {"ok": True, "ebay_item_id": ebay_item_id, "new_available_qty": shopify_qty}

@app.post("/api/ebay-sync/push-to-shopify")
async def ebay_sync_push_to_shopify(body: EbaySyncPush, request: Request):
    """Reverse of /api/ebay-sync/push: sets SHOPIFY's quantity to match eBay's
    CURRENT live quantity for a confirmed-HDID pair. eBay is the source of
    truth here and is never written to. Exists for the case where eBay's
    number is trusted and Shopify's is the one that's stale/wrong -- the
    mirror-image direction of the main queue action, same live-recheck
    discipline (re-confirms HDID, re-reads eBay live right before writing,
    never trusts what the table on screen showed)."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    import asyncio

    match_res = (supabase.table("inventory_match").select("ebay_id,shopify_id,title,hd_id")
                 .eq("business_id", business_id).eq("shopify_id", body.shopify_variant_id).execute())
    match = (match_res.data or [None])[0]
    if not match or not match.get("hd_id") or not match.get("ebay_id"):
        raise HTTPException(400, "No confirmed HDID match for this Shopify variant — refusing to push.")
    ebay_item_id = match["ebay_id"]
    title = match.get("title") or ""

    log_row = {"business_id": business_id, "ebay_item_id": str(ebay_item_id),
               "shopify_variant_id": str(body.shopify_variant_id), "title": title,
               "requested_available_qty": 0, "ebay_status": "not_touched"}

    # Fresh eBay live check -- the source of truth this push is copying from.
    ebay_token = await asyncio.to_thread(get_ebay_access_token, business_id)
    try:
        status_resp = await asyncio.to_thread(_ebay_get_item_status, ebay_token, ebay_item_id)
    except Exception as e:
        log_row.update({"shopify_status": "error", "shopify_error": f"eBay GetItem failed: {e}"})
        supabase.table("ebay_quantity_push_log").insert(log_row).execute()
        raise HTTPException(502, f"Could not check eBay's live listing status: {e}")
    item = status_resp.get("Item") or {}
    selling_status = item.get("SellingStatus") or {}
    listing_status = selling_status.get("ListingStatus")
    if status_resp.get("Ack") not in ("Success", "Warning") or listing_status != "Active":
        log_row.update({"shopify_status": "refused", "shopify_error": f"Listing not Active (status={listing_status})"})
        supabase.table("ebay_quantity_push_log").insert(log_row).execute()
        raise HTTPException(409, f"eBay listing {ebay_item_id} is not currently Active "
                                  f"(status: {listing_status or 'unknown'}) — refusing to push.")

    quantity = _safe_int(item.get("Quantity")) or 0
    quantity_sold = _safe_int(selling_status.get("QuantitySold")) or 0
    ebay_available = max(quantity - quantity_sold, 0)
    log_row["requested_available_qty"] = ebay_available
    log_row["live_quantity_sold"] = quantity_sold

    try:
        settings = get_ebay_settings(business_id)
        domain = (settings.get("SHOPIFY_STORE_DOMAIN", "") or "").strip().replace("https://", "").replace("http://", "").strip("/")
        if not domain:
            raise Exception("Shopify not connected")
        shopify_token = await asyncio.to_thread(get_shopify_access_token, business_id)
        sp_headers = {"X-Shopify-Access-Token": shopify_token, "Content-Type": "application/json"}
        location_id = await asyncio.to_thread(_get_shopify_primary_location_id, domain, sp_headers)
        if not location_id:
            raise Exception("Could not determine Shopify location")
        variant_data = await asyncio.to_thread(_fetch_shopify_variants_by_ids, domain, shopify_token, location_id, [str(body.shopify_variant_id)])
        inv_item_id = (variant_data.get(str(body.shopify_variant_id)) or {}).get("inventory_item_id")
        if not inv_item_id:
            raise Exception(f"Could not resolve Shopify variant {body.shopify_variant_id} to a live inventory item")
        import requests as _req
        mutation = {
            "query": """mutation setQty($input: InventorySetQuantitiesInput!) {
                inventorySetQuantities(input: $input) {
                    inventoryAdjustmentGroup { changes { name delta quantityAfterChange } }
                    userErrors { field message }
                }
            }""",
            "variables": {"input": {
                "reason": "correction", "name": "available", "ignoreCompareQuantity": True,
                "quantities": [{"inventoryItemId": inv_item_id, "locationId": location_id, "quantity": ebay_available}],
            }},
        }
        r = await asyncio.to_thread(_req.post, f"https://{domain}/admin/api/2024-10/graphql.json", headers=sp_headers, json=mutation, timeout=20)
        resp = r.json() if r.status_code == 200 else {}
        gql_errors = resp.get("errors") or []
        user_errors = (resp.get("data", {}) or {}).get("inventorySetQuantities", {}).get("userErrors", []) if not gql_errors else gql_errors
        if r.status_code != 200 or gql_errors or user_errors:
            raise Exception("; ".join(e.get("message", "") for e in user_errors) or f"HTTP {r.status_code}")
        log_row["shopify_status"] = "success"
    except Exception as e:
        log_row.update({"shopify_status": "error", "shopify_error": str(e)})
        supabase.table("ebay_quantity_push_log").insert(log_row).execute()
        raise HTTPException(502, f"Shopify quantity update failed: {e}")

    try:
        supabase.table("shopify_inventory").update({"quantity": ebay_available}).eq("business_id", business_id).eq("id", str(body.shopify_variant_id)).execute()
    except Exception as e:
        print(f"ebay_sync_push_to_shopify: local shopify_inventory update failed (Shopify itself succeeded): {e}")

    try:
        import datetime as _dt
        supabase.table("ebay_sync_queue").update({
            "status": "pushed_to_shopify", "pushed_at": _dt.datetime.utcnow().isoformat(), "new_available_qty": ebay_available,
        }).eq("business_id", business_id).eq("shopify_variant_id", body.shopify_variant_id).eq("status", "pending").execute()
    except Exception as e:
        print(f"ebay_sync_push_to_shopify: queue status update failed (Shopify itself succeeded): {e}")

    supabase.table("ebay_quantity_push_log").insert(log_row).execute()
    return {"ok": True, "shopify_variant_id": body.shopify_variant_id, "new_available_qty": ebay_available}

async def shopify_shipping_rule_worker():
    """Fully independent, fully automatic -- no button, and by DESIGN never
    calls eBay at all (explicit instruction: this must never be bundled with
    or trigger anything that touches eBay). Reads whatever shipping_cost is
    ALREADY sitting in ebay_listing_status (kept fresh by the separate eBay
    sync workers) and pushes any $69/$195 item to its matched Shopify product.
    Runs on its own clock every 15 minutes -- catches new matches, price-tier
    changes, etc. without waiting on or depending on an eBay sync ever running."""
    import asyncio
    while True:
        try:
            res = supabase.table("app_settings").select("business_id").eq("key", "EBAY_REFRESH_TOKEN").execute()
            business_ids = list(set(r["business_id"] for r in (res.data or [])))
            for biz_id in business_ids:
                try:
                    result = await asyncio.to_thread(apply_shopify_shipping_rule, biz_id)
                    if result.get("pushed"):
                        print(f"shopify_shipping_rule_worker: pushed {result['pushed']} item(s) for business {biz_id}")
                except Exception as e:
                    print(f"shopify_shipping_rule_worker: business {biz_id} failed: {e}")
        except Exception as e:
            print(f"shopify_shipping_rule_worker error: {e}")
        await asyncio.sleep(900)  # 15 min

async def order_sync_worker():
    """Keeps the local `orders` table fresh automatically, so Financials never has to
    hit eBay/Shopify live. The FIRST time a business is seen, does a full historical
    backfill (365 days) so the table starts complete, not just from whenever this
    feature was turned on. After that, syncs a rolling 14-day window every 20 minutes
    (catches late-arriving fee/refund/tracking data, not just brand-new orders)."""
    import asyncio
    while True:
        try:
            res = supabase.table("app_settings").select("business_id").eq("key", "EBAY_REFRESH_TOKEN").execute()
            business_ids = list(set(r["business_id"] for r in (res.data or [])))
            for biz_id in business_ids:
                try:
                    settings = get_ebay_settings(biz_id)
                    backfilled = settings.get("ORDERS_BACKFILLED", "") == "true"
                    days_back = 14 if backfilled else 720  # ~2 years — eBay's getOrders hard-caps creationdate at 2 years, stricter than Finance API's 5-year limit
                    result = await asyncio.to_thread(sync_orders_for_business, biz_id, days_back)
                    had_errors = bool(result.get("errors"))
                    if result.get("upserted"):
                        # New/updated order rows landed -- Total_Payouts/Profit/ROI on the
                        # Lots page depend on this data but had no automatic trigger of
                        # their own before this; every OTHER path that touches acquisitions
                        # (edits, cash payments, CSV import) already auto-recalculates,
                        # this was the one gap, which is exactly why the manual Recalculate
                        # button was still load-bearing. Isolated in its own try/except so a
                        # recalc failure can never mark the order sync itself as failed.
                        try:
                            # Attribute orphaned multi-box labels + refresh shipping costs
                            # BEFORE recalculating lot profits, so profits always reflect
                            # full shipping. No button pressing required — this is the
                            # same automatic 20-minute cycle as the order sync itself.
                            apply_shipping_matches(biz_id)
                        except Exception as e:
                            print(f"order_sync_worker: shipping match failed for business {biz_id} (order sync unaffected): {e}")
                        try:
                            apply_acquisition_profits(biz_id)
                        except Exception as e:
                            print(f"order_sync_worker: recalculate failed for business {biz_id} (order sync unaffected): {e}")
                    if not backfilled:
                        if had_errors and result["upserted"] == 0:
                            # Don't mark complete — the eBay pull itself failed, so we got
                            # little/nothing. Retry the full backfill again next cycle instead
                            # of silently settling for an incomplete result.
                            print(f"order_sync_worker: business {biz_id} initial backfill FAILED "
                                  f"(will retry next cycle) — errors: {result['errors']}")
                        else:
                            save_ebay_setting(biz_id, "ORDERS_BACKFILLED", "true")
                            print(f"order_sync_worker: business {biz_id} completed initial backfill "
                                  f"({result['upserted']} order line(s))"
                                  + (f", errors: {result['errors']}" if result.get("errors") else ""))
                    else:
                        print(f"order_sync_worker: business {biz_id} synced {result['upserted']} order line(s)"
                              + (f", errors: {result['errors']}" if result.get("errors") else ""))
                except Exception as e:
                    print(f"order_sync_worker: business {biz_id} failed: {e}")
        except Exception as e:
            print(f"order_sync_worker error: {e}")
        await asyncio.sleep(1200)  # 20 minutes

async def analytics_snapshot_worker():
    """Runs once a day, at a fixed time (00:10 UTC), for EVERY business -- unlike
    shopify_sync_worker above, this has nothing to do with Shopify; it loops over
    the businesses table directly, not app_settings' Shopify-connected subset,
    since Inventory Snapshot Value already accounts for eBay-only, Shopify-only,
    or both. Only reusing the same 'sleep until next time boundary' scheduling
    shape as shopify_sync_worker, not its logic.

    Exists because Inventory Snapshot Value has no historical source at all --
    every other Analytics metric can be recomputed for a past date directly from
    orders/listings' own timestamps, but inventory value is a live aggregate with
    no record of what it used to be. This is what actually builds that history,
    one row per business per day, from here on."""
    import asyncio, datetime as _dt
    while True:
        now = _dt.datetime.now(_dt.timezone.utc)
        next_run = now.replace(hour=0, minute=10, second=0, microsecond=0)
        if next_run <= now:
            next_run += _dt.timedelta(days=1)
        await asyncio.sleep(max((next_run - now).total_seconds(), 5))
        try:
            biz_res = supabase.table("businesses").select("id").execute()
            business_ids = [b["id"] for b in (biz_res.data or [])]
            for biz_id in business_ids:
                try:
                    result = await asyncio.to_thread(run_daily_analytics_snapshot_for_business, biz_id)
                    print(f"analytics_snapshot_worker: business {biz_id} -> {result}")
                except Exception as e:
                    print(f"analytics_snapshot_worker: business {biz_id} failed: {e}")
        except Exception as e:
            print(f"analytics_snapshot_worker error: {e}")

async def analytics_cache_refresh_worker():
    """Runs every 2 hours, for every business, proactively recomputing the
    Analytics page's own DEFAULT view (last 30 days for /api/analytics,
    2025-01 through the current month for monthly-trend) and saving it into
    analytics_query_cache -- so opening the page with the default filters
    (the overwhelming majority of page loads) hits a warm cache instead of
    recomputing everything live, every single time. A genuinely different
    date range (a custom filter change) still computes live on that first
    request and gets cached from there, same read-through pattern -- this
    worker just makes sure the COMMON case never has to pay that cost."""
    import asyncio, datetime as _dt
    while True:
        await asyncio.sleep(2 * 60 * 60)  # every 2 hours
        try:
            biz_res = supabase.table("businesses").select("id").execute()
            business_ids = [b["id"] for b in (biz_res.data or [])]
            for biz_id in business_ids:
                try:
                    now = _dt.datetime.utcnow()
                    a_end = now.strftime("%Y-%m-%d")
                    a_start = (now - _dt.timedelta(days=30)).strftime("%Y-%m-%d")
                    a_payload = await asyncio.to_thread(_compute_analytics_payload, biz_id, a_start, a_end)
                    _set_analytics_cache(biz_id, "analytics", a_start, a_end, a_payload)

                    mt_start = "2025-01-01"
                    mt_end = now.strftime("%Y-%m-01")
                    mt_payload = await asyncio.to_thread(_compute_monthly_trend_payload, biz_id, mt_start, mt_end)
                    _set_analytics_cache(biz_id, "monthly_trend", mt_start, mt_end, mt_payload)

                    print(f"analytics_cache_refresh_worker: refreshed business {biz_id}")
                except Exception as e:
                    print(f"analytics_cache_refresh_worker: business {biz_id} failed: {e}")
        except Exception as e:
            print(f"analytics_cache_refresh_worker error: {e}")

def run_daily_analytics_snapshot_for_business(business_id: str) -> dict:
    """One row per business per day in analytics_snapshots -- inventory value
    (today's live figure) plus that single day's sale price/new-listings numbers,
    so a future trend view has real daily granularity instead of just 'now'.

    Also computes YTD net revenue and YTD inventory spend as of today, for Net
    Cash Yield -- simplified per explicit request to drop cash entirely from
    this one (just post-fee orders minus inventory spend) and precomputed here
    instead of recalculated live on every page load, since summing an entire
    year's rows on every request was the actual slowdown."""
    import datetime as _dt
    from zoneinfo import ZoneInfo

    mt_today = _dt.datetime.now(ZoneInfo("America/Denver")).strftime("%Y-%m-%d")
    year_start = f"{mt_today[:4]}-01-01"

    inventory_value = _compute_inventory_snapshot_value(business_id)
    green_revenue = _compute_green_revenue(business_id, mt_today, mt_today)

    order_rows = _fetch_all_for_business(business_id, "orders", "gross_revenue,quantity,order_date")
    order_rows = [r for r in order_rows if r.get("order_date") == mt_today]
    total_revenue = sum(r.get("gross_revenue") or 0 for r in order_rows)
    total_qty_sold = sum(r.get("quantity") or 0 for r in order_rows)
    avg_sale_price_day = round(total_revenue / total_qty_sold, 2) if total_qty_sold else 0

    listing_rows = _fetch_all_for_business(business_id, "listings", "price,created_at")
    listing_rows = [r for r in listing_rows if mt_today <= (r.get("created_at") or "") <= mt_today + "T23:59:59"]
    new_listings_count_day = len(listing_rows)
    new_listings_value_day = round(sum(r.get("price") or 0 for r in listing_rows), 2)

    # FIXED a real bug here: these two were plain unpaginated .execute() calls,
    # which Supabase caps around 1000 rows -- silently undercounting if there
    # are more orders/acquisitions YTD than that (confirmed: the app showed
    # $162,967 revenue while a proper full query returned $479,667.45). Every
    # other query in this codebase already uses _fetch_all_for_business for
    # exactly this reason; these two just weren't written that way originally.
    ytd_order_rows = _fetch_all_for_business(business_id, "orders", "final_net,order_date")
    ytd_order_rows = [r for r in ytd_order_rows if year_start <= (r.get("order_date") or "") <= mt_today]
    ytd_net_revenue = round(sum(r.get("final_net") or 0 for r in ytd_order_rows), 2)

    ytd_acq_rows = _fetch_all_for_business(business_id, "acquisitions", "cost,date")
    ytd_acq_rows = [r for r in ytd_acq_rows if year_start <= (r.get("date") or "") <= mt_today]
    ytd_inventory_spend = round(sum(r.get("cost") or 0 for r in ytd_acq_rows), 2)

    ytd_cash = _compute_cash_for_year(business_id, int(mt_today[:4]))
    ytd_net_cash_yield = round(ytd_net_revenue + ytd_cash - ytd_inventory_spend, 2)

    # YTD Inventory Appreciation: today's inventory value minus whatever it was
    # on/near Jan 1 -- reuses the same nearest-snapshot lookup Inventory Growth
    # Multiple already relies on, since it's the same underlying question
    # (what did inventory look like at the start of the year). Net Business
    # Appreciation = Net Cash Yield + Inventory Appreciation: cash generated
    # AND the change in what's still sitting in inventory, over the same
    # period -- the fuller picture of whether the business grew this year.
    jan1_value, jan1_date = _nearest_inventory_snapshot(business_id, year_start)
    ytd_inventory_appreciation = round(inventory_value - jan1_value, 2) if jan1_value is not None else None
    ytd_net_business_appreciation = (
        round(ytd_net_cash_yield + ytd_inventory_appreciation, 2)
        if ytd_inventory_appreciation is not None else None
    )

    record = {
        "business_id": business_id,
        "snapshot_date": mt_today,
        "inventory_snapshot_value": inventory_value,
        "green_revenue": green_revenue,
        "avg_sale_price_day": avg_sale_price_day,
        "new_listings_count_day": new_listings_count_day,
        "new_listings_value_day": new_listings_value_day,
        "ytd_net_revenue": ytd_net_revenue,
        "ytd_inventory_spend": ytd_inventory_spend,
        "ytd_cash": ytd_cash,
        "ytd_net_cash_yield": ytd_net_cash_yield,
        "ytd_inventory_appreciation": ytd_inventory_appreciation,
        "ytd_net_business_appreciation": ytd_net_business_appreciation,
        # Count of active eBay listings per flat-shipping tier as of today —
        # requested so shipping composition is tracked historically, not just
        # queryable "now". Keys are dollar amounts as strings + 'unknown' for
        # listings eBay returns no flat cost for (calculated/freight).
        "active_shipping_breakdown": (lambda rows_: (lambda d: d)(
            __import__('collections').Counter(
                ("unknown" if r.get("shipping_cost") is None else str(int(r["shipping_cost"]) if float(r["shipping_cost"]).is_integer() else r["shipping_cost"]))
                for r in rows_ if r.get("listing_status") == "Active"
            )))(_fetch_all_for_business(business_id, "ebay_listing_status", "listing_status,shipping_cost")),
    }
    existing = supabase.table("analytics_snapshots").select("id").eq("business_id", business_id)\
        .eq("snapshot_date", mt_today).limit(1).execute()
    if existing.data:
        supabase.table("analytics_snapshots").update(record).eq("id", existing.data[0]["id"]).execute()
    else:
        supabase.table("analytics_snapshots").insert(record).execute()
    return record

EBAY_DESCRIPTION = "Shipped primarily with UPS and sometimes USPS. If you have special packing or shipping needs, please send a message. This item is sold in as-is condition. The seller assumes no liability for the use, operation, or installation of this product. Due to the technical nature of this equipment, the buyer is responsible for having the item professionally inspected and installed by a certified technician prior to use."

def photo_url(photo_id: str, thumb: bool = False) -> str:
    if not photo_id or photo_id in ("", "nan", "0"):
        return ""
    if thumb:
        return f"{SUPABASE_URL}/storage/v1/render/image/public/part-photos/{photo_id}?width=500&height=500&resize=cover&quality=80"
    return f"{SUPABASE_URL}/storage/v1/object/public/part-photos/{photo_id}"

def _photo_seq(photo_id: str):
    """Photos written by /api/photos/upload carry their original capture/
    selection order as a trailing _<idx> before the extension (e.g.
    030826_143022_3.jpg — see the batch_ts fix in that endpoint, and the
    matching helper in mantle-scanner's scanner_service.py which renames
    scanned photos while preserving this same suffix). Returns None for
    anything that doesn't match (legacy single-photo scans, _extra add-on
    photos), which callers should sort to the end rather than error on."""
    m = re.search(r'_(\d+)\.[^.]+$', str(photo_id or ""))
    return int(m.group(1)) if m else None

def get_all_photo_ids(primary_photo_id: str) -> list:
    """A listing only stores its primary photo_id, but scans capture multiple photos
    per group (group_photos table). Marketplace listings should use ALL of them,
    not just the primary — this looks up the rest via the shared group_id.

    Always returns primary_photo_id first. group_photos has no defined order, so
    without this, "Make Main Photo" (which only ever updates the listing's
    photo_id) had no actual effect on eBay/Shopify submissions — the photo array
    sent to both marketplaces came back in whatever order Supabase happened to
    return group_photos, ignoring which one was chosen as primary.

    The remaining (non-primary) photos are further sorted by _photo_seq so they
    match original capture order too, instead of Supabase's arbitrary row order
    — real bug this fixes: that arbitrary order was a second, separate source of
    out-of-order images beyond which photo became primary."""
    pid = str(primary_photo_id or "")
    if not pid:
        return []
    try:
        group_row = supabase.table("group_photos").select("group_id").eq("photo_id", pid).limit(1).execute()
        group_id = (group_row.data or [{}])[0].get("group_id", "")
        if not group_id:
            return [pid]
        gp = supabase.table("group_photos").select("photo_id").eq("group_id", group_id).execute()
        all_pids = [r["photo_id"] for r in (gp.data or []) if r.get("photo_id")]
        if not all_pids:
            return [pid]
        rest = [p for p in all_pids if p != pid]
        rest.sort(key=lambda p: (0, _photo_seq(p)) if _photo_seq(p) is not None else (1, 0))
        return [pid] + rest
    except Exception:
        return [pid]

# ── EBAY INVENTORY API ───────────────────────────────────────── #
EBAY_API_BASE = "https://api.ebay.com"

# Alternate shipping policies a listing can be assigned instead of the account default.
EBAY_SHIPPING_POLICY_OPTIONS = [
    {"id": "251042655020", "label": "Free Shipping"},
    {"id": "251441449020", "label": "$18 Shipping"},
    {"id": "251094371020", "label": "$69 Shipping"},
    {"id": "251094542020", "label": "$195 Shipping"},
]
def _parse_ship_label_cost(label: str):
    label = (label or "").strip()
    if label.lower().startswith("free"):
        return 0
    import re as _re
    m = _re.search(r'\$(\d+(?:\.\d+)?)', label)
    return float(m.group(1)) if m else None
_EBAY_SHIP_POLICY_COST_BY_ID = {
    opt["id"]: _parse_ship_label_cost(opt["label"]) for opt in EBAY_SHIPPING_POLICY_OPTIONS
}

EBAY_ENV_KEYS = [
    "EBAY_USER_TOKEN", "EBAY_APP_ID", "EBAY_DEV_ID", "EBAY_CERT_ID", "EBAY_RUNAME",
    "EBAY_PAYMENT_POLICY_ID", "EBAY_RETURN_POLICY_ID", "EBAY_FULFILLMENT_POLICY_ID",
    "EBAY_MERCHANT_LOCATION_KEY", "EBAY_LOCATION_ZIP", "EBAY_LOCATION_COUNTRY",
    "EBAY_LOCATION_CITY_STATE", "EBAY_DEFAULT_CATEGORY_ID",
    "EBAY_DEFAULT_MOTORS_CATEGORY_ID",
]

EBAY_OAUTH_SCOPES = (
    "https://api.ebay.com/oauth/api_scope "
    "https://api.ebay.com/oauth/api_scope/sell.inventory "
    "https://api.ebay.com/oauth/api_scope/sell.fulfillment "
    "https://api.ebay.com/oauth/api_scope/sell.finances "
    "https://api.ebay.com/oauth/api_scope/sell.analytics.readonly"
)

def _lot_prefix(sku: str) -> str:
    """Transposes a SKU to its lot: 'RJ-123' belongs to lot 'RJ' (prefix before the
    first '-'); a bare SKU with no '-' at all belongs to that same lot directly.
    Shared across acquisitions, financials, and Shopify-profit code so the rule
    can't drift between them."""
    return sku.split("-", 1)[0] if "-" in sku else sku

def _require_valid_lot_sku_for_publish(business_id: str, listing: dict):
    """Server-side twin of the frontend's validateLotSkusForPublish/lotSkuFor check —
    that one only lives in JavaScript, so it protects the normal buttons but nothing
    that might ever call these publish endpoints directly. Raises 400 with the same
    'must match a real lot' message unless the listing's ebay_sku prefix matches an
    actual row on the Lots page. Called by all three publish endpoints (eBay v1,
    eBay v2, Shopify) so a bad SKU can never reach a live platform listing no matter
    which path is used to get there.

    FIXED: was ONLY ever checking listings.ebay_sku -- but a listing matched via the
    Inventory tab (see listing_by_ebay_id in list_inventory) can have a real,
    perfectly valid, lot-tagged SKU sitting on the live eBay listing itself
    (ebay_listing_status.sku, e.g. 'RB66-US') while its own local listings row never
    had ebay_sku filled in at all. That's not a missing SKU, it's the wrong column
    being checked. Now falls back to the real live eBay SKU via ebay_item_id before
    concluding there's genuinely no SKU anywhere."""
    sku = (listing.get("ebay_sku") or "").strip()
    if not sku and listing.get("ebay_item_id"):
        live = (supabase.table("ebay_listing_status").select("sku")
                .eq("business_id", business_id).eq("item_id", listing["ebay_item_id"])
                .limit(1).execute().data or [])
        if live and live[0].get("sku"):
            sku = live[0]["sku"].strip()
    if not sku:
        raise HTTPException(400, "This item has no Lot SKU set — assign one from the Lots page before publishing.")
    prefix = _lot_prefix(sku)
    known_lot_skus = {a["sku"] for a in (supabase.table("acquisitions").select("sku")
                       .eq("business_id", business_id).execute().data or []) if a.get("sku")}
    if prefix not in known_lot_skus:
        raise HTTPException(400, f"SKU '{sku}' doesn't match any real lot on the Lots page — "
                                  f"fix it there or on this item's Lot SKU field before publishing.")

def _sku_needs_assignment(sku: str) -> bool:
    """True for a blank SKU or a bare 'PREFIX-' placeholder (nothing after the dash)
    — the v2 eBay push path allows duplicate bare SKUs like 'AM1-' on purpose (see
    /assign-lot), so these haven't actually been assigned to a specific item yet,
    even though their prefix matches a real lot."""
    return (not sku) or sku.endswith("-")

def _validate_sku_for_assignment(business_id: str, sku: str):
    """Per direct instruction for the Uncat page's SKU tools (single-row Save AND
    the bulk multi-select assign): a SKU being SAVED here must be a real, complete
    assignment, not just a step toward one. Two rules, both hard failures:
      1. Can't be blank or a bare 'PREFIX-' with nothing after the dash -- that's
         still effectively uncategorized, just wearing a SKU-shaped string.
      2. The prefix must match an actual row on the Lots page -- same rule
         _require_valid_lot_sku_for_publish already enforces at publish time,
         applied here too so a bad SKU can't even get saved in the first place,
         not just get caught later when publishing."""
    sku = (sku or "").strip()
    if _sku_needs_assignment(sku):
        raise HTTPException(400, f"'{sku or '(blank)'}' isn't a complete SKU — it still needs the part after the dash.")
    prefix = _lot_prefix(sku)
    known_lot_skus = {a["sku"] for a in (supabase.table("acquisitions").select("sku")
                       .eq("business_id", business_id).execute().data or []) if a.get("sku")}
    if prefix not in known_lot_skus:
        raise HTTPException(400, f"'{prefix}' doesn't match any real lot on the Lots page.")

def _is_blank_sku(s):
    """True for blank, '(no SKU)', or the 'lister-{id}' fallback — the set of values
    that count as 'not a real SKU yet' for order-sync SKU-preservation purposes.
    Shared so eBay and Shopify order sync can't drift on what counts as blank."""
    s = (s or "").strip().lower()
    return s in ("", "(no sku)") or s.startswith("lister-")

def get_ebay_settings(business_id: str) -> dict:
    res = supabase.table("app_settings").select("*").eq("business_id", business_id).execute()
    settings = {row["key"]: row["value"] for row in (res.data or [])}
    return settings

def save_ebay_setting(business_id: str, key: str, value: str):
    existing = supabase.table("app_settings").select("key").eq("business_id", business_id).eq("key", key).limit(1).execute()
    if existing.data:
        supabase.table("app_settings").update({"value": value}).eq("business_id", business_id).eq("key", key).execute()
    else:
        supabase.table("app_settings").insert({"business_id": business_id, "key": key, "value": value}).execute()

def get_ebay_access_token(business_id: str) -> str:
    """Returns a valid eBay access token, transparently refreshing it via the stored
    18-month refresh token if the cached access token is missing or expired.
    Raises a clear error if the account has never completed OAuth."""
    import requests as _req, time, base64 as _b64

    settings = get_ebay_settings(business_id)
    refresh_token = settings.get("EBAY_REFRESH_TOKEN", "")
    access_token = settings.get("EBAY_USER_TOKEN", "")
    expires_at = float(settings.get("EBAY_TOKEN_EXPIRES_AT", "0") or 0)

    if not refresh_token:
        raise Exception("eBay account not connected — go to Settings and click 'Connect eBay Account'")

    # 60s safety margin before expiry
    if access_token and time.time() < (expires_at - 60):
        return access_token

    app_id = settings.get("EBAY_APP_ID", "")
    cert_id = settings.get("EBAY_CERT_ID", "")
    if not app_id or not cert_id:
        raise Exception("Missing EBAY_APP_ID / EBAY_CERT_ID in Settings")

    basic = _b64.b64encode(f"{app_id}:{cert_id}".encode()).decode()
    r = _req.post(
        "https://api.ebay.com/identity/v1/oauth2/token",
        headers={"Authorization": f"Basic {basic}", "Content-Type": "application/x-www-form-urlencoded"},
        data={
            "grant_type": "refresh_token",
            "refresh_token": refresh_token,
            "scope": EBAY_OAUTH_SCOPES,
        },
        timeout=15,
    )
    if r.status_code != 200:
        raise Exception(f"eBay token refresh failed ({r.status_code}): {r.text[:300]} — try reconnecting your eBay account in Settings")

    data = r.json()
    new_token = data["access_token"]
    new_expires_at = time.time() + int(data.get("expires_in", 7200))
    save_ebay_setting(business_id, "EBAY_USER_TOKEN", new_token)
    save_ebay_setting(business_id, "EBAY_TOKEN_EXPIRES_AT", str(new_expires_at))
    return new_token

def ebay_headers(token: str, content_language: bool = True) -> dict:
    h = {"Authorization": f"Bearer {token}", "Accept": "application/json"}
    if content_language:
        h["Content-Type"] = "application/json"
        h["Content-Language"] = "en-US"
    return h

def ensure_ebay_location(token: str, location_key: str, zip_code: str, country: str):
    """Create the merchant inventory location if it doesn't already exist. Required before publishing."""
    import requests as _req
    check = _req.get(f"{EBAY_API_BASE}/sell/inventory/v1/location/{location_key}",
                      headers=ebay_headers(token, content_language=False), timeout=15)
    if check.status_code == 200:
        return
    body = {
        "location": {"address": {"postalCode": zip_code, "country": country or "US"}},
        "locationTypes": ["WAREHOUSE"],
        "name": "Primary Location",
    }
    r = _req.post(f"{EBAY_API_BASE}/sell/inventory/v1/location/{location_key}",
                   headers=ebay_headers(token), json=body, timeout=15)
    if r.status_code not in (200, 201, 204):
        raise Exception(f"Failed to create eBay inventory location: {r.status_code} {r.text}")

def get_ebay_item_location_text(token: str, location_key: str) -> str:
    """Unused for now — v2's Item.Location field turned out to be plain free text
    (see EBAY_LOCATION_CITY_STATE setting), not something stored on the v1 merchant
    location resource. Left here in case a future path needs to read that resource."""
    import requests as _req
    if not location_key:
        return ""
    r = _req.get(f"{EBAY_API_BASE}/sell/inventory/v1/location/{location_key}",
                 headers=ebay_headers(token, content_language=False), timeout=15)
    if r.status_code != 200:
        return ""
    addr = (r.json().get("location") or {}).get("address") or {}
    city = addr.get("city") or ""
    state = addr.get("stateOrProvince") or ""
    return ", ".join(p for p in (city, state) if p)

def ebay_condition(cond: str) -> str:
    return "NEW" if (cond or "").lower() == "new" else "USED_EXCELLENT"

def push_listing_to_ebay(listing: dict, mode: str, hours_from_now: float = None, brand_override: str = None, mpn_override: str = None) -> dict:
    """
    mode: 'draft' | 'now' | 'schedule'
    Returns dict with offer_id, item_id (if published), status, scheduled_at
    """
    import requests as _req
    from datetime import timedelta

    biz_id = listing.get("business_id")
    if not biz_id:
        raise Exception("Listing has no business_id — cannot look up eBay settings safely")
    settings = get_ebay_settings(biz_id)
    token = get_ebay_access_token(biz_id)

    payment_policy    = settings.get("EBAY_PAYMENT_POLICY_ID", "")
    return_policy      = settings.get("EBAY_RETURN_POLICY_ID", "")
    fulfillment_policy = listing.get("ebay_fulfillment_policy_id") or settings.get("EBAY_FULFILLMENT_POLICY_ID", "")
    location_key       = settings.get("EBAY_MERCHANT_LOCATION_KEY", "")
    location_zip       = settings.get("EBAY_LOCATION_ZIP", "")
    location_country   = settings.get("EBAY_LOCATION_COUNTRY", "US")
    # A category_id of "0" (string) is a real, recognized "not yet categorized"
    # sentinel elsewhere in this codebase (see the .or_(...eq.0...) filter used
    # to find uncategorized items) -- but Python truthiness doesn't catch it,
    # since a non-empty string like "0" is always truthy. Without this explicit
    # check, "0" would sail straight past both the fallback AND the "is it set"
    # guard below, and get submitted to eBay as a literal, invalid category ID
    # -- confirmed as the actual cause of a real "Category is not valid" failure.
    raw_category_id = listing.get("ebay_category_id")
    category_id = raw_category_id if raw_category_id and str(raw_category_id) != "0" else settings.get("EBAY_DEFAULT_CATEGORY_ID", "")

    if not (payment_policy and return_policy and fulfillment_policy and location_key):
        raise Exception("Missing eBay business policy IDs or location key — set these in Settings first")
    if not category_id or str(category_id) == "0":
        raise Exception("This item has no eBay category set")

    sku = listing.get("ebay_sku") or f"lister-{listing['id']}"
    title = (listing.get("title") or "Untitled item")[:80]
    desc  = listing.get("description") or EBAY_DESCRIPTION
    qty   = int(listing.get("quantity") or 1)
    price = float(listing.get("price") or 0)
    pid   = str(listing.get("photo_id") or "")
    images = [photo_url(p) for p in get_all_photo_ids(pid) if photo_url(p)] if pid else []

    # 1. Create/replace inventory item
    brand = (brand_override or listing.get("brand") or "").strip()
    if not brand:
        first_word = title.split()[0].strip(",.;:-") if title else ""
        brand = first_word if first_word else "Unbranded"

    # Guess MPN: the longest alphanumeric (letters+digits) token in the title that isn't the brand —
    # usually the true part number rather than a shorter model class label.
    # eBay requires SOME MPN value in many categories — if we can't guess one, send their
    # standard "Does Not Apply" placeholder rather than omitting the aspect entirely, since
    # a missing BrandMPN aspect makes publishOffer fail outright in those categories.
    mpn = (mpn_override or "").strip() or None
    mpn_is_fallback = False
    if not mpn:
        alnum_tokens = [
            w.strip(",.;:-") for w in title.split()
            if w.lower().strip(",.;:-") != brand.lower()
            and any(c.isdigit() for c in w) and any(c.isalpha() for c in w)
        ]
        if alnum_tokens:
            mpn = max(alnum_tokens, key=len)
    if not mpn:
        mpn = "Does Not Apply"
        mpn_is_fallback = True

    product_data = {
        "title": title,
        "description": desc,
        "imageUrls": images,
        "aspects": {"Brand": [brand], "MPN": [mpn]},
        "brand": brand,
        "mpn": mpn,
    }

    inv_body = {
        "condition": ebay_condition(listing.get("condition")),
        "product": product_data,
        "availability": {"shipToLocationAvailability": {"quantity": qty}},
    }
    r = _req.put(f"{EBAY_API_BASE}/sell/inventory/v1/inventory_item/{sku}",
                 headers=ebay_headers(token), json=inv_body, timeout=20)
    if r.status_code not in (200, 201, 204):
        raise Exception(f"createInventoryItem failed: {r.status_code} {r.text}")

    ensure_ebay_location(token, location_key, location_zip, location_country)

    # 2. Create offer (or reuse existing offer_id if already drafted)
    offer_id = listing.get("ebay_offer_id")
    scheduled_at_iso = None
    offer_body = {
        "sku": sku,
        "marketplaceId": "EBAY_US",
        "format": "FIXED_PRICE",
        "availableQuantity": qty,
        "categoryId": str(category_id),
        "listingDescription": desc,
        "listingPolicies": {
            "paymentPolicyId": payment_policy,
            "returnPolicyId": return_policy,
            "fulfillmentPolicyId": fulfillment_policy,
            "bestOfferTerms": {"bestOfferEnabled": True},  # always on, per your standing preference
        },
        "pricingSummary": {"price": {"value": f"{price:.2f}", "currency": "USD"}},
        "merchantLocationKey": location_key,
    }
    if mode == "schedule":
        hrs = float(hours_from_now or 1)
        scheduled_dt = datetime.utcnow() + timedelta(hours=hrs)
        scheduled_at_iso = scheduled_dt.strftime("%Y-%m-%dT%H:%M:%SZ")
        offer_body["listingStartDate"] = scheduled_at_iso

    if offer_id:
        r = _req.put(f"{EBAY_API_BASE}/sell/inventory/v1/offer/{offer_id}",
                      headers=ebay_headers(token), json=offer_body, timeout=20)
        if r.status_code not in (200, 204):
            raise Exception(f"updateOffer failed: {r.status_code} {r.text}")
    else:
        r = _req.post(f"{EBAY_API_BASE}/sell/inventory/v1/offer",
                       headers=ebay_headers(token), json=offer_body, timeout=20)
        if r.status_code == 400 and "already exists" in r.text.lower():
            # A previous attempt for this SKU already created an offer on eBay's side
            # (e.g. it died at the publish step afterward) but our DB never learned the
            # offer_id. eBay's error body includes it — recover it and update instead of failing.
            existing_offer_id = None
            try:
                for err in r.json().get("errors", []):
                    for p in err.get("parameters", []):
                        if p.get("name") == "offerId":
                            existing_offer_id = p.get("value")
            except Exception:
                pass
            if existing_offer_id:
                offer_id = existing_offer_id
                r2 = _req.put(f"{EBAY_API_BASE}/sell/inventory/v1/offer/{offer_id}",
                               headers=ebay_headers(token), json=offer_body, timeout=20)
                if r2.status_code not in (200, 204):
                    raise Exception(f"updateOffer (recovered offerId) failed: {r2.status_code} {r2.text}")
            else:
                raise Exception(f"createOffer failed: {r.status_code} {r.text}")
        elif r.status_code not in (200, 201):
            raise Exception(f"createOffer failed: {r.status_code} {r.text}")
        else:
            offer_id = r.json().get("offerId")

    result = {"offer_id": offer_id, "sku": sku, "item_id": None, "status": "draft", "scheduled_at": None, "brand": brand, "mpn": mpn, "mpn_is_fallback": mpn_is_fallback}

    if mode == "draft":
        return result

    # 3. Publish
    r = _req.post(f"{EBAY_API_BASE}/sell/inventory/v1/offer/{offer_id}/publish",
                   headers=ebay_headers(token, content_language=False), timeout=20)
    if r.status_code not in (200, 201):
        raise Exception(f"publishOffer failed: {r.status_code} {r.text}")
    listing_id = r.json().get("listingId")
    result["item_id"] = listing_id
    result["status"] = "scheduled" if mode == "schedule" else "published"
    result["scheduled_at"] = scheduled_at_iso
    return result

def push_listing_to_ebay_v2(listing: dict, mode: str, hours_from_now: float = None, brand_override: str = None, mpn_override: str = None) -> dict:
    """Creates a listing via eBay's classic Trading API (AddFixedPriceItem) instead of
    the Sell Inventory API used by push_listing_to_ebay. On the Trading API, SKU/Custom
    Label is a plain free-text field on the item with no account-wide uniqueness
    requirement — unlike the Inventory API, where SKU IS the resource's identifier and
    a colliding SKU silently overwrites another item's data. This is the "Intake 2"
    parallel push path; existing Inventory-API listings are untouched and keep using
    push_listing_to_ebay. No offer_id in the result — that's an Inventory API concept
    that doesn't exist here; ebay_offer_id is left null for v2 listings on purpose, so
    other code can tell the two systems' listings apart (offer_id set = old system,
    offer_id null + item_id set = this one)."""
    import requests as _req
    import xml.etree.ElementTree as ET
    from xml.sax.saxutils import escape as _xesc
    from datetime import timedelta

    biz_id = listing.get("business_id")
    if not biz_id:
        raise Exception("Listing has no business_id — cannot look up eBay settings safely")
    settings = get_ebay_settings(biz_id)
    token = get_ebay_access_token(biz_id)

    payment_policy      = settings.get("EBAY_PAYMENT_POLICY_ID", "")
    return_policy        = settings.get("EBAY_RETURN_POLICY_ID", "")
    fulfillment_policy   = listing.get("ebay_fulfillment_policy_id") or settings.get("EBAY_FULFILLMENT_POLICY_ID", "")
    raw_category_id      = listing.get("ebay_category_id")
    category_id          = raw_category_id if raw_category_id and str(raw_category_id) != "0" else settings.get("EBAY_DEFAULT_CATEGORY_ID", "")
    location_zip        = settings.get("EBAY_LOCATION_ZIP", "")
    location_country     = settings.get("EBAY_LOCATION_COUNTRY", "US")
    location_city_state = settings.get("EBAY_LOCATION_CITY_STATE", "")

    if not (payment_policy and return_policy and fulfillment_policy):
        raise Exception("Missing eBay business policy IDs — set these in Settings first")
    if not category_id or str(category_id) == "0":
        raise Exception("This item has no eBay category set")
    if not location_zip:
        raise Exception("Missing EBAY_LOCATION_ZIP — set this in Settings first")
    if not location_city_state:
        raise Exception("Missing EBAY_LOCATION_CITY_STATE — set this in Settings first (e.g. 'Loveland, CO'). This is separate from Merchant Location Key, which only the older v1 publish path uses.")

    sku = listing.get("ebay_sku") or f"lister-{listing['id']}"
    title = (listing.get("title") or "Untitled item")[:80]
    desc  = listing.get("description") or EBAY_DESCRIPTION
    qty   = int(listing.get("quantity") or 1)
    price = float(listing.get("price") or 0)
    pid   = str(listing.get("photo_id") or "")
    images = [photo_url(p) for p in get_all_photo_ids(pid) if photo_url(p)] if pid else []

    brand = (brand_override or listing.get("brand") or "").strip()
    if not brand:
        first_word = title.split()[0].strip(",.;:-") if title else ""
        brand = first_word if first_word else "Unbranded"

    mpn = (mpn_override or "").strip() or None
    mpn_is_fallback = False
    if not mpn:
        alnum_tokens = [
            w.strip(",.;:-") for w in title.split()
            if w.lower().strip(",.;:-") != brand.lower()
            and any(c.isdigit() for c in w) and any(c.isalpha() for c in w)
        ]
        if alnum_tokens:
            mpn = max(alnum_tokens, key=len)
    if not mpn:
        mpn = "Does Not Apply"
        mpn_is_fallback = True

    condition_id = "1000" if ebay_condition(listing.get("condition")) == "NEW" else "3000"

    picture_xml = "".join(f"<PictureURL>{_xesc(u)}</PictureURL>" for u in images[:12])
    item_specifics_xml = (
        f"<NameValueList><Name>Brand</Name><Value>{_xesc(brand)}</Value></NameValueList>"
        f"<NameValueList><Name>MPN</Name><Value>{_xesc(mpn)}</Value></NameValueList>"
    )
    if listing.get("category_mode") == "motors":
        # Auto parts specifically: eBay Motors expects "Manufacturer Part Number" as
        # its own separate item specific, distinct from the "MPN" aspect used
        # elsewhere -- both need the same value sent, not just one or the other.
        item_specifics_xml += f"<NameValueList><Name>Manufacturer Part Number</Name><Value>{_xesc(mpn)}</Value></NameValueList>"

    result = {"offer_id": None, "sku": sku, "item_id": None, "status": "draft", "scheduled_at": None, "brand": brand, "mpn": mpn, "mpn_is_fallback": mpn_is_fallback}

    if mode == "draft":
        # No real draft concept on the Trading API — nothing is sent to eBay yet, this
        # is purely a local "not published" marker until you choose Publish/Schedule.
        return result

    schedule_xml = ""
    scheduled_at_iso = None
    if mode == "schedule":
        hrs = float(hours_from_now or 1)
        scheduled_dt = datetime.utcnow() + timedelta(hours=hrs)
        scheduled_at_iso = scheduled_dt.strftime("%Y-%m-%dT%H:%M:%SZ")
        schedule_xml = f"<ScheduleTime>{scheduled_at_iso}</ScheduleTime>"

    # Previously skipped for motors-mode items after eBay's error said Best
    # Offer was "unavailable...for the specified category" -- but that error
    # came from BEFORE the Site ID fix, when the whole request was going to
    # the wrong marketplace and eBay likely couldn't properly resolve the
    # category's real feature set at all. Re-enabled now that Site is correct
    # -- if a specific category genuinely doesn't support Best Offer, eBay
    # will say so on its own (as a warning, not a blocking error) and it'll be
    # visible in the modal's failure/status text either way.
    best_offer_xml = '<BestOfferDetails><BestOfferEnabled>true</BestOfferEnabled></BestOfferDetails>'

    xml_body = (
        '<?xml version="1.0" encoding="utf-8"?>'
        '<AddFixedPriceItemRequest xmlns="urn:ebay:apis:eBLBaseComponents">'
        f'<RequesterCredentials><eBayAuthToken>{token}</eBayAuthToken></RequesterCredentials>'
        '<Item>'
        f'<Title>{_xesc(title)}</Title>'
        f'<Description><![CDATA[{desc}]]></Description>'
        f'<PrimaryCategory><CategoryID>{_xesc(str(category_id))}</CategoryID></PrimaryCategory>'
        f'<StartPrice>{price:.2f}</StartPrice>'
        '<CategoryMappingAllowed>true</CategoryMappingAllowed>'
        f'<ConditionID>{condition_id}</ConditionID>'
        f'<Country>{_xesc(location_country)}</Country><Currency>USD</Currency>'
        f'<PostalCode>{_xesc(location_zip)}</PostalCode>'
        f'<Location>{_xesc(location_city_state)}</Location>'
        '<DispatchTimeMax>3</DispatchTimeMax>'
        '<ListingDuration>GTC</ListingDuration>'
        '<ListingType>FixedPriceItem</ListingType>'
        f'<Quantity>{qty}</Quantity>'
        f'<SKU>{_xesc(sku)}</SKU>'
        f'{best_offer_xml}'
        f'<PictureDetails>{picture_xml}</PictureDetails>'
        f'<ItemSpecifics>{item_specifics_xml}</ItemSpecifics>'
        '<SellerProfiles>'
        f'<SellerPaymentProfile><PaymentProfileID>{_xesc(payment_policy)}</PaymentProfileID></SellerPaymentProfile>'
        f'<SellerReturnProfile><ReturnProfileID>{_xesc(return_policy)}</ReturnProfileID></SellerReturnProfile>'
        f'<SellerShippingProfile><ShippingProfileID>{_xesc(fulfillment_policy)}</ShippingProfileID></SellerShippingProfile>'
        '</SellerProfiles>'
        f'{schedule_xml}'
        '</Item>'
        '</AddFixedPriceItemRequest>'
    )
    # THE actual root cause of 'Category is not valid' on Motors items, found by
    # checking this directly: X-EBAY-API-SITEID was hardcoded to "0" (the
    # standard US site) for every publish, motors or not. eBay Motors is a
    # SEPARATE site (ID 100) with its own category tree -- a Motors category
    # simply doesn't exist under Site 0, so no valid Motors category ID could
    # ever have worked here, regardless of which one got chosen. This is why
    # fixing the category-matching logic alone never resolved it.
    site_id = "100" if listing.get("category_mode") == "motors" else "0"
    headers = {
        "X-EBAY-API-COMPATIBILITY-LEVEL": "1193",
        "X-EBAY-API-CALL-NAME": "AddFixedPriceItem",
        "X-EBAY-API-SITEID": site_id,
        "Content-Type": "text/xml",
    }
    r = _req.post("https://api.ebay.com/ws/api.dll", headers=headers, data=xml_body.encode("utf-8"), timeout=30)
    resp = _ebay_xml_to_dict(ET.fromstring(r.content))
    ack = resp.get("Ack")
    if ack not in ("Success", "Warning"):
        raise Exception(f"AddFixedPriceItem failed (Ack={ack}): {resp.get('Errors')}")

    result["item_id"] = resp.get("ItemID")
    result["status"] = "scheduled" if mode == "schedule" else "published"
    result["scheduled_at"] = scheduled_at_iso
    return result

def _ebay_revise_sku_only(business_id: str, ebay_item_id: str, new_sku: str):
    """Trading API ReviseItem, sending ONLY ItemID + SKU — deliberately nothing else.
    eBay's classic Revise calls only change fields present in the request, so leaving
    out PictureDetails/ItemSpecifics/price/etc. here means they're left completely
    alone. Never expand this call to include other fields for that same reason — the
    whole point is a relabel that can't touch anything else on a live listing."""
    import requests as _req
    import xml.etree.ElementTree as ET
    from xml.sax.saxutils import escape as _xesc

    token = get_ebay_access_token(business_id)
    xml_body = (
        '<?xml version="1.0" encoding="utf-8"?>'
        '<ReviseItemRequest xmlns="urn:ebay:apis:eBLBaseComponents">'
        f'<RequesterCredentials><eBayAuthToken>{token}</eBayAuthToken></RequesterCredentials>'
        '<Item>'
        f'<ItemID>{_xesc(ebay_item_id)}</ItemID>'
        f'<SKU>{_xesc(new_sku)}</SKU>'
        '</Item>'
        '</ReviseItemRequest>'
    )
    headers = {
        "X-EBAY-API-COMPATIBILITY-LEVEL": "1193",
        "X-EBAY-API-CALL-NAME": "ReviseItem",
        "X-EBAY-API-SITEID": "0",
        "Content-Type": "text/xml",
    }
    r = _req.post("https://api.ebay.com/ws/api.dll", headers=headers, data=xml_body.encode("utf-8"), timeout=20)
    resp = _ebay_xml_to_dict(ET.fromstring(r.content))
    ack = resp.get("Ack")
    if ack not in ("Success", "Warning"):
        raise Exception(f"ReviseItem (SKU only) failed (Ack={ack}): {resp.get('Errors')}")
    return resp

def _ebay_revise_quantity_only(business_id: str, ebay_item_id: str, new_total_quantity: int):
    """Trading API ReviseItem, sending ONLY ItemID + Quantity — deliberately nothing
    else, same proven pattern as _ebay_revise_sku_only above. eBay's classic Revise
    calls only change fields present in the request, so leaving out price/title/
    ItemSpecifics/etc. here means they're left completely alone. Never expand this
    call to include other fields for that same reason.

    IMPORTANT: new_total_quantity must be the LIFETIME total (what ReviseItem
    expects), not "how many available right now" -- eBay computes available as
    Quantity - QuantitySold itself. The caller is responsible for adding the
    live QuantitySold before calling this (see push_inventory_quantity below);
    this function does no such adjustment itself and will send exactly the
    number it's given."""
    import requests as _req
    import xml.etree.ElementTree as ET
    from xml.sax.saxutils import escape as _xesc

    token = get_ebay_access_token(business_id)
    xml_body = (
        '<?xml version="1.0" encoding="utf-8"?>'
        '<ReviseItemRequest xmlns="urn:ebay:apis:eBLBaseComponents">'
        f'<RequesterCredentials><eBayAuthToken>{token}</eBayAuthToken></RequesterCredentials>'
        '<Item>'
        f'<ItemID>{_xesc(ebay_item_id)}</ItemID>'
        f'<Quantity>{int(new_total_quantity)}</Quantity>'
        '</Item>'
        '</ReviseItemRequest>'
    )
    headers = {
        "X-EBAY-API-COMPATIBILITY-LEVEL": "1193",
        "X-EBAY-API-CALL-NAME": "ReviseItem",
        "X-EBAY-API-SITEID": "0",
        "Content-Type": "text/xml",
    }
    r = _req.post("https://api.ebay.com/ws/api.dll", headers=headers, data=xml_body.encode("utf-8"), timeout=20)
    resp = _ebay_xml_to_dict(ET.fromstring(r.content))
    ack = resp.get("Ack")
    if ack not in ("Success", "Warning"):
        raise Exception(f"ReviseItem (Quantity only) failed (Ack={ack}): {resp.get('Errors')}")
    return resp

def _ebay_end_item(business_id: str, ebay_item_id: str, reason: str = "NotAvailable"):
    """Trading API EndItem -- pulls a listing down entirely. eBay's ReviseItem
    flatly refuses a Quantity that computes to 0 available (error 515: "The
    quantity must be a valid number greater than 0."), confirmed against
    eBay's own docs: 0 isn't a quantity state to them, it's "this item should
    no longer be for sale," which is EndItem's job, not ReviseItem's. Callers
    pushing a target quantity of 0 (e.g. sold out on Shopify) must call this
    instead of _ebay_revise_quantity_only. EndingReason "NotAvailable" is
    eBay's documented code for exactly this case -- the item sold out
    elsewhere, not a listing mistake or a sale to the high bidder."""
    import requests as _req
    import xml.etree.ElementTree as ET
    from xml.sax.saxutils import escape as _xesc

    token = get_ebay_access_token(business_id)
    xml_body = (
        '<?xml version="1.0" encoding="utf-8"?>'
        '<EndItemRequest xmlns="urn:ebay:apis:eBLBaseComponents">'
        f'<RequesterCredentials><eBayAuthToken>{token}</eBayAuthToken></RequesterCredentials>'
        f'<ItemID>{_xesc(ebay_item_id)}</ItemID>'
        f'<EndingReason>{_xesc(reason)}</EndingReason>'
        '</EndItemRequest>'
    )
    headers = {
        "X-EBAY-API-COMPATIBILITY-LEVEL": "1193",
        "X-EBAY-API-CALL-NAME": "EndItem",
        "X-EBAY-API-SITEID": "0",
        "Content-Type": "text/xml",
    }
    r = _req.post("https://api.ebay.com/ws/api.dll", headers=headers, data=xml_body.encode("utf-8"), timeout=20)
    resp = _ebay_xml_to_dict(ET.fromstring(r.content))
    ack = resp.get("Ack")
    if ack not in ("Success", "Warning"):
        raise Exception(f"EndItem failed (Ack={ack}): {resp.get('Errors')}")
    return resp

class UpdateSkuV2(BaseModel):
    new_sku: str


@app.post("/api/listings/{item_id}/ebay-v2/sku")
async def update_ebay_v2_sku(item_id: str, body: UpdateSkuV2, request: Request):
    """Relabels the SKU on a live Intake-2/Trading-API listing without touching
    anything else — the actual point of switching to this system. Refuses items that
    went through the old Inventory-API flow, since their SKU can never be changed via
    any API call, full stop (that's an eBay platform restriction, not a gap here)."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    new_sku = (body.new_sku or "").strip()
    if not new_sku:
        raise HTTPException(400, "new_sku is required")
    _validate_sku_for_assignment(business_id, new_sku)
    try:
        res = supabase.table("listings").select("id,ebay_item_id,ebay_offer_id,business_id").eq("id", item_id).limit(1).execute()
        if not res.data:
            raise HTTPException(404, "Listing not found")
        listing = res.data[0]
        if listing.get("ebay_offer_id"):
            raise HTTPException(400, "This item was published via the old eBay system — its SKU can't be changed via any API call, on eBay's side.")
        ebay_item_id = listing.get("ebay_item_id")
        if not ebay_item_id:
            raise HTTPException(400, "This item hasn't been published to eBay yet")
        _ebay_revise_sku_only(business_id, ebay_item_id, new_sku)
        supabase.table("listings").update({"ebay_sku": new_sku}).eq("id", item_id).execute()
        # Also update ebay_listing_status.sku -- that's the table every
        # "uncategorized listings" view (Lots, the Uncat tab) actually reads
        # from, not listings.ebay_sku. Without this, a successful, verified-
        # on-eBay SKU push would still show up as uncategorized in Lister's
        # own views until the next active-listings sync (up to a day away),
        # even though eBay itself was already correct.
        try:
            supabase.table("ebay_listing_status").update({"sku": new_sku}).eq("business_id", business_id).eq("item_id", ebay_item_id).execute()
        except Exception as e:
            print(f"update_ebay_v2_sku: local ebay_listing_status update failed (eBay itself succeeded): {e}")
        return {"ok": True, "ebay_sku": new_sku}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))

# ── PAGES ─────────────────────────────────────────────────────── #

@app.get("/auction/research", response_class=HTMLResponse)
@app.get("/auction/research/{share_id}", response_class=HTMLResponse)
async def auction_research_page(request: Request, share_id: str = None):
    """Same page whether visited plain (the owner's own working session) or with
    a /share_id suffix (a shared, read-only view sent to someone else) -- the
    page's own tryLoadShared() JS already parses the URL and calls
    /api/auction/load-research/{share_id} correctly on its own. This route was
    missing the {share_id} path entirely before, so a shared link 404'd for
    anyone who clicked it despite the frontend logic being fully built and
    ready -- the actual bug was just this one missing route."""
    import os
    with open(os.path.join(os.path.dirname(__file__), "templates", "auction_research.html")) as f:
        html = f.read()
    return HTMLResponse(content=html, headers={
        "Content-Security-Policy": "default-src * blob: data:; script-src * blob: data: 'unsafe-inline' 'unsafe-eval'; style-src * 'unsafe-inline'; img-src * blob: data:;"
    })

@app.get("/auction/share/{session_id}", response_class=HTMLResponse)
async def auction_capture_share_page(session_id: str, request: Request):
    """Public, no-auth page for sharing one Live Lot Capture session (a single catalog)
    externally, e.g. a link to send a friend. Read-only — the page fetches its own data
    from /api/auction/capture/public/{session_id}, no login required."""
    import os
    with open(os.path.join(os.path.dirname(__file__), "templates", "auction_share.html")) as f:
        html = f.read()
    return HTMLResponse(content=html, headers={
        "Content-Security-Policy": "default-src * blob: data:; script-src * blob: data: 'unsafe-inline' 'unsafe-eval'; style-src * 'unsafe-inline'; img-src * blob: data:;"
    })

@app.get("/auction", response_class=HTMLResponse)
async def auction_page(request: Request):
    business_id = require_auth(request)
    if not business_id:
        from fastapi.responses import RedirectResponse
        return RedirectResponse("/login", status_code=302)
    import os
    with open(os.path.join(os.path.dirname(__file__), "templates", "auction.html")) as f:
        html = f.read()
    return HTMLResponse(content=html, headers={
        "Content-Security-Policy": "default-src * blob: data:; script-src * blob: data: 'unsafe-inline' 'unsafe-eval'; style-src * 'unsafe-inline'; img-src * blob: data:;"
    })

@app.get("/v2", response_class=HTMLResponse)
async def dashboard_v2(request: Request):
    from fastapi.responses import HTMLResponse
    import os
    with open(os.path.join(os.path.dirname(__file__), "templates", "v2.html")) as f:
        html = f.read()
    return HTMLResponse(content=html, headers={
        "Content-Security-Policy": "default-src * blob: data:; script-src * blob: data: 'unsafe-inline' 'unsafe-eval'; style-src * 'unsafe-inline'; img-src * blob: data:;"
    })

@app.get("/admin", response_class=HTMLResponse)
async def admin_page(request: Request):
    business_id, is_admin = get_business_info(request)
    if not business_id or not is_admin:
        from fastapi.responses import RedirectResponse
        return RedirectResponse("/login", status_code=302)
    with open(os.path.join(os.path.dirname(__file__), "templates", "admin.html")) as f:
        return HTMLResponse(f.read())

@app.patch("/api/admin/businesses/{business_id}")
async def admin_update_business(business_id: str, request: Request):
    auth_business_id, is_admin = get_business_info(request)
    if not auth_business_id or not is_admin:
        raise HTTPException(401, "Unauthorized")
    try:
        body = await request.json()
        allowed = {k: v for k, v in body.items() if k in ("scan_limit", "scan_count", "is_admin")}
        supabase.table("businesses").update(allowed).eq("id", business_id).execute()
        return {"ok": True}
    except Exception as e:
        raise HTTPException(500, str(e))

@app.delete("/api/admin/businesses/{business_id}")
async def admin_delete_business(business_id: str, request: Request):
    auth_business_id, is_admin = get_business_info(request)
    if not auth_business_id or not is_admin:
        raise HTTPException(401, "Unauthorized")
    try:
        supabase.table("sessions").delete().eq("business_id", business_id).execute()
        supabase.table("listings").delete().eq("business_id", business_id).execute()
        supabase.table("listing_groups").delete().eq("business_id", business_id).execute()
        supabase.table("businesses").delete().eq("id", business_id).execute()
        return {"ok": True}
    except Exception as e:
        raise HTTPException(500, str(e))

@app.get("/api/admin/businesses")
async def admin_businesses(request: Request):
    auth_business_id, is_admin = get_business_info(request)
    if not auth_business_id or not is_admin:
        raise HTTPException(401, "Unauthorized")
    try:
        biz = supabase.table("businesses").select("id,name,email,created_at,scan_count,scan_limit,is_admin").order("created_at", desc=True).execute()
        businesses = biz.data or []
        for b in businesses:
            lid = supabase.table("listings").select("id", count="exact").eq("business_id", b["id"]).execute()
            b["listing_count"] = lid.count or 0
            sid = supabase.table("auction_research_sessions").select("id,created_at", count="exact").eq("business_id", b["id"]).order("created_at", desc=True).limit(1).execute()
            b["scan_count"] = sid.count or 0
            b["last_active"] = sid.data[0]["created_at"] if sid.data else None
        return {"businesses": businesses}
    except Exception as e:
        raise HTTPException(500, str(e))

@app.get("/paywall", response_class=HTMLResponse)
async def paywall_page(request: Request):
    with open(os.path.join(os.path.dirname(__file__), "templates", "paywall.html")) as f:
        return HTMLResponse(f.read())

@app.get("/team", response_class=HTMLResponse)
async def team_portal(request: Request):
    with open(os.path.join(os.path.dirname(__file__), "templates", "team_portal.html")) as f:
        return HTMLResponse(f.read())

@app.get("/portal", response_class=HTMLResponse)
async def portal(request: Request):
    import os
    with open(os.path.join(os.path.dirname(__file__), "templates", "portal.html")) as f:
        html = f.read()
    from fastapi.responses import HTMLResponse
    return HTMLResponse(content=html, headers={
        "Content-Security-Policy": "default-src * blob: data:; script-src * blob: data: 'unsafe-inline' 'unsafe-eval'; style-src * 'unsafe-inline'; img-src * blob: data:;"
    })


def require_auth(request: Request):
    """Returns business_id if authenticated, else None."""
    token = request.cookies.get("session_id")
    if not token:
        return None
    try:
        res = supabase.table("sessions").select("business_id").eq("token", token).execute()
        if res.data:
            return res.data[0]["business_id"]
    except Exception:
        pass
    return None

def get_business_info(request: Request):
    """Returns (business_id, is_admin) or (None, False)."""
    token = request.cookies.get("session_id")
    if not token:
        return None, False
    try:
        res = supabase.table("sessions").select("business_id").eq("token", token).execute()
        if not res.data:
            return None, False
        bid = res.data[0]["business_id"]
        biz = supabase.table("businesses").select("is_admin").eq("id", bid).execute()
        is_admin = bool(biz.data[0]["is_admin"]) if biz.data else False
        return bid, is_admin
    except Exception:
        pass
    return None, False

def get_nav_context(request: Request):
    """Returns the shared context every page's nav bar (templates/_nav.html) needs
    — business_id, is_admin, account_label — or None if not logged in. Consolidates
    what used to be near-identical boilerplate repeated in every page route."""
    business_id, is_admin = get_business_info(request)
    if not business_id:
        return None
    account_label = ""
    biz = supabase.table("businesses").select("name,email").eq("id", business_id).limit(1).execute()
    if biz.data:
        account_label = biz.data[0].get("email") or biz.data[0].get("name") or ""
    return {"business_id": business_id, "is_admin": is_admin, "account_label": account_label}

@app.get("/", response_class=HTMLResponse)
async def dashboard(request: Request):
    nav = get_nav_context(request)
    if nav is None:
        from fastapi.responses import RedirectResponse
        return RedirectResponse("/login", status_code=302)
    return templates.TemplateResponse("index.html", {"request": request, "is_admin": nav["is_admin"], "account_label": nav["account_label"], "active_tab": "intake"})

# ── API: LISTINGS ─────────────────────────────────────────────── #

@app.get("/api/ebay/shipping-policy-options")
async def get_shipping_policy_options(request: Request):
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    settings = get_ebay_settings(business_id)
    return {
        "default_policy_id": settings.get("EBAY_FULFILLMENT_POLICY_ID", ""),
        "options": EBAY_SHIPPING_POLICY_OPTIONS,
    }

@app.get("/api/listings")
async def get_listings(request: Request, archived: bool = False):
    business_id = require_auth(request)
    if not business_id:
        return JSONResponse({"error": "unauthorized"}, status_code=401)
    try:
        q = supabase.table("listings").select("*").eq("business_id", business_id)
        q = q.eq("status", "archived") if archived else q.neq("status", "archived")
        res = q.order("created_at", desc=True).execute()
        listings = res.data or []

        # Re-sort by the listing's originating group's created_at (true, strictly
        # sequential submission order) instead of the listing's own created_at.
        # A listing's own created_at is set by the external scanner service
        # whenever ITS processing happens to finish — and since photos of
        # different complexity take different amounts of time to analyze, that
        # can complete out of order even though photos were submitted in a
        # strict sequence. This was very likely the actual cause of items
        # appearing to land at the top, bottom, or middle unpredictably.
        all_pids = [str(l.get("photo_id") or "") for l in listings if l.get("photo_id")]
        group_created_at = {}
        if all_pids:
            try:
                pid_to_gid_full = {}
                for i in range(0, len(all_pids), 200):
                    chunk = all_pids[i:i+200]
                    gp_all = supabase.table("group_photos").select("group_id, photo_id").in_("photo_id", chunk).execute()
                    pid_to_gid_full.update({row["photo_id"]: row["group_id"] for row in (gp_all.data or [])})
                all_group_ids = list(set(pid_to_gid_full.values()))
                gid_to_created = {}
                for i in range(0, len(all_group_ids), 200):
                    chunk = all_group_ids[i:i+200]
                    lg_res = supabase.table("listing_groups").select("id,created_at").in_("id", chunk).execute()
                    gid_to_created.update({row["id"]: row["created_at"] for row in (lg_res.data or [])})
                for pid, gid in pid_to_gid_full.items():
                    if gid in gid_to_created:
                        group_created_at[pid] = gid_to_created[gid]
            except Exception as e:
                print(f"group-based listing sort failed, falling back to listings.created_at: {e}")

        listings.sort(key=lambda l: group_created_at.get(str(l.get("photo_id") or ""), l.get("created_at") or ""), reverse=True)

        # Batch fetch all group photos for these listings
        primary_pids = [str(l.get("photo_id") or "") for l in listings if l.get("photo_id")]
        group_photo_map = {}  # photo_id -> [all photo_ids in same group]
        if primary_pids:
            try:
                gp_res = supabase.table("group_photos")                    .select("group_id, photo_id")                    .in_("photo_id", primary_pids[:100])                    .execute()
                # Map primary photo -> group_id
                pid_to_gid = {row["photo_id"]: row["group_id"] for row in (gp_res.data or [])}
                group_ids = list(set(pid_to_gid.values()))
                if group_ids:
                    all_gp = supabase.table("group_photos")                        .select("group_id, photo_id")                        .in_("group_id", group_ids)                        .execute()
                    # Build group_id -> [photo_ids]
                    gid_to_photos = {}
                    for row in (all_gp.data or []):
                        gid_to_photos.setdefault(row["group_id"], []).append(row["photo_id"])
                    # Map primary photo_id -> all photos in its group
                    for pid, gid in pid_to_gid.items():
                        group_photo_map[pid] = gid_to_photos.get(gid, [pid])
            except Exception as search_err:
                print(f"   Search grounding failed: {search_err}")


        # Batch-fetch readable category paths for any listings that have a category set
        cat_ids = list({str(l["ebay_category_id"]) for l in listings if l.get("ebay_category_id")})
        cat_path_map = {}
        if cat_ids:
            try:
                cat_res = supabase.table("ebay_categories").select("category_id,path").in_("category_id", cat_ids[:200]).execute()
                cat_path_map = {row["category_id"]: row["path"] for row in (cat_res.data or [])}
            except Exception as e:
                print(f"category path lookup failed: {e}")

        default_cat_id = str(get_ebay_settings(business_id).get("EBAY_DEFAULT_CATEGORY_ID", "") or "")
        default_motors_cat_id = str(get_ebay_settings(business_id).get("EBAY_DEFAULT_MOTORS_CATEGORY_ID", "") or "") or "26261"

        for l in listings:
            pid = str(l.get("photo_id") or "")
            raw_photos = group_photo_map.get(pid, [pid] if pid else [])
            # group_photos has no defined order — always put the listing's actual
            # photo_id first, so "Make Main Photo" (which only ever updates
            # photo_id) actually sticks after a reload instead of reverting to
            # whatever order Supabase happened to return.
            all_photos = ([pid] + [p for p in raw_photos if p != pid]) if pid else raw_photos
            l["thumb_url"]  = photo_url(pid, thumb=True)
            l["full_url"]   = photo_url(pid)
            l["all_photos"] = [{"id": p, "thumb": photo_url(p, thumb=True), "full": photo_url(p)} for p in all_photos if p]
            l["ebay_category_path"] = cat_path_map.get(str(l.get("ebay_category_id") or ""), "")
            l["ebay_category_is_default"] = (
                (l.get("category_mode") == "motors" and str(l.get("ebay_category_id") or "") == default_motors_cat_id)
                or (l.get("category_mode") != "motors" and bool(default_cat_id) and str(l.get("ebay_category_id") or "") == default_cat_id)
            )
            # Coerce types
            l["price"]      = float(l.get("price") or 0)
            l["price_used"] = float(l.get("price_used") or 0)
            l["price_new"]  = float(l.get("price_new") or 0)
            l["quantity"]   = int(l.get("quantity") or 1)
            # Normalize condition — default to "used" if blank/null
            cond = str(l.get("condition") or "").strip().lower()
            l["condition"] = cond if cond in ("new", "used") else "used"
            # Normalize listing_type — items from batch upload are not auctions
            lt = str(l.get("listing_type") or "").strip().lower()
            if lt not in ("auction", "fixed"):
                l["listing_type"] = "fixed"
        return JSONResponse(listings)
    except Exception as e:
        import traceback; traceback.print_exc()
        raise HTTPException(500, str(e))

class UpdateField(BaseModel):
    field: str
    value: object



@app.get("/api/export/ebay-csv")
async def export_ebay_csv(request: Request):
    import csv, io
    from fastapi.responses import StreamingResponse
    from datetime import datetime
    business_id = require_auth(request)
    try:
        q = supabase.table("listings").select(
            "title,description,price,price_used,price_new,quantity,condition,photo_id,ebay_category_id,description"
        ).neq("status", "archived")
        if business_id:
            q = q.eq("business_id", business_id)
        res = q.execute()
        items = res.data or []
    except Exception as e:
        raise HTTPException(500, str(e))

    output = io.StringIO()

    # eBay draft flat file headers — same format as the working version
    output.write('#INFO,Version=0.0.2,Template= eBay-draft-listings-template_US,,,,,,,,\n')
    output.write('#INFO Action and Category ID are required fields.,,,,,,,,,,\n')
    output.write('#INFO,,,,,,,,,,\n')
    output.write('Action(SiteID=US|Country=US|Currency=USD|Version=1193|CC=UTF-8),Custom label (SKU),Category ID,Title,UPC,Price,Quantity,Item photo URL,Condition ID,Description,Format\n')

    writer = csv.writer(output, quoting=csv.QUOTE_ALL)

    for item in items:
        cond = str(item.get("condition") or "used").strip().lower()
        cond_id = "NEW" if cond == "new" else "USED"
        pid = str(item.get("photo_id") or "")
        pic = "|".join(photo_url(p) for p in get_all_photo_ids(pid) if photo_url(p))
        category_id = "12576"
        price = float(item.get("price") or item.get("price_used") or 0)
        writer.writerow([
            "Draft",
            "",
            category_id,
            str(item.get("title",""))[:80],
            "",
            f"{price:.2f}",
            str(int(item.get("quantity") or 1)),
            pic,
            cond_id,
            str(item.get('description','') or EBAY_DESCRIPTION),
            "FixedPrice",
        ])

    csv_bytes = output.getvalue().encode("utf-8")
    fn = f"ebay_export_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
    return StreamingResponse(
        io.BytesIO(csv_bytes),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={fn}"}
    )


@app.get("/api/photos/view/{photo_id}")
async def view_photo(photo_id: str, t: str = ""):
    from fastapi.responses import Response
    img_bytes = supabase.storage.from_("part-photos").download(photo_id)
    return Response(content=img_bytes, media_type="image/jpeg", headers={
        "Cache-Control": "no-store, no-cache, must-revalidate, max-age=0",
        "Pragma": "no-cache",
        "Expires": "0"
    })

@app.post("/api/photos/rotate")
async def rotate_photo(request: Request):
    from PIL import Image
    import io
    body = await request.json()
    photo_id = body.get("photo_id", "")
    if not photo_id:
        raise HTTPException(400, "photo_id required")
    try:
        img_bytes = supabase.storage.from_("part-photos").download(photo_id)
        img = Image.open(io.BytesIO(img_bytes))
        direction = body.get('direction', 'cw')
        img = img.rotate(90 if direction == 'ccw' else -90, expand=True)
        buf = io.BytesIO()
        img.save(buf, format="JPEG", quality=90)
        buf.seek(0)
        supabase.storage.from_("part-photos").upload(
            path=photo_id,
            file=buf.read(),
            file_options={"content-type": "image/jpeg", "upsert": "true"}
        )
        return {"ok": True}
    except Exception as e:
        raise HTTPException(500, str(e))

@app.post("/api/reset-queue")
async def reset_queue():
    try:
        supabase.table("listing_groups").update({"status": "waiting"}).in_("status", ["processing", "pending"]).execute()
        return {"ok": True}
    except Exception as e:
        raise HTTPException(500, str(e))

@app.post("/api/stuck-groups/{group_id}/clear")
async def clear_stuck_group(group_id: str, request: Request):
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    try:
        group_res = supabase.table("listing_groups").select("created_at").eq("id", group_id).limit(1).execute()
        created_at = group_res.data[0]["created_at"] if group_res.data else None

        # Stop it from ever being retried or counted as "processing" again
        supabase.table("listing_groups").update({"status": "archived"}).eq("id", group_id).eq("business_id", business_id).execute()

        # If a listing already exists for this group's photo, archive it. If not (scanning never
        # got far enough to create one), create a minimal archived placeholder so it's actually
        # visible on the Archive page instead of silently vanishing.
        gp = supabase.table("group_photos").select("photo_id").eq("group_id", group_id).limit(1).execute()
        if gp.data:
            pid = gp.data[0]["photo_id"]
            title = "Failed scan (unknown date)"
            if created_at:
                try:
                    dt = datetime.fromisoformat(created_at.replace("Z", "+00:00"))
                    title = f"Failed scan — {dt.strftime('%b %d, %Y %I:%M %p')}"
                except Exception:
                    pass

            existing = supabase.table("listings").select("id,title").eq("photo_id", pid).eq("business_id", business_id).limit(1).execute()
            if existing.data:
                update = {"status": "archived"}
                cur_title = (existing.data[0].get("title") or "").strip()
                if not cur_title or cur_title == "Scanning...":
                    update["title"] = title
                supabase.table("listings").update(update).eq("id", existing.data[0]["id"]).execute()
            else:
                supabase.table("listings").insert({
                    "business_id": business_id,
                    "photo_id": pid,
                    "title": title,
                    "status": "archived",
                    "price": 0,
                    "quantity": 1,
                    "condition": "used",
                }).execute()
        return {"ok": True}
    except Exception as e:
        raise HTTPException(500, str(e))

@app.get("/api/stuck-groups")
async def get_stuck_groups(request: Request):
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    try:
        # Same fix as /api/stats above: "error" means the scan failed or was
        # deliberately abandoned -- it's done trying, not still stuck/scanning.
        # This is a SEPARATE endpoint from /api/stats (powers the detail list
        # you get from clicking into the scanning count, not the summary
        # number itself), so it had the exact same bug independently and
        # needed its own fix -- confirmed: the summary correctly showed 8,
        # but this list still showed 100+ of the same abandoned groups.
        groups = supabase.table("listing_groups").select("id,status,created_at")\
            .eq("business_id", business_id).in_("status", ["pending", "processing"]).execute()
        out = []
        for g in (groups.data or []):
            gp = supabase.table("group_photos").select("photo_id").eq("group_id", g["id"]).limit(1).execute()
            pid = gp.data[0]["photo_id"] if gp.data else ""
            out.append({
                "group_id": g["id"],
                "status": g["status"],
                "created_at": g.get("created_at"),
                "thumb_url": photo_url(pid, thumb=True) if pid else "",
            })
        return {"stuck": out}
    except Exception as e:
        raise HTTPException(500, str(e))

@app.get("/api/stats")
async def get_stats(request: Request):
    try:
        business_id = require_auth(request)
        if not business_id:
            return {"total": 0, "processing": 0, "value": 0}
        res = supabase.table("listings").select("price, status").eq("business_id", business_id).neq("status", "archived").execute()
        items = res.data or []
        total = len(items)
        value = sum(float(i.get("price") or 0) for i in items)
        # "error" means the scan failed or was deliberately abandoned -- it's
        # done trying, not still working. Counting it as "processing" here
        # made the progress bar show a wildly inflated, permanently-stuck
        # number any time a batch of groups got marked error (confirmed: 132
        # old abandoned groups showed up as "141 scanning" alongside the 9
        # that were genuinely active).
        pending = supabase.table("listing_groups").select("id").eq("business_id", business_id).in_("status", ["pending", "processing"]).execute()
        processing = len(pending.data or [])
        return {"total": total, "processing": processing, "value": round(value, 2)}
    except Exception as e:
        raise HTTPException(500, str(e))

@app.patch("/api/listings/{item_id}")
async def update_listing(item_id: str, body: UpdateField, request: Request):
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    try:
        supabase.table("listings").update({body.field: body.value}).eq("id", item_id).execute()
        return {"ok": True}
    except Exception as e:
        raise HTTPException(500, str(e))

class AssignLot(BaseModel):
    lot_sku: str
    suffix: Optional[str] = None

@app.post("/api/listings/{item_id}/assign-lot")
async def assign_lot_sku(item_id: str, body: AssignLot, request: Request):
    """Sets ebay_sku to '{lot}-' (e.g. "AM1-") by default — no number appended. If a
    suffix is supplied (fully optional, from the intake page's Active Suffix field),
    uses '{lot}-{suffix}' instead for that call — lets someone key in a custom second
    half (e.g. a batch/day code) that applies to everything being intaked in that
    session, without requiring one. This endpoint is only called from the v2/Trading-
    API intake page, where SKU/Custom Label is plain free text with no account-wide
    uniqueness requirement (unlike the old v1 Inventory API, where SKU is the
    resource's own identifier and a collision would silently overwrite another item's
    live listing data). If this endpoint is ever reused from a v1-publishing page, the
    old per-lot numbering (see git history) needs to come back for that case."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    lot = (body.lot_sku or "").strip()
    if not lot:
        raise HTTPException(400, "lot_sku is required")
    suffix = (body.suffix or "").strip()
    new_sku = f"{lot}-{suffix}" if suffix else f"{lot}-"
    try:
        supabase.table("listings").update({"ebay_sku": new_sku}).eq("id", item_id).execute()
        return {"ok": True, "ebay_sku": new_sku}
    except Exception as e:
        raise HTTPException(500, str(e))

class EbaySubmit(BaseModel):
    mode: str  # 'draft' | 'now' | 'schedule'
    hours_from_now: Optional[float] = None
    brand: Optional[str] = None
    mpn: Optional[str] = None


@app.post("/api/listings/{item_id}/find-mpn")
async def find_mpn(item_id: str, request: Request):
    """Auto parts specifically: looks at the item's actual photos for a real
    manufacturer part number physically marked on the part (stamp, sticker,
    casting, tag) -- auto parts almost always have this somewhere. Returns
    the value to fill into the MPN field, which then gets sent to eBay as
    BOTH the 'MPN' and 'Manufacturer Part Number' aspects at publish time
    (see push_listing_to_ebay_v2) -- two genuinely separate fields in eBay's
    item specifics for this category, both need the same value."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")

    res = supabase.table("listings").select("title,photo_id").eq("id", item_id).limit(1).execute()
    if not res.data:
        raise HTTPException(404, "Listing not found")
    listing = res.data[0]
    title = listing.get("title") or ""
    pid = str(listing.get("photo_id") or "")
    photo_ids = get_all_photo_ids(pid) if pid else []
    if not photo_ids:
        raise HTTPException(400, "No photos available for this item")

    import os, json
    import google.generativeai as genai
    gemini_key = os.getenv("GEMINI_API_KEY", "")
    if not gemini_key:
        raise HTTPException(400, "GEMINI_API_KEY not set")
    genai.configure(api_key=gemini_key)
    model = genai.GenerativeModel("gemini-2.5-flash")

    prompt = f"""This is an auto part being listed for sale on eBay. Its listing title is: "{title}"

Look carefully at the photos for a manufacturer part number physically marked on the
part or its packaging -- stamped, etched, printed, or on a label/tag/sticker. Auto
parts almost always have this marked somewhere directly on the item. This is
different from a generic model or product name -- it's usually a specific
alphanumeric code unique to this exact part.

Return ONLY a JSON object, no other text, in this exact shape:
{{"mpn": "the part number you found" or null if you genuinely cannot find one, "confidence": "high", "medium", or "low"}}

Only return a real value if you can actually see it in the photos or are highly
confident from context -- never guess or make one up."""

    parts = [prompt]
    for photo_id in photo_ids[:4]:
        try:
            img_bytes = supabase.storage.from_("part-photos").download(photo_id)
            parts.append({"mime_type": "image/jpeg", "data": img_bytes})
        except Exception as e:
            print(f"find_mpn: failed to fetch photo {photo_id}: {e}")
            continue

    try:
        response = model.generate_content(parts, generation_config={"max_output_tokens": 500})
        text = response.text.strip()
        if text.startswith("```"):
            text = text.split("```")[1]
            if text.startswith("json"):
                text = text[4:]
        parsed = json.loads(text.strip())
        mpn = (parsed.get("mpn") or "").strip()
        if not mpn:
            return {"ok": True, "found": False}
        return {"ok": True, "found": True, "mpn": mpn, "confidence": parsed.get("confidence", "low")}
    except Exception as e:
        raise HTTPException(500, f"Find MPN failed: {e}")

@app.post("/api/listings/{item_id}/rematch-category")
async def rematch_category(item_id: str, request: Request, mode: str = "industrial", exclude_ids: str = ""):
    """Re-runs category matching for one listing in the given lane (from the Intake
    toggle). Exists because the background auto_fill_worker's revalidation sweep
    only re-checks listings categorized OUTSIDE Business & Industrial / eBay Motors —
    a listing sitting at the generic 26261 fallback is technically still inside that
    root, so it passes the sweep's check and is never revisited automatically. This
    is the only way to fix already-scanned items after a category-matching bug fix
    ships; new scans self-correct via the worker."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    if mode not in ("industrial", "motors"):
        mode = "industrial"
    res = supabase.table("listings").select("id,title,ebay_category_id").eq("id", item_id).eq("business_id", business_id).execute()
    if not res.data:
        raise HTTPException(404, "Listing not found")
    title = res.data[0].get("title") or ""
    current_category_id = str(res.data[0].get("ebay_category_id") or "")
    if not title or title == "Scanning...":
        raise HTTPException(400, "No title yet")
    # Exclude the current category PLUS every category this item has already landed
    # on this session (passed in by the frontend, which tracks it per-item in memory)
    # -- excluding only the current one lets repeated presses ping-pong forever
    # between just two candidates (the generic fallback and one wrong specific match)
    # since each press only ever rules out the single most-recent answer. Excluding
    # the full history instead means each press rotates to a genuinely new candidate
    # until the real options are exhausted, at which point it settles on the fallback.
    exclude_set = {c for c in (exclude_ids or "").split(",") if c}
    if current_category_id:
        exclude_set.add(current_category_id)
    suggestion = suggest_ebay_category(title, business_id, restrict=True, exclude_ids=list(exclude_set), mode=mode)
    if suggestion and suggestion.get("category_id"):
        supabase.table("listings").update({
            "ebay_category_id": suggestion["category_id"],
            "category_mode": mode,
        }).eq("id", item_id).execute()
        # Real full breadcrumb, same source every normal listing's displayed
        # category comes from (ebay_categories.path) -- NOT just the bare leaf
        # name. Root cause of a real bug: this endpoint never returned this,
        # so both the bulk "Re-match Category" button and any per-tile recalc
        # button could only ever show a stale or incomplete path afterward,
        # unlike a normally-matched item.
        category_path = None
        try:
            path_res = (supabase.table("ebay_categories").select("path")
                        .eq("category_id", str(suggestion["category_id"])).limit(1).execute())
            if path_res.data:
                category_path = path_res.data[0].get("path")
        except Exception as e:
            print(f"rematch_category: path lookup failed: {e}")
        return {"ok": True, "category_id": suggestion["category_id"], "name": suggestion.get("name"),
                "category_path": category_path,
                "tree_id": suggestion.get("tree_id"), "is_fallback": suggestion.get("is_fallback", False)}
    else:
        return {"ok": True, "category_id": None, "name": None, "note": "still no B&I/Motors match"}

@app.post("/api/listings/{item_id}/ebay")
async def submit_to_ebay(item_id: str, body: EbaySubmit, request: Request):
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    try:
        res = supabase.table("listings").select("*").eq("id", item_id).limit(1).execute()
        if not res.data:
            raise HTTPException(404, "Listing not found")
        listing = res.data[0]
        _require_valid_lot_sku_for_publish(business_id, listing)
        result = push_listing_to_ebay(listing, body.mode, body.hours_from_now, body.brand, body.mpn)
        update = {
            "ebay_offer_id": result["offer_id"],
            "ebay_sku": result["sku"],
            "ebay_status": result["status"],
            "ebay_error": None,
            "brand": result.get("brand"),
        }
        if result["item_id"]:
            update["ebay_item_id"] = result["item_id"]
        if result["scheduled_at"]:
            update["ebay_scheduled_at"] = result["scheduled_at"]
        try:
            update["ebay_mpn"] = result.get("mpn")
            update["ebay_mpn_is_fallback"] = result.get("mpn_is_fallback", False)
            supabase.table("listings").update(update).eq("id", item_id).execute()
        except Exception:
            # ebay_mpn / ebay_mpn_is_fallback columns may not exist on this Supabase project yet —
            # retry without them so the actual eBay submission still gets saved
            update.pop("ebay_mpn", None)
            update.pop("ebay_mpn_is_fallback", None)
            supabase.table("listings").update(update).eq("id", item_id).execute()
        try:
            _maybe_confirm_inventory_match(business_id, item_id)
        except Exception:
            pass
        if result.get("item_id"):
            _mark_needs_post_publish_sync(business_id)
        return {"ok": True, **result}
    except HTTPException:
        raise
    except Exception as e:
        supabase.table("listings").update({"ebay_status": "failed", "ebay_error": str(e)}).eq("id", item_id).execute()
        raise HTTPException(500, str(e))

@app.post("/api/listings/{item_id}/ebay-v2")
async def submit_to_ebay_v2(item_id: str, body: EbaySubmit, request: Request):
    """Intake 2's push button — same shape as /api/listings/{id}/ebay, but goes through
    push_listing_to_ebay_v2 (Trading API) instead. Kept fully separate from the original
    endpoint so nothing about the existing Inventory-API flow changes underneath it."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    try:
        res = supabase.table("listings").select("*").eq("id", item_id).limit(1).execute()
        if not res.data:
            raise HTTPException(404, "Listing not found")
        listing = res.data[0]
        _require_valid_lot_sku_for_publish(business_id, listing)
        if listing.get("ebay_item_id") and not listing.get("ebay_offer_id") and body.mode != "draft":
            raise HTTPException(400, "Already published — re-submitting isn't supported yet. Use the SKU field to relabel it instead.")
        result = push_listing_to_ebay_v2(listing, body.mode, body.hours_from_now, body.brand, body.mpn)
        update = {
            "ebay_sku": result["sku"],
            "ebay_status": result["status"],
            "ebay_error": None,
            "brand": result.get("brand"),
            # v2 (Trading API) never sets ebay_offer_id — that's purely a v1 (Sell
            # Inventory API) concept. A listing that once had a v1 attempt keeps that
            # field forever otherwise, which the Uncategorized view reads as "still
            # v1, SKU can't be edited" — incorrectly locking a listing that's actually
            # live via v2 right now. A successful v2 publish is authoritative that v2
            # is the current mechanism, so always clear it here.
            "ebay_offer_id": None,
        }
        if result["item_id"]:
            update["ebay_item_id"] = result["item_id"]
        if result["scheduled_at"]:
            update["ebay_scheduled_at"] = result["scheduled_at"]
        try:
            update["ebay_mpn"] = result.get("mpn")
            update["ebay_mpn_is_fallback"] = result.get("mpn_is_fallback", False)
            supabase.table("listings").update(update).eq("id", item_id).execute()
        except Exception:
            update.pop("ebay_mpn", None)
            update.pop("ebay_mpn_is_fallback", None)
            supabase.table("listings").update(update).eq("id", item_id).execute()
        try:
            _maybe_confirm_inventory_match(business_id, item_id)
        except Exception:
            pass
        if result.get("item_id"):
            _mark_needs_post_publish_sync(business_id)
        return {"ok": True, **result}
    except HTTPException:
        raise
    except Exception as e:
        supabase.table("listings").update({"ebay_status": "failed", "ebay_error": str(e)}).eq("id", item_id).execute()
        raise HTTPException(500, str(e))

def get_gemini_key(business_id: str) -> str:
    settings = get_ebay_settings(business_id)
    return settings.get("GEMINI_API_KEY", "") or os.getenv("GEMINI_API_KEY", "")

def sync_ebay_categories(token: str) -> dict:
    """Download eBay's full category tree(s) and save them locally — EVERY node, not
    just leaves, so parent/grouping categories show their real IDs too. is_leaf marks
    which ones are actually valid for listing an item (eBay requires a leaf category).
    Fetches BOTH tree_id 0 (the standard US marketplace — Business & Industrial,
    Electronics, etc.) AND tree_id 100 (eBay Motors — a genuinely SEPARATE category
    tree, confirmed directly: syncing tree 0 alone returned 34 real top-level
    branches with zero 'Motors' among them, which is why every real auto part kept
    defaulting to the generic Business & Industrial fallback)."""
    import requests as _req
    headers = {"Authorization": f"Bearer {token}", "Accept": "application/json"}

    all_nodes = []
    all_top_level_names = []
    total_raw_bytes = 0

    for tree_id, tree_label in [("0", "US"), ("100", "Motors")]:
        r = _req.get(f"https://api.ebay.com/commerce/taxonomy/v1/category_tree/{tree_id}",
                      headers=headers, timeout=120)
        if r.status_code != 200:
            print(f"sync_ebay_categories: tree {tree_id} ({tree_label}) failed: {r.status_code} {r.text[:300]}")
            continue
        tree = r.json()
        total_raw_bytes += len(r.content)

        def walk(node, path):
            cat = node.get("category", {})
            name = cat.get("categoryName", "")
            cid = cat.get("categoryId", "")
            new_path = path + [name] if name else path
            children = node.get("childCategoryTreeNodes", [])
            if cid:
                all_nodes.append({
                    "category_id": str(cid), "name": name, "path": " > ".join(new_path),
                    "is_leaf": not bool(children), "tree_id": tree_id,
                })
            for child in children:
                walk(child, new_path)

        root = tree.get("rootCategoryNode", {})
        top_level_children = root.get("childCategoryTreeNodes", [])
        top_level_names = [c.get("category", {}).get("categoryName", "") for c in top_level_children]
        all_top_level_names.extend(f"{n} (tree {tree_id})" for n in top_level_names)
        print(f"sync_ebay_categories: tree {tree_id} ({tree_label}) raw response {len(r.content)} bytes, "
              f"{len(top_level_children)} top-level branches: {top_level_names}")

        for child in top_level_children:
            walk(child, [])

    # Upsert in batches so we don't blow request size limits
    for i in range(0, len(all_nodes), 500):
        batch = all_nodes[i:i+500]
        supabase.table("ebay_categories").upsert(batch, on_conflict="category_id").execute()
    return {"count": len(all_nodes), "top_level_names": all_top_level_names, "raw_bytes": total_raw_bytes}

def _motors_fallback_id(business_id: str) -> str:
    """The eBay Motors catch-all leaf, mirroring 26261's role in Business & Industrial.

    Was previously: require a manually-configured Settings value
    (EBAY_DEFAULT_MOTORS_CATEGORY_ID), and fall back to the WRONG id (26261, which
    is actually a Business & Industrial category) if it was never set. That's
    exactly what caused a real publish failure ('Category is not valid') on a
    user's very first Motors listing -- the setting was never configured, so a
    Business & Industrial category got submitted for a Motors-mode item.

    Fixed to require zero manual setup: if a Settings override isn't present,
    automatically queries ebay_categories for a real "Other ___" leaf category
    from tree_id 100 (Motors) specifically -- now that sync_ebay_categories
    stores which tree each category came from. Only falls back to the
    (deliberately wrong-looking but non-crashing) 26261 if no Motors categories
    have been synced at all yet, which needs a one-time /api/ebay/sync-categories
    run to resolve, not any manual category lookup."""
    try:
        mid = str(get_ebay_settings(business_id).get("EBAY_DEFAULT_MOTORS_CATEGORY_ID", "") or "").strip()
        if mid:
            return mid
    except Exception:
        pass

    try:
        res = supabase.table("ebay_categories").select("category_id,path").eq("tree_id", "100")\
            .eq("is_leaf", True).ilike("name", "Other %").limit(1).execute()
        if res.data:
            return str(res.data[0]["category_id"])
    except Exception as e:
        print(f"_motors_fallback_id: auto-lookup failed: {e}")

    print("_motors_fallback_id: no Motors 'Other' leaf found in ebay_categories (tree_id 100) -- "
          "run Sync Categories once to populate it. Falling back to 26261, which is WRONG for "
          "Motors and will fail at publish time.")
    return "26261"


def suggest_ebay_category(title: str, business_id: str, restrict: bool = True,
                           exclude_ids: list = None, mode: str = "industrial") -> dict:
    """Match an item title to an eBay leaf category using eBay's own suggestion engine.

    mode is the intake toggle and is absolute — the two lanes never cross:
        "industrial" -> tree 0, Business & Industrial branch only, falls back to 26261
        "motors"     -> tree 100 only, falls back to EBAY_DEFAULT_MOTORS_CATEGORY_ID

    Background on the long-running bug: querying both trees was already correct, but
    results were appended tree-0-first and the caller took results[0]. Tree 0 answers
    almost any query with SOMETHING under Business & Industrial, so tree 100 was
    unreachable in practice. eBay ranks each tree independently and gives no
    comparable cross-tree score, so there is no sound way to auto-arbitrate between
    them — hence an explicit mode instead of a heuristic.
    """
    import requests as _req

    mode = mode if mode in ("industrial", "motors") else "industrial"
    tree_id = "100" if mode == "motors" else "0"

    def _fallback():
        if mode == "motors":
            fid = _motors_fallback_id(business_id)
            return {"category_id": fid, "name": None, "path": None,
                    "tree_id": "100", "is_fallback": True}
        return {"category_id": "26261", "name": "Other Business & Industrial",
                "path": "Business & Industrial > Other Business & Industrial",
                "tree_id": "0", "is_fallback": True}

    try:
        token = get_ebay_access_token(business_id)
    except Exception as e:
        print(f"suggest_ebay_category: {e}")
        return _fallback()

    try:
        r = _req.get(
            f"https://api.ebay.com/commerce/taxonomy/v1/category_tree/{tree_id}/get_category_suggestions",
            headers={"Authorization": f"Bearer {token}", "Accept": "application/json"},
            params={"q": title}, timeout=15
        )
    except Exception as e:
        print(f"suggest_ebay_category: tree {tree_id} request failed: {e}")
        return _fallback()

    if r.status_code != 200:
        print(f"suggest_ebay_category: tree {tree_id} returned {r.status_code}: {r.text[:300]}")
        return _fallback()

    results = []
    for s in r.json().get("categorySuggestions", []):
        cat = s.get("category", {})
        ancestors = s.get("categoryTreeNodeAncestors", [])
        path = " > ".join(a.get("categoryName", "") for a in ancestors[::-1])
        full_path = f"{path} > {cat.get('categoryName','')}" if path else cat.get("categoryName", "")
        results.append({"category_id": cat.get("categoryId"), "name": cat.get("categoryName"),
                         "path": full_path, "tree_id": tree_id, "is_fallback": False})

    if restrict and mode == "industrial":
        # Literal TOP-LEVEL segment must be exactly "Business & Industrial" — not a
        # substring match anywhere in the path, which is what previously let Home &
        # Garden results through on words like "Motors".
        results = [x for x in results
                   if (x["path"] or "").split(" > ")[0].strip() == "Business & Industrial"]
    # mode == "motors" needs no filter: tree 100 IS eBay Motors by definition.

    if exclude_ids:
        exclude_set = {str(x) for x in exclude_ids if x}
        results = [x for x in results if str(x["category_id"]) not in exclude_set]

    # Cross-check each candidate against our OWN synced category data before
    # trusting it. eBay's live suggestion API can hand back a category that's
    # since been retired, merged, or isn't actually a leaf -- their suggestion
    # service and their publish validation aren't always in sync with each
    # other, especially with eBay actively restructuring the Motors taxonomy
    # right now. Real failure this was built for: a driveshaft U-joint kit got
    # 'Category is not valid' from AddFixedPriceItem despite (presumably)
    # matching something specific, not the generic fallback. If our own synced
    # ebay_categories doesn't confirm a candidate as a real leaf in the right
    # tree, skip it rather than submit something that's likely to fail --
    # falling back to the known-safe generic category is better than a
    # plausible-but-wrong specific one.
    if results:
        ids = [str(x["category_id"]) for x in results if x.get("category_id")]
        try:
            known = supabase.table("ebay_categories").select("category_id,is_leaf,tree_id")\
                .in_("category_id", ids).execute().data or []
            valid_leaf_ids = {str(k["category_id"]) for k in known if k.get("is_leaf") and str(k.get("tree_id")) == tree_id}
        except Exception as e:
            print(f"suggest_ebay_category: leaf cross-check failed, trusting eBay's suggestion as-is: {e}")
            valid_leaf_ids = None  # sync table unreachable -- don't block on it, just skip the extra check

        if valid_leaf_ids is not None:
            confirmed = [x for x in results if str(x["category_id"]) in valid_leaf_ids]
            if confirmed:
                results = confirmed
            else:
                print(f"suggest_ebay_category: top suggestion(s) for '{title}' "
                      f"({[x['category_id'] for x in results[:3]]}) not confirmed as a leaf in tree {tree_id} "
                      f"by our own synced data -- falling back instead of risking a failed publish. "
                      f"If ebay_categories hasn't been re-synced recently, run Sync Categories.")
                results = []

    return results[0] if results else _fallback()

_category_sync_job_status = {}  # business_id -> {"running": bool, "result": dict|None, "started_at": iso, "finished_at": iso|None}

async def _run_category_sync_background(business_id: str, token: str):
    import asyncio, datetime as _dt
    try:
        result = await asyncio.to_thread(sync_ebay_categories, token)
        _category_sync_job_status[business_id] = {
            "running": False, "result": result,
            "started_at": _category_sync_job_status.get(business_id, {}).get("started_at"),
            "finished_at": _dt.datetime.utcnow().isoformat(),
        }
    except Exception as e:
        _category_sync_job_status[business_id] = {
            "running": False, "result": {"error": str(e)},
            "started_at": _category_sync_job_status.get(business_id, {}).get("started_at"),
            "finished_at": _dt.datetime.utcnow().isoformat(),
        }

@app.get("/api/ebay/other-leaf-candidates")
async def api_other_leaf_candidates(request: Request):
    """One-time helper: lists every synced leaf category whose name starts with
    "Other " so you can find the eBay Motors catch-all without digging through
    eBay's site. Paste the right ID into Settings > EBAY_DEFAULT_MOTORS_CATEGORY_ID."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    res = supabase.table("ebay_categories").select("category_id,name,path") \
        .ilike("name", "Other %").eq("is_leaf", True).limit(500).execute()
    rows = sorted(res.data or [], key=lambda r: r.get("path") or "")
    return {"count": len(rows), "candidates": rows}

@app.post("/api/ebay/sync-categories")
async def api_sync_categories(request: Request):
    """Kicks off the category tree pull (both tree 0 and tree 100 — see
    sync_ebay_categories) as a background job instead of one long blocking request.
    Pulling two full category trees synchronously was slow enough to hit read
    timeouts depending on how eBay's API responded that moment — same fix pattern
    already used for the similarly slow active-listings sync."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    if _category_sync_job_status.get(business_id, {}).get("running"):
        return {"started": False, "already_running": True}
    settings = get_ebay_settings(business_id)
    try:
        token = get_ebay_access_token(business_id)
    except Exception as e:
        raise HTTPException(400, str(e))
    import asyncio, datetime as _dt
    _category_sync_job_status[business_id] = {
        "running": True, "result": None,
        "started_at": _dt.datetime.utcnow().isoformat(), "finished_at": None,
    }
    asyncio.create_task(_run_category_sync_background(business_id, token))
    return {"started": True}

@app.get("/api/ebay/sync-categories-status")
async def api_sync_categories_status(request: Request, response: Response):
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    response.headers["Cache-Control"] = "no-store"
    return _category_sync_job_status.get(business_id, {"running": False, "result": None, "started_at": None, "finished_at": None})

@app.post("/api/listings/{item_id}/auto-category")
async def api_auto_category(item_id: str, request: Request, broad: bool = False,
                             exclude: str = None, query: str = None, mode: str = None):
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    try:
        res = supabase.table("listings").select("title,category_mode").eq("id", item_id).limit(1).execute()
        if not res.data:
            raise HTTPException(404, "Listing not found")
        title = query if query else res.data[0].get("title", "")
        # mode comes from the Intake toggle (sent by the frontend); if the caller
        # didn't pass one, fall back to whatever this listing was scanned under.
        effective_mode = mode if mode in ("industrial", "motors") else (res.data[0].get("category_mode") or "industrial")
        # restrict is always True now — this business only ever sells in Business &
        # Industrial / eBay Motors, full stop, no "search all categories" escape hatch.
        exclude_list = [c for c in (exclude or "").split(",") if c]
        suggestion = suggest_ebay_category(title, business_id, restrict=True, exclude_ids=exclude_list, mode=effective_mode)
        if not suggestion or not suggestion.get("category_id"):
            fallback_mode_id = _motors_fallback_id(business_id) if effective_mode == "motors" else \
                (get_ebay_settings(business_id).get("EBAY_DEFAULT_CATEGORY_ID", "") or "26261")
            suggestion = {"category_id": fallback_mode_id, "name": None, "path": ""}
        supabase.table("listings").update({
            "ebay_category_id": suggestion["category_id"],
            "category_mode": effective_mode,
        }).eq("id", item_id).execute()
        return {"ok": True, **suggestion}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))

@app.get("/api/ebay/categories-tree")
async def categories_tree(request: Request, root: str = None):
    """Build a nested tree from the locally synced ebay_categories table, so the
    Categories page can render a collapsible tree instead of a flat list."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    try:
        # Detect whether is_leaf column exists yet (older syncs won't have it) — probe once up front
        use_is_leaf = True
        try:
            supabase.table("ebay_categories").select("is_leaf").limit(1).execute()
        except Exception:
            use_is_leaf = False

        select_fields = "category_id,name,path,is_leaf" if use_is_leaf else "category_id,name,path"
        all_rows = []
        offset = 0
        page_size = 1000
        while True:
            q = supabase.table("ebay_categories").select(select_fields)
            if root:
                q = q.ilike("path", f"{root}%")
            res = q.range(offset, offset + page_size - 1).execute()
            batch = res.data or []
            all_rows.extend(batch)
            if len(batch) < page_size:
                break
            offset += page_size

        tree = {}
        for row in all_rows:
            parts = [p.strip() for p in (row.get("path") or "").split(">") if p.strip()]
            if not parts:
                continue
            node = tree
            for i, part in enumerate(parts):
                if part not in node:
                    node[part] = {"__children__": {}, "__id__": None, "__leaf__": False}
                if i == len(parts) - 1:
                    node[part]["__id__"] = row["category_id"]
                    # Old schema (no is_leaf column) only ever stored leaf categories, so default True there
                    node[part]["__leaf__"] = bool(row.get("is_leaf")) if use_is_leaf else True
                node = node[part]["__children__"]
        return {"tree": tree}
    except Exception as e:
        raise HTTPException(500, str(e))

@app.get("/categories", response_class=HTMLResponse)
async def categories_page(request: Request):
    nav = get_nav_context(request)
    if nav is None:
        from fastapi.responses import RedirectResponse
        return RedirectResponse("/login", status_code=302)
    return templates.TemplateResponse("categories.html", {"request": request, "is_admin": nav["is_admin"], "account_label": nav["account_label"], "active_tab": "categories"})

@app.get("/offers", response_class=HTMLResponse)
async def offers_page(request: Request):
    """Watcher/view leaderboard for active listings -- the foundation for the
    incoming-Best-Offer decision hub described in watcher_view pricing notes
    (hold firm on offers when watch_count is elevated/rising, respond fast
    when flat). NOTE: this page surfaces watch_count/view_count only -- there
    is no eBay integration yet that pulls actual incoming Best Offer amounts
    (that's a separate build, likely the Negotiation API or GetBestOffers-
    equivalent). watch_count/view_count only start populating after
    IncludeWatchCount was added to the active-listings sync (8/7) -- existing
    rows show null until the next full sync cycle touches them."""
    nav = get_nav_context(request)
    if nav is None:
        from fastapi.responses import RedirectResponse
        return RedirectResponse("/login", status_code=302)
    return templates.TemplateResponse("offers.html", {"request": request, "is_admin": nav["is_admin"], "account_label": nav["account_label"], "active_tab": "offers"})

@app.get("/notes", response_class=HTMLResponse)
async def notes_page(request: Request):
    nav = get_nav_context(request)
    if nav is None:
        from fastapi.responses import RedirectResponse
        return RedirectResponse("/login", status_code=302)
    return templates.TemplateResponse("notes.html", {"request": request, "is_admin": nav["is_admin"], "account_label": nav["account_label"], "active_tab": "notes"})

@app.get("/uncat", response_class=HTMLResponse)
async def uncat_page(request: Request):
    """Quick-access dashboard combining two views that already exist elsewhere
    (Financials' Uncategorized Orders, Lots' Uncategorized Listings) -- both
    stay fully intact in their original homes; this just pulls the same two
    existing endpoints (/api/financials, /api/acquisitions) so both action-item
    lists are reachable from one tab instead of navigating to two different
    pages."""
    nav = get_nav_context(request)
    if nav is None:
        from fastapi.responses import RedirectResponse
        return RedirectResponse("/login", status_code=302)
    return templates.TemplateResponse("uncat.html", {"request": request, "is_admin": nav["is_admin"], "account_label": nav["account_label"], "active_tab": "uncat"})

_OFFERS_SORT_COLUMNS = {
    "title": "title", "sku": "sku", "price": "price",
    "watchers": "watch_count", "views": "view_count",
    "watch_view": "watch_per_view_pct",
    "views_day": "view_count",  # views_per_day = view_count / constant -- same order, no separate column needed
}

@app.get("/api/offers")
async def list_offers(request: Request, limit: int = 150, sort: str = "watchers", dir: str = "desc",
                       min_price: float = None, max_price: float = None):
    """Backs the lower (watcher/view leaderboard) section of the Offers page:
    top `limit` Active listings, sorted by whichever column the user clicked,
    optionally filtered to a price range. REAL BUG FIXED 8/8: this used to
    pull every single Active listing (~4,700 rows) into Python and sort
    there -- Postgres does ORDER BY + LIMIT (and now the price WHERE)
    directly, so this only ever transfers `limit` rows, never the full
    table. `sort` is a friendly column key (see _OFFERS_SORT_COLUMNS), not a
    raw DB column name, so the frontend can't accidentally sort on something
    unindexed/unsafe. watch_per_view_pct is a Postgres GENERATED column
    (stored) specifically so it can be ORDER BY'd directly like any other
    column instead of requiring a full-table Python sort for that one case."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    limit = max(1, min(limit, 500))
    sort_col = _OFFERS_SORT_COLUMNS.get(sort, "watch_count")
    descending = dir != "asc"
    # Tiebreaker keeps ordering stable/sensible when the primary column has
    # many equal values (e.g. many items tied on 0 watchers).
    tiebreak = "view_count" if sort_col != "view_count" else "watch_count"

    def _apply_filters(q):
        if min_price is not None:
            q = q.gte("price", min_price)
        if max_price is not None:
            q = q.lte("price", max_price)
        return q

    query = (supabase.table("ebay_listing_status")
             .select("item_id,sku,title,price,watch_count,view_count,watch_per_view_pct,gallery_url")
             .eq("business_id", business_id).eq("listing_status", "Active"))
    query = _apply_filters(query)
    # REAL BUG FIXED 8/8, TWO SEPARATE ISSUES:
    # (1) nullsfirst=False in this Supabase client does NOT mean "nulls last" --
    #     it only omits the nulls directive entirely when False, leaving
    #     Postgres's own actual default in place: NULL sorts as larger than
    #     any real value, in BOTH directions -- so DESC puts nulls first,
    #     ASC puts nulls last. With most listings missing watch_count/
    #     view_count data, "Sort by Watchers" was showing null-data rows at
    #     the very top of a 150-row leaderboard, burying every item that
    #     actually had watcher data. Correct behavior for a metric column is
    #     nulls-always-last regardless of direction ("no data" isn't a value
    #     to rank as highest OR lowest, it just shouldn't be at the top) --
    #     so NULLS LAST is made explicit on both directions, not left to
    #     Postgres's default, and not by dropping the rows either: a
    #     null-metric row still shows in the table, just trailing the real
    #     values instead of leading them.
    # (2) Chaining .order() twice sends TWO separate order= query params
    #     (verified directly: produces "order=a.desc&order=b.desc" in the
    #     request), not PostgREST's documented single comma-joined
    #     multi-column syntax ("order=a.desc,b.desc") -- meaning the
    #     "tiebreak" second .order() call was never confirmed to actually
    #     combine with the first as intended. Built as one explicit combined
    #     order string instead, matching PostgREST's real documented format,
    #     with nullslast made explicit on the tiebreak column too.
    order_str = f"{sort_col}.{'desc' if descending else 'asc'}.nullslast,{tiebreak}.desc.nullslast"
    query.params = query.params.add("order", order_str)
    top_rows = (query.limit(limit).execute()).data or []

    total_query = _apply_filters(supabase.table("ebay_listing_status")
                                  .select("item_id", count="exact")
                                  .eq("business_id", business_id).eq("listing_status", "Active"))
    total_active = (total_query.limit(1).execute()).count or 0

    populated_query = _apply_filters(supabase.table("ebay_listing_status")
                                      .select("item_id", count="exact")
                                      .eq("business_id", business_id).eq("listing_status", "Active")
                                      .or_("watch_count.not.is.null,view_count.not.is.null"))
    populated = (populated_query.limit(1).execute()).count or 0

    # views_per_day still derived here, on the already-limited top_rows only.
    for r in top_rows:
        vc = r.get("view_count")
        r["views_per_day"] = round(vc / _ANALYTICS_WINDOW_DAYS, 2) if vc is not None else None

    return {"offers": top_rows, "total_active": total_active, "populated": populated, "shown": len(top_rows)}

class NoteCreate(BaseModel):
    section: str
    content: str

@app.get("/api/notes")
async def list_notes(request: Request, section: str):
    """Notes are a plain, permanent log -- date + whatever was typed, nothing
    fancier. Newest first, scoped to one section at a time (starting with
    'audit' -- more sections can be added later just by using a new section
    key, no schema change needed)."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    res = (supabase.table("notes").select("id,content,created_at")
           .eq("business_id", business_id).eq("section", section)
           .order("created_at", desc=True).execute())
    return {"notes": res.data or []}

@app.post("/api/notes")
async def add_note(body: NoteCreate, request: Request):
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    content = (body.content or "").strip()
    if not content:
        raise HTTPException(400, "content is required")
    section = (body.section or "").strip()
    if not section:
        raise HTTPException(400, "section is required")
    res = (supabase.table("notes")
           .insert({"business_id": business_id, "section": section, "content": content})
           .execute())
    return {"ok": True, "note": (res.data or [{}])[0]}

@app.delete("/api/notes/{note_id}")
async def delete_note(note_id: int, request: Request):
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    supabase.table("notes").delete().eq("id", note_id).eq("business_id", business_id).execute()
    return {"ok": True}

@app.post("/api/notes/files")
async def upload_audit_file(request: Request, file: UploadFile = File(...), section: str = "audit"):
    """PDF attachments for the Notes page (Audit Notes etc.) — stored in the
    private 'audit-docs' Supabase bucket, indexed in audit_files, downloadable
    any time later. Same storage pattern as part-photos/auction-pdfs."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    fname = (file.filename or "document.pdf").strip()
    if not fname.lower().endswith(".pdf"):
        raise HTTPException(400, "Only PDF files are accepted here")
    content = await file.read()
    if not content:
        raise HTTPException(400, "File is empty")
    import uuid as _uuid
    file_id = str(_uuid.uuid4())
    storage_path = f"{business_id}/{file_id}.pdf"
    supabase.storage.from_("audit-docs").upload(
        storage_path, content, {"content-type": "application/pdf"}
    )
    res = supabase.table("audit_files").insert({
        "id": file_id, "business_id": business_id, "section": section,
        "filename": fname, "storage_path": storage_path, "size_bytes": len(content),
    }).execute()
    return {"ok": True, "file": (res.data or [{}])[0]}

@app.get("/api/notes/files")
async def list_audit_files(request: Request, section: str = "audit"):
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    res = (supabase.table("audit_files").select("id,filename,size_bytes,uploaded_at")
           .eq("business_id", business_id).eq("section", section)
           .order("uploaded_at", desc=True).execute())
    return {"files": res.data or []}

@app.get("/api/notes/files/{file_id}/download")
async def download_audit_file(file_id: str, request: Request):
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    res = (supabase.table("audit_files").select("filename,storage_path")
           .eq("id", file_id).eq("business_id", business_id).limit(1).execute())
    if not res.data:
        raise HTTPException(404, "File not found")
    row = res.data[0]
    data = supabase.storage.from_("audit-docs").download(row["storage_path"])
    from fastapi.responses import Response as _Resp
    safe_name = row["filename"].replace('"', "")
    return _Resp(content=data, media_type="application/pdf",
                 headers={"Content-Disposition": f'inline; filename="{safe_name}"'})

@app.delete("/api/notes/files/{file_id}")
async def delete_audit_file(file_id: str, request: Request):
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    res = (supabase.table("audit_files").select("storage_path")
           .eq("id", file_id).eq("business_id", business_id).limit(1).execute())
    if res.data:
        try:
            supabase.storage.from_("audit-docs").remove([res.data[0]["storage_path"]])
        except Exception:
            pass
        supabase.table("audit_files").delete().eq("id", file_id).eq("business_id", business_id).execute()
    return {"ok": True}

@app.get("/api/ebay/category-search")
async def ebay_category_search(q: str, request: Request):
    """Look up valid LEAF category IDs by keyword, using the token already saved in Settings."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    import requests as _req
    settings = get_ebay_settings(business_id)
    try:
        token = get_ebay_access_token(business_id)
    except Exception as e:
        raise HTTPException(400, str(e))
    headers = {"Authorization": f"Bearer {token}", "Accept": "application/json"}
    r = _req.get(
        "https://api.ebay.com/commerce/taxonomy/v1/category_tree/0/get_category_suggestions",
        headers=headers, params={"q": q}, timeout=15
    )
    if r.status_code != 200:
        raise HTTPException(500, f"{r.status_code}: {r.text}")
    data = r.json()
    out = []
    for s in data.get("categorySuggestions", []):
        cat = s.get("category", {})
        path = " > ".join(a.get("categoryName","") for a in s.get("categoryTreeNodeAncestors", [])[::-1])
        out.append({"id": cat.get("categoryId"), "name": cat.get("categoryName"), "path": path})
    return {"results": out}

@app.get("/api/ebay/policies")
async def list_ebay_policies(request: Request):
    """Fetch payment/return/fulfillment policy IDs using the token already saved in Settings."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    import requests as _req
    settings = get_ebay_settings(business_id)
    try:
        token = get_ebay_access_token(business_id)
    except Exception as e:
        raise HTTPException(400, str(e))
    headers = {"Authorization": f"Bearer {token}", "Accept": "application/json"}
    base = "https://api.ebay.com/sell/account/v1"
    out = {}
    mapping = {
        "payment_policy": ("paymentPolicies", "paymentPolicyId"),
        "return_policy": ("returnPolicies", "returnPolicyId"),
        "fulfillment_policy": ("fulfillmentPolicies", "fulfillmentPolicyId"),
    }
    for kind, (list_key, id_key) in mapping.items():
        r = _req.get(f"{base}/{kind}?marketplace_id=EBAY_US", headers=headers, timeout=15)
        if r.status_code != 200:
            out[kind] = {"error": f"{r.status_code}: {r.text}"}
            continue
        data = r.json()
        out[kind] = [{"name": p.get("name"), "id": p.get(id_key)} for p in data.get(list_key, [])]
    return out

@app.post("/api/listings/{item_id}/rescan")
async def rescan_listing(item_id: str, request: Request):
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    try:
        res = supabase.table("listings").select("photo_id").eq("id", item_id).limit(1).execute()
        pid = (res.data[0].get("photo_id", "") if res.data else "")
        if pid:
            grp = supabase.table("group_photos").select("group_id").eq("photo_id", pid).limit(1).execute()
            if grp.data:
                gid = grp.data[0]["group_id"]
                supabase.table("listing_groups").update({"status": "pending"}).eq("id", gid).execute()
        supabase.table("listings").update({"status": "pending", "title": "Scanning..."}).eq("id", item_id).execute()
        return {"ok": True}
    except Exception as e:
        raise HTTPException(500, str(e))

class RestoreItems(BaseModel):
    ids: Optional[list] = None  # None/omitted = restore ALL archived items

@app.post("/api/listings/restore")
async def restore_listings(request: Request, body: RestoreItems):
    business_id = require_auth(request)
    if not business_id:
        return JSONResponse({"error": "unauthorized"}, status_code=401)
    try:
        q = supabase.table("listings").update({"status": "pending"}).eq("business_id", business_id).eq("status", "archived")
        if body.ids:
            q = q.in_("id", [str(i) for i in body.ids])
        res = q.execute()
        return {"ok": True, "count": len(res.data or [])}
    except Exception as e:
        raise HTTPException(500, str(e))

@app.get("/financials", response_class=HTMLResponse)
async def financials_page(request: Request):
    nav = get_nav_context(request)
    if nav is None:
        from fastapi.responses import RedirectResponse
        return RedirectResponse("/login", status_code=302)
    return templates.TemplateResponse("financials.html", {"request": request, "is_admin": nav["is_admin"], "account_label": nav["account_label"], "active_tab": "financials"})

@app.get("/analytics", response_class=HTMLResponse)
async def analytics_page(request: Request):
    nav = get_nav_context(request)
    if nav is None:
        from fastapi.responses import RedirectResponse
        return RedirectResponse("/login", status_code=302)
    return templates.TemplateResponse("analytics.html", {"request": request, "is_admin": nav["is_admin"], "account_label": nav["account_label"], "active_tab": "analytics"})

@app.get("/acquisitions", response_class=HTMLResponse)
async def acquisitions_page(request: Request):
    nav = get_nav_context(request)
    if nav is None:
        from fastapi.responses import RedirectResponse
        return RedirectResponse("/login", status_code=302)
    return templates.TemplateResponse("acquisitions.html", {"request": request, "is_admin": nav["is_admin"], "account_label": nav["account_label"], "active_tab": "lots"})

@app.get("/inventory", response_class=HTMLResponse)
async def inventory_page(request: Request):
    nav = get_nav_context(request)
    if nav is None:
        from fastapi.responses import RedirectResponse
        return RedirectResponse("/login", status_code=302)
    return templates.TemplateResponse("inventory.html", {"request": request, "is_admin": nav["is_admin"], "account_label": nav["account_label"], "active_tab": "inventory"})

@app.get("/shopify-sync", response_class=HTMLResponse)
async def shopify_sync_page(request: Request):
    nav = get_nav_context(request)
    if nav is None:
        from fastapi.responses import RedirectResponse
        return RedirectResponse("/login", status_code=302)
    return templates.TemplateResponse("shopify_sync.html", {"request": request, "is_admin": nav["is_admin"], "account_label": nav["account_label"], "active_tab": "shopify"})


@app.get("/browse-search", response_class=HTMLResponse)
async def browse_search_page(request: Request):
    nav = get_nav_context(request)
    return templates.TemplateResponse("browse_search.html", {"request": request, "is_admin": nav["is_admin"], "account_label": nav["account_label"], "active_tab": "browse_search"})


@app.get("/api/saved-searches")
async def saved_searches_list(request: Request, response: Response):
    """The permanent list of saved search terms — this is the primary content of the tab."""
    business_id = get_business_id(request)
    if not business_id:
        raise HTTPException(401, "Not logged in")
    res = supabase.table("saved_searches").select("*").eq("business_id", business_id) \
        .order("created_at", desc=True).execute()
    return {"searches": res.data}


@app.post("/api/saved-searches/run")
async def saved_searches_run(request: Request, body: dict = Body(...)):
    """Adds the term to the permanent saved-searches list (if new) and runs it against
    eBay's Browse API, filtered server-side to items listed on the given date via the
    itemStartDate filter — eBay only returns that day's listings, nothing is pulled and
    filtered afterward. Results are stored in browse_search_results."""
    import requests as _req
    from datetime import datetime, timezone

    business_id = get_business_id(request)
    if not business_id:
        raise HTTPException(401, "Not logged in")

    query = (body.get("query") or "").strip()
    if not query:
        raise HTTPException(400, "query is required")

    listed_date = (body.get("listed_date") or "").strip()
    if not listed_date:
        listed_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    min_price = body.get("min_price")
    if min_price in ("", None):
        min_price = None
    else:
        try:
            min_price = float(min_price)
        except (TypeError, ValueError):
            raise HTTPException(400, "min_price must be a number")

    start_iso = f"{listed_date}T00:00:00Z"
    end_iso = f"{listed_date}T23:59:59Z"

    token = get_ebay_access_token(business_id)
    headers = {
        "Authorization": f"Bearer {token}",
        "X-EBAY-C-MARKETPLACE-ID": "EBAY_US",
    }

    filter_parts = [f"itemStartDate:[{start_iso}..{end_iso}]", "itemLocationCountry:US"]
    if min_price is not None:
        filter_parts.append(f"price:[{min_price}..]")
        filter_parts.append("priceCurrency:USD")
    filter_str = ",".join(filter_parts)

    all_items = []
    offset = 0
    page_size = 200
    max_items = 1000  # hard ceiling per run regardless of how many eBay reports
    ebay_warnings = []
    while len(all_items) < max_items:
        r = _req.get(
            f"{EBAY_API_BASE}/buy/browse/v1/item_summary/search",
            headers=headers,
            params={
                "q": query,
                "limit": page_size,
                "offset": offset,
                "sort": "newlyListed",
                "filter": filter_str,
            },
            timeout=20,
        )
        if r.status_code != 200:
            raise HTTPException(502, f"eBay Browse API error ({r.status_code}): {r.text[:300]}")
        data = r.json()
        if data.get("warnings"):
            ebay_warnings.extend(data["warnings"])
        items = data.get("itemSummaries", [])
        if not items:
            break
        all_items.extend(items)
        offset += page_size
        if len(items) < page_size:
            break

    # If eBay rejected a filter, it silently falls back to unfiltered results
    # rather than erroring — surface that instead of storing results that don't
    # actually match the requested date/price.
    if ebay_warnings:
        raise HTTPException(502, f"eBay rejected a filter: {ebay_warnings}")

    rows = []
    for it in all_items[:max_items]:
        creation_date = it.get("itemCreationDate") or ""
        if not creation_date.startswith(listed_date):
            continue  # extra guard: only store items actually listed on the requested date
        price = it.get("price", {}) or {}
        price_value = price.get("value")
        if min_price is not None and price_value is not None and float(price_value) < min_price:
            continue  # extra guard: only store items at/above the requested min price
        image = it.get("image", {}) or {}
        seller = it.get("seller", {}) or {}
        categories = it.get("categories", []) or []
        rows.append({
            "business_id": business_id,
            "query": query,
            "item_id": it.get("itemId"),
            "title": it.get("title"),
            "price": price_value,
            "currency": price.get("currency"),
            "condition": it.get("condition"),
            "item_web_url": it.get("itemWebUrl"),
            "image_url": image.get("imageUrl"),
            "seller_username": seller.get("username"),
            "category_id": categories[0].get("categoryId") if categories else None,
            "item_creation_date": it.get("itemCreationDate"),
        })

    if rows:
        supabase.table("browse_search_results").upsert(rows, on_conflict="business_id,query,item_id").execute()

    now_iso = datetime.now(timezone.utc).isoformat()
    supabase.table("saved_searches").upsert({
        "business_id": business_id,
        "query": query,
        "last_run_at": now_iso,
        "last_run_date": listed_date,
        "last_result_count": len(rows),
        "min_price": min_price,
    }, on_conflict="business_id,query").execute()

    return {"query": query, "listed_date": listed_date, "min_price": min_price, "count": len(rows)}


@app.get("/api/saved-searches/results")
async def saved_searches_results(request: Request, response: Response, q: str, listed_date: str = None):
    """Stored results for one saved search — optionally narrowed to a specific listed date."""
    business_id = get_business_id(request)
    if not business_id:
        raise HTTPException(401, "Not logged in")

    query_builder = supabase.table("browse_search_results").select("*") \
        .eq("business_id", business_id).eq("query", q)
    if listed_date:
        query_builder = query_builder.gte("item_creation_date", f"{listed_date}T00:00:00.000Z") \
            .lte("item_creation_date", f"{listed_date}T23:59:59.999Z")
    res = query_builder.order("fetched_at", desc=True).limit(500).execute()
    return {"rows": res.data}


@app.delete("/api/saved-searches/{search_id}")
async def saved_searches_delete(request: Request, search_id: int):
    business_id = get_business_id(request)
    if not business_id:
        raise HTTPException(401, "Not logged in")
    supabase.table("saved_searches").delete().eq("id", search_id).eq("business_id", business_id).execute()
    return {"deleted": search_id}

class AcquisitionCreate(BaseModel):
    sku: str
    name: Optional[str] = None
    payment_method: Optional[str] = None
    date: Optional[str] = None
    cost: Optional[float] = None
    cash: Optional[float] = None
    notes: Optional[str] = None

def apply_acquisition_profits(business_id: str) -> dict:
    """Computes eBay/Total_Payouts/Profit/ROI for every acquisition — done ENTIRELY
    inside Postgres via the recalculate_acquisition_profits() function (see
    create_recalculate_function.sql), not fetched into Python and looped over.
    This is what 'calculated in the database' should have meant from the start —
    one atomic SQL statement, no pagination limits, no partial-upsert failures."""
    res = supabase.rpc("recalculate_acquisition_profits", {"biz_id": business_id}).execute()
    updated = res.data if isinstance(res.data, int) else 0
    _apply_cash_to_profit(business_id)
    _apply_shopify_sales_to_profit(business_id)
    return {"updated": updated}

def _apply_cash_to_profit(business_id: str):
    """Total_Payouts is meant to be the genuine grand total of money received
    across every channel -- eBay, Shopify, and Cash -- but Cash was previously
    folded into Profit only, never into the Total_Payouts column itself, so a
    lot with real cash income (e.g. AM1: $5,400 cash) showed a Total_Payouts
    figure that silently excluded it. Fixed: cash now gets added into
    total_payouts here too, before _apply_shopify_sales_to_profit adds Shopify
    on top of it. Profit math is unchanged (total_payouts - cost, with cash
    now already included in total_payouts instead of added separately) --
    same final number, just an honest Total_Payouts column.

    Still safe to run any number of times: reads the RPC's fresh eBay-only
    total_payouts each time (never reads back a previously-cash-adjusted
    value), so repeated Recalculate runs can't compound cash on top of itself."""
    start = 0
    while True:
        page = supabase.table("acquisitions").select("id,cost,cash,total_payouts")\
            .eq("business_id", business_id).range(start, start + 999).execute().data or []
        for row in page:
            cost = row.get("cost") or 0
            cash = row.get("cash") or 0
            ebay_total_payouts = row.get("total_payouts") or 0
            total_payouts = ebay_total_payouts + cash
            profit = total_payouts - cost
            roi_pct = round(profit / cost * 100, 2) if cost else None
            supabase.table("acquisitions").update({
                "total_payouts": total_payouts, "profit": profit, "roi_pct": roi_pct,
            }).eq("id", row["id"]).execute()
        if len(page) < 1000:
            break
        start += 1000

def _apply_shopify_sales_to_profit(business_id: str):
    """recalculate_acquisition_profits() only sums eBay payouts into total_payouts
    (the Postgres function has no SQL source in this repo, but its own naming and
    behavior confirm it's eBay-only), so Shopify sales never counted toward a lot's
    Total_Payouts/Profit/ROI even though every Shopify order is already synced into
    the same local `orders` table eBay orders use. This adds each lot's matching
    Shopify orders.final_net (already shipping-cost-adjusted) both as its own visible
    shopify_payouts column (for the Lots page's Shopify column, shown between eBay
    and Total_Payouts) and folded into total_payouts itself, so Total_Payouts is a
    genuine eBay+Shopify grand total. Same layering approach as _apply_cash_to_profit
    — reads whatever total_payouts/profit that function just left (RPC baseline +
    cash), so this always adds Shopify on top of a fresh, not previously-corrected,
    number rather than compounding across repeated Recalculate runs."""
    acquisitions = []
    import datetime as _dt
    start = 0
    while True:
        page = supabase.table("acquisitions").select("id,sku,cost,profit,total_payouts,became_green_at")\
            .eq("business_id", business_id).range(start, start + 999).execute().data or []
        acquisitions.extend(page)
        if len(page) < 1000:
            break
        start += 1000
    if not acquisitions:
        return

    known_lot_skus = {a["sku"] for a in acquisitions if a.get("sku")}
    shopify_sales_by_prefix = {}
    start = 0
    while True:
        page = supabase.table("orders").select("sku,final_net").eq("business_id", business_id)\
            .eq("platform", "Shopify").range(start, start + 999).execute().data or []
        for row in page:
            sku = row.get("sku") or ""
            if not sku:
                continue
            prefix = _lot_prefix(sku)
            if prefix in known_lot_skus:
                shopify_sales_by_prefix[prefix] = shopify_sales_by_prefix.get(prefix, 0) + (row.get("final_net") or 0)
        if len(page) < 1000:
            break
        start += 1000

    for a in acquisitions:
        sales = shopify_sales_by_prefix.get(a.get("sku")) or 0
        cost = a.get("cost") or 0
        profit = (a.get("profit") or 0) + sales
        total_payouts = (a.get("total_payouts") or 0) + sales
        roi_pct = round(profit / cost * 100, 2) if cost else None

        # "Green" = profit > 1 (the user's own simple flag). became_green_at marks
        # the most recent time this lot crossed from not-green into green -- set
        # once on the transition, kept as-is while it stays green, and CLEARED if
        # profit drops back to <= 1 (so a later re-crossing gets a fresh, correct
        # timestamp rather than an old stale one). This is what lets Green Revenue
        # be filtered by date range correctly without a full daily snapshot table:
        # only count a lot's sales from became_green_at onward, never sales that
        # happened while it was still red.
        was_green_at = a.get("became_green_at")
        if profit > 1:
            became_green_at = was_green_at or _dt.datetime.utcnow().isoformat()
        else:
            became_green_at = None

        supabase.table("acquisitions").update({
            "profit": profit,
            "roi_pct": roi_pct,
            "shopify_payouts": sales,
            "total_payouts": total_payouts,
            "became_green_at": became_green_at,
        }).eq("id", a["id"]).execute()

@app.post("/api/analytics/refresh-snapshot-now")
async def refresh_snapshot_now(request: Request):
    """Manually runs today's analytics_snapshot computation immediately, instead
    of waiting for the overnight job (00:10 UTC). Useful right after adding a
    new column to analytics_snapshots (like ytd_cash) -- the existing row for
    today won't have real data in that new column until this runs again."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    result = run_daily_analytics_snapshot_for_business(business_id)
    return result

# ── Cash Payments Ledger ──────────────────────────────────────────
# Replaces the single acquisitions.cash field with a real dated ledger, per the
# user's explicit ask for a proper edit-history/date field for cash rather than
# spreading estimates around a value with no real date attached.
#
# A payment row is EITHER:
#  - a precise payment: payment_date is set, spread_start_date is null. Used
#    for every NEW cash entry added via "Add Cash Pmt" going forward, and for
#    the three lots the user explicitly named as entered this month (11R, R66,
#    AM1) during the one-time migration below.
#  - a spread entry: spread_start_date is set (the lot's own acquisitions.date),
#    payment_date is null. Used only by the one-time historical migration, for
#    every OTHER lot's pre-existing cash value -- since we don't know the real
#    date that cash arrived, it's spread evenly per day from the lot's purchase
#    date through a FIXED CUTOFF (CASH_SPREAD_CUTOFF_DATE below), not an
#    ever-rolling "today". Originally this really did roll forward forever (a
#    lot purchased Oct 1 would spread Oct 1 -> whenever it's queried, growing
#    every day) -- deliberately changed after the user decided that was the
#    actual source of the confusion: every new period kept picking up a few
#    more dollars from ~80 old lots' spread payments, with no way to tell "real
#    new cash" apart from "old spread still trickling in" without a breakdown
#    tool. Freezing the spread window at a fixed date ends this permanently:
#    from CASH_SPREAD_CUTOFF_DATE onward, spread payments contribute ZERO to
#    any range -- only new precise payments (via Add Cash Pmt) count for
#    anything after that date, no more ambiguity, ever.
CASH_SPREAD_CUTOFF_DATE = "2026-06-30"

def _cash_contribution_in_range(amount: float, spread_start_date: str, range_start: str, range_end: str, today_str: str, extra_min_date: str = None) -> float:
    """For a spread-type payment: how much of it falls inside [range_start,
    range_end], given it's spread evenly across [spread_start_date, today].
    extra_min_date (optional) further restricts which days count -- e.g. only
    days on/after a lot's became_green_at, for Green Revenue -- without
    changing the daily RATE itself, which is still based on the lot's full
    original spread window."""
    import datetime as _dt3
    start = max(spread_start_date, "0001-01-01")
    total_days = (_dt3.date.fromisoformat(today_str) - _dt3.date.fromisoformat(start)).days + 1
    if total_days <= 0:
        return 0.0
    daily_rate = amount / total_days
    overlap_start = max(start, range_start, extra_min_date) if extra_min_date else max(start, range_start)
    overlap_end = min(today_str, range_end)
    if overlap_start > overlap_end:
        return 0.0
    overlap_days = (_dt3.date.fromisoformat(overlap_end) - _dt3.date.fromisoformat(overlap_start)).days + 1
    return daily_rate * overlap_days

def _compute_cash_in_range(business_id: str, start_date_str: str, end_date_str: str, only_skus: set = None, min_date_by_sku: dict = None) -> float:
    """only_skus + min_date_by_sku (both optional) restrict this to a specific
    set of lots (e.g. currently-green ones) and, per lot, an extra lower bound
    on which days count -- used by Green Revenue to only count cash from a
    lot's became_green_at onward, same rule as its order revenue."""
    import datetime as _dt3
    today_str = min(_dt3.datetime.utcnow().strftime("%Y-%m-%d"), CASH_SPREAD_CUTOFF_DATE)
    rows = _fetch_all_for_business(business_id, "cash_payments", "sku,amount,payment_date,spread_start_date")
    total = 0.0
    for r in rows:
        sku = r.get("sku")
        if only_skus is not None and sku not in only_skus:
            continue
        extra_min = (min_date_by_sku or {}).get(sku)
        amount = r.get("amount") or 0
        if r.get("payment_date"):
            if start_date_str <= r["payment_date"] <= end_date_str and (not extra_min or r["payment_date"] >= extra_min):
                total += amount
        elif r.get("spread_start_date"):
            total += _cash_contribution_in_range(amount, r["spread_start_date"], start_date_str, end_date_str, today_str, extra_min_date=extra_min)
    return round(total, 2)

def _compute_cash_for_year(business_id: str, year: int) -> float:
    """Cash attributed to a specific calendar year, per explicit request:
    - a PRECISE payment counts toward the year of its own payment_date
    - a SPREAD payment (an old lot's cash with no real date attached) counts
      toward the year the LOT was purchased (spread_start_date's year) --
      using the FULL amount, not a prorated slice. This is only correct
      because cash spreading is now frozen at CASH_SPREAD_CUTOFF_DATE: past
      that date every spread payment's contribution is fully vested (its
      total across all time equals its full amount, nothing left to prorate
      further), so attributing the whole thing to the lot's purchase year is
      exact, not an approximation."""
    rows = _fetch_all_for_business(business_id, "cash_payments", "amount,payment_date,spread_start_date")
    total = 0.0
    for r in rows:
        amount = r.get("amount") or 0
        d = r.get("payment_date") or r.get("spread_start_date")
        if d and int(str(d)[:4]) == year:
            total += amount
    return round(total, 2)

@app.get("/api/analytics/cash-breakdown")
async def api_cash_breakdown(request: Request, start: str, end: str):
    """Every single cash_payments row and its actual computed contribution to
    [start, end] -- built specifically to answer 'how are you getting $X' with
    real row-level data instead of more explanation. Also surfaces a
    likely_duplicate_rows list: pairs of rows for the SAME sku, SAME amount,
    SAME type, created within a few seconds of each other -- the actual
    signature of an accidental double-insert (e.g. the one-time migration
    running twice). A lot legitimately having several DIFFERENT cash payments
    over time is normal, not flagged -- confirmed against real data that the
    old 'any sku with >1 row' check was too blunt and flagged legitimate
    multi-payment lots as false positives."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    import datetime as _dt3
    today_str = min(_dt3.datetime.utcnow().strftime("%Y-%m-%d"), CASH_SPREAD_CUTOFF_DATE)
    rows = _fetch_all_for_business(business_id, "cash_payments", "sku,amount,payment_date,spread_start_date,created_at")

    detail = []
    for r in rows:
        sku = r.get("sku")
        amount = r.get("amount") or 0
        if r.get("payment_date"):
            contribution = amount if start <= r["payment_date"] <= end else 0.0
            row_type = "precise"
            date_info = r["payment_date"]
        else:
            contribution = _cash_contribution_in_range(amount, r.get("spread_start_date") or today_str, start, end, today_str)
            row_type = "spread"
            date_info = f"{r.get('spread_start_date')} -> {today_str} (rolling)"
        detail.append({
            "sku": sku, "type": row_type, "full_amount": amount, "date_info": date_info,
            "contribution_to_range": round(contribution, 2), "created_at": r.get("created_at"),
        })

    # Real duplicate signature: same sku + same amount + same type, created
    # within 10 seconds of each other (a migration or double-click firing twice
    # looks like this; a genuinely different payment added later never does).
    likely_duplicate_rows = []
    by_key = {}
    for r in rows:
        key = (r.get("sku"), r.get("amount"), "precise" if r.get("payment_date") else "spread")
        by_key.setdefault(key, []).append(r.get("created_at") or "")
    for (sku, amount, row_type), timestamps in by_key.items():
        timestamps = sorted(t for t in timestamps if t)
        for i in range(1, len(timestamps)):
            t1 = _dt3.datetime.fromisoformat(timestamps[i-1].replace("Z", "+00:00"))
            t2 = _dt3.datetime.fromisoformat(timestamps[i].replace("Z", "+00:00"))
            if (t2 - t1).total_seconds() < 10:
                likely_duplicate_rows.append({"sku": sku, "amount": amount, "type": row_type})
                break

    detail.sort(key=lambda d: d["contribution_to_range"], reverse=True)
    return {
        "start": start, "end": end,
        "total_contribution": round(sum(d["contribution_to_range"] for d in detail), 2),
        "likely_duplicate_rows": likely_duplicate_rows,
        "rows": detail,
    }

class AddCashPayment(BaseModel):
    sku: str
    amount: float
    date: str  # "YYYY-MM-DD"

@app.post("/api/acquisitions/add-cash-payment")
async def add_cash_payment(body: AddCashPayment, request: Request):
    """The 'Add Cash Pmt' action: sku must match a real, existing lot (binary --
    rejected otherwise, can't attach cash to a lot that doesn't exist). Inserts
    a precise, dated payment, then updates that lot's acquisitions.cash to the
    new running total and re-runs the normal profit recalculation so
    profit/total_payouts/roi_pct/became_green_at all reflect it immediately --
    same pipeline every other cash change already goes through."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")

    lot = supabase.table("acquisitions").select("id,sku,cash").eq("business_id", business_id).eq("sku", body.sku).limit(1).execute()
    if not lot.data:
        raise HTTPException(400, f"No lot found with SKU '{body.sku}' -- cash can only be added to an existing lot")

    supabase.table("cash_payments").insert({
        "business_id": business_id, "sku": body.sku, "amount": body.amount, "payment_date": body.date,
    }).execute()

    payment_rows = supabase.table("cash_payments").select("amount").eq("business_id", business_id).eq("sku", body.sku).execute().data or []
    new_total_cash = round(sum(r.get("amount") or 0 for r in payment_rows), 2)
    supabase.table("acquisitions").update({"cash": new_total_cash}).eq("id", lot.data[0]["id"]).execute()

    apply_acquisition_profits(business_id)
    return {"ok": True, "sku": body.sku, "new_total_cash": new_total_cash}

@app.get("/api/acquisitions/debug-sku/{sku}")
async def debug_acquisition_sku(sku: str, request: Request):
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")

    def _fetch_all(table, select_cols):
        all_rows = []
        page_size = 1000
        start = 0
        while True:
            res = supabase.table(table).select(select_cols).eq("business_id", business_id)\
                .range(start, start + page_size - 1).execute()
            page = res.data or []
            all_rows.extend(page)
            if len(page) < page_size:
                break
            start += page_size
        return all_rows

    acquisitions = _fetch_all("acquisitions", "*")
    matching_acq_rows = [a for a in acquisitions if a.get("sku") == sku]

    skus = list(set(a["sku"] for a in acquisitions if a.get("sku")))
    orders = _fetch_all("orders", "sku,final_net")

    matched_orders = []
    for row in orders:
        order_sku = row.get("sku") or ""
        if "-" not in order_sku:
            continue
        prefix = order_sku.split("-", 1)[0]
        if prefix == sku:
            matched_orders.append(row)

    computed_total = sum((r.get("final_net") or 0) for r in matched_orders)

    return {
        "sku_in_acquisitions_skus_set": sku in skus,
        "total_acquisitions_fetched": len(acquisitions),
        "total_orders_fetched": len(orders),
        "matching_acquisition_rows": matching_acq_rows,
        "matched_order_count": len(matched_orders),
        "computed_total": round(computed_total, 2),
        "sample_matched_orders": matched_orders[:5],
    }

class AcquisitionEdit(BaseModel):
    id: str
    sku: str
    name: Optional[str] = None
    payment_method: Optional[str] = None
    date: Optional[str] = None
    cost: Optional[float] = None
    cash: Optional[float] = None

ALLOWED_ACQ_EDIT_FIELDS = {"sku", "name", "payment_method", "date", "cost", "cash"}

class AcquisitionSingleEdit(BaseModel):
    id: str
    field: str
    value: Optional[str] = None

@app.post("/api/acquisitions/edit-single")
async def edit_acquisition_single(request: Request, body: AcquisitionSingleEdit):
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    if body.field not in ALLOWED_ACQ_EDIT_FIELDS:
        raise HTTPException(400, f"Field '{body.field}' is not editable")
    import uuid as _uuid

    res = supabase.table("acquisitions").select("*").eq("business_id", business_id).eq("id", body.id).limit(1).execute()
    if not res.data:
        raise HTTPException(404, "Row not found")
    cur = res.data[0]

    batch_id = str(_uuid.uuid4())
    try:
        supabase.table("acquisitions_history").insert({
            "business_id": business_id, "acquisition_id": body.id, "batch_id": batch_id,
            "sku": cur.get("sku"), "name": cur.get("name"), "payment_method": cur.get("payment_method"),
            "date": cur.get("date"), "cost": cur.get("cost"), "cash": cur.get("cash"),
        }).execute()
    except Exception as e:
        raise HTTPException(500, f"Could not save undo history, aborting to be safe: {e}")

    value = body.value
    if body.field in ("cost", "cash"):
        try:
            value = float(value) if value not in (None, "") else None
        except (TypeError, ValueError):
            value = None
    elif value == "":
        value = None

    try:
        supabase.table("acquisitions").update({body.field: value}).eq("id", body.id).execute()
    except Exception as e:
        raise HTTPException(500, str(e))

    recalc_error = None
    try:
        apply_acquisition_profits(business_id)
    except Exception as e:
        recalc_error = str(e)

    return {"ok": True, "field": body.field, "value": value, "recalc_error": recalc_error}

@app.post("/api/acquisitions/edit-batch")
async def edit_acquisitions_batch(request: Request, edits: List[AcquisitionEdit] = Body(...)):
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    if not edits:
        return {"ok": True, "updated": 0}
    import uuid as _uuid

    ids = [e.id for e in edits]
    res = supabase.table("acquisitions").select("*").eq("business_id", business_id).in_("id", ids).execute()
    current_by_id = {r["id"]: r for r in (res.data or [])}

    batch_id = str(_uuid.uuid4())
    history_rows = []
    for e in edits:
        cur = current_by_id.get(e.id)
        if not cur:
            continue
        history_rows.append({
            "business_id": business_id, "acquisition_id": e.id, "batch_id": batch_id,
            "sku": cur.get("sku"), "name": cur.get("name"), "payment_method": cur.get("payment_method"),
            "date": cur.get("date"), "cost": cur.get("cost"), "cash": cur.get("cash"),
        })
    if history_rows:
        try:
            supabase.table("acquisitions_history").insert(history_rows).execute()
        except Exception as e:
            raise HTTPException(500, f"Could not save undo history, aborting to be safe: {e}")

    updated = 0
    for e in edits:
        cur = current_by_id.get(e.id)
        if not cur:
            continue
        record = dict(cur)
        record["sku"] = e.sku
        record["name"] = e.name
        record["payment_method"] = e.payment_method
        record["date"] = e.date
        record["cost"] = e.cost
        record["cash"] = e.cash
        try:
            supabase.table("acquisitions").upsert(record).execute()
            updated += 1
        except Exception as ex:
            print(f"edit_acquisitions_batch: failed to update {e.id}: {ex}")

    recalc_error = None
    try:
        apply_acquisition_profits(business_id)
    except Exception as e:
        recalc_error = str(e)

    return {"ok": True, "updated": updated, "batch_id": batch_id, "recalc_error": recalc_error}

@app.get("/api/acquisitions/cash-history")
async def get_cash_history(request: Request, start: Optional[str] = None, end: Optional[str] = None):
    """Every acquisitions_history row is a snapshot of a lot's fields right BEFORE
    an edit was made — this reconstructs the actual change events (old cash -> new
    cash, with a timestamp) by comparing each snapshot against whatever came next
    for that same lot (either a later edit, or its current live value if that was
    the last edit). Lets you see cash changes over a period even though the
    underlying table was originally built as an undo buffer, not an audit log."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")

    hist_rows = []
    start_i = 0
    while True:
        page = (supabase.table("acquisitions_history").select("*")
                .eq("business_id", business_id)
                .order("acquisition_id").order("created_at")
                .range(start_i, start_i + 999).execute().data or [])
        hist_rows.extend(page)
        if len(page) < 1000:
            break
        start_i += 1000

    current_by_id = {}
    if hist_rows:
        acq_ids = list({r["acquisition_id"] for r in hist_rows})
        for i in range(0, len(acq_ids), 200):
            chunk = acq_ids[i:i+200]
            res = supabase.table("acquisitions").select("id,cash,sku,name").in_("id", chunk).execute()
            for a in (res.data or []):
                current_by_id[a["id"]] = a

    # Group by acquisition_id, walk chronologically, pair each snapshot's cash with
    # whatever the NEXT snapshot (or current live value) shows.
    by_acq = {}
    for r in hist_rows:
        by_acq.setdefault(r["acquisition_id"], []).append(r)

    events = []
    for acq_id, rows in by_acq.items():
        rows.sort(key=lambda r: r["created_at"])
        current = current_by_id.get(acq_id, {})
        for i, r in enumerate(rows):
            old_cash = r.get("cash")
            new_cash = rows[i + 1].get("cash") if i + 1 < len(rows) else current.get("cash")
            if old_cash == new_cash:
                continue  # this edit didn't actually touch cash — skip it
            events.append({
                "acquisition_id": acq_id, "sku": r.get("sku"), "name": r.get("name"),
                "old_cash": old_cash, "new_cash": new_cash,
                "delta": round((new_cash or 0) - (old_cash or 0), 2),
                "changed_at": rows[i + 1]["created_at"] if i + 1 < len(rows) else r["created_at"],
            })

    if start:
        events = [e for e in events if e["changed_at"] >= start]
    if end:
        events = [e for e in events if e["changed_at"] <= end + "T23:59:59"]
    events.sort(key=lambda e: e["changed_at"], reverse=True)

    return {"events": events, "total_delta": round(sum(e["delta"] for e in events), 2)}

@app.post("/api/acquisitions/undo")
async def undo_acquisitions_edit(request: Request):
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    latest = supabase.table("acquisitions_history").select("batch_id").eq("business_id", business_id)\
        .order("created_at", desc=True).limit(1).execute()
    if not latest.data:
        return {"ok": True, "restored": 0, "message": "No edits to undo"}
    batch_id = latest.data[0]["batch_id"]

    history_rows = supabase.table("acquisitions_history").select("*").eq("business_id", business_id)\
        .eq("batch_id", batch_id).execute()
    restored = 0
    for h in (history_rows.data or []):
        try:
            supabase.table("acquisitions").update({
                "sku": h.get("sku"), "name": h.get("name"), "payment_method": h.get("payment_method"),
                "date": h.get("date"), "cost": h.get("cost"), "cash": h.get("cash"),
            }).eq("id", h["acquisition_id"]).execute()
            restored += 1
        except Exception as e:
            print(f"undo_acquisitions_edit: failed to restore {h['acquisition_id']}: {e}")

    supabase.table("acquisitions_history").delete().eq("business_id", business_id).eq("batch_id", batch_id).execute()

    try:
        apply_acquisition_profits(business_id)
    except Exception:
        pass

    return {"ok": True, "restored": restored}

@app.post("/api/acquisitions/recalculate")
async def recalculate_acquisitions(request: Request):
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    try:
        result = apply_acquisition_profits(business_id)
        return {"ok": True, **result}
    except Exception as e:
        raise HTTPException(500, str(e))

class SkuOverrideUpdate(BaseModel):
    item_id: str
    new_sku: str

@app.post("/api/acquisitions/sku-override")
async def set_sku_override(body: SkuOverrideUpdate, request: Request):
    """Sets a LOCAL-ONLY SKU correction for a listing whose real eBay SKU can never
    be changed (v1/Inventory-API listings — see update_ebay_v2_sku for why). Stored
    in its own column so the active-listings sync (which upserts only item_id, sku,
    title, price, quantity_available, etc.) never touches it and can't stomp a
    correction on the next refresh, unlike editing the plain 'sku' column directly."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    new_sku = (body.new_sku or "").strip()
    if not new_sku:
        raise HTTPException(400, "new_sku is required")
    _validate_sku_for_assignment(business_id, new_sku)
    res = (supabase.table("ebay_listing_status")
           .update({"sku_override": new_sku})
           .eq("business_id", business_id).eq("item_id", body.item_id).execute())
    if not res.data:
        raise HTTPException(404, "Listing not found")
    return {"ok": True, "item_id": body.item_id, "sku_override": new_sku}

@app.get("/api/acquisitions/sku-overrides")
async def list_sku_overrides(request: Request):
    """Every currently-set sku_override, permanently visible here regardless of
    whether it's since matched a lot. The Uncategorized view drops an item the
    moment its override resolves to a known lot (it's no longer 'needs
    attention'), which means there was previously no way to look back at
    corrections already made -- this exists specifically to be pulled up any
    time, since sku_override is the only place this correction lives."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    rows = _fetch_all_for_business(business_id, "ebay_listing_status",
                                    "item_id,title,sku,sku_override,price,quantity_available,updated_at,start_time")
    overrides = [r for r in rows if r.get("sku_override")]

    lot_rows = _fetch_all_for_business(business_id, "acquisitions", "sku,name,cost,date")
    lots_by_sku = {l["sku"]: l for l in lot_rows if l.get("sku")}

    out = []
    for r in overrides:
        prefix = _lot_prefix(r["sku_override"])
        lot = lots_by_sku.get(prefix)
        out.append({
            "item_id": r.get("item_id"),
            "title": r.get("title"),
            "original_ebay_sku": r.get("sku"),
            "your_corrected_sku": r.get("sku_override"),
            "lot_sku": prefix,
            "lot_name": lot.get("name") if lot else None,
            "lot_matched": lot is not None,
            "price": r.get("price"),
            "quantity_available": r.get("quantity_available"),
            "updated_at": r.get("updated_at"),
            "created_at": r.get("start_time"),
        })
    out.sort(key=lambda x: x.get("created_at") or "", reverse=True)
    return {"overrides": out}

@app.get("/api/inventory/uncategorized-listings")
async def api_uncategorized_listings_only(request: Request):
    """Lean, standalone version of the 'uncategorized' section of /api/acquisitions --
    same classification logic, but skips two things that section pays for on every
    load that the Uncat tab never uses: acquisitions.select('*') (only the sku
    column is actually needed here -- known_lot_skus), and a SECOND full paginated
    fetch of the entire orders table (thousands of rows) that exists only to compute
    a sales-total summary number this page doesn't display. The active_rows and
    listings fetches are kept as-is -- those are the actual, unavoidable work
    classification depends on.

    ACTIVE-ONLY ON PURPOSE: this page's real job is flagging newly-listed items
    still missing a storage-location suffix (e.g. 'AM1-' with nothing after the
    dash) so a real location can be assigned before it sells -- not a general
    audit of every SKU ever assigned. An Ended listing's location no longer
    matters. (8/8: briefly widened this to include Ended/Unsold, on a wrong
    guess about what this page was for -- reverted.)"""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")

    known_lot_skus = {r["sku"] for r in
                       supabase.table("acquisitions").select("sku").eq("business_id", business_id).execute().data or []
                       if r.get("sku")}

    active_rows = []
    start = 0
    while True:
        page = supabase.table("ebay_listing_status").select("item_id,sku,sku_override,title,price,quantity_available,updated_at,gallery_url,start_time")\
            .eq("business_id", business_id).eq("listing_status", "Active")\
            .range(start, start + 999).execute().data or []
        active_rows.extend(page)
        if len(page) < 1000:
            break
        start += 1000

    listing_by_item_id = {}
    start = 0
    while True:
        page = supabase.table("listings").select("id,ebay_item_id,ebay_offer_id")\
            .eq("business_id", business_id).not_.is_("ebay_item_id", "null")\
            .range(start, start + 999).execute().data or []
        for l in page:
            listing_by_item_id[l["ebay_item_id"]] = l
        if len(page) < 1000:
            break
        start += 1000

    uncategorized_value, uncategorized_count, uncategorized_items = 0, 0, []
    for row in active_rows:
        raw_sku = row.get("sku") or ""
        override_sku = row.get("sku_override") or ""
        sku = override_sku or raw_sku
        value = (row.get("price") or 0) * (row.get("quantity_available") or 0)
        prefix = _lot_prefix(sku) if sku else None
        needs_sku = _sku_needs_assignment(sku)
        matched_listing = listing_by_item_id.get(row.get("item_id"))
        sku_editable = bool(matched_listing) and not matched_listing.get("ebay_offer_id")
        needs_review = needs_sku or (matched_listing is not None and not sku_editable and not override_sku)
        if sku and not needs_review and prefix in known_lot_skus:
            continue
        uncategorized_value += value
        uncategorized_count += 1
        uncategorized_items.append({
            "item_id": row.get("item_id"), "sku": sku, "raw_sku": raw_sku, "title": row.get("title"),
            "price": row.get("price"), "quantity_available": row.get("quantity_available"),
            "value": round(value, 2), "needs_sku": needs_sku,
            "listing_id": matched_listing.get("id") if matched_listing else None,
            "sku_editable": sku_editable, "has_override": bool(override_sku),
            "has_matched_listing": matched_listing is not None,
            "gallery_url": row.get("gallery_url"), "created_at": row.get("start_time"),
        })
    uncategorized_items.sort(key=lambda r: r.get("created_at") or "", reverse=True)
    return {"active_listings_value": round(uncategorized_value, 2), "active_listings_count": uncategorized_count, "items": uncategorized_items}

@app.get("/api/acquisitions")
async def list_acquisitions(request: Request):
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    res = supabase.table("acquisitions").select("*").eq("business_id", business_id)\
        .order("date", desc=True).execute()
    acquisitions = res.data or []

    # Active-listings dollar value per lot — purely a local read against the last
    # eBay ActiveList sync (see /api/acquisitions/sync-active-listings), zero live
    # calls on page load. Matched by SKU prefix before the first '-' when there is
    # one (e.g. 'RJ-123' belongs to lot 'RJ'); a bare SKU with no '-' at all
    # (e.g. just 'RJ') belongs to that same lot directly — both forms transpose to
    # the same lot SKU. quantity_available is eBay's own real remaining-count field.
    # Paginated fetch — a single unpaginated .execute() silently caps at Supabase's
    # default row limit (1000), which was quietly truncating this business's ~4,700
    # active rows down to a fraction of the real total (confirmed: direct SQL on the
    # table showed the full, correct numbers while this endpoint showed far less).
    active_rows = []
    start = 0
    while True:
        page = supabase.table("ebay_listing_status").select("item_id,sku,sku_override,title,price,quantity_available,updated_at")\
            .eq("business_id", business_id).eq("listing_status", "Active")\
            .range(start, start + 999).execute().data or []
        active_rows.extend(page)
        if len(page) < 1000:
            break
        start += 1000
    known_lot_skus = {a["sku"] for a in acquisitions if a.get("sku")}

    value_by_prefix = {}
    count_by_prefix = {}
    uncategorized_value, uncategorized_count = 0, 0
    uncategorized_items = []

    # Join against the internal listings table by eBay item ID so the frontend knows
    # which of these can actually have their SKU edited. Only Trading-API (v2) listings
    # support that — Inventory-API (v1) listings can never have their SKU changed via
    # any API call, an eBay platform restriction (see update_ebay_v2_sku). Paginated
    # for the same reason as everything else here — thousands of rows possible.
    listing_by_item_id = {}
    start = 0
    while True:
        page = supabase.table("listings").select("id,ebay_item_id,ebay_offer_id")\
            .eq("business_id", business_id).not_.is_("ebay_item_id", "null")\
            .range(start, start + 999).execute().data or []
        for l in page:
            listing_by_item_id[l["ebay_item_id"]] = l
        if len(page) < 1000:
            break
        start += 1000

    for row in active_rows:
        raw_sku = row.get("sku") or ""
        override_sku = row.get("sku_override") or ""
        sku = override_sku or raw_sku  # a manual override always wins — it exists
                                        # specifically because the real eBay SKU can
                                        # never be corrected for locked v1 listings
        value = (row.get("price") or 0) * (row.get("quantity_available") or 0)
        prefix = _lot_prefix(sku) if sku else None
        # A bare 'PREFIX-' SKU (nothing after the dash) hasn't actually been assigned
        # to a specific item yet — it's a placeholder from the v2 eBay push path,
        # which allows duplicate bare SKUs on purpose (see /assign-lot). Even though
        # its prefix matches a real lot, it still needs an individual SKU, so it
        # belongs here alongside blank SKUs — not silently counted as "matched."
        needs_sku = _sku_needs_assignment(sku)

        matched_listing = listing_by_item_id.get(row.get("item_id"))
        sku_editable = bool(matched_listing) and not matched_listing.get("ebay_offer_id")
        # FIXED a real blind spot: a locked (v1/Inventory-API) listing whose raw eBay
        # SKU happens to superficially match a real lot's prefix format (e.g. an old
        # eBay-generated SKU like "AM-1024" that coincidentally looks like a valid
        # lot-prefixed SKU) was being counted as correctly matched -- but that match
        # was never actually verified by anyone through Lister's own lot system for
        # legacy v1 listings. A locked listing (one Lister actually tracks via a real
        # listings row, published through the v1 pusher) with NO explicit sku_override
        # on record needs review, regardless of whether its raw SKU looks plausible.
        #
        # BUG FIXED HERE: the first version of this check fired for ANY item with no
        # matched_listing at all -- which is the completely normal state for most
        # eBay-native items Lister never touched, not a sign of anything wrong. That
        # made nearly everything show up as "needs review" (confirmed: user reported
        # Uncategorized was suddenly showing basically the whole inventory). Only
        # apply this check when matched_listing actually exists -- i.e. Lister
        # genuinely knows about this item via a real listing row, which is the actual
        # case this was built for.
        needs_review = needs_sku or (matched_listing is not None and not sku_editable and not override_sku)

        if sku and not needs_review and prefix in known_lot_skus:
            value_by_prefix[prefix] = value_by_prefix.get(prefix, 0) + value
            count_by_prefix[prefix] = count_by_prefix.get(prefix, 0) + 1
        else:
            uncategorized_value += value
            uncategorized_count += 1
            uncategorized_items.append({
                "item_id": row.get("item_id"), "sku": sku, "raw_sku": raw_sku, "title": row.get("title"),
                "price": row.get("price"), "quantity_available": row.get("quantity_available"),
                "value": round(value, 2), "needs_sku": needs_sku,
                "listing_id": matched_listing.get("id") if matched_listing else None,
                "sku_editable": sku_editable, "has_override": bool(override_sku),
                "has_matched_listing": matched_listing is not None,
            })
    uncategorized_items.sort(key=lambda r: r["value"], reverse=True)
    active_synced_at = max((r.get("updated_at") for r in active_rows if r.get("updated_at")), default=None)

    # Same transpose rule applied to sales, for the "Uncategorized" row's sales total.
    # Paginated fetch — this business has thousands of orders, well past Supabase's
    # default single-request row cap.
    all_orders = []
    start = 0
    while True:
        page = supabase.table("orders").select("sku,final_net").eq("business_id", business_id)\
            .range(start, start + 999).execute().data or []
        all_orders.extend(page)
        if len(page) < 1000:
            break
        start += 1000

    uncategorized_sales = 0
    for o in all_orders:
        sku = o.get("sku") or ""
        if not sku or sku.lower() in ("(no sku)",) or sku.lower().startswith("lister-"):
            continue
        if _lot_prefix(sku) not in known_lot_skus:
            uncategorized_sales += o.get("final_net") or 0

    for a in acquisitions:
        a["active_listings_value"] = round(value_by_prefix.get(a.get("sku"), 0), 2)
        a["active_listings_count"] = count_by_prefix.get(a.get("sku"), 0)
        a["active_listings_synced_at"] = active_synced_at

    uncategorized = {
        "active_listings_value": round(uncategorized_value, 2),
        "active_listings_count": uncategorized_count,
        "sales_total": round(uncategorized_sales, 2),
        "items": uncategorized_items,
    }

    return {"acquisitions": acquisitions, "uncategorized": uncategorized}

@app.post("/api/acquisitions")
async def create_acquisition(request: Request, body: AcquisitionCreate):
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    try:
        record = body.dict()
        record["business_id"] = business_id
        res = supabase.table("acquisitions").insert(record).execute()
        apply_acquisition_profits(business_id)
        return {"ok": True, "acquisition": (res.data or [{}])[0]}
    except Exception as e:
        raise HTTPException(500, str(e))

@app.delete("/api/acquisitions/{acquisition_id}")
async def delete_acquisition(acquisition_id: str, request: Request):
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    try:
        supabase.table("acquisitions").delete().eq("id", acquisition_id).eq("business_id", business_id).execute()
        return {"ok": True}
    except Exception as e:
        raise HTTPException(500, str(e))

def _parse_money(val) -> Optional[float]:
    """Handles '$1,550 ', '$0 ', '', None -> float or None"""
    if val is None:
        return None
    s = str(val).strip().replace("$", "").replace(",", "").strip()
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None

def _parse_acq_date(val) -> Optional[str]:
    """Handles the mixed formats seen in the existing spreadsheet:
    'May-23' (Mon-YY), '4/10/2024' (M/D/YYYY), '1/1/9999' (placeholder/unknown -> None)."""
    import datetime as _dt
    s = str(val).strip()
    if not s or s.startswith("1/1/9999"):
        return None
    for fmt in ("%b-%y", "%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d"):
        try:
            return _dt.datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None

@app.post("/api/orders/backfill-sku")
async def backfill_order_skus(request: Request, file: UploadFile = File(...)):
    """ONE-TIME historical backfill: matches orders that never got a real SKU
    (blank, '(no SKU)', or the 'lister-{id}' fallback) against a CSV export of
    (Listing title, Lot Name, Net sales) — using the listing title as the join key,
    since this export has no order ID. The Lot Name gets resolved against Acquisitions'
    sku field first, then its name field, to find the real SKU code. Best-effort:
    if multiple orders share the exact same title, they're paired off in order,
    which is as precise as this data allows."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    import io, csv as _csv

    content = await file.read()
    text = content.decode("utf-8-sig", errors="replace")
    reader = _csv.DictReader(io.StringIO(text))
    csv_rows = list(reader)

    # Build lot-name -> real sku lookup from Acquisitions (sku match takes priority over name match)
    def _fetch_all(table, select_cols):
        all_rows = []
        page_size = 1000
        start = 0
        while True:
            res = supabase.table(table).select(select_cols).eq("business_id", business_id)\
                .range(start, start + page_size - 1).execute()
            page = res.data or []
            all_rows.extend(page)
            if len(page) < page_size:
                break
            start += page_size
        return all_rows

    acquisitions = _fetch_all("acquisitions", "sku,name")
    sku_by_lower_sku = {a["sku"].strip().lower(): a["sku"] for a in acquisitions if a.get("sku")}
    sku_by_lower_name = {}
    for a in acquisitions:
        if a.get("name") and a.get("sku"):
            key = a["name"].strip().lower()
            sku_by_lower_name.setdefault(key, a["sku"])  # first match wins if ambiguous

    def resolve_sku(lot_name: str):
        key = (lot_name or "").strip().lower()
        if not key:
            return None
        return sku_by_lower_sku.get(key) or sku_by_lower_name.get(key)

    # Fetch every order still missing a real SKU
    blank_orders = _fetch_all("orders", "id,title,sku")
    blank_orders = [o for o in blank_orders if not o.get("sku") or o["sku"] in ("", "(no SKU)") or o["sku"].lower().startswith("lister-")]

    by_title = {}
    for o in blank_orders:
        key = (o.get("title") or "").strip().lower()
        by_title.setdefault(key, []).append(o)

    updates = []
    matched, no_lot_match, no_order_match, ambiguous_lots = 0, 0, 0, 0
    for row in csv_rows:
        title = (row.get("Listing title") or "").strip()
        lot_name = (row.get("Lot Name") or "").strip()
        if not title or not lot_name:
            continue
        real_sku = resolve_sku(lot_name)
        if not real_sku:
            no_lot_match += 1
            continue
        candidates = by_title.get(title.lower())
        if not candidates:
            no_order_match += 1
            continue
        order = candidates.pop(0)  # consume one match so the next identical title gets a different order
        updates.append({"id": order["id"], "sku": f"{real_sku}-BF"})  # -BF marks it as backfilled, not a real location code
        matched += 1

    updated = 0
    for i in range(0, len(updates), 500):
        chunk = updates[i:i+500]
        try:
            supabase.table("orders").upsert(chunk, on_conflict="id").execute()
            updated += len(chunk)
        except Exception as e:
            print(f"backfill_order_skus: batch {i}-{i+len(chunk)} failed: {e}")

    return {
        "ok": True, "csv_rows": len(csv_rows), "blank_orders_before": len(blank_orders),
        "matched": matched, "updated": updated,
        "no_lot_match": no_lot_match, "no_order_match": no_order_match,
    }

@app.post("/api/acquisitions/upload-csv")
async def upload_acquisitions_csv(request: Request, file: UploadFile = File(...)):
    """Only imports the true manual-input columns (SKU, Name, Payment_Method, Date, Cost,
    Cash) — eBay, Total_Payouts, and Profit from the old spreadsheet are NOT imported,
    since those are now computed live instead of stored as a static snapshot."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    import io, csv as _csv

    content = await file.read()
    text = content.decode("utf-8-sig", errors="replace")
    reader = _csv.DictReader(io.StringIO(text))

    inserted = 0
    skipped = 0
    for row in reader:
        sku = str(row.get("SKU") or "").strip()
        if not sku:
            skipped += 1
            continue
        record = {
            "business_id": business_id,
            "sku": sku,
            "name": str(row.get("Name") or "").strip() or None,
            "payment_method": str(row.get("Payment_Method") or "").strip() or None,
            "date": _parse_acq_date(row.get("Date")),
            "cost": _parse_money(row.get("Cost")),
            "cash": _parse_money(row.get("Cash")),
        }
        try:
            supabase.table("acquisitions").insert(record).execute()
            inserted += 1
        except Exception:
            skipped += 1

    try:
        match_result = apply_acquisition_profits(business_id)
    except Exception as e:
        match_result = {"updated": 0, "error": str(e)}

    return {"ok": True, "inserted": inserted, "skipped": skipped, "recalculated": match_result.get("updated", 0)}

def fetch_ebay_orders(business_id: str, start_iso: str, end_iso: str) -> list:
    """Returns normalized order-line rows from eBay's Fulfillment API for the given date range."""
    import requests as _req
    token = get_ebay_access_token(business_id)
    rows = []
    offset = 0
    max_pages = 25  # 25 * 200 = 5,000 orders safety cap
    for _ in range(max_pages):
        r = _req.get(
            f"{EBAY_API_BASE}/sell/fulfillment/v1/order",
            headers=ebay_headers(token, content_language=False),
            params={
                "filter": f"creationdate:[{start_iso}..{end_iso}]",
                "limit": 200,
                "offset": offset,
            },
            timeout=20,
        )
        if r.status_code != 200:
            raise Exception(f"eBay getOrders failed ({r.status_code}): {r.text[:300]}")
        data = r.json()
        for order in data.get("orders", []):
            created = order.get("creationDate", "")
            order_delivery_cost = float((order.get("pricingSummary") or {}).get("deliveryCost", {}).get("value", 0) or 0)
            tax_addr = ((order.get("buyer") or {}).get("taxAddress") or {})
            buyer_state = tax_addr.get("stateOrProvince", "")
            buyer_zip = tax_addr.get("postalCode", "")
            buyer_country = tax_addr.get("countryCode", "")

            # totalDueSeller is eBay's own authoritative "what you actually get after fees
            # AND refunds" figure — more reliable than reconstructing net from fee/refund
            # pieces separately, which can drift (eBay adjusts fees on partial refunds in
            # ways that don't always show up cleanly in the separate fee/refund fields).
            # It's order-level, so for multi-item orders it's prorated by gross revenue share.
            order_total_due_seller = float((order.get("paymentSummary") or {}).get("totalDueSeller", {}).get("value", 0) or 0)

            order_refund_total = sum(
                float((r.get("amount") or {}).get("value", 0) or 0)
                for r in (order.get("paymentSummary") or {}).get("refunds", []) or []
            )
            line_items = order.get("lineItems", [])
            has_line_level_refunds = any(li.get("refunds") for li in line_items)
            order_gross_subtotal = sum(
                float((li.get("lineItemCost") or {}).get("value", 0) or 0)
                + float((li.get("deliveryCost") or {}).get("shippingCost", {}).get("value", 0) or 0)
                for li in line_items
            )
            order_subtotal_for_refund_proration = sum(
                float((li.get("lineItemCost") or {}).get("value", 0) or 0) for li in line_items
            ) if order_refund_total and not has_line_level_refunds else 0

            for li in line_items:
                item_price = float((li.get("lineItemCost") or {}).get("value", 0) or 0)
                buyer_shipping = float((li.get("deliveryCost") or {}).get("shippingCost", {}).get("value", 0) or 0)
                if has_line_level_refunds:
                    refund = sum(float((r.get("amount") or {}).get("value", 0) or 0) for r in (li.get("refunds") or []))
                elif order_refund_total and order_subtotal_for_refund_proration > 0:
                    refund = order_refund_total * (item_price / order_subtotal_for_refund_proration)
                else:
                    refund = 0.0
                line_gross = item_price + buyer_shipping
                net_from_ebay = (order_total_due_seller * (line_gross / order_gross_subtotal)
                                  if order_gross_subtotal > 0 else order_total_due_seller)
                rows.append({
                    "platform": "eBay",
                    "sku": li.get("sku") or "(no SKU)",
                    "title": li.get("title", ""),
                    "quantity": int(li.get("quantity", 1)),
                    "revenue": line_gross,
                    "refund": refund,
                    "net_from_ebay": net_from_ebay,
                    "buyer_shipping": buyer_shipping,
                    "order_delivery_cost": order_delivery_cost,
                    "buyer_state": buyer_state, "buyer_zip": buyer_zip, "buyer_country": buyer_country,
                    "order_date": created[:10] if created else "",
                    "order_id": order.get("orderId", ""),
                    "line_item_id": li.get("lineItemId", ""),
                    "legacy_item_id": li.get("legacyItemId", ""),
                })
        total = data.get("total", 0)
        offset += 200
        if offset >= total or not data.get("orders"):
            break
    return rows

def fetch_ebay_fees_by_line_item(business_id: str, start_iso: str, end_iso: str) -> dict:
    """Returns {(order_id, line_item_id): net_fee_amount} for marketplace fees, AND
    {order_id: label_cost} for eBay-purchased shipping labels (SHIPPING_LABEL transaction
    type) — both captured from the SAME Finance API pass, no extra API calls needed.
    Note: eBay only associates a SHIPPING_LABEL transaction with a specific orderId when
    the label was purchased individually — bulk/batch label purchases only report one
    lump amount with no per-order breakdown, so those can't be attributed here.
    eBay hard-caps each query's date span at 36 months and total retention at 5 years,
    so a wide range (e.g. a full historical backfill) gets auto-split into ~30-month
    chunks here rather than failing outright."""
    import datetime as _dt

    def _fetch_window(window_start_iso: str, window_end_iso: str) -> tuple:
        import requests as _req
        token = get_ebay_access_token(business_id)
        window_fees = {}
        window_labels = {}
        offset = 0
        max_pages = 25  # 25 * 200 = 5,000 transactions safety cap per window
        for _ in range(max_pages):
            r = _req.get(
                "https://apiz.ebay.com/sell/finances/v1/transaction",
                headers=ebay_headers(token, content_language=False),
                params={
                    "filter": f"transactionDate:[{window_start_iso}..{window_end_iso}]",
                    "limit": 200,
                    "offset": offset,
                },
                timeout=20,
            )
            if r.status_code != 200:
                raise Exception(f"eBay Finance transactions failed ({r.status_code}): {r.text[:300]}")
            data = r.json()
            for txn in data.get("transactions", []):
                order_id = txn.get("orderId", "")
                if txn.get("transactionType") == "SHIPPING_LABEL" and order_id:
                    amt = float((txn.get("amount") or {}).get("value", 0) or 0)
                    window_labels[order_id] = window_labels.get(order_id, 0.0) + amt
                    continue
                for li in txn.get("orderLineItems", []) or []:
                    line_item_id = li.get("lineItemId", "")
                    if not order_id or not line_item_id:
                        continue
                    key = (order_id, line_item_id)
                    fee_total = 0.0
                    for fee in li.get("marketplaceFees", []) or []:
                        fee_total += float((fee.get("amount") or {}).get("value", 0) or 0)
                    window_fees[key] = window_fees.get(key, 0.0) + fee_total
            total = data.get("total", 0)
            offset += 200
            if offset >= total or not data.get("transactions"):
                break
        return window_fees, window_labels

    start_dt = _dt.datetime.strptime(start_iso, "%Y-%m-%dT%H:%M:%S.%fZ")
    end_dt = _dt.datetime.strptime(end_iso, "%Y-%m-%dT%H:%M:%S.%fZ")
    fees = {}
    ebay_labels_by_order = {}
    window_start = start_dt
    while window_start < end_dt:
        window_end = min(window_start + _dt.timedelta(days=900), end_dt)  # ~30 months per chunk
        window_fees, window_labels = _fetch_window(
            window_start.strftime("%Y-%m-%dT%H:%M:%S.%f")[:-3] + "Z",
            window_end.strftime("%Y-%m-%dT%H:%M:%S.%f")[:-3] + "Z",
        )
        fees.update(window_fees)
        ebay_labels_by_order.update(window_labels)
        window_start = window_end
    return {"fees_by_line": fees, "ebay_labels_by_order": ebay_labels_by_order}

def fetch_shopify_orders(business_id: str, start_iso: str, end_iso: str) -> list:
    """Returns normalized order-line rows from Shopify's Admin API for the given date range,
    with refunds subtracted (from the embedded order payload) and an ESTIMATED payment-processing
    fee (Shopify's standard 2.9%+$0.30 rate, applied only when payment_gateway_names shows
    Shopify Payments/Shop Pay was used) prorated across line items. This is an estimate, not
    the exact fee — getting the real figure requires one extra API call per order, which made
    this page unusably slow on any real order volume."""
    import requests as _req
    settings = get_ebay_settings(business_id)
    domain = (settings.get("SHOPIFY_STORE_DOMAIN", "") or "").strip().replace("https://", "").replace("http://", "").strip("/")
    if not domain:
        return []
    token = get_shopify_access_token(business_id)
    rows = []
    url = f"https://{domain}/admin/api/2024-10/orders.json"
    params = {"status": "any", "created_at_min": start_iso, "created_at_max": end_iso, "limit": 250}
    headers = {"X-Shopify-Access-Token": token}
    while url:
        r = _req.get(url, headers=headers, params=params, timeout=20)
        if r.status_code == 401:
            token = get_shopify_access_token(business_id, force_refresh=True)
            headers["X-Shopify-Access-Token"] = token
            r = _req.get(url, headers=headers, params=params, timeout=20)
        if r.status_code != 200:
            raise Exception(f"Shopify orders fetch failed ({r.status_code}): {r.text[:300]}")
        data = r.json()
        for order in data.get("orders", []):
            created = order.get("created_at", "")
            order_id = order.get("id")

            # Refunds are embedded in the order payload already — no extra call needed.
            refunded_by_line = {}
            for refund in order.get("refunds", []) or []:
                for rli in refund.get("refund_line_items", []) or []:
                    lid = rli.get("line_item_id")
                    amt = float(rli.get("subtotal", 0) or 0) + float(rli.get("total_tax", 0) or 0)
                    refunded_by_line[lid] = refunded_by_line.get(lid, 0.0) + amt

            # Payment-processing fee: getting the EXACT fee requires one extra API call per
            # order (the transactions sub-resource), which was making this page painfully slow
            # on any real order volume. Instead, estimate using Shopify's standard online rate
            # (2.9% + $0.30) when the order actually went through Shopify Payments — no extra
            # call needed, since payment_gateway_names is already in the order payload.
            gateways = order.get("payment_gateway_names", []) or []
            used_shopify_payments = any("shopify_payments" in g.lower() or "shop_pay" in g.lower() for g in gateways)
            order_total = float(order.get("total_price", 0) or 0)
            order_fee_total = (order_total * 0.029 + 0.30) if (used_shopify_payments and order_total > 0) else 0.0

            # Fulfillment tracking numbers — embedded in the order payload already, same as
            # refunds above. Needed so Shopify orders can be matched against shipping_labels
            # the same way eBay orders already are (see apply_shipping_matches).
            trackings = []
            for f in order.get("fulfillments", []) or []:
                tn = f.get("tracking_number")
                if tn:
                    trackings.append(tn)
                for tn2 in (f.get("tracking_numbers") or []):
                    if tn2 and tn2 not in trackings:
                        trackings.append(tn2)

            line_items = order.get("line_items", [])
            order_subtotal = sum(float(li.get("price", 0) or 0) * int(li.get("quantity", 1)) for li in line_items)

            for li in line_items:
                gross = float(li.get("price", 0) or 0) * int(li.get("quantity", 1))
                refund_amt = refunded_by_line.get(li.get("id"), 0.0)
                fee_share = (order_fee_total * (gross / order_subtotal)) if order_subtotal > 0 else 0.0
                rows.append({
                    "platform": "Shopify",
                    "sku": li.get("sku") or "(no SKU)",
                    "title": li.get("title", ""),
                    "variant_id": str(li.get("variant_id")) if li.get("variant_id") else None,
                    "quantity": int(li.get("quantity", 1)),
                    "line_item_id": li.get("id"),
                    "revenue": gross,
                    "refund": refund_amt,
                    "fee": fee_share,
                    "net": gross - refund_amt - fee_share,
                    "order_date": created[:10] if created else "",
                    "order_id": order.get("name", ""),
                    "tracking_numbers": trackings,
                })
        # Shopify cursor pagination via Link header
        link = r.headers.get("Link", "")
        next_url = None
        for part in link.split(","):
            if 'rel="next"' in part:
                next_url = part.split(";")[0].strip().strip("<>")
        url = next_url
        params = None  # next_url already has query params baked in
    return rows

def apply_shipping_matches(business_id: str) -> dict:
    """Matches every order's tracking_number against shipping_labels and writes the
    result directly into orders.shipping_cost / orders.final_net — a stored, one-time
    computation, not something recalculated on every page read. Pure local DB work,
    no eBay/Shopify API calls, batched so it's fast even across thousands of orders.
    Both fetches are paginated — Supabase silently caps a single query at ~1000 rows,
    and with thousands of orders/labels that was causing real matches to be missed."""
    def _fetch_all(table, select_cols, extra_filter=None):
        all_rows = []
        page_size = 1000
        start = 0
        while True:
            q = supabase.table(table).select(select_cols).eq("business_id", business_id)
            if extra_filter:
                q = extra_filter(q)
            res = q.range(start, start + page_size - 1).execute()
            page = res.data or []
            all_rows.extend(page)
            if len(page) < page_size:
                break
            start += page_size
        return all_rows

    orders = _fetch_all("orders", "*", lambda q: q.not_.is_("tracking_number", "null"))
    labels = _fetch_all("shipping_labels", "tracking_number,cost,recipient,created_date")
    if not orders:
        return {"updated": 0, "debug_orders_fetched": 0, "debug_labels_fetched": len(labels)}

    cost_by_tracking = {row["tracking_number"]: (row.get("cost") or 0) for row in labels}

    # ── Multi-box label attribution ─────────────────────────────────────────
    # A multi-box shipment produces N Pirate Ship labels but eBay only ever
    # knows the ONE tracking number that was uploaded to it, so the other N-1
    # labels sat orphaned in shipping_labels and their cost was never counted
    # against any order. Attribution rule: an orphan label (attached to no
    # order at all) is attributed to an order when that order already has a
    # matched label with the SAME recipient bought within ±36h of the orphan.
    # Each orphan attaches to at most one order (closest label time wins), so
    # cost can never be double-counted. The attachment is persisted onto
    # orders.tracking_number, and sync merges rather than overwrites tracking
    # numbers, so the attribution is durable across every future sync cycle.
    from datetime import datetime as _dt
    def _parse_label_dt(txt):
        txt = (txt or "").replace(" MDT", "").replace(" MST", "").replace(" MT", "").strip()
        for fmt in ("%m/%d/%y %I:%M %p", "%m/%d/%Y %I:%M %p"):
            try:
                return _dt.strptime(txt, fmt)
            except ValueError:
                continue
        return None

    label_info = {row["tracking_number"]: row for row in labels}
    assigned = set()
    for order in orders:
        for tn in (order.get("tracking_number") or "").split(","):
            tn = tn.strip()
            if tn:
                assigned.add(tn)

    orphans_by_recipient = {}
    for row in labels:
        tn = row["tracking_number"]
        if tn in assigned or not (row.get("cost") or 0):
            continue
        rcpt = (row.get("recipient") or "").strip().lower()
        ts = _parse_label_dt(row.get("created_date"))
        if rcpt and ts:
            orphans_by_recipient.setdefault(rcpt, []).append((tn, ts))

    attached_orphans = 0
    if orphans_by_recipient:
        for order in orders:
            own = [tn.strip() for tn in (order.get("tracking_number") or "").split(",") if tn.strip()]
            anchor_times = []
            for tn in own:
                li = label_info.get(tn)
                if li:
                    ts = _parse_label_dt(li.get("created_date"))
                    rcpt = (li.get("recipient") or "").strip().lower()
                    if ts and rcpt:
                        anchor_times.append((rcpt, ts))
            if not anchor_times:
                continue
            added = []
            for rcpt, ts in anchor_times:
                for tn, ots in list(orphans_by_recipient.get(rcpt, [])):
                    if abs((ots - ts).total_seconds()) <= 36 * 3600 and tn not in assigned:
                        added.append(tn)
                        assigned.add(tn)
                        orphans_by_recipient[rcpt].remove((tn, ots))
            if added:
                order["tracking_number"] = ",".join(own + added)
                attached_orphans += len(added)

    updates = []
    for order in orders:
        trackings = (order.get("tracking_number") or "").split(",")
        shipping_cost = round(sum(cost_by_tracking.get(tn.strip(), 0) or 0 for tn in trackings if tn.strip()), 2)
        if shipping_cost <= 0:
            continue  # nothing to update — leave existing stored values alone
        if shipping_cost == round(float(order.get("shipping_cost") or 0), 2) and order.get("final_net") is not None:
            continue  # already correct — skip the write
        base_net = order.get("net") or 0
        # Copy the FULL existing row and just overwrite the two fields that changed —
        # a partial payload fails Postgres's NOT NULL check on the insert-half of the
        # upsert statement even when the row already exists and would only be updated.
        record = dict(order)
        record["shipping_cost"] = shipping_cost
        record["final_net"] = round(base_net - shipping_cost, 2)
        updates.append(record)

    updated = 0
    for i in range(0, len(updates), 500):
        chunk = updates[i:i+500]
        try:
            supabase.table("orders").upsert(chunk, on_conflict="id").execute()
            updated += len(chunk)
        except Exception as e:
            print(f"apply_shipping_matches: batch {i}-{i+len(chunk)} failed: {e}")

    return {"updated": updated, "orders_with_tracking": len(orders), "debug_labels_fetched": len(labels), "debug_matches_found": len(updates), "orphan_labels_attached": attached_orphans}

@app.post("/api/financials/apply-shipping-matches")
async def apply_shipping_matches_now(request: Request):
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    try:
        result = apply_shipping_matches(business_id)
        return {"ok": True, **result}
    except Exception as e:
        raise HTTPException(500, str(e))

@app.post("/api/shipping-labels/upload")
async def upload_shipping_labels(request: Request, file: UploadFile = File(...)):
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    import io, csv as _csv

    content = await file.read()
    rows = []
    filename = (file.filename or "").lower()

    if filename.endswith(".csv"):
        text = content.decode("utf-8-sig", errors="replace")
        reader = _csv.DictReader(io.StringIO(text))
        rows = list(reader)
    else:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        raw_rows = list(ws.iter_rows(values_only=True))
        if not raw_rows:
            raise HTTPException(400, "File appears empty")
        headers = [str(h).strip() if h else "" for h in raw_rows[0]]
        for r in raw_rows[1:]:
            if not any(r):
                continue
            rows.append({headers[i]: r[i] for i in range(len(headers)) if i < len(r)})

    records = []
    skipped = 0
    for row in rows:
        tracking = str(row.get("Tracking Number") or "").strip()
        if not tracking:
            skipped += 1
            continue
        cost = row.get("Cost")
        try:
            cost = float(cost) if cost not in (None, "") else None
        except (ValueError, TypeError):
            cost = None
        # Refunded labels (misprints that were voided in Pirate Ship) are not real
        # spend -- their money came back. Importing them at face value inflated
        # shipping cost whenever they matched or got attributed to an order
        # (seen live: 3 refunded reprints added $112.41 to one order). Store them
        # at cost 0 so they can never count, while keeping the row so a re-upload
        # of an older export can't resurrect the original cost either.
        if str(row.get("Status") or "").strip().lower() == "refunded":
            cost = 0
        records.append({
            "tracking_number": tracking,
            "business_id": business_id,
            "recipient": str(row.get("Recipient") or ""),
            "cost": cost,
            "created_date": str(row.get("Created Date") or ""),
            "ship_from": str(row.get("Ship From") or ""),
            "source": str(row.get("Source") or ""),
        })

    inserted = 0
    for i in range(0, len(records), 500):
        chunk = records[i:i+500]
        try:
            supabase.table("shipping_labels").upsert(chunk).execute()
            inserted += len(chunk)
        except Exception as e:
            print(f"upload_shipping_labels: batch {i}-{i+len(chunk)} failed: {e}")
            skipped += len(chunk)

    match_result = apply_shipping_matches(business_id)

    return {"ok": True, "inserted": inserted, "skipped": skipped, "total_rows": len(rows), "matched": match_result.get("updated", 0)}

@app.post("/api/financials/match-shipping")
async def match_shipping_costs(request: Request, body: dict = Body(...)):
    """Given a list of eBay order IDs (from the currently loaded Financials view),
    fetches each order's tracking number from eBay and matches it against uploaded
    Pirate Ship data by tracking number. Deliberately opt-in (not automatic) since
    it's one extra API call per order."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    import requests as _req
    order_ids = list(set(body.get("order_ids", [])))[:300]  # sane cap per click
    token = get_ebay_access_token(business_id)

    tracking_by_order = {}
    for oid in order_ids:
        try:
            r = _req.get(
                f"{EBAY_API_BASE}/sell/fulfillment/v1/order/{oid}/shipping_fulfillment",
                headers=ebay_headers(token, content_language=False),
                timeout=15,
            )
            if r.status_code == 200:
                fulfillments = r.json().get("fulfillments", [])
                for f in fulfillments:
                    tn = f.get("shipmentTrackingNumber")
                    if tn:
                        tracking_by_order.setdefault(oid, []).append(tn)
        except Exception:
            continue

    all_trackings = [tn for lst in tracking_by_order.values() for tn in lst]
    cost_by_tracking = {}
    if all_trackings:
        res = supabase.table("shipping_labels").select("tracking_number,cost")\
            .eq("business_id", business_id).in_("tracking_number", all_trackings).execute()
        for row in (res.data or []):
            cost_by_tracking[row["tracking_number"]] = row.get("cost")

    result = {}
    for oid, trackings in tracking_by_order.items():
        cost = sum(cost_by_tracking.get(tn, 0) or 0 for tn in trackings)
        result[oid] = {"tracking_numbers": trackings, "shipping_cost": round(cost, 2)}

    return {"by_order": result, "matched": len(result), "requested": len(order_ids)}

def _sync_orders_window(business_id: str, start_iso: str, end_iso: str) -> dict:
    """Does the actual API pulling + upserting for ONE date window. Called repeatedly,
    once per month, by sync_orders_for_business below — never call this directly with
    a huge range, since each call blocks for as long as that window takes."""
    import requests as _req, datetime as _dt

    upserted = 0
    errors = {}

    # --- eBay: orders + real fees + tracking numbers ---
    try:
        ebay_rows = fetch_ebay_orders(business_id, start_iso, end_iso)
        try:
            fee_data = fetch_ebay_fees_by_line_item(business_id, start_iso, end_iso)
            fees_by_line = fee_data["fees_by_line"]
            ebay_labels_by_order = fee_data["ebay_labels_by_order"]
        except Exception as e:
            fees_by_line = {}
            ebay_labels_by_order = {}
            errors["ebay_fees"] = str(e)

        # Tracking numbers, one call per unique order (needed for shipping cost matching)
        token = get_ebay_access_token(business_id)
        order_ids = list(set(r["order_id"] for r in ebay_rows if r.get("order_id")))
        tracking_by_order = {}
        for oid in order_ids[:500]:  # safety cap per sync cycle
            try:
                r = _req.get(
                    f"{EBAY_API_BASE}/sell/fulfillment/v1/order/{oid}/shipping_fulfillment",
                    headers=ebay_headers(token, content_language=False), timeout=15,
                )
                if r.status_code == 200:
                    for f in r.json().get("fulfillments", []):
                        tn = f.get("shipmentTrackingNumber")
                        if tn:
                            tracking_by_order.setdefault(oid, []).append(tn)
            except Exception:
                continue

        all_trackings = [tn for lst in tracking_by_order.values() for tn in lst]
        cost_by_tracking = {}
        # Chunk the .in_() lookup — a huge tracking-number list in one query can produce
        # an oversized request that Supabase rejects outright.
        for i in range(0, len(all_trackings), 200):
            chunk = all_trackings[i:i+200]
            try:
                res = supabase.table("shipping_labels").select("tracking_number,cost")\
                    .eq("business_id", business_id).in_("tracking_number", chunk).execute()
                for row in (res.data or []):
                    cost_by_tracking[row["tracking_number"]] = row.get("cost") or 0
            except Exception as e:
                errors["shipping_match"] = str(e)

        def _safe(n):
            """NaN/Infinity are valid Python floats but not valid JSON — a single bad
            fee value here would otherwise silently corrupt the whole upsert payload."""
            try:
                n = float(n)
                return round(n, 2) if (n == n and abs(n) != float("inf")) else 0.0  # n==n is False only for NaN
            except (TypeError, ValueError):
                return 0.0

        # Preserve any real SKU already stored (manually backfilled, or otherwise better
        # than what eBay's API returns) — a routine re-sync should never downgrade a
        # good SKU back to blank/"(no SKU)"/the lister-{id} fallback, and should never
        # overwrite a deliberate manual correction with eBay's raw value either, even
        # if that raw value also happens to be non-blank — a manual correction is meant
        # to be final, one-and-done, not silently re-derived on every future sync.
        candidate_ids = [f"ebay:{row['order_id']}:{row['line_item_id']}" for row in ebay_rows]
        existing_sku_by_id = {}
        existing_tracking_by_id = {}
        for i in range(0, len(candidate_ids), 200):
            chunk = candidate_ids[i:i+200]
            try:
                res = supabase.table("orders").select("id,sku,tracking_number").in_("id", chunk).execute()
                for r in (res.data or []):
                    existing_sku_by_id[r["id"]] = r.get("sku")
                    existing_tracking_by_id[r["id"]] = r.get("tracking_number") or ""
            except Exception:
                pass

        # Multi-box orders: eBay only knows the ONE tracking number uploaded to it,
        # but a 6-box shipment has 6 Pirate Ship labels — the other 5 live only in
        # shipping_labels (source blank) and previously could never attach to the
        # order, undercounting shipping cost (seen live: qty-6 order, $235 of labels,
        # $39.19 counted). Fix: MERGE stored trackings with eBay's instead of
        # overwriting, so extra trackings attached to the order (manually or by any
        # future tooling) survive every 20-minute sync. Their label costs are looked
        # up right here so shipping_cost reflects the full merged set.
        extra_trackings = []
        for rid in candidate_ids:
            for tn in (existing_tracking_by_id.get(rid) or "").split(","):
                tn = tn.strip()
                if tn and tn not in cost_by_tracking:
                    extra_trackings.append(tn)
        for i in range(0, len(extra_trackings), 200):
            chunk = extra_trackings[i:i+200]
            try:
                res = supabase.table("shipping_labels").select("tracking_number,cost")\
                    .eq("business_id", business_id).in_("tracking_number", chunk).execute()
                for row in (res.data or []):
                    cost_by_tracking[row["tracking_number"]] = row.get("cost") or 0
            except Exception as e:
                errors["shipping_match_extras"] = str(e)

        skipped_rows = 0
        for row in ebay_rows:
            fee = _safe(fees_by_line.get((row["order_id"], row["line_item_id"]), 0.0))
            revenue = _safe(row["revenue"])
            refund_amt = _safe(row.get("refund", 0))
            # net_from_ebay (derived from paymentSummary.totalDueSeller) is authoritative —
            # it already correctly reflects fees AND refunds exactly as eBay computed them,
            # including edge cases (like fee adjustments on partial refunds) that a manual
            # revenue-refund-fee reconstruction can drift from. Fall back to reconstruction
            # only if that field is missing for some reason.
            net = _safe(row["net_from_ebay"]) if "net_from_ebay" in row else _safe(revenue - refund_amt - fee)
            trackings = list(tracking_by_order.get(row["order_id"], []))
            _rid = f"ebay:{row['order_id']}:{row['line_item_id']}"
            for tn in (existing_tracking_by_id.get(_rid) or "").split(","):
                tn = tn.strip()
                if tn and tn not in trackings:
                    trackings.append(tn)
            pirate_ship_cost = sum(cost_by_tracking.get(tn, 0) or 0 for tn in trackings)
            # Pirate Ship match takes priority (it's the common case); eBay-purchased
            # label cost fills in only when there's no Pirate Ship match for this order.
            shipping_cost = _safe(pirate_ship_cost if pirate_ship_cost > 0 else ebay_labels_by_order.get(row["order_id"], 0))
            record_id = f"ebay:{row['order_id']}:{row['line_item_id']}"
            existing_sku = existing_sku_by_id.get(record_id)
            final_sku = existing_sku if (existing_sku and not _is_blank_sku(existing_sku)) else row["sku"]
            record = {
                "id": record_id,
                "business_id": business_id, "platform": "eBay", "order_id": row["order_id"],
                "sku": final_sku, "title": row["title"], "quantity": row["quantity"],
                "legacy_item_id": row.get("legacy_item_id", ""),
                "order_date": row["order_date"], "gross_revenue": revenue,
                "buyer_shipping": _safe(row.get("buyer_shipping", 0)),
                "refund": _safe(row.get("refund", 0)),
                "order_delivery_cost": _safe(row.get("order_delivery_cost", 0)),
                "buyer_state": row.get("buyer_state", ""), "buyer_zip": row.get("buyer_zip", ""), "buyer_country": row.get("buyer_country", ""),
                "fee": fee, "net": net,
                "tracking_number": ",".join(trackings) if trackings else None,
                "shipping_cost": shipping_cost,
                "final_net": _safe(net - shipping_cost),
            }
            try:
                supabase.table("orders").upsert(record, on_conflict="id").execute()
                upserted += 1
            except Exception as e:
                skipped_rows += 1
                print(f"sync_orders_for_business: skipped bad row {record.get('id')}: {e}")
        if skipped_rows:
            errors["ebay_skipped_rows"] = f"{skipped_rows} row(s) failed to upsert — see logs for details"
    except Exception as e:
        errors["ebay"] = str(e)

    # --- Shopify: orders + estimated fee + embedded refunds ---
    try:
        def _safe2(n):
            try:
                n = float(n)
                return round(n, 2) if (n == n and abs(n) != float("inf")) else 0.0
            except (TypeError, ValueError):
                return 0.0

        shopify_rows = fetch_shopify_orders(business_id, start_iso, end_iso)

        # Shipping cost matching — same shipping_labels table eBay orders already match
        # against (Pirate Ship CSV uploads), keyed by tracking number. An order's total
        # shipping cost gets prorated across its line items the same way the payment
        # fee already is (by revenue share), since Shopify orders can have multiple lines.
        shopify_trackings = list({tn for row in shopify_rows for tn in row.get("tracking_numbers", [])})
        cost_by_tracking_shopify = {}
        for i in range(0, len(shopify_trackings), 200):
            chunk = shopify_trackings[i:i+200]
            try:
                res = supabase.table("shipping_labels").select("tracking_number,cost")\
                    .eq("business_id", business_id).in_("tracking_number", chunk).execute()
                for row in (res.data or []):
                    cost_by_tracking_shopify[row["tracking_number"]] = row.get("cost") or 0
            except Exception as e:
                errors["shopify_shipping_match"] = str(e)

        order_subtotal_by_order = {}
        order_shipping_cost_by_order = {}
        for row in shopify_rows:
            oid = row["order_id"]
            order_subtotal_by_order[oid] = order_subtotal_by_order.get(oid, 0.0) + row["revenue"]
            if oid not in order_shipping_cost_by_order:
                order_shipping_cost_by_order[oid] = sum(
                    cost_by_tracking_shopify.get(tn, 0) or 0 for tn in row.get("tracking_numbers", [])
                )

        # Preserve any real SKU already stored for this exact line item — matched by
        # the same stable composite id used for the upsert itself (order_id +
        # Shopify's real line-item id), same rule as the eBay side above.
        # PREVIOUSLY matched by (order_id, title) instead, based on a since-removed
        # id scheme where the id used to embed the sku itself. That approach has a
        # real collision risk with the CURRENT id scheme: an order with two line
        # items sharing the same title (confirmed this happens — e.g. multiple
        # identical units bought as separate cart lines) would only remember one
        # corrected SKU per title and could silently apply it to the wrong line.
        # Matching by the exact id removes that risk entirely.
        candidate_shopify_ids = [
            f"shopify:{row['order_id']}:{row.get('line_item_id') or f'noid-{i}'}"
            for i, row in enumerate(shopify_rows)
        ]
        existing_sku_by_shopify_id = {}
        for i in range(0, len(candidate_shopify_ids), 200):
            chunk = candidate_shopify_ids[i:i+200]
            try:
                res = supabase.table("orders").select("id,sku").in_("id", chunk).execute()
                for r in (res.data or []):
                    existing_sku_by_shopify_id[r["id"]] = r.get("sku")
            except Exception:
                pass

        shopify_skipped = 0
        for i, row in enumerate(shopify_rows):
            net = _safe2(row.get("net", row["revenue"]))
            oid = row["order_id"]
            order_subtotal = order_subtotal_by_order.get(oid, 0.0)
            order_shipping_cost = order_shipping_cost_by_order.get(oid, 0.0)
            shipping_share = _safe2(order_shipping_cost * (row["revenue"] / order_subtotal)) if order_subtotal > 0 else 0.0
            final_net = _safe2(net - shipping_share)
            trackings = row.get("tracking_numbers", [])
            # Stable ID uses Shopify's own real line-item ID, not a positional index —
            # a positional 'i' shifts whenever the fetched list's order or length
            # changes between runs (a new order appearing earlier, pagination
            # differences, date-window edges moving), silently minting a NEW id for
            # the SAME logical line item every sync cycle and leaving the old row
            # behind forever. That was the actual cause of orders appearing
            # duplicated many times over — not something introduced today, a
            # pre-existing bug this line_item_id fixes at the root.
            line_item_id = row.get("line_item_id") or f"noid-{i}"  # fallback only for
            # rows Shopify somehow returned without a line item id at all (shouldn't
            # normally happen) — keeps behavior no worse than before for that edge case
            record_id = f"shopify:{row['order_id']}:{line_item_id}"
            existing_sku = existing_sku_by_shopify_id.get(record_id)
            final_sku = existing_sku if (existing_sku and not _is_blank_sku(existing_sku)) else row["sku"]
            record = {
                "id": record_id,
                "business_id": business_id, "platform": "Shopify", "order_id": row["order_id"],
                "sku": final_sku, "title": row["title"], "quantity": row["quantity"],
                "order_date": row["order_date"], "gross_revenue": _safe2(row["revenue"]),
                "fee": _safe2(row.get("fee", 0)), "net": net,
                "tracking_number": ",".join(trackings) if trackings else None,
                "shipping_cost": shipping_share,
                "final_net": final_net,
            }
            try:
                supabase.table("orders").upsert(record, on_conflict="id").execute()
                upserted += 1
            except Exception as e:
                shopify_skipped += 1
                print(f"sync_orders_for_business: skipped bad Shopify row {record.get('id')}: {e}")
        if shopify_skipped:
            errors["shopify_skipped_rows"] = f"{shopify_skipped} row(s) failed to upsert — see logs for details"
    except Exception as e:
        errors["shopify"] = str(e)

    return {"upserted": upserted, "errors": errors}

def sync_orders_for_business(business_id: str, days_back: int = 90, resume: bool = True) -> dict:
    """Pulls orders (with real fees, refunds, and shipping cost) from eBay + Shopify and
    upserts them into the local `orders` table — one MONTH at a time, not the whole
    range in one shot. Each month is its own independent call: if one month fails or
    times out, the others still land, and the next run just continues from wherever
    it left off (upserts are idempotent, so re-running already-synced months is harmless).
    Also checkpoints progress in app_settings (ORDERS_SYNC_CHECKPOINT) so an interruption
    — a deploy restart, a dropped connection, anything — resumes from the last completed
    month instead of re-walking from the start every time. Pass resume=False to force a
    full re-walk regardless of checkpoint (needed after a code change that affects fields
    on months already marked complete).
    This is the ONLY place that hits the live APIs for financial data — Financials
    itself just queries the local table, so filtering by date range is instant."""
    import datetime as _dt

    now = _dt.datetime.utcnow()
    range_start = now - _dt.timedelta(days=days_back)
    total_upserted = 0
    all_errors = {}

    month_start = range_start.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    checkpoint_str = None
    if resume:
        settings = get_ebay_settings(business_id)
        checkpoint_str = settings.get("ORDERS_SYNC_CHECKPOINT", "") or None
    if checkpoint_str:
        try:
            checkpoint_dt = _dt.datetime.strptime(checkpoint_str, "%Y-%m-%d")
            if checkpoint_dt >= month_start:
                # Resume from the month AFTER the last one that fully completed
                resumed_month = (checkpoint_dt.replace(day=28) + _dt.timedelta(days=4)).replace(day=1)
                if resumed_month > month_start:
                    month_start = resumed_month
                    print(f"sync_orders_for_business: resuming from checkpoint, starting at {month_start.strftime('%Y-%m')}")
        except ValueError:
            pass

    while month_start < now:
        next_month = (month_start.replace(day=28) + _dt.timedelta(days=4)).replace(day=1)
        window_end = min(next_month, now) - _dt.timedelta(seconds=1)
        window_start_iso = max(month_start, range_start).strftime("%Y-%m-%dT00:00:00.000Z")
        window_end_iso = window_end.strftime("%Y-%m-%dT%H:%M:%S.%f")[:-3] + "Z"

        label = month_start.strftime("%Y-%m")
        try:
            result = _sync_orders_window(business_id, window_start_iso, window_end_iso)
            total_upserted += result["upserted"]
            if result.get("errors"):
                all_errors[label] = result["errors"]
            print(f"sync_orders_for_business: month {label} -> {result['upserted']} row(s)"
                  + (f", errors: {result['errors']}" if result.get("errors") else ""))
            # Checkpoint AFTER each month succeeds, so an interruption mid-next-month
            # still resumes correctly rather than re-doing this completed one.
            try:
                save_ebay_setting(business_id, "ORDERS_SYNC_CHECKPOINT", month_start.strftime("%Y-%m-%d"))
            except Exception:
                pass
        except Exception as e:
            all_errors[label] = str(e)
            print(f"sync_orders_for_business: month {label} FAILED entirely: {e}")

        month_start = next_month

    # Reached the end of the range successfully — clear the checkpoint so a future
    # full re-walk (e.g. after a code change) starts fresh instead of resuming past everything.
    try:
        save_ebay_setting(business_id, "ORDERS_SYNC_CHECKPOINT", "")
    except Exception:
        pass

    return {"upserted": total_upserted, "errors": all_errors}


_sync_status = {}  # business_id -> {"running": bool, "result": dict|None, "started_at": iso, "finished_at": iso|None}

async def _run_sync_background(business_id: str, days_back: int, resume: bool = True):
    import asyncio, datetime as _dt
    try:
        result = await asyncio.to_thread(sync_orders_for_business, business_id, days_back, resume)
        _sync_status[business_id] = {
            "running": False, "result": result,
            "started_at": _sync_status.get(business_id, {}).get("started_at"),
            "finished_at": _dt.datetime.utcnow().isoformat(),
        }
    except Exception as e:
        _sync_status[business_id] = {
            "running": False, "result": {"upserted": 0, "errors": {"fatal": str(e)}},
            "started_at": _sync_status.get(business_id, {}).get("started_at"),
            "finished_at": _dt.datetime.utcnow().isoformat(),
        }

def backfill_tracking_numbers(business_id: str) -> dict:
    """Fills in tracking_number for orders that don't have one, WITHOUT re-fetching
    orders or fee data at all — just the per-order shipping_fulfillment lookup, which
    is the only piece that was ever wrong. Much faster than a full sync_orders_for_business
    pass since it skips getOrders and the Finance API entirely."""
    import requests as _req

    res = supabase.table("orders").select("id,order_id").eq("business_id", business_id)\
        .eq("platform", "eBay").is_("tracking_number", "null").execute()
    rows = res.data or []
    order_ids = list(set(r["order_id"] for r in rows if r.get("order_id")))

    token = get_ebay_access_token(business_id)
    tracking_by_order = {}
    for oid in order_ids:
        try:
            r = _req.get(
                f"{EBAY_API_BASE}/sell/fulfillment/v1/order/{oid}/shipping_fulfillment",
                headers=ebay_headers(token, content_language=False), timeout=15,
            )
            if r.status_code == 200:
                for f in r.json().get("fulfillments", []):
                    tn = f.get("shipmentTrackingNumber")
                    if tn:
                        tracking_by_order.setdefault(oid, []).append(tn)
        except Exception:
            continue

    updated = 0
    for row in rows:
        trackings = tracking_by_order.get(row["order_id"])
        if trackings:
            try:
                supabase.table("orders").update({"tracking_number": ",".join(trackings)}).eq("id", row["id"]).execute()
                updated += 1
            except Exception:
                pass

    return {"updated": updated, "orders_checked": len(order_ids), "rows_missing": len(rows)}

def backfill_legacy_item_ids(business_id: str) -> dict:
    """Fills in legacy_item_id for orders that need it — specifically only orders that
    ALSO still have a blank SKU, since that's the only reason this field matters. Scoping
    to just those (~600-900 orders, not all 11,000+) is what actually makes this finishable —
    trying to backfill every historical order took hours and kept getting killed by
    routine deploys before finishing."""
    import requests as _req

    rows = []
    page_size = 1000
    start = 0
    while True:
        res = supabase.table("orders").select("id,order_id,sku").eq("business_id", business_id)\
            .eq("platform", "eBay").is_("legacy_item_id", "null")\
            .range(start, start + page_size - 1).execute()
        page = res.data or []
        rows.extend(page)
        if len(page) < page_size:
            break
        start += page_size

    rows = [r for r in rows if _is_blank_sku(r.get("sku"))]

    order_ids = list(set(r["order_id"] for r in rows if r.get("order_id")))

    token = get_ebay_access_token(business_id)
    legacy_by_id = {}  # our composite id -> legacyItemId
    for oid in order_ids:
        try:
            r = _req.get(f"{EBAY_API_BASE}/sell/fulfillment/v1/order/{oid}",
                         headers=ebay_headers(token, content_language=False), timeout=15)
            if r.status_code == 200:
                order = r.json()
                for li in order.get("lineItems", []):
                    record_id = f"ebay:{oid}:{li.get('lineItemId', '')}"
                    legacy_by_id[record_id] = li.get("legacyItemId", "")
        except Exception:
            continue

    updated = 0
    for row in rows:
        legacy_id = legacy_by_id.get(row["id"])
        if legacy_id:
            try:
                supabase.table("orders").update({"legacy_item_id": legacy_id}).eq("id", row["id"]).execute()
                updated += 1
            except Exception:
                pass

    return {"updated": updated, "orders_checked": len(order_ids), "rows_missing": len(rows)}

@app.post("/api/financials/backfill-legacy-item-ids")
async def backfill_legacy_item_ids_now(request: Request):
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    import asyncio
    if _sync_status.get(business_id, {}).get("running"):
        return {"ok": True, "already_running": True}
    _sync_status[business_id] = {"running": True, "result": None, "started_at": __import__("datetime").datetime.utcnow().isoformat(), "finished_at": None}
    async def _run():
        import datetime as _dt
        try:
            result = await asyncio.to_thread(backfill_legacy_item_ids, business_id)
            _sync_status[business_id] = {"running": False, "result": result, "started_at": _sync_status.get(business_id, {}).get("started_at"), "finished_at": _dt.datetime.utcnow().isoformat()}
        except Exception as e:
            _sync_status[business_id] = {"running": False, "result": {"error": str(e)}, "started_at": _sync_status.get(business_id, {}).get("started_at"), "finished_at": _dt.datetime.utcnow().isoformat()}
    asyncio.create_task(_run())
    return {"ok": True, "started": True}

@app.post("/api/orders/backfill-sku-by-item-id")
async def backfill_sku_by_item_id(request: Request, file: UploadFile = File(...)):
    """The RELIABLE version of the SKU backfill — matches on exact legacy_item_id
    instead of fuzzy title text. Expects a CSV with 'Lot Name' and 'ITEM_ID' columns."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    import io, csv as _csv

    content = await file.read()
    text = content.decode("utf-8-sig", errors="replace")
    reader = _csv.DictReader(io.StringIO(text))
    csv_rows = list(reader)

    def _fetch_all(table, select_cols):
        all_rows = []
        page_size = 1000
        start = 0
        while True:
            res = supabase.table(table).select(select_cols).eq("business_id", business_id)\
                .range(start, start + page_size - 1).execute()
            page = res.data or []
            all_rows.extend(page)
            if len(page) < page_size:
                break
            start += page_size
        return all_rows

    acquisitions = _fetch_all("acquisitions", "sku,name")
    sku_by_lower_sku = {a["sku"].strip().lower(): a["sku"] for a in acquisitions if a.get("sku")}
    sku_by_lower_name = {}
    for a in acquisitions:
        if a.get("name") and a.get("sku"):
            sku_by_lower_name.setdefault(a["name"].strip().lower(), a["sku"])

    def resolve_sku(lot_name):
        key = (lot_name or "").strip().lower()
        return (sku_by_lower_sku.get(key) or sku_by_lower_name.get(key)) if key else None

    orders = _fetch_all("orders", "id,legacy_item_id,sku")
    order_by_legacy_id = {o["legacy_item_id"]: o for o in orders if o.get("legacy_item_id")}

    updates = []
    matched, no_lot_match, no_order_match, already_tagged = 0, 0, 0, 0
    for row in csv_rows:
        item_id = (row.get("ITEM_ID") or "").strip()
        lot_name = (row.get("Lot Name") or "").strip()
        if not item_id or not lot_name:
            continue
        order = order_by_legacy_id.get(item_id)
        if not order:
            no_order_match += 1
            continue
        existing = (order.get("sku") or "").strip().lower()
        if existing and existing not in ("", "(no sku)") and not existing.startswith("lister-"):
            already_tagged += 1
            continue
        real_sku = resolve_sku(lot_name)
        if not real_sku:
            no_lot_match += 1
            continue
        updates.append({"id": order["id"], "sku": f"{real_sku}-BF"})
        matched += 1

    updated = 0
    for i in range(0, len(updates), 500):
        chunk = updates[i:i+500]
        try:
            supabase.table("orders").upsert(chunk, on_conflict="id").execute()
            updated += len(chunk)
        except Exception as e:
            print(f"backfill_sku_by_item_id: batch {i}-{i+len(chunk)} failed: {e}")

    return {
        "ok": True, "csv_rows": len(csv_rows), "matched": matched, "updated": updated,
        "no_lot_match": no_lot_match, "no_order_match": no_order_match, "already_tagged": already_tagged,
    }

@app.post("/api/financials/backfill-tracking")
async def backfill_tracking_now(request: Request):
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    import asyncio
    if _sync_status.get(business_id, {}).get("running"):
        return {"ok": True, "already_running": True}
    _sync_status[business_id] = {"running": True, "result": None, "started_at": __import__("datetime").datetime.utcnow().isoformat(), "finished_at": None}
    async def _run():
        import datetime as _dt
        try:
            result = await asyncio.to_thread(backfill_tracking_numbers, business_id)
            _sync_status[business_id] = {"running": False, "result": result, "started_at": _sync_status.get(business_id, {}).get("started_at"), "finished_at": _dt.datetime.utcnow().isoformat()}
        except Exception as e:
            _sync_status[business_id] = {"running": False, "result": {"error": str(e)}, "started_at": _sync_status.get(business_id, {}).get("started_at"), "finished_at": _dt.datetime.utcnow().isoformat()}
    asyncio.create_task(_run())
    return {"ok": True, "started": True}

@app.post("/api/financials/sync-now")
async def sync_now(request: Request, days_back: int = None, resume: bool = True):
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    import asyncio, datetime as _dt
    if _sync_status.get(business_id, {}).get("running"):
        return {"ok": True, "already_running": True}
    if days_back is None:
        # Safety net for any caller that doesn't explicitly choose -- once the
        # historical backfill is done, a normal sync only needs to catch up
        # recent data (matches the automated worker's own steady-state
        # window), not walk the full ~2-year range every single time.
        settings = get_ebay_settings(business_id)
        backfilled = settings.get("ORDERS_BACKFILLED", "") == "true"
        days_back = 14 if backfilled else 720
    _sync_status[business_id] = {"running": True, "result": None, "started_at": _dt.datetime.utcnow().isoformat(), "finished_at": None}
    asyncio.create_task(_run_sync_background(business_id, days_back, resume))
    return {"ok": True, "started": True}

@app.get("/api/financials/sync-status")
async def sync_status(request: Request):
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    return _sync_status.get(business_id, {"running": False, "result": None, "started_at": None, "finished_at": None})

class OrderSkuUpdate(BaseModel):
    order_row_id: str
    new_sku: str

@app.post("/api/orders/sku")
async def update_order_sku(body: OrderSkuUpdate, request: Request):
    """Manually sets an order line's SKU directly — used by the Financials
    Uncategorized view's inline edit. Writes straight to the same column every
    sync's preservation check reads from, so once set here, it's protected by
    that same 'always keep an existing good SKU' rule — no different from a
    correction made via bulk SQL or the CSV backfill tools."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    new_sku = (body.new_sku or "").strip()
    if not new_sku:
        raise HTTPException(400, "new_sku is required")
    res = (supabase.table("orders").update({"sku": new_sku})
           .eq("business_id", business_id).eq("id", body.order_row_id).execute())
    if not res.data:
        raise HTTPException(404, "Order line not found")
    known_lot_skus = {a["sku"] for a in (supabase.table("acquisitions").select("sku")
                       .eq("business_id", business_id).execute().data or []) if a.get("sku")}
    recategorized = _lot_prefix(new_sku) in known_lot_skus
    return {"ok": True, "id": body.order_row_id, "sku": new_sku, "recategorized": recategorized}

@app.get("/api/debug/sku-lookup")
async def debug_sku_lookup(request: Request, order_id: str = None, title: str = None):
    """TEMPORARY diagnostic endpoint — remove after use. Shows the raw order
    row and any ebay_listing_status row(s) that could match it by item_id or
    title, so we can see directly whether a real sku_override exists or not.
    Also reports the true total row count in ebay_listing_status, unpaginated,
    to directly confirm or rule out the 1000-row Supabase cap theory."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    out = {}
    if order_id:
        out["order"] = (supabase.table("orders").select("*")
                         .eq("business_id", business_id).eq("order_id", order_id).execute().data)
    if title:
        out["listing_status_matches"] = (
            supabase.table("ebay_listing_status")
            .select("item_id,title,sku,sku_override,updated_at")
            .eq("business_id", business_id).ilike("title", f"%{title}%").execute().data)
    uncapped_res = (supabase.table("ebay_listing_status").select("item_id", count="exact")
                    .eq("business_id", business_id).limit(1).execute())
    out["total_ebay_listing_status_rows"] = uncapped_res.count
    return out


@app.get("/api/financials/uncategorized-orders")
async def api_financials_uncategorized_orders(request: Request):
    """Lean, standalone version of the 'uncategorized' section of /api/financials --
    queries finance_uncategorized_order_skus directly (confirmed ~1.5s even across
    full history, since the view does its own filtering/joining in Postgres) with
    NO date range restriction, instead of going through the full endpoint above,
    which fetches and processes the ENTIRE orders table for whatever date range is
    given first (11,000+ rows for a multi-year range) just to throw almost all of
    it away and extract this same small subset at the end. Built for the Uncat tab,
    which only ever needed this piece and was paying for the other, much heavier
    90% of that endpoint's work on every load for nothing."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    view_res = supabase.table("finance_uncategorized_order_skus").select("*").eq("business_id", business_id).execute()
    items = []
    net_total = 0
    for r in (view_res.data or []):
        net_total += r.get("net") or 0
        items.append({
            "id": r["id"], "sku": r.get("raw_sku") or "", "sku_override": r.get("sku_override"),
            "matched_ebay_sku": r.get("matched_ebay_sku"),
            "title": r.get("title"), "platform": r.get("platform"),
            "order_id": r.get("order_id"), "order_date": r.get("order_date", ""),
            "quantity": r.get("quantity"), "revenue": r.get("revenue"),
            "net": r.get("net"), "buyer_shipping": r.get("buyer_shipping") or 0,
            "shipping_cost": r.get("shipping_cost") or 0, "needs_sku": not (r.get("raw_sku") or ""),
        })
    items.sort(key=lambda r: r.get("order_date", ""), reverse=True)
    return {"count": len(items), "net": round(net_total, 2), "items": items}

@app.get("/api/financials")
async def api_financials(request: Request, start: str = None, end: str = None, include_shopify: bool = True):
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    import datetime as _dt
    now = _dt.datetime.utcnow()
    end_dt = _dt.datetime.fromisoformat(end) if end else now
    start_dt = _dt.datetime.fromisoformat(start) if start else (end_dt - _dt.timedelta(days=30))
    start_date_str = start_dt.strftime("%Y-%m-%d")
    end_date_str = end_dt.strftime("%Y-%m-%d")

    # A manual SKU override (set via the Uncategorized/My SKU Overrides tools)
    # lives on ebay_listing_status, keyed by the eBay item id -- it's applied
    # to the LISTING, but a historical order row's own `sku` column never
    # automatically updates to match, since those are two separate places
    # the SKU lives. Build the mapping once (legacy_item_id -> override) and
    # prefer it everywhere below an order's sku would otherwise be used, so
    # a correction made after the fact is actually reflected here instead of
    # showing the stale original value forever.
    override_rows = _fetch_all_for_business(business_id, "ebay_listing_status", "item_id,title,sku_override")
    override_by_item_id = {str(r["item_id"]): r["sku_override"] for r in override_rows
                            if r.get("item_id") and r.get("sku_override")}
    # Real gap found: legacy_item_id comes straight from eBay's API response
    # (li.get("legacyItemId", "")) and can legitimately come back empty for
    # some orders -- when that happens, the lookup above has nothing to match
    # on at all, no matter how it's cast. Title is virtually always present on
    # both sides, so it's a reliable second path to the same override when the
    # item_id link isn't there.
    def _norm_title(t):
        import re
        # Aggressive on purpose: strips everything except letters and digits,
        # lowercased. orders.title and ebay_listing_status.title come from two
        # separate eBay API calls (order sync vs listing sync) -- a plain
        # .strip().lower() match kept silently failing, which points at some
        # punctuation/whitespace/formatting difference between the two that a
        # simpler normalization wasn't catching. These titles are long and
        # specific enough that this level of normalization won't cause false
        # matches between genuinely different items.
        return re.sub(r'[^a-z0-9]', '', (t or '').lower())

    override_by_title = {_norm_title(r["title"]): r["sku_override"] for r in override_rows
                          if r.get("title") and r.get("sku_override")}

    def _lookup_override(r):
        legacy_id = r.get("legacy_item_id")
        override = override_by_item_id.get(str(legacy_id)) if legacy_id else None
        if not override:
            override = override_by_title.get(_norm_title(r.get("title")))
        return override

    def _effective_sku(r):
        override = _lookup_override(r)
        return override if override else (r.get("sku") or "")

    # Financials reads straight from the local `orders` table — shipping_cost and final_net
    # are already-computed, stored columns (written by apply_shipping_matches whenever a
    # Pirate Ship CSV is uploaded, and by the sync jobs). No computation happens here.
    query = supabase.table("orders").select("*").eq("business_id", business_id)\
        .gte("order_date", start_date_str).lte("order_date", end_date_str)
    if not include_shopify:
        query = query.eq("platform", "eBay")
    res = query.execute()
    rows = res.data or []
    rows.sort(key=lambda r: r.get("order_date", ""), reverse=True)

    order_lines = [{
        "id": r["id"], "sku": _effective_sku(r), "title": r["title"], "platform": r["platform"], "order_id": r["order_id"],
        "quantity": r["quantity"], "revenue": r["gross_revenue"], "net": r["final_net"],
        "buyer_shipping": r.get("buyer_shipping") or 0,
        "shipping_cost": r.get("shipping_cost") or 0, "order_date": r.get("order_date", ""),
    } for r in rows]

    last_sync_res = supabase.table("orders").select("synced_at").eq("business_id", business_id)\
        .order("synced_at", desc=True).limit(1).execute()
    last_synced_at = (last_sync_res.data or [{}])[0].get("synced_at")

    # Uncategorized breakdown is now sourced directly from the
    # finance_uncategorized_order_skus view -- it does the same
    # override-lookup + lot-prefix join in Postgres itself, so it can never
    # be silently truncated by a REST-layer row cap the way the old
    # Python-side loop (which had to pull the *entire* listings table into
    # memory to build its lookup) could be. Scoped to the same date range
    # as everything else on this page.
    view_res = (supabase.table("finance_uncategorized_order_skus").select("*")
                .eq("business_id", business_id)
                .gte("order_date", start_date_str).lte("order_date", end_date_str).execute())
    uncategorized_order_items = []
    uncategorized_order_net = 0
    for r in (view_res.data or []):
        uncategorized_order_net += r.get("net") or 0
        uncategorized_order_items.append({
            "id": r["id"], "sku": r.get("raw_sku") or "", "sku_override": r.get("sku_override"),
            "matched_ebay_sku": r.get("matched_ebay_sku"),
            "title": r.get("title"), "platform": r.get("platform"),
            "order_id": r.get("order_id"), "order_date": r.get("order_date", ""),
            "quantity": r.get("quantity"), "revenue": r.get("revenue"),
            "net": r.get("net"), "buyer_shipping": r.get("buyer_shipping") or 0,
            "shipping_cost": r.get("shipping_cost") or 0, "needs_sku": not (r.get("raw_sku") or ""),
        })
    uncategorized_order_items.sort(key=lambda r: r.get("order_date", ""), reverse=True)

    return {
        "start": start_date_str,
        "end": end_date_str,
        "order_lines": order_lines,
        "last_synced_at": last_synced_at,
        "totals": {
            "orders": len(rows),
            "quantity": sum(r["quantity"] for r in rows),
            "revenue": round(sum(r["gross_revenue"] for r in rows), 2),
            "net": round(sum(r["final_net"] for r in rows), 2),
        },
        "uncategorized": {
            "count": len(uncategorized_order_items),
            "net": round(uncategorized_order_net, 2),
            "items": uncategorized_order_items,
        },
        "errors": {},
    }

def _fetch_all_for_business(business_id: str, table: str, select_cols: str) -> list:
    all_rows, start_i, page_size = [], 0, 1000
    q_base = supabase.table(table).select(select_cols).eq("business_id", business_id)
    while True:
        res = q_base.range(start_i, start_i + page_size - 1).execute()
        page = res.data or []
        all_rows.extend(page)
        if len(page) < page_size:
            break
        start_i += page_size
    return all_rows

def _compute_inventory_snapshot_value(business_id: str) -> float:
    """Current total (price x qty) across matched eBay/Shopify inventory, avoiding
    double-counting a matched pair. Always 'as of right now' -- shared by both the
    live /api/analytics endpoint and the daily snapshot worker below, so there's
    exactly one place this math lives."""
    inv_rows = _fetch_all_for_business(business_id, "inventory_match", "ebay_id,shopify_id")
    ebay_by_id = {r["id"]: r for r in _fetch_all_for_business(business_id, "ebay_inventory", "id,price,quantity")}
    shopify_by_id = {r["id"]: r for r in _fetch_all_for_business(business_id, "shopify_inventory", "id,price,quantity,product_id")}
    shopify_by_product_id = {r["product_id"]: r for r in shopify_by_id.values() if r.get("product_id")}
    value = 0.0
    for row in inv_rows:
        e = ebay_by_id.get(row.get("ebay_id")) or {}
        s = shopify_by_id.get(row.get("shopify_id")) or shopify_by_product_id.get(row.get("shopify_id")) or {}
        price = e.get("price") if e.get("price") is not None else s.get("price")
        qty = e.get("quantity") if e.get("quantity") is not None else s.get("quantity")
        if price is not None and qty is not None:
            value += float(price) * float(qty)
    return round(value, 2)

def _nearest_inventory_snapshot(business_id: str, date_str: str):
    """Closest recorded Inventory Snapshot Value on or before date_str, using the
    analytics_snapshots history (CSV backfill + daily tracking). Extracted from
    the live /api/analytics endpoint's Inventory Growth Multiple logic so the
    daily snapshot worker can share the exact same lookup for YTD Inventory
    Appreciation, instead of duplicating it. Must exclude NULL rows (the CSV
    backfill's gap months) -- otherwise 'latest row on or before this date'
    could BE a gap row, silently returning None even though a perfectly good
    earlier value exists (a real bug found and fixed once already)."""
    res = supabase.table("analytics_snapshots").select("snapshot_date,inventory_snapshot_value")\
        .eq("business_id", business_id).lte("snapshot_date", date_str)\
        .not_.is_("inventory_snapshot_value", "null")\
        .order("snapshot_date", desc=True).limit(1).execute()
    row = (res.data or [None])[0]
    return (row.get("inventory_snapshot_value"), row.get("snapshot_date")) if row else (None, None)

# ── Generic Analytics query cache ─────────────────────────────────────────
# Why this exists: the Analytics page's heaviest cost isn't any one figure --
# it's that /api/analytics and /api/analytics/monthly-trend recompute
# everything live, on every single page load, even though most loads are
# just "open the page and look at the default view" (same date range as
# last time, nothing actually changed). This is the same 'precompute it,
# don't recalculate live' principle already applied to Net Cash Yield,
# generalized to any date range instead of just YTD.
#
# Design: a plain read-through cache keyed by (business_id, endpoint,
# start, end). A request first checks for a cached row fresher than
# ANALYTICS_CACHE_MAX_AGE_SECONDS; if found, returns it instantly with zero
# computation. If not (first time, or stale), it falls through to computing
# live -- exactly as before -- then SAVES that result into the cache before
# returning, so the next request for the same range is fast. On top of that,
# analytics_cache_refresh_worker (below) proactively recomputes the page's
# own DEFAULT range every 2 hours regardless of whether anyone's looking, so
# it's essentially always warm for the common case.
ANALYTICS_CACHE_MAX_AGE_SECONDS = 2 * 60 * 60  # 2 hours

def _get_analytics_cache(business_id: str, endpoint: str, start: str, end: str):
    import datetime as _dt4
    res = supabase.table("analytics_query_cache").select("payload,computed_at")\
        .eq("business_id", business_id).eq("endpoint", endpoint)\
        .eq("start_date", start).eq("end_date", end).limit(1).execute()
    row = (res.data or [None])[0]
    if not row:
        return None
    computed_at = _dt4.datetime.fromisoformat(row["computed_at"].replace("Z", "+00:00"))
    age = (_dt4.datetime.now(_dt4.timezone.utc) - computed_at).total_seconds()
    if age > ANALYTICS_CACHE_MAX_AGE_SECONDS:
        return None
    return row["payload"]

def _set_analytics_cache(business_id: str, endpoint: str, start: str, end: str, payload: dict):
    import datetime as _dt4
    record = {
        "business_id": business_id, "endpoint": endpoint, "start_date": start, "end_date": end,
        "payload": payload, "computed_at": _dt4.datetime.now(_dt4.timezone.utc).isoformat(),
    }
    existing = supabase.table("analytics_query_cache").select("id").eq("business_id", business_id)\
        .eq("endpoint", endpoint).eq("start_date", start).eq("end_date", end).limit(1).execute()
    if existing.data:
        supabase.table("analytics_query_cache").update(record).eq("id", existing.data[0]["id"]).execute()
    else:
        supabase.table("analytics_query_cache").insert(record).execute()

def _compute_green_revenue(business_id: str, start_date_str: str, end_date_str: str) -> float:
    """'Green Revenue' -- sum of actual order revenue (gross_revenue, per the
    user's own flag: is lot profit > 1 = Green) PLUS cash payments belonging to
    green lots, within the SELECTED DATE RANGE, for lots that are currently
    green.

    Rebuilt after the first version was wrong: it summed each green lot's
    all-time total_payouts regardless of date range, so it always showed a
    lifetime figure no matter what filter was applied (confirmed by the user --
    showed $981k for "last month", which was actually all-time revenue across
    every profitable lot ever).

    The real problem underneath, which the user identified themselves: a lot's
    "green" status is evaluated NOW, but we're filtering PAST sales -- without
    knowing when it actually turned green, a sale from before it was profitable
    could get miscounted as "green revenue". Solved without a daily snapshot
    table: acquisitions.became_green_at is now stamped once, at the moment a
    lot's profit crosses above 1 (see _apply_shopify_sales_to_profit), and
    cleared if it drops back down. A sale only counts here if the lot is
    currently green AND the sale happened on or after that lot's became_green_at
    -- so sales from while it was still red are correctly excluded.

    Cash from a green lot follows the exact same became_green_at boundary --
    a spread payment's contribution is bounded to days on/after that date, and
    a precise payment only counts if dated on/after it, same as orders."""
    lot_rows = _fetch_all_for_business(business_id, "acquisitions", "sku,profit,became_green_at")
    green_lots = {r["sku"]: r["became_green_at"] for r in lot_rows if r.get("sku") and (r.get("profit") or 0) > 1 and r.get("became_green_at")}
    if not green_lots:
        return 0.0

    order_rows = _fetch_all_for_business(business_id, "orders", "sku,gross_revenue,order_date")
    order_rows = [r for r in order_rows if start_date_str <= (r.get("order_date") or "") <= end_date_str]
    total = 0.0
    for r in order_rows:
        sku = r.get("sku") or ""
        prefix = _lot_prefix(sku) if sku else None
        became_green_at = green_lots.get(prefix)
        if not became_green_at:
            continue
        order_date = r.get("order_date") or ""
        # became_green_at is a full timestamp, order_date is just a date -- compare
        # on date portion only, so the transition day itself still counts
        if order_date >= became_green_at[:10]:
            total += float(r.get("gross_revenue") or 0)

    min_date_by_sku = {sku: ts[:10] for sku, ts in green_lots.items()}
    total += _compute_cash_in_range(business_id, start_date_str, end_date_str, only_skus=set(green_lots.keys()), min_date_by_sku=min_date_by_sku)
    return round(total, 2)

@app.get("/api/analytics/green-revenue-breakdown")
async def api_green_revenue_breakdown(request: Request, start: str, end: str):
    """Per-lot detail behind a Green Revenue figure, for actually verifying a
    number that looks surprising instead of just asserting it's correct --
    same exact logic as _compute_green_revenue, but returns each contributing
    lot's sku, became_green_at, and how much of its revenue fell in this range,
    instead of collapsing straight to one total."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    lot_rows = _fetch_all_for_business(business_id, "acquisitions", "sku,profit,became_green_at")
    green_lots = {r["sku"]: r["became_green_at"] for r in lot_rows if r.get("sku") and (r.get("profit") or 0) > 1 and r.get("became_green_at")}

    order_rows = _fetch_all_for_business(business_id, "orders", "sku,gross_revenue,order_date")
    order_rows = [r for r in order_rows if start <= (r.get("order_date") or "") <= end]

    by_lot = {}
    for r in order_rows:
        sku = r.get("sku") or ""
        prefix = _lot_prefix(sku) if sku else None
        became_green_at = green_lots.get(prefix)
        if not became_green_at:
            continue
        order_date = r.get("order_date") or ""
        if order_date >= became_green_at[:10]:
            entry = by_lot.setdefault(prefix, {"sku": prefix, "became_green_at": became_green_at, "revenue": 0.0, "order_count": 0})
            entry["revenue"] += float(r.get("gross_revenue") or 0)
            entry["order_count"] += 1

    rows = sorted(by_lot.values(), key=lambda x: x["revenue"], reverse=True)
    for r in rows:
        r["revenue"] = round(r["revenue"], 2)
    return {"start": start, "end": end, "total": round(sum(r["revenue"] for r in rows), 2), "lots": rows}

@app.get("/api/analytics")
async def api_analytics(request: Request, start: str = None, end: str = None):
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    import datetime as _dt
    now = _dt.datetime.utcnow()
    end_dt = _dt.datetime.fromisoformat(end) if end else now
    start_dt = _dt.datetime.fromisoformat(start) if start else (end_dt - _dt.timedelta(days=30))
    start_date_str = start_dt.strftime("%Y-%m-%d")
    end_date_str = end_dt.strftime("%Y-%m-%d")

    cached = _get_analytics_cache(business_id, "analytics", start_date_str, end_date_str)
    if cached is not None:
        return cached

    payload = _compute_analytics_payload(business_id, start_date_str, end_date_str)
    _set_analytics_cache(business_id, "analytics", start_date_str, end_date_str, payload)
    return payload

def _compute_analytics_payload(business_id: str, start_date_str: str, end_date_str: str) -> dict:
    """The actual heavy computation behind /api/analytics -- extracted out of the
    route handler so both the live endpoint (on a cache miss) and
    analytics_cache_refresh_worker (proactively, every 2 hours) can call the
    exact same logic instead of duplicating it."""
    import datetime as _dt
    start_dt = _dt.date.fromisoformat(start_date_str)
    end_dt = _dt.date.fromisoformat(end_date_str)
    num_days = max((end_dt - start_dt).days + 1, 1)

    # --- Inventory Snapshot Value ---
    # This is a CURRENT, as-of-right-now figure -- there's no historical daily
    # snapshot stored anywhere, so it does NOT change based on the date range
    # above (a past inventory value would need data we don't capture). Shown
    # alongside the period-filtered metrics but not filtered by them; the UI
    # labels it "as of now" rather than implying otherwise.
    inventory_snapshot_value = _compute_inventory_snapshot_value(business_id)
    green_revenue = _compute_green_revenue(business_id, start_date_str, end_date_str)

    # --- Order-level figures (within the selected date range) ---
    order_rows = _fetch_all_for_business(business_id, "orders", "gross_revenue,final_net,quantity,order_date")
    order_rows = [r for r in order_rows if start_date_str <= (r.get("order_date") or "") <= end_date_str]
    total_order_revenue = sum(r.get("gross_revenue") or 0 for r in order_rows)
    total_net_revenue = sum(r.get("final_net") or 0 for r in order_rows)
    total_qty_sold = sum(r.get("quantity") or 0 for r in order_rows)
    avg_sales_per_day = round(len(order_rows) / num_days, 2)

    # --- Cash in this range (once, shared by Revenue, Net Sales, and Avg Sales/Day($) below) ---
    cash_in_range = _compute_cash_in_range(business_id, start_date_str, end_date_str)

    # --- Revenue: eBay + Shopify order revenue (both already in the same orders
    # table, no platform filter anywhere -- confirmed) PLUS cash in this range.
    total_revenue = round(total_order_revenue + cash_in_range, 2)
    # FIXED: this previously used order-only revenue (no cash), inconsistent
    # with every other revenue-like figure on this page now including cash.
    avg_sales_per_day_dollars = round(total_revenue / num_days, 2)

    # Green Revenue as a % of Revenue: the metric that actually distinguishes
    # "portfolio genuinely getting more profitable over time" from "raw sales
    # dollars just moved because of seasonality/listing activity" -- the dollar
    # figure alone tracks overall sales volume mechanically (it's a subset of
    # the same sales), so it can look like it "just follows Revenue" even when
    # the underlying share of green lots is climbing. This is the number that
    # actually tests that.
    green_revenue_pct = round(green_revenue / total_revenue * 100, 1) if total_revenue else None

    # --- Avg Order Price: renamed from "Average Sale Price" and redefined per
    # explicit request -- Revenue (including cash, same figure as the Revenue
    # card above) divided by ORDER COUNT, not order-only revenue divided by
    # units sold. This is deliberately the "higher number" the user did the
    # math for themselves (Revenue $102,666 / 432 orders = ~$237).
    avg_order_price = round(total_revenue / len(order_rows), 2) if order_rows else 0

    # --- Net Sales: revenue AFTER eBay/Shopify fees (final_net, not gross_revenue)
    # plus cash payments dated (or spread) within the range ---
    # FIXED a real bug here: this previously added cash on top of gross_revenue,
    # which is exactly backwards for something called "Net Sales" -- it could
    # (and did) show a number HIGHER than gross Revenue itself, which never made
    # sense. Net Sales should be smaller than gross Revenue before cash is even
    # added, since fees come out first. Now uses final_net (the same authoritative
    # post-fee figure Financials already uses) as the base.
    net_sales = round(total_net_revenue + cash_in_range, 2)

    # --- Inventory Growth Multiple: % change in Inventory Snapshot Value across the
    # selected range, using the shared _nearest_inventory_snapshot lookup (the
    # closest snapshot ON OR BEFORE each end of the range stands in for that
    # date, since snapshots aren't recorded for every single day).
    start_inv_value, start_inv_date = _nearest_inventory_snapshot(business_id, start_date_str)
    end_inv_value, end_inv_date = _nearest_inventory_snapshot(business_id, end_date_str)
    inventory_growth_pct = None
    if start_inv_value and end_inv_value is not None:
        inventory_growth_pct = round((end_inv_value - start_inv_value) / start_inv_value * 100, 1)

    # --- Average New Listings Per Day: count and $ (within the selected date range) ---
    listing_rows = _fetch_all_for_business(business_id, "listings", "price,created_at")
    listing_rows = [r for r in listing_rows if start_date_str <= (r.get("created_at") or "") <= end_date_str + "T23:59:59"]
    total_new_listings = len(listing_rows)
    total_new_listings_value = sum(r.get("price") or 0 for r in listing_rows)
    avg_new_listings_per_day_count = round(total_new_listings / num_days, 2)
    avg_new_listings_per_day_value = round(total_new_listings_value / num_days, 2)

    return {
        "start": start_date_str,
        "end": end_date_str,
        "num_days": num_days,
        "inventory_snapshot_value": inventory_snapshot_value,
        "green_revenue": green_revenue,
        "green_revenue_pct": green_revenue_pct,
        "avg_sales_per_day": avg_sales_per_day,
        "avg_sales_per_day_dollars": avg_sales_per_day_dollars,
        "avg_new_listings_per_day_count": avg_new_listings_per_day_count,
        "avg_new_listings_per_day_value": avg_new_listings_per_day_value,
        "totals_in_range": {
            "orders": len(order_rows),
            "quantity_sold": total_qty_sold,
            "revenue": round(total_revenue, 2),
            "cash": round(cash_in_range, 2),
            "net_sales": net_sales,
            "new_listings": total_new_listings,
            "new_listings_value": round(total_new_listings_value, 2),
            "inventory_growth_pct": inventory_growth_pct,
            "inventory_growth_start_date": start_inv_date,
            "inventory_growth_end_date": end_inv_date,
        },
    }

@app.get("/api/analytics/history")
async def api_analytics_history(request: Request, days: int = 90):
    """Returns the accumulated daily snapshots from analytics_snapshots, oldest
    to newest, for trend display. Empty/short at first -- one row gets added per
    business per day going forward by analytics_snapshot_worker; there's no
    backfill for dates before this feature existed, since that data was never
    captured."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    import datetime as _dt
    cutoff = (_dt.datetime.utcnow() - _dt.timedelta(days=days)).strftime("%Y-%m-%d")
    res = supabase.table("analytics_snapshots").select("*").eq("business_id", business_id)\
        .gte("snapshot_date", cutoff).order("snapshot_date").execute()
    return {"snapshots": res.data or []}

@app.get("/api/analytics/net-cash-yield")
async def api_net_cash_yield(request: Request, year: int = None):
    """Net Cash Yield for a given year: post-fee order revenue (orders.final_net,
    eBay+Shopify) PLUS cash attributed to that year PLUS all inventory purchases
    dated that year, subtracted.

    Cash attribution, per explicit request: a precise cash payment counts
    toward the year it's dated in, going forward. An old spread payment (a
    lot's cash with no real date attached) counts toward the year the LOT was
    purchased -- see _compute_cash_for_year for why the FULL amount is correct
    here, not a prorated slice, now that spreading is frozen.

    For the CURRENT year, reads the precomputed ytd_net_revenue/
    ytd_inventory_spend/ytd_cash columns the daily snapshot worker already
    writes, instead of re-summing a whole year of rows live on every page load
    (which was the actual slowdown). For a PAST, fully-closed year, computes it
    live -- a one-time bounded sum over a year that's already over and won't
    change, not something hit on every load the way 'current year so far' is."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    import datetime as _dt
    now = _dt.datetime.utcnow()
    target_year = year or now.year
    start_date_str = f"{target_year:04d}-01-01"
    is_ytd = target_year == now.year

    if is_ytd:
        snap = supabase.table("analytics_snapshots").select("snapshot_date,ytd_net_revenue,ytd_inventory_spend,ytd_cash,ytd_net_cash_yield,ytd_inventory_appreciation,ytd_net_business_appreciation")\
            .eq("business_id", business_id).gte("snapshot_date", start_date_str)\
            .order("snapshot_date", desc=True).limit(1).execute()
        row = (snap.data or [None])[0]
        total_net_revenue = (row or {}).get("ytd_net_revenue") or 0
        total_spend = (row or {}).get("ytd_inventory_spend") or 0
        total_cash = (row or {}).get("ytd_cash") or 0
        stored_yield = (row or {}).get("ytd_net_cash_yield")
        stored_appreciation = (row or {}).get("ytd_inventory_appreciation")
        stored_business_appreciation = (row or {}).get("ytd_net_business_appreciation")
        end_date_str = (row or {}).get("snapshot_date") or now.strftime("%Y-%m-%d")
    else:
        end_date_str = f"{target_year:04d}-12-31"
        order_rows = _fetch_all_for_business(business_id, "orders", "final_net,order_date")
        order_rows = [r for r in order_rows if start_date_str <= (r.get("order_date") or "") <= end_date_str]
        total_net_revenue = sum(r.get("final_net") or 0 for r in order_rows)
        acq_rows = _fetch_all_for_business(business_id, "acquisitions", "cost,date")
        acq_rows = [r for r in acq_rows if start_date_str <= (r.get("date") or "") <= end_date_str]
        total_spend = sum(r.get("cost") or 0 for r in acq_rows)
        total_cash = _compute_cash_for_year(business_id, target_year)
        stored_yield = None
        # Past year: appreciation is Dec 31 value minus Jan 1 value, both from
        # whatever real snapshot history exists near those dates.
        jan1_value, _ = _nearest_inventory_snapshot(business_id, start_date_str)
        dec31_value, _ = _nearest_inventory_snapshot(business_id, end_date_str)
        stored_appreciation = round(dec31_value - jan1_value, 2) if (jan1_value is not None and dec31_value is not None) else None
        stored_business_appreciation = None  # computed below once net_cash_yield is known

    net_sales = total_net_revenue + total_cash
    # Use the stored value directly when we have it (the current year, read
    # straight from the daily snapshot) instead of recomputing -- only falls
    # back to computing it here for a past year, which has no stored snapshot.
    net_cash_yield = stored_yield if stored_yield is not None else round(net_sales - total_spend, 2)
    inventory_appreciation = stored_appreciation
    net_business_appreciation = (
        stored_business_appreciation if stored_business_appreciation is not None
        else (round(net_cash_yield + inventory_appreciation, 2) if inventory_appreciation is not None else None)
    )

    return {
        "year": target_year,
        "start": start_date_str,
        "end": end_date_str,
        "is_ytd": is_ytd,
        "net_sales": round(net_sales, 2),
        "net_revenue_only": round(total_net_revenue, 2),
        "cash": round(total_cash, 2),
        "total_spend": round(total_spend, 2),
        "net_cash_yield": net_cash_yield,
        "inventory_appreciation": inventory_appreciation,
        "net_business_appreciation": net_business_appreciation,
    }

@app.get("/api/analytics/monthly-trend")
async def api_analytics_monthly_trend(request: Request, start: str = None, end: str = None):
    """See _compute_monthly_trend_payload for what this actually computes --
    this route is now just a thin cache-check wrapper around it (same
    read-through cache pattern as /api/analytics), since Green Revenue by
    month alone makes one live computation PER MONTH in range, which was
    almost certainly the single biggest source of the page feeling slow."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    import datetime as _dt
    now = _dt.datetime.utcnow()
    start_str = f"{start}-01" if start else "2025-01-01"
    end_str = f"{end}-01" if end else now.strftime("%Y-%m-01")

    cached = _get_analytics_cache(business_id, "monthly_trend", start_str, end_str)
    if cached is not None:
        return cached

    payload = _compute_monthly_trend_payload(business_id, start_str, end_str)
    _set_analytics_cache(business_id, "monthly_trend", start_str, end_str, payload)
    return payload

def _compute_monthly_trend_payload(business_id: str, start_str: str, end_str: str) -> dict:
    """Inventory spend vs. sales, grouped by month -- straight from acquisitions.cost
    (grouped by acquisitions.date), orders.gross_revenue (grouped by
    orders.order_date), and cash payments in that same month (via
    _compute_cash_in_range) -- Sales now includes cash, matching Revenue on the
    main Analytics view. Both eBay and Shopify order revenue were already
    captured here (no platform filter anywhere in these order queries, confirmed).

    Green Revenue by month is computed the SAME way as the live /api/analytics
    endpoint (_compute_green_revenue), once per month -- NOT read from
    analytics_snapshots. Realized after shipping the daily-snapshot version:
    there was never a real reason to limit this to the shallow snapshot history
    -- a lot's became_green_at plus dated order history already lets Green
    Revenue be computed for any past month directly, exactly like Inventory
    Spend vs. Sales. Full historical depth back to `start`, not just "since
    daily tracking began."

    Inventory Snapshot Value by month IS still sourced from analytics_snapshots
    -- unlike the other two, it's a point-in-time aggregate with no dated
    transaction history to reconstruct it from, so it genuinely can only go as
    far back as real snapshots exist (the CSV backfill + daily tracking).

    start_str/end_str are "YYYY-MM-01" full date strings, already resolved by
    the route handler above (defaults: start=2025-01, end=the current month)."""
    import datetime as _dt
    now = _dt.datetime.utcnow()

    # Full calendar month list between start and end (not just months that
    # happen to have acquisitions/order activity), so all three charts share a
    # clean, continuous x-axis with no gaps.
    all_months = []
    y, m = int(start_str[:4]), int(start_str[5:7])
    end_y, end_m = int(end_str[:4]), int(end_str[5:7])
    while (y, m) <= (end_y, end_m):
        all_months.append(f"{y:04d}-{m:02d}")
        m += 1
        if m > 12:
            m = 1
            y += 1

    acq_rows = _fetch_all_for_business(business_id, "acquisitions", "cost,date")
    order_rows = _fetch_all_for_business(business_id, "orders", "gross_revenue,order_date")
    snapshot_rows = _fetch_all_for_business(business_id, "analytics_snapshots", "snapshot_date,inventory_snapshot_value")

    spend_by_month, sales_by_month = {}, {}
    for r in acq_rows:
        d = (r.get("date") or "")[:7]
        if d in all_months:
            spend_by_month[d] = spend_by_month.get(d, 0) + (r.get("cost") or 0)
    for r in order_rows:
        d = (r.get("order_date") or "")[:7]
        if d in all_months:
            sales_by_month[d] = sales_by_month.get(d, 0) + (r.get("gross_revenue") or 0)

    inventory_spend = [round(spend_by_month.get(m, 0), 2) for m in all_months]

    today_str = now.strftime("%Y-%m-%d")
    current_month_str = now.strftime("%Y-%m")

    def _month_bounds(month):
        next_m_y, next_m_m = (int(month[:4]) + 1, 1) if month[5:7] == "12" else (int(month[:4]), int(month[5:7]) + 1)
        month_end = (_dt.date(next_m_y, next_m_m, 1) - _dt.timedelta(days=1)).strftime("%Y-%m-%d")
        if month == current_month_str:
            # The current month is still in progress -- a bar for it shouldn't
            # include days that haven't happened yet. Capping at today is also
            # what makes this match the top card's "This Month" quick-range,
            # which stops at today rather than the actual last day of the month.
            month_end = min(month_end, today_str)
        return f"{month}-01", month_end

    # Sales now includes cash in that same month, matching Revenue on the main
    # Analytics view -- both eBay and Shopify order revenue were already
    # captured here (no platform filter anywhere in these queries, confirmed).
    sales = []
    for month in all_months:
        month_start, month_end = _month_bounds(month)
        order_total = sales_by_month.get(month, 0)
        cash_total = _compute_cash_in_range(business_id, month_start, month_end)
        sales.append(round(order_total + cash_total, 2))

    # Green Revenue: one real call per month, same math as the live endpoint.
    green_revenue_by_month = []
    for month in all_months:
        month_start, month_end = _month_bounds(month)
        green_revenue_by_month.append(_compute_green_revenue(business_id, month_start, month_end))

    # Inventory Snapshot Value by month -- deliberately NOT tied to the start/end
    # filter above. Runs its own independent range: from whichever month has the
    # EARLIEST real data through the current month, always, regardless of what's
    # selected in Start Month/End Month. This chart's x-axis is compressed to
    # only the span where real data actually exists, per explicit request.
    #
    # Real bug found while building the gap-filling below: the old version could
    # pick up a gap row (inventory_snapshot_value = NULL, from months the user's
    # CSV backfill genuinely had no number for) as its "latest seen" value,
    # silently erasing a perfectly good earlier value instead of skipping past
    # it -- that's why January and everything after briefly went blank. Now:
    #  - an exact snapshot dated within the month wins outright
    #  - a gap with a REAL value on both sides gets the average of those two
    #    (a real midpoint, per explicit request for exactly this)
    #  - a gap with no earlier real value at all is left blank rather than
    #    guessed at, since there's nothing to average against -- but this only
    #    happens now if data is missing from the MIDDLE of the real range, since
    #    the leading all-blank months are trimmed off the x-axis entirely below
    real_points = sorted(
        [(r.get("snapshot_date"), r.get("inventory_snapshot_value")) for r in snapshot_rows if r.get("inventory_snapshot_value") is not None],
        key=lambda p: p[0],
    )
    if real_points:
        inv_start_month = real_points[0][0][:7]
        inv_months = []
        y, m = int(inv_start_month[:4]), int(inv_start_month[5:7])
        while (y, m) <= (now.year, now.month):
            inv_months.append(f"{y:04d}-{m:02d}")
            m += 1
            if m > 12:
                m = 1
                y += 1
    else:
        inv_months = []

    inventory_value_by_month = []
    for month in inv_months:
        month_start, month_end = _month_bounds(month)
        exact = None
        for d, v in real_points:
            if month_start <= d <= month_end:
                exact = v
        if exact is not None:
            inventory_value_by_month.append(round(exact, 2))
            continue
        before, after = None, None
        for d, v in real_points:
            if d <= month_end:
                before = v
            elif after is None:
                after = v
        if before is not None and after is not None:
            inventory_value_by_month.append(round((before + after) / 2, 2))
        elif before is not None:
            inventory_value_by_month.append(round(before, 2))
        else:
            inventory_value_by_month.append(None)

    # "Inventory multiple": spend as a % of sales for that month, e.g. $50k spend /
    # $100k sales = 50%. Precomputed here (not in JS) so the toggle is just a
    # dataset swap, not a recalculation.
    pct_multiple = [round((s / sa * 100), 1) if sa else None for s, sa in zip(inventory_spend, sales)]
    # "$ Variance": sales - inventory spend for that month, as a single summed
    # number per month rather than two separate series -- same pattern as the
    # % Multiple view, just a different combination of the same two numbers.
    dollar_variance = [round(sa - s, 2) for s, sa in zip(inventory_spend, sales)]

    # Green Revenue as a % of that month's Revenue -- same "sales" array already
    # represents total Revenue per month (order revenue + cash, matching the
    # live Revenue figure), so no new data fetch needed for this. Precomputed
    # here, same pattern as pct_multiple above, so the chart's toggle is just a
    # dataset swap.
    green_revenue_pct_by_month = [
        round((g / sa * 100), 1) if sa else None
        for g, sa in zip(green_revenue_by_month, sales)
    ]

    return {
        "start": start_str[:7],
        "end": end_str[:7],
        "months": all_months,
        "inventory_spend": inventory_spend,
        "green_revenue_by_month": green_revenue_by_month,
        "green_revenue_pct_by_month": green_revenue_pct_by_month,
        "inventory_value_months": inv_months,
        "inventory_value_by_month": inventory_value_by_month,
        "sales": sales,
        "pct_multiple": pct_multiple,
        "dollar_variance": dollar_variance,
    }

@app.get("/archive", response_class=HTMLResponse)
async def archive_page(request: Request):
    nav = get_nav_context(request)
    if nav is None:
        from fastapi.responses import RedirectResponse
        return RedirectResponse("/login", status_code=302)
    return templates.TemplateResponse("archive.html", {"request": request, "is_admin": nav["is_admin"], "account_label": nav["account_label"], "active_tab": "archive"})

@app.post("/api/listings/archive-batch")
async def archive_batch(request: Request):
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    try:
        res = supabase.table("listings").select("*").eq("business_id", business_id).neq("status", "archived").execute()
        items = res.data or []
        if items:
            ids = [str(i["id"]) for i in items]
            supabase.table("listings").update({"status": "archived"}).in_("id", ids).execute()
        return {"ok": True, "count": len(items)}
    except Exception as e:
        raise HTTPException(500, str(e))

# ── API: BATCH UPLOAD ─────────────────────────────────────────── #

class CreateGroup(BaseModel):
    session_id:    str
    condition:     str
    category_mode: str = "industrial"
    pricing_mode:  str = "always_search"
    lot_sku:       str = ""
    lot_suffix:    str = ""

@app.post("/api/groups")
async def create_group(body: CreateGroup, request: Request):
    try:
        business_id = require_auth(request)
        # Check scan limit
        biz = supabase.table("businesses").select("scan_count,scan_limit").eq("id", business_id).execute()
        if biz.data:
            scan_count = biz.data[0].get("scan_count") or 0
            scan_limit = biz.data[0].get("scan_limit") or 50
            if scan_count >= scan_limit:
                raise HTTPException(402, "Scan limit reached. Please contact us to upgrade your plan.")
            # Increment scan count
            supabase.table("businesses").update({"scan_count": scan_count + 1}).eq("id", business_id).execute()
        category_mode = body.category_mode if body.category_mode in ("industrial", "motors") else "industrial"
        pricing_mode = body.pricing_mode if body.pricing_mode in ("always_search", "api_first") else "always_search"
        res = supabase.table("listing_groups").insert({
            "session_id": body.session_id,
            "status":     "waiting",
            "quantity":   1,
            "condition":  body.condition,
            "category_mode": category_mode,
            "pricing_mode": pricing_mode,
            "lot_sku":    (body.lot_sku or "").strip(),
            "lot_suffix": (body.lot_suffix or "").strip(),
            "business_id": business_id,
            "created_at": datetime.utcnow().isoformat(),
        }).execute()
        import traceback
        print(f"Group insert result: {res}")
        data = res.data
        gid = data[0]["id"] if isinstance(data, list) and data else (data.get("id") if isinstance(data, dict) else None)
        if not gid:
            raise Exception(f"No ID returned. data={data}")
        return {"id": gid}
    except Exception as e:
        import traceback; traceback.print_exc()
        raise HTTPException(500, str(e))

class SubmitGroup(BaseModel):
    group_id:  str
    condition: str
    quantity:  int

@app.post("/api/listings/{item_id}/assign-lot-from-group")
async def assign_lot_from_group(item_id: str, request: Request):
    """Root-cause fix for lot SKUs coming in under the WRONG lot: the old
    auto-assign used whatever lot was active in whichever open browser tab
    happened to poll first — a stale tab from an earlier session could stamp
    its old lot onto brand-new items. This endpoint instead reads the lot that
    was stored ON THE GROUP at upload time (listing → photo → group), so the
    result is the same no matter which tab (or how many tabs) trigger it.
    Returns lot_sku="" if the group predates lot capture — caller may fall
    back to its own active lot in that case."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    res = supabase.table("listings").select("photo_id,ebay_sku").eq("id", item_id).limit(1).execute()
    if not res.data:
        raise HTTPException(404, "Listing not found")
    photo_id = res.data[0].get("photo_id")
    lot, suffix = "", ""
    if photo_id:
        gp = supabase.table("group_photos").select("group_id").eq("photo_id", photo_id).limit(1).execute()
        if gp.data:
            grp = (supabase.table("listing_groups").select("lot_sku,lot_suffix")
                   .eq("id", gp.data[0]["group_id"]).limit(1).execute())
            if grp.data:
                lot = (grp.data[0].get("lot_sku") or "").strip()
                suffix = (grp.data[0].get("lot_suffix") or "").strip()
    if not lot:
        return {"ok": True, "ebay_sku": None, "lot_sku": ""}
    new_sku = f"{lot}-{suffix}" if suffix else f"{lot}-"
    supabase.table("listings").update({"ebay_sku": new_sku}).eq("id", item_id).execute()
    return {"ok": True, "ebay_sku": new_sku, "lot_sku": lot}


@app.post("/api/groups/{group_id}/rescan")
async def rescan_group(group_id: str, request: Request):
    business_id = require_auth(request)
    try:
        supabase.table("listing_groups").update({"status": "pending"}).eq("id", group_id).execute()
        return {"ok": True}
    except Exception as e:
        raise HTTPException(500, str(e))

@app.get("/api/groups/{group_id}/listing")
async def get_group_listing(group_id: str, request: Request):
    try:
        gp = supabase.table("group_photos").select("photo_id").eq("group_id", group_id).limit(1).execute()
        if not gp.data:
            return {"listing": None}
        photo_id = gp.data[0]["photo_id"]
        res = supabase.table("listings").select("id,status,title,price").eq("photo_id", photo_id).limit(1).execute()
        return {"listing": res.data[0] if res.data else None}
    except Exception as e:
        raise HTTPException(500, str(e))

@app.post("/api/groups/submit")
async def submit_group(body: SubmitGroup, request: Request):
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    try:
        supabase.table("listing_groups").update({
            "condition": body.condition,
            "quantity":  body.quantity,
            "status":    "pending",
        }).eq("id", body.group_id).execute()
        return {"ok": True}
    except Exception as e:
        raise HTTPException(500, str(e))

@app.post("/api/listings/{item_id}/add-photos")
async def add_photos_to_listing(item_id: str, request: Request):
    """Adds extra photos to an already-scanned listing's photo group — for detail
    shots taken after the fact, not meant to go through mantle-scanner's AI pipeline.
    Safe to do post-scan: mantle-scanner only re-scans groups with status='pending'
    and explicitly skips any photo already linked in group_photos, so this can never
    trigger a re-scan. New photos just get pulled in automatically the next time this
    listing publishes to eBay/Shopify, since publishing already uses every photo in
    the group, not just the primary one."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    try:
        res = supabase.table("listings").select("photo_id").eq("id", item_id).limit(1).execute()
        if not res.data:
            raise HTTPException(404, "Listing not found")
        primary_pid = str(res.data[0].get("photo_id") or "")
        if not primary_pid:
            raise HTTPException(400, "Listing has no primary photo to attach a group to")

        group_row = supabase.table("group_photos").select("group_id").eq("photo_id", primary_pid).limit(1).execute()
        group_id = (group_row.data or [{}])[0].get("group_id", "")
        if not group_id:
            raise HTTPException(400, "Could not find this listing's photo group")

        form = await request.form()
        files = form.getlist("files")
        if not files:
            raise HTTPException(400, "No files provided")

        uploaded = []
        for idx, file in enumerate(files):
            contents = await file.read()
            dt = datetime.now()
            fn = f"{dt.strftime('%d%m%y')}_{dt.strftime('%H%M%S')}_{idx}_extra.jpg"
            supabase.storage.from_("part-photos").upload(
                path=fn, file=contents, file_options={"content-type": "image/jpeg", "upsert": "true"}
            )
            supabase.table("group_photos").insert({"group_id": group_id, "photo_id": fn}).execute()
            uploaded.append({"photo_id": fn, "url": photo_url(fn, thumb=True)})

        return {"ok": True, "uploaded": uploaded}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))

@app.delete("/api/listings/{item_id}/photos/{photo_id}")
async def delete_listing_photo(item_id: str, photo_id: str, request: Request):
    """Unlinks one photo from a listing's photo group (doesn't touch storage — other
    rows could still reference the same file). If the deleted photo was the listing's
    primary photo_id, promotes another remaining photo to primary automatically.
    Refuses to delete a listing's only remaining photo."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    try:
        res = supabase.table("listings").select("photo_id").eq("id", item_id).limit(1).execute()
        if not res.data:
            raise HTTPException(404, "Listing not found")
        primary_pid = str(res.data[0].get("photo_id") or "")
        if not primary_pid:
            raise HTTPException(400, "Listing has no primary photo")

        group_row = supabase.table("group_photos").select("group_id").eq("photo_id", primary_pid).limit(1).execute()
        group_id = (group_row.data or [{}])[0].get("group_id", "")
        if not group_id:
            raise HTTPException(400, "Could not find this listing's photo group")

        group_photos = supabase.table("group_photos").select("photo_id").eq("group_id", group_id).execute().data or []
        photo_ids = [p["photo_id"] for p in group_photos]
        if photo_id not in photo_ids:
            raise HTTPException(404, "Photo not found in this listing's group")
        if len(photo_ids) <= 1:
            raise HTTPException(400, "Can't delete the only photo on a listing")

        supabase.table("group_photos").delete().eq("group_id", group_id).eq("photo_id", photo_id).execute()

        new_primary = primary_pid
        if photo_id == primary_pid:
            new_primary = next(p for p in photo_ids if p != photo_id)
            supabase.table("listings").update({"photo_id": new_primary}).eq("id", item_id).execute()
        return {"ok": True, "photo_id": new_primary}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))

@app.post("/api/photos/upload")
async def upload_photo(request: Request):
    try:
        form     = await request.form()
        file     = form["file"]
        gid      = str(form["group_id"])
        idx      = int(form.get("index", 0))
        contents = await file.read()
        # batch_ts (sent once per Done-click, shared by every photo in that
        # batch) replaces calling datetime.now() independently per request.
        # Real bug this fixes: concurrent uploads (since the earlier speed
        # fix) can be PROCESSED by the server in a different order than they
        # were sent -- each independent datetime.now() call would then embed
        # a different second into the filename, so sorting by filename (which
        # is how photo order/primary gets determined downstream) no longer
        # matched selection order. A shared timestamp makes idx the only
        # thing that differs between photos in one batch, so sort order is
        # exactly selection order regardless of network/arrival timing.
        batch_ts = form.get("batch_ts")
        dt = datetime.strptime(batch_ts, "%Y%m%d%H%M%S") if batch_ts else datetime.now()
        fn  = f"{dt.strftime('%d%m%y')}_{dt.strftime('%H%M%S')}_{idx}.jpg"
        print(f"Uploading photo: {fn}, size={len(contents)}, group={gid}")
        supabase.storage.from_("part-photos").upload(
            path=fn,
            file=contents,
            file_options={"content-type": "image/jpeg", "upsert": "true"}
        )
        supabase.table("group_photos").insert({"group_id": gid, "photo_id": fn}).execute()
        return {"ok": True, "photo_id": fn, "url": photo_url(fn, thumb=True)}
    except Exception as e:
        import traceback; traceback.print_exc()
        raise HTTPException(500, str(e))



# ── API: FULL DEEP RESEARCH ──────────────────────────────────── #

@app.post("/api/auction/deep-research-full")
async def deep_research_full(request: Request):
    import os, json, base64, asyncio, fitz
    from concurrent.futures import ThreadPoolExecutor
    import google.generativeai as genai

    form = await request.form()
    items_json = form.get("items", "[]")
    items = json.loads(items_json)
    gemini_key = os.getenv("GEMINI_API_KEY", "")
    if not gemini_key:
        raise HTTPException(400, "GEMINI_API_KEY not set")

    pdf_bytes = None
    pdf_file = form.get("pdf")
    if pdf_file and hasattr(pdf_file, "read"):
        pdf_bytes = await pdf_file.read()

    genai.configure(api_key=gemini_key)
    model = genai.GenerativeModel("gemini-2.5-flash")
    loop = asyncio.get_event_loop()
    executor = ThreadPoolExecutor(max_workers=1)

    def extract_single_image(pdf_bytes, img_index):
        try:
            import fitz
            doc = fitz.open(stream=pdf_bytes, filetype="pdf")
            seen = set()
            all_images = []
            for page in doc:
                for img in page.get_images(full=True):
                    xref = img[0]
                    if xref in seen: continue
                    seen.add(xref)
                    bi = doc.extract_image(xref)
                    if bi and len(bi.get("image","")) > 8000:
                        all_images.append(bi["image"])
            doc.close()
            if img_index < len(all_images):
                return [all_images[img_index]]
            return []
        except Exception as e:
            print(f"extract_single_image error: {e}")
            return []

    def extract_page_image(pdf_bytes, page_start, page_end):
        try:
            doc = fitz.open(stream=pdf_bytes, filetype="pdf")
            images = []
            for page_num in range(page_start - 1, min(page_end, len(doc))):
                page = doc[page_num]
                mat = fitz.Matrix(2.0, 2.0)
                pix = page.get_pixmap(matrix=mat)
                images.append(pix.tobytes("jpeg"))
            doc.close()
            return images
        except Exception as e:
            print(f"Image extract error: {e}")
            return []

    def identify_item_from_image(images, title):
        if not images:
            return title
        try:
            id_prompt = f"You are an auction appraiser. Use BOTH the image AND the listing title equally to identify this item. Title: {title}. Look at the image for: exact model numbers on labels/nameplates, brand logos, condition, visible accessories. Combine both sources and return one precise description. Do not ignore the title — it may contain info not visible in the image."
            parts = [id_prompt] + [{"mime_type": "image/jpeg", "data": img} for img in images[:1]]
            r = model.generate_content(parts, generation_config={"max_output_tokens": 150})
            result = r.text.strip().strip('"')
            return result if result else title
        except Exception as e:
            print(f"Image ID error: {e}")
            return title

    def clean_title(raw_title):
        """Strip address fragments, company boilerplate, and catalog noise from auction titles."""
        import re
        t = raw_title
        # Remove street addresses like "Siemensstrasse 7", "123 Main St"
        t = re.sub(r'\d+\s+[A-Z][a-z]+(?:strasse|street|ave|blvd|rd|st|dr|ln|way)', '', t, flags=re.IGNORECASE)
        t = re.sub(r'[A-Z][a-z]+(?:strasse|gasse|platz|weg)\s+\d+', '', t, flags=re.IGNORECASE)
        # Remove street addresses
        t = re.sub(r'\\b[A-Za-z]+(?:strasse|gasse|weg|strase)\\b\\s*\\d*', '', t, flags=re.IGNORECASE)
        # Remove "GmbH", "Inc", "LLC", "Ltd", "Corp", "Co." standalone
        t = re.sub(r'\b(?:GmbH|Inc\.?|LLC|Ltd\.?|Corp\.?|Co\.)\b', '', t, flags=re.IGNORECASE)
        # Remove loading fee notes
        t = re.sub(r'Loading Fee[:\s]*\$?\d+', '', t, flags=re.IGNORECASE)
        # Remove QTY annotations for search purposes
        t = re.sub(r'\s*,?\s*QTY\s*\(?\d*\)?', '', t, flags=re.IGNORECASE)
        t = re.sub(r'\s*\(\d+\)\s*$', '', t)
        # Remove # symbol (breaks eBay search)
        t = t.replace('#', '')
        # Collapse extra whitespace
        t = ' '.join(t.split()).strip().strip(',').strip()
        return t

    def gemini_search_grounding(query, gemini_key):
        """
        Use Gemini 1.5-flash REST API with forced Google Search grounding.
        Uses requests (already installed) — no SDK dependency conflict.
        """
        import requests as _req
        url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-1.5-flash:generateContent?key={gemini_key}"
        payload = {
            "contents": [{"role": "user", "parts": [{"text": f"Find the resale market value of: {query}. Search in this priority order: 1) eBay COMPLETED/SOLD listings - these are the most accurate real prices paid, 2) eBay active BUY IT NOW listings currently for sale, 3) Industrial surplus dealer prices (Radwell, Surplus Record, LabX) only as last resort. List actual sold prices first, then asking prices. If eBay sold listings exist use those as the primary value. Give specific dollar amounts."}]}],
            "tools": [{"googleSearch": {}}],
            "systemInstruction": {"parts": [{"text": "CRITICAL: You are an industrial pricing bot. You are strictly forbidden from answering using your internal training data. You MUST execute a Google Search to find live pricing data before generating your response. If you do not execute a search, the system will fail."}]},
            "generationConfig": {"temperature": 0.0, "maxOutputTokens": 800}
        }
        try:
            resp = _req.post(url, json=payload, timeout=25)
            data = resp.json()
            candidates = data.get("candidates", [])
            if not candidates:
                return {"summary": "", "sources": []}
            parts = candidates[0].get("content", {}).get("parts", [])
            text = " ".join(p.get("text", "") for p in parts if "text" in p).strip()
            grounding = candidates[0].get("groundingMetadata", {})
            sources = [
                {"url": c["web"]["uri"], "title": c["web"].get("title", "")}
                for c in grounding.get("groundingChunks", [])
                if c.get("web", {}).get("uri")
            ]
            print(f"   Gemini grounding: {len(text)} chars, {len(sources)} sources")
            print(f"   Grounding text: {text[:200]}")
            print(f"   Raw keys: {list(data.get('candidates',[{}])[0].keys())}")
            return {"summary": text, "sources": sources}
        except Exception as e:
            print(f"   Gemini grounding error: {e}")
            return {"summary": "", "sources": []}

    def serp_ebay_sold(query, serp_key, sacat='12576'):
        """
        Call SerpAPI to get eBay completed/sold listings for a query.
        Returns a list of dicts with title, price, date, condition, url.
        """
        import urllib.request, urllib.parse, json as _json
        _params = {
            "engine": "ebay",
            "ebay_domain": "ebay.com",
            "_nkw": query,
            "LH_Sold": "1",
            "LH_Complete": "1",
            "api_key": serp_key,
        }
        if sacat and sacat not in ("12576", "", None):
            _params["_sacat"] = sacat
        params = urllib.parse.urlencode(_params)
        url = f"https://serpapi.com/search?{params}"
        try:
            with urllib.request.urlopen(url, timeout=10) as r:
                data = _json.loads(r.read())
            results = []
            print(f"   SerpAPI response keys: {list(data.keys())}")
            results_list = data.get("organic_results") or data.get("ebay_results") or data.get("shopping_results") or []
            for item in results_list[:8]:
                price_raw = item.get("price", {})
                price = price_raw.get("extracted") or price_raw.get("raw") or 0
                try:
                    price = float(str(price).replace("$","").replace(",",""))
                except Exception:
                    price = 0
                if price > 0:
                    results.append({
                        "title": item.get("title","")[:80],
                        "price": price,
                        "condition": item.get("condition","Used"),
                        "date": item.get("selling_states",{}).get("sold_date","") or "",
                        "url": item.get("link",""),
                    })
            return results
        except Exception as e:
            print(f"   SerpAPI error: {e}")
            return []

    def research_item(item, images):
        import os as _os
        title = item.get("title", "")
        lot = item.get("lot", "")
        current_val = item.get("your_value", 0) or 0
        serp_key = _os.getenv("SERP_API_KEY", "")

        # Clean title before research — remove address/company junk
        clean = clean_title(title)
        if clean != title:
            print(f"Lot {lot} title cleaned: '{title}' → '{clean}'")

        # Step 1: identify exact model from image
        identified = identify_item_from_image(images, clean)
        if identified != clean:
            print(f"Lot {lot} image ID: {identified}")

        # Step 2: Gemini Search Grounding for real market pricing
        serp_results = []
        serp_context = ""
        _gsummary = ""
        if gemini_key:
            _grounding = gemini_search_grounding(clean, gemini_key)
            _gsummary = _grounding.get("summary", "")
            if _gsummary:
                serp_context = f"""MARKET RESEARCH DATA (from live Google Search — use as PRIMARY pricing source):
{_gsummary}

Extract specific dollar amounts from the above. Base revised_value on actual prices found.
Do NOT ignore this data. Do NOT use your training knowledge if this data contradicts it.
"""
        # Legacy SerpAPI block (disabled — kept for fallback reference)
        if False and serp_key:
            # --- Pre-classification: get eBay _sacat and negative keywords ---
            _sacat_map = {
                "12576": "Business & Industrial - Other",
                "58058": "Lasers & Laser Optics (industrial/scientific lasers, fiber lasers, CO2 lasers, laser systems)",
                "105595": "Laser Accessories & Parts",
                "11804": "CNC, Metalworking & Manufacturing",
                "11808": "Electrical Equipment & Supplies",
                "11803": "Semiconductor & PCB Equipment",
                "78989": "Test, Measurement & Inspection",
                "4666":  "Pumps & Plumbing",
                "11816": "Hydraulics, Pneumatics & Plumbing",
                "11815": "Healthcare, Lab & Dental",
                "3673":  "Computers & Networking",
                "58058": "Lasers & Laser Accessories",
                "11700": "Consumer Electronics",
                "26230": "Hand Tools",
                "92074": "Power Tools",
            }
            _cat_prompt = f"""You are an eBay category classifier for industrial equipment.

Item: {clean}

Choose the single best eBay category ID from this list:
{chr(10).join(f'  {k}: {v}' for k,v in _sacat_map.items())}

Also decide if negative keywords are needed to filter out medical/consumer results.
Negative keywords to consider: -medical -dental -cosmetic -hair -aesthetic -salon

Respond ONLY with valid JSON, no markdown:
{{"sacat": "12576", "negative_keywords": "-medical -dental", "is_industrial": true}}

If unsure about negative keywords, use empty string for negative_keywords."""

            _sacat = "12576"
            _negative_kw = ""
            _is_industrial = True
            try:
                _cat_response = model.generate_content(
                    _cat_prompt,
                    generation_config={"max_output_tokens": 100, "temperature": 0}
                )
                _cat_text = _cat_response.text.strip()
                if "```" in _cat_text:
                    _cat_text = _cat_text.split("```")[1]
                    if _cat_text.startswith("json"):
                        _cat_text = _cat_text[4:]
                _cat_text = _cat_text.strip()
                # Find the JSON object
                _s = _cat_text.find("{")
                _e = _cat_text.rfind("}") + 1
                if _s >= 0 and _e > _s:
                    _cat_text = _cat_text[_s:_e]
                import json as _json2
                from json_repair import repair_json as _rj
                _cat_data = _json2.loads(_rj(_cat_text))
                _sacat = str(_cat_data.get("sacat", "12576"))
                _negative_kw = str(_cat_data.get("negative_keywords", ""))
                _is_industrial = bool(_cat_data.get("is_industrial", True))
                print(f"   Category: {_sacat_map.get(_sacat, _sacat)}, industrial={_is_industrial}, negatives='{_negative_kw}'")
            except Exception as _ce:
                print(f"   Category pre-classification failed: {_ce}, using default sacat=12576")

            # Build search query with phrase matching + negative keywords
            _w = clean.split()
            _base_query = '"' + clean + '"' if len(_w) >= 3 else clean
            search_query = (_base_query + " " + _negative_kw).strip()
            print(f"   SerpAPI eBay sold search: '{search_query}' (sacat={_sacat})")
            serp_results = serp_ebay_sold(search_query, serp_key, sacat=_sacat)

            # --- IQR variance-based sanity check ---
            if serp_results:
                prices = sorted([r["price"] for r in serp_results])
                n = len(prices)
                if n >= 4:
                    q1 = prices[n // 4]
                    q3 = prices[(3 * n) // 4]
                    iqr = q3 - q1
                    median = prices[n // 2]
                    cv = (iqr / median) if median > 0 else 1
                    if cv > 1.5:
                        print(f"   SerpAPI results discarded — high variance (CV={cv:.2f}), mixed categories likely")
                        serp_results = []
                    else:
                        print(f"   SerpAPI variance OK: IQR=${iqr:.0f}, CV={cv:.2f}, median=${median:.0f}")
                elif n >= 2:
                    # Small sample: check if range is >5x spread
                    _spread = prices[-1] / prices[0] if prices[0] > 0 else 10
                    if _spread > 5:
                        print(f"   SerpAPI results discarded — spread too wide ({prices[0]:.0f}-{prices[-1]:.0f})")
                        serp_results = []

            if serp_results:
                prices = [r["price"] for r in serp_results]
                avg = sum(prices) / len(prices)
                low = min(prices)
                high = max(prices)
                lines = [f"  - ${r['price']:.0f} — {r['title']} ({r['condition']}) {r['date']}" for r in serp_results]
                serp_context = f"""
REAL EBAY SOLD DATA (from live eBay completed listings — use this as primary pricing source):
Found {len(serp_results)} sold comps: low ${low:.0f}, high ${high:.0f}, avg ${avg:.0f}
{chr(10).join(lines)}

Base your revised_value on these actual sold prices. Do not override this with guesses.
"""
                print(f"   SerpAPI: {len(serp_results)} comps, avg ${avg:.0f}, range ${low:.0f}-${high:.0f}")
            else:
                serp_context = "No eBay sold comps found via SerpAPI - use web search grounding for pricing."
                print(f"   SerpAPI: no results for '{search_query}'")

        prompt = f"""You are an expert industrial machinery appraiser and secondary market researcher.
Your job is to determine the actual cash value of an industrial asset at auction.

--- ITEM DETAILS ---
Lot: #{lot}
Clean Title: {clean}
Image-Identified Model: {identified}
Initial Estimate: ${current_val}

--- MARKET RESEARCH DATA ---
{serp_context}

--- PRICING HIERARCHY RULES (CRITICAL) ---
You must evaluate the MARKET RESEARCH DATA using this strict waterfall hierarchy. Do NOT skip tiers.

TIER 1: SOLD/COMPLETED LISTINGS (Highest Priority)
If the data contains actual verified sold prices, base your estimate entirely on these. Ignore all asking prices.
-> pricing_tier = "SOLD_COMPS"

TIER 2: ACTIVE MARKETPLACE LISTINGS (The Ceiling)
If no sold data exists, look for active listings on open marketplaces (eBay, etc).
Rule: The LOWEST reasonable active listing establishes the absolute CEILING of value. A buyer will not pay $8,000 if they can buy it right now on eBay for $3,995.
Calculation: Find the lowest active price. Apply a 15-25% discount to estimate actual sell price. Ignore high-priced outliers.
-> pricing_tier = "ASKING_PRICES"

TIER 3: INDUSTRIAL DEALER ASKING PRICES (Last Resort Anchor)
If NO marketplace data exists, use retail/surplus dealer asking prices (Radwell, PLC Center, etc).
Rule: Dealers charge massive premiums. Apply a 40-60% discount to find auction/resale cash value.
-> pricing_tier = "ASKING_PRICES"

TIER 4: NO DATA
If the MARKET RESEARCH DATA contains no dollar values relevant to this item, admit it.
-> pricing_tier = "NO_DATA"

--- HALLUCINATION GUARDRAILS ---
- You are FORBIDDEN from using pricing_tier "SOLD_COMPS" unless the word "sold" or "completed" is explicitly in the data.
- Do NOT average a $3,995 eBay listing with a $15,000 dealer listing. The $3,995 becomes the absolute ceiling.
- Do NOT fabricate comps. Only list prices explicitly found in the MARKET RESEARCH DATA above.
- confidence must be "high" only with 3+ verified sold comps, otherwise "medium" or "low".

SHIPPING WEIGHT: Estimate from item type and visible size.

Return ONLY valid JSON, no markdown:
{{"revised_value": 3200, "confidence": "medium", "pricing_tier": "ASKING_PRICES", "pricing_flag": "Based on lowest active eBay listing $3,995 minus 20% discount", "comps": [{{"title": "Item name", "price": 3995, "date": "Apr 2025", "source": "eBay Active"}}], "image_notes": "What the image shows", "recommendation": "watch", "rec_reason": "One active eBay listing at $3,995 sets ceiling, estimated sell price $3,200", "notes": "Market summary with sources", "weight_item_lbs": 50.0, "weight_packaged_lbs": 55.0, "weight_note": "Estimated", "liquidity_score": 2, "liquidity_note": "Limited market data", "sold_30d": 0, "sold_90d": 0, "active_listings": 1}}

pricing_tier values: SOLD_COMPS | ASKING_PRICES | MSRP_ONLY | COMPARABLE_ITEMS | NO_DATA
weight fields: use null if truly unknown"""

        # Use Gemini with search grounding if available
        try:
            from google import genai as _gc
            from google.genai import types as _gt
            _client = _gc.Client(api_key=gemini_key)
            _parts = [prompt]
            for img_bytes in images[:2]:
                _parts.append(_gt.Part.from_bytes(data=img_bytes, mime_type="image/jpeg"))
            _cfg = _gt.GenerateContentConfig(
                tools=[_gt.Tool(google_search=_gt.GoogleSearch())],
                max_output_tokens=1500
            )
            _resp = _client.models.generate_content(
                model="gemini-2.5-flash",
                contents=_parts,
                config=_cfg
            )
            response = _resp
            # Extract grounding from new SDK response
            try:
                gm = _resp.candidates[0].grounding_metadata
                if gm and gm.search_entry_point:
                    ai_overview_html = gm.search_entry_point.rendered_content or ""
                for chunk in (gm.grounding_chunks or []):
                    if hasattr(chunk, "web") and chunk.web:
                        grounding_sources.append({"title": chunk.web.title or "", "uri": chunk.web.uri or ""})
            except Exception:
                pass
            # Make .text work for downstream parsing
            class _Wrap:
                def __init__(self, r): self._r = r
                @property
                def text(self): return self._r.text
                @property
                def candidates(self): return self._r.candidates
            response = _Wrap(_resp)
        except Exception as search_err:
            print(f"   Search grounding failed: {search_err}")
            parts = [prompt]
            for img_bytes in images[:2]:
                parts.append({"mime_type": "image/jpeg", "data": img_bytes})
            response = model.generate_content(parts, generation_config={"max_output_tokens": 1500})
        # Extract AI overview + sources from grounding metadata
        ai_overview_html = ""
        grounding_sources = []
        try:
            candidates = response.candidates
            if candidates:
                gm = getattr(candidates[0], "grounding_metadata", None)
                if gm:
                    sep = getattr(gm, "search_entry_point", None)
                    if sep:
                        ai_overview_html = getattr(sep, "rendered_content", "") or ""
                    chunks = getattr(gm, "grounding_chunks", []) or []
                    for chunk in chunks:
                        web = getattr(chunk, "web", None)
                        if web:
                            grounding_sources.append({
                                "title": getattr(web, "title", ""),
                                "uri":   getattr(web, "uri", ""),
                            })
        except Exception as gm_err:
            print(f"   Grounding metadata error: {gm_err}")

        raw = response.text.strip()
        print(f"   Deep research raw response (lot {lot}): {raw[:2000]}")
        try:
            _d = json.loads(raw if raw.startswith("{") else raw[raw.find("{"):raw.rfind("}")+1])
            print(f"   notes: {_d.get(chr(110)+chr(111)+chr(116)+chr(101)+chr(115),chr(101)+chr(109)+chr(112)+chr(116)+chr(121))}")
        except: pass
        print(f"   AI overview chars: {len(ai_overview_html)}, sources: {len(grounding_sources)}")
        # Strip markdown fences
        if "```" in raw:
            raw = raw.split("```")[1]
            if raw.startswith("json"):
                raw = raw[4:]
            raw = raw.strip()
        raw = " ".join(raw.splitlines())

        raw = " ".join(raw.splitlines())
        # Find JSON object boundaries
        start = raw.find("{")
        end = raw.rfind("}") + 1
        if start >= 0 and end > start:
            raw = raw[start:end]
        try:
            data = json.loads(raw)
        except json.JSONDecodeError:
            from json_repair import repair_json
            data = json.loads(repair_json(raw))
        # Sanitize string fields to prevent SSE encoding issues
        for key in ["image_notes", "rec_reason", "notes", "confidence", "recommendation", "pricing_tier", "pricing_flag", "liquidity_note", "weight_note"]:
            if key in data:
                data[key] = str(data[key]).replace("\n", " ").replace("\r", " ")
        if "comps" in data:
            for comp in data["comps"]:
                for k in comp:
                    comp[k] = str(comp[k]).replace("\n", " ") if isinstance(comp[k], str) else comp[k]
        data["ai_overview_html"] = ai_overview_html
        # Inject grounding summary into notes if available
        if _gsummary and not data.get("notes"):
            data["notes"] = _gsummary
        data["grounding_sources"] = grounding_sources
        # If both SerpAPI and grounding failed, override any hallucinated high confidence
        if not serp_results and not ai_overview_html and not _gsummary:
            if data.get("pricing_tier") == "SOLD_COMPS" and data.get("confidence") == "high":
                data["confidence"] = "low"
                data["pricing_tier"] = "NO_DATA"
                data["pricing_flag"] = "No verified data sources available - estimate may not reflect actual market"

        # --- Hybrid Escalation: call gemini-2.5-pro for hard items ---
        escalate_tiers = {"NO_DATA", "COMPARABLE_ITEMS", "MSRP_ONLY"}
        if data.get("pricing_tier") in escalate_tiers and gemini_key:
            print(f"   Escalating lot {lot} to gemini-2.5-pro (tier={data.get('pricing_tier')})")
            try:
                import requests as _req
                pro_url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-pro:generateContent?key={gemini_key}"
                pro_payload = {
                    "contents": [{"role": "user", "parts": [{"text": prompt}]}],
                    "generationConfig": {"temperature": 0.0, "maxOutputTokens": 2000}
                }
                pro_resp = _req.post(pro_url, json=pro_payload, timeout=60)
                pro_data = pro_resp.json()
                pro_parts = pro_data.get("candidates", [{}])[0].get("content", {}).get("parts", [])
                pro_raw = " ".join(p.get("text", "") for p in pro_parts if "text" in p).strip()
                if pro_raw:
                    if "```" in pro_raw:
                        pro_raw = pro_raw.split("```")[1]
                        if pro_raw.startswith("json"): pro_raw = pro_raw[4:]
                    pro_raw = pro_raw.strip()
                    s = pro_raw.find("{"); e = pro_raw.rfind("}") + 1
                    if s >= 0 and e > s:
                        pro_raw = pro_raw[s:e]
                    try:
                        pro_result = json.loads(pro_raw)
                    except Exception:
                        from json_repair import repair_json
                        pro_result = json.loads(repair_json(pro_raw))
                    # Sanitize and merge
                    for key in ["image_notes","rec_reason","notes","confidence","recommendation","pricing_tier","pricing_flag","liquidity_note","weight_note"]:
                        if key in pro_result:
                            pro_result[key] = str(pro_result[key]).replace("\n"," ").replace("\r"," ")
                    pro_result["model_used"] = "gemini-2.5-pro"
                    pro_result["ai_overview_html"] = ai_overview_html
                    pro_result["grounding_sources"] = grounding_sources
                    if _gsummary and not pro_result.get("notes"):
                        pro_result["notes"] = _gsummary
                    print(f"   Pro escalation result: tier={pro_result.get('pricing_tier')}, value={pro_result.get('revised_value')}")
                    return pro_result
            except Exception as _pe:
                print(f"   Pro escalation failed: {_pe}")

        return data

    async def generate():
        total = len(items)
        for i, item in enumerate(items):
            yield {"data": json.dumps({"type": "start", "lot": item.get("lot"), "index": i, "total": total})}
            try:
                images = []
                # Try to get image from uploaded PDF first
                if pdf_bytes and item.get("_page_start"):
                    images = await loop.run_in_executor(
                        executor, extract_page_image, pdf_bytes,
                        item["_page_start"], item.get("_page_end", item["_page_start"])
                    )
                # Fall back to fetching from stored PDF via scan_id
                if not images and item.get("_page_img"):
                    try:
                        img_url = item["_page_img"]
                        # Extract scan_id and img_index from URL like /api/auction/page-image/{scan_id}/{idx}
                        parts_url = img_url.strip("/").split("/")
                        if len(parts_url) >= 2:
                            sid = parts_url[-2]
                            idx = int(parts_url[-1])
                            stored_pdf = supabase.storage.from_("auction-pdfs").download(f"{sid}.pdf")
                            images = await loop.run_in_executor(
                                executor, lambda: extract_single_image(stored_pdf, idx)
                            )
                    except Exception as img_e:
                        print(f"Auto image fetch error: {img_e}")
                result = await loop.run_in_executor(executor, research_item, item, images)
                yield {"data": json.dumps({
                    "type": "result",
                    "lot": item.get("lot"),
                    "index": i,
                    "total": total,
                    "has_image": len(images) > 0,
                    **result
                })}
            except json.JSONDecodeError as e:
                print(f"JSON parse error for lot {item.get('lot')}: {e}")
                yield {"data": json.dumps({
                    "type": "result",
                    "lot": item.get("lot"),
                    "index": i,
                    "total": total,
                    "has_image": len(images) > 0,
                    "revised_value": item.get("your_value", 0),
                    "confidence": "low",
                    "comps": [],
                    "image_notes": "Research completed but response parsing failed",
                    "recommendation": "watch",
                    "rec_reason": "Could not parse research results — try again",
                    "notes": ""
                })}
            except Exception as e:
                print(f"Deep research error for lot {item.get('lot')}: {e}")
                yield {"data": json.dumps({
                    "type": "error",
                    "lot": item.get("lot"),
                    "index": i,
                    "total": total,
                    "error": str(e)
                })}
            await asyncio.sleep(0.1)
        yield {"data": json.dumps({"type": "done", "total": total})}

    return EventSourceResponse(generate())




@app.get("/api/auction/research-items/{scan_id}")
async def get_research_items(scan_id: str):
    """Return watchlisted items for a scan so any client can load them."""
    import json
    try:
        row = supabase.table("auction_research_sessions")             .select("items,results,title").eq("share_id", scan_id).single().execute()
        data = row.data
        return {
            "scan_id": scan_id,
            "title":   data.get("title",""),
            "items":   json.loads(data.get("items","[]")),
            "results": json.loads(data.get("results","{}")),
        }
    except Exception:
        return {"scan_id": scan_id, "items": [], "results": {}}


@app.post("/api/auction/save-research")
async def save_research(request: Request):
    import json, uuid
    body = await request.json()
    share_id = body.get("share_id") or str(uuid.uuid4())[:8]
    title    = body.get("title", "Auction Research")
    items    = body.get("items", [])
    results  = body.get("results", {})
    try:
        supabase.table("auction_research_sessions").upsert({
            "share_id": share_id,
            "title":    title,
            "items":    json.dumps(items),
            "results":  json.dumps(results),
        }, on_conflict="share_id").execute()
    except Exception as e:
        raise HTTPException(500, f"Save failed: {e}")
    return {"share_id": share_id}


@app.get("/api/auction/scans")
async def list_scans():
    try:
        res = supabase.table("auction_research_sessions")            .select("share_id, title, items, created_at")            .order("created_at", desc=True)            .limit(50)            .execute()
        scans = []
        for row in (res.data or []):
            import json as _j
            items = row.get("items") or []
            if isinstance(items, str):
                try: items = _j.loads(items)
                except: items = []
            scans.append({
                "id": row["share_id"],
                "name": row.get("title", row["share_id"]),
                "items": items,
                "ts": row.get("created_at", "")
            })
        return {"scans": scans}
    except Exception as e:
        raise HTTPException(500, str(e))

@app.get("/api/auction/scans/{scan_id}")
async def get_scan(scan_id: str):
    try:
        import json as _j
        res = supabase.table("auction_research_sessions")            .select("share_id, title, items")            .eq("share_id", scan_id)            .limit(1)            .execute()
        if not res.data:
            raise HTTPException(404, "Scan not found")
        row = res.data[0]
        items = row.get("items") or []
        if isinstance(items, str):
            try: items = _j.loads(items)
            except: items = []
        return {"id": row["share_id"], "name": row.get("title", row["share_id"]), "items": items}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))

@app.post("/api/auction/scans")
async def save_scan(request: Request):
    try:
        import json as _j
        body = await request.json()
        scan_id = body.get("id")
        name = body.get("name", scan_id)
        items = body.get("items", [])
        import json as _j
        supabase.table("auction_research_sessions").upsert({
            "share_id": scan_id,
            "title": name,
            "items": _j.dumps(items),
        }, on_conflict="share_id").execute()
        return {"ok": True, "id": scan_id}
    except Exception as e:
        raise HTTPException(500, str(e))

@app.delete("/api/auction/scans/{scan_id}")
async def delete_scan(scan_id: str):
    try:
        supabase.table("auction_research_sessions")            .delete()            .eq("share_id", scan_id)            .execute()
        return {"ok": True}
    except Exception as e:
        raise HTTPException(500, str(e))

@app.get("/api/auction/load-research/{share_id}")
async def load_research(share_id: str):
    import json
    try:
        row = supabase.table("auction_research_sessions")             .select("*").eq("share_id", share_id).single().execute()
        data = row.data
        return {
            "share_id": data["share_id"],
            "title":    data.get("title", ""),
            "items":    json.loads(data.get("items", "[]")),
            "results":  json.loads(data.get("results", "{}")),
        }
    except Exception as e:
        raise HTTPException(404, f"Session not found: {e}")


@app.post("/api/auction/research-export")
async def research_export(request: Request):
    import io, json
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from fastapi.responses import StreamingResponse

    form = await request.form()
    items = json.loads(form.get("items", "[]"))

    wb = Workbook()
    ws = wb.active
    ws.title = "Deep Research"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1E2535")
    amber_fill = PatternFill("solid", fgColor="412402")
    green_fill = PatternFill("solid", fgColor="052E16")
    alt_fill = PatternFill("solid", fgColor="161B28")
    center = Alignment(horizontal="center", vertical="center")
    wrap = Alignment(wrap_text=True, vertical="center")
    thin = Border(
        bottom=Side(style="thin", color="2D3348"),
        right=Side(style="thin", color="2D3348")
    )

    headers = ["Lot", "Title", "Original Value", "Revised Value", "Confidence", "Recommendation", "Notes", "Your Notes", "eBay Search"]
    col_widths = [8, 45, 15, 15, 12, 16, 40, 30, 20]

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.row_dimensions[1].height = 22

    for ri, item in enumerate(items, 2):
        row_data = [
            item.get("lot", ""),
            item.get("title", ""),
            item.get("original_value", 0),
            item.get("revised_value", 0),
            item.get("confidence", "").capitalize(),
            item.get("recommendation", "").capitalize(),
            item.get("rec_reason") or item.get("image_notes", ""),
            item.get("user_note", ""),
            "View eBay Sold"
        ]
        rec = item.get("recommendation", "").lower()
        fill = green_fill if rec == "buy" else amber_fill if rec == "watch" else (alt_fill if ri % 2 == 0 else None)

        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            if fill:
                cell.fill = fill
            cell.border = thin
            cell.alignment = wrap if ci in (2, 7) else center
            if ci in (3, 4) and isinstance(val, (int, float)):
                cell.number_format = '"$"#,##0'
            if ci == 8 and item.get("ebay_search"):
                ws.cell(row=ri, column=ci).hyperlink = item["ebay_search"]
                ws.cell(row=ri, column=ci).font = Font(color="4A9EFF", underline="single")
        ws.row_dimensions[ri].height = 20

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=deep_research.xlsx"}
    )


@app.post("/api/auction/research-export")
async def research_export(request: Request):
    import io, json
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from fastapi.responses import StreamingResponse

    form = await request.form()
    items = json.loads(form.get("items", "[]"))

    wb = Workbook()
    ws = wb.active
    ws.title = "Deep Research"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1E2535")
    amber_fill = PatternFill("solid", fgColor="412402")
    green_fill = PatternFill("solid", fgColor="052E16")
    alt_fill = PatternFill("solid", fgColor="161B28")
    center = Alignment(horizontal="center", vertical="center")
    wrap = Alignment(wrap_text=True, vertical="center")
    thin = Border(
        bottom=Side(style="thin", color="2D3348"),
        right=Side(style="thin", color="2D3348")
    )

    headers = ["Lot", "Title", "Original Value", "Revised Value", "Confidence", "Recommendation", "Notes", "eBay Search"]
    col_widths = [8, 45, 15, 15, 12, 16, 40, 20]

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.row_dimensions[1].height = 22

    for ri, item in enumerate(items, 2):
        row_data = [
            item.get("lot", ""),
            item.get("title", ""),
            item.get("original_value", 0),
            item.get("revised_value", 0),
            item.get("confidence", "").capitalize(),
            item.get("recommendation", "").capitalize(),
            item.get("rec_reason") or item.get("image_notes", ""),
            item.get("user_note", ""),
            "View eBay Sold"
        ]
        rec = item.get("recommendation", "").lower()
        fill = green_fill if rec == "buy" else amber_fill if rec == "watch" else (alt_fill if ri % 2 == 0 else None)

        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            if fill:
                cell.fill = fill
            cell.border = thin
            cell.alignment = wrap if ci in (2, 7) else center
            if ci in (3, 4) and isinstance(val, (int, float)):
                cell.number_format = '"$"#,##0'
            if ci == 8 and item.get("ebay_search"):
                ws.cell(row=ri, column=ci).hyperlink = item["ebay_search"]
                ws.cell(row=ri, column=ci).font = Font(color="4A9EFF", underline="single")
        ws.row_dimensions[ri].height = 20

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=deep_research.xlsx"}
    )

# ── API: AUCTION DEEP RESEARCH ───────────────────────────────── #

class DeepResearch(BaseModel):
    title: str
    current_value: float = 0

@app.post("/api/auction/deep-research")
async def deep_research(body: DeepResearch):
    import os, asyncio, json
    from concurrent.futures import ThreadPoolExecutor
    import google.generativeai as genai
    gemini_key = os.getenv("GEMINI_API_KEY", "")
    if not gemini_key:
        raise HTTPException(400, "GEMINI_API_KEY not set")
    genai.configure(api_key=gemini_key)
    model = genai.GenerativeModel("gemini-2.5-flash")
    prompt = f"""You are an expert industrial equipment appraiser.
Research this auction item thoroughly: "{body.title}"
Current estimate: ${body.current_value}

Check eBay sold listings, industrial dealers, and recent auction results.
Assume working used condition.

Return ONLY a JSON object (no markdown):
{{"your_value": 5000, "notes": "Sold $4,500-$6,000 on eBay 2024"}}

your_value must be an integer."""

    loop = asyncio.get_event_loop()
    executor = ThreadPoolExecutor(max_workers=1)
    try:
        response = await loop.run_in_executor(executor, lambda: model.generate_content(prompt, generation_config={"max_output_tokens": 300}))
        raw = response.text.strip().replace("```json","").replace("```","").strip()
        try:
            data = json.loads(raw)
        except json.JSONDecodeError:
            from json_repair import repair_json
            data = json.loads(repair_json(raw))
        data["ai_overview_html"] = ai_overview_html
        data["grounding_sources"] = grounding_sources
        return data
    except Exception as e:
        raise HTTPException(500, str(e))


# ── API: EXCEL EXPORT ─────────────────────────────────────────── #

@app.post("/api/auction/export-excel")
async def export_excel(request: Request):
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    body = await request.json()
    items = body.get("items", [])
    name = body.get("name", "Auction Scan")

    wb = Workbook()
    ws = wb.active
    ws.title = name[:31]

    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill("solid", fgColor="1A1F2E")
    hv_fill = PatternFill("solid", fgColor="5C3A00")
    hv_font = Font(color="FAC775", bold=True)

    headers = ["Lot", "Title", "Est. Value", "Notes", "Deep Scan", "Watchlisted"]
    ws.append(headers)
    for col in range(1, 7):
        cell = ws.cell(row=1, column=col)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")

    for item in items:
        val = int(item.get("your_value", 0) or 0)
        row = [
            str(item.get("lot", "")),
            str(item.get("title", "")),
            f"${val:,}",
            str(item.get("notes", "")),
            "Yes" if item.get("_deep") else "",
            "Yes" if item.get("_watch") else "",
        ]
        ws.append(row)
        if val >= 500:
            r = ws.max_row
            for col in range(1, 7):
                ws.cell(row=r, column=col).fill = hv_fill
                ws.cell(row=r, column=col).font = hv_font

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 45
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from datetime import datetime
    fn = f"auction_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fn}"}
    )


# ── API: AUCTION PAGE IMAGE ───────────────────────────────────── #

@app.get("/api/auction/page-image/{scan_id}/{img_index}")
async def get_page_image(scan_id: str, img_index: int):
    import fitz
    from fastapi.responses import Response
    try:
        pdf_data = supabase.storage.from_("auction-pdfs").download(f"{scan_id}.pdf")
        doc = fitz.open(stream=pdf_data, filetype="pdf")
        # Collect all large embedded images (skip logos/watermarks < 5KB)
        all_images = []
        seen_xrefs = set()
        for page in doc:
            for img in page.get_images(full=True):
                xref = img[0]
                if xref in seen_xrefs:
                    continue
                seen_xrefs.add(xref)
                try:
                    base_image = doc.extract_image(xref)
                    if base_image and base_image.get("image") and len(base_image["image"]) > 8000:
                        all_images.append(base_image)
                except Exception as search_err:
                    print(f"   Search grounding failed: {search_err}")
                    pass
        # Fallback: render page and crop item image area for image-only PDFs
        if not all_images or img_index < 0 or img_index >= len(all_images):
            doc2 = fitz.open(stream=pdf_data, filetype="pdf")
            items_per_page = 3
            page_num = img_index // items_per_page
            slot = img_index % items_per_page
            if page_num >= len(doc2):
                page_num = len(doc2) - 1
            page = doc2[page_num]
            pw, ph = page.rect.width, page.rect.height
            slot_h = ph / items_per_page
            clip = fitz.Rect(0, slot * slot_h, pw * 0.28, (slot + 1) * slot_h)
            mat = fitz.Matrix(2.0, 2.0)
            pix = page.get_pixmap(matrix=mat, clip=clip)
            doc2.close()
            return Response(content=pix.tobytes("jpeg"), media_type="image/jpeg",
                          headers={"Cache-Control": "public, max-age=86400"})
        img_data = all_images[img_index]
        ext = img_data.get("ext", "jpeg")
        mime = "image/jpeg" if ext in ("jpg", "jpeg") else "image/" + ext
        return Response(content=img_data["image"], media_type=mime, headers={"Cache-Control": "public, max-age=86400"})
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))

# ── API: PDF AUCTION SCAN ─────────────────────────────────────── #

from sse_starlette.sse import EventSourceResponse

@app.post("/api/auction/scan-txt")
async def scan_txt_auction(file: UploadFile = File(...)):
    import os, json, asyncio
    import google.generativeai as genai

    contents = await file.read()
    text = contents.decode("utf-8", errors="ignore")

    gemini_key = os.getenv("GEMINI_API_KEY", "")
    if not gemini_key:
        raise HTTPException(400, "GEMINI_API_KEY not set")

    genai.configure(api_key=gemini_key)
    model = genai.GenerativeModel("gemini-2.5-flash")

    prompt_template = """You are a world-class auction appraiser with deep expertise in industrial equipment, lab instruments, and commercial goods.

Extract EVERY auction lot from this catalog text.

For each lot return a JSON object:
- lot: lot number as string
- title: full item title as written
- description: one sentence description
- estimate_low: integer dollar amount
- estimate_high: integer dollar amount
- your_value: integer (your single best estimate - total lot value)
- notes: brief market note with price source

PRICING RULES:
- All values MUST be plain integers (no $, no text)
- Base on ACTUAL used market values from eBay sold listings
- If no lots found, return: []
- Return ONLY a JSON array, no markdown"""

    # Split text into chunks of ~8000 chars
    chunk_size = 8000
    chunks = [text[i:i+chunk_size] for i in range(0, len(text), chunk_size)]
    if not chunks:
        chunks = [text]

    from sse_starlette.sse import EventSourceResponse

    async def generate():
        loop = asyncio.get_event_loop()
        from concurrent.futures import ThreadPoolExecutor
        executor = ThreadPoolExecutor(max_workers=1)
        all_items = []
        total = len(chunks)

        for i, chunk in enumerate(chunks):
            try:
                def call_gemini(c=chunk, idx=i):
                    response = model.generate_content(
                        [prompt_template, f"\nCATALOG SECTION {idx+1}/{total}:\n{c}"],
                        generation_config={"max_output_tokens": 16000}
                    )
                    return response.text

                raw = await loop.run_in_executor(executor, call_gemini)
                raw = " ".join(raw.splitlines())
                if "```" in raw:
                    raw = raw.split("```")[1]
                    if raw.startswith("json"):
                        raw = raw[4:]
                    raw = raw.strip()
                start = raw.find("[")
                end = raw.rfind("]") + 1
                if start >= 0 and end > start:
                    raw = raw[start:end]
                try:
                    items = json.loads(raw)
                except Exception:
                    from json_repair import repair_json
                    items = json.loads(repair_json(raw))
                all_items.extend(items)
                yield {
                    "data": json.dumps({
                        "chunk": i + 1,
                        "total_chunks": total,
                        "items": items,
                        "scan_id": None,
                        "done": False
                    }, separators=(',', ':'))
                }
            except Exception as e:
                print(f"TXT chunk {i+1} error: {e}")
            await asyncio.sleep(0.1)

        yield {"data": json.dumps({"done": True, "total": len(all_items), "scan_id": None})}

    return EventSourceResponse(generate())


@app.post("/api/auction/scan-pdf")
async def scan_pdf_auction(file: UploadFile = File(...)):
    import os, base64, json, fitz, asyncio, uuid
    import google.generativeai as genai

    contents = await file.read()
    gemini_key = os.getenv("GEMINI_API_KEY", "")
    if not gemini_key:
        raise HTTPException(400, "GEMINI_API_KEY not set")

    # Store PDF in Supabase for later image retrieval
    scan_id = str(uuid.uuid4())[:8]
    try:
        supabase.storage.from_("auction-pdfs").upload(
            path=f"{scan_id}.pdf",
            file=contents,
            file_options={"content-type": "application/pdf", "upsert": "true"}
        )
    except Exception as upload_err:
        print(f"PDF storage warning: {upload_err}")
        scan_id = None

    # Extract text chunks — fall back to image rendering for image-only PDFs
    try:
        doc = fitz.open(stream=contents, filetype="pdf")
        total_pages = len(doc)
        chunk_size = 2
        page_chunks = []
        page_images = []  # list of (page_num, jpeg_bytes) for image fallback
        for start in range(0, total_pages, chunk_size):
            end = min(start + chunk_size, total_pages)
            chunk_text = ""
            for page_num in range(start, end):
                chunk_text += doc[page_num].get_text() + "\n"
            if chunk_text.strip():
                page_chunks.append(chunk_text)

        # If no text found, render pages as images
        if not page_chunks:
            print(f"PDF has no text — switching to image scan ({total_pages} pages)")
            for page_num in range(total_pages):
                page = doc[page_num]
                mat = fitz.Matrix(2.0, 2.0)
                pix = page.get_pixmap(matrix=mat)
                page_images.append((page_num, pix.tobytes("jpeg")))
        doc.close()
    except Exception as e:
        raise HTTPException(500, f"PDF read error: {e}")

    genai.configure(api_key=gemini_key)
    model = genai.GenerativeModel("gemini-2.5-flash")

    prompt_template = """You are a world-class auction appraiser with deep expertise in industrial equipment, lab instruments, and commercial goods.

Extract EVERY auction lot from this catalog section.

For each lot return a JSON object:
- lot: lot number as string
- title: full item title as written
- description: one sentence description
- estimate_low: integer dollar amount
- estimate_high: integer dollar amount
- your_value: integer (your single best estimate - total lot value)
- notes: brief market note with price source

EXPERT AUCTION TITLE INTERPRETATION:
- Quantities: "(2)", "QTY (3)", "SET OF 4", "PAIR", "x3" = price TOTAL for ALL units combined
- Vague lots: "SHELF OF...", "PALLET OF...", "BOX OF..." = estimate total resale of all contents
- Condition notes like "AS-IS", "UNTESTED", "ACTIVATION NOT GUARANTEED" = still price as normal working condition
- Always search for the SPECIFIC brand + model for accurate pricing
- Ignore auction house names, catalog numbers, location references in titles

PRICING RULES:
- All values MUST be plain integers (no $, no text)
- Base on ACTUAL used market values from eBay sold listings
- If no lots found, return: []
- Return ONLY a JSON array, no markdown

Example: [{"lot":"5","title":"Oakton pH Meter","description":"Portable pH/ORP meter with case","estimate_low":80,"estimate_high":150,"your_value":100,"notes":"Sells $80-150 used on eBay"}]"""

    def call_gemini(chunk_text, i, total):
        response = model.generate_content(
            [prompt_template, f"\nCATALOG SECTION {i+1}/{total}:\n{chunk_text[:10000]}"],
            generation_config={"max_output_tokens": 16000}
        )
        return response.text

    def call_gemini_image(page_num, img_bytes, total):
        """Send a rendered page image to Gemini Vision for lot extraction."""
        print(f"   Image scan page {page_num+1}/{total}")
        img_prompt = prompt_template + f"\n\nThis is page {page_num+1} of {total} of an auction catalog. Extract all lots visible in this image."
        response = model.generate_content(
            [img_prompt, {"mime_type": "image/jpeg", "data": img_bytes}],
            generation_config={"max_output_tokens": 16000}
        )
        return response.text

    async def generate():
        import asyncio
        from concurrent.futures import ThreadPoolExecutor
        loop = asyncio.get_event_loop()
        executor = ThreadPoolExecutor(max_workers=1)
        all_items = []

        # Image-only PDF path
        if page_images and not page_chunks:
            total_chunks = len(page_images)
            for i, (page_num, img_bytes) in enumerate(page_images):
                try:
                    raw = await loop.run_in_executor(executor, call_gemini_image, page_num, img_bytes, total_chunks)
                    raw = " ".join(raw.splitlines())
                    if "```" in raw:
                        raw = raw.split("```")[1]
                        if raw.startswith("json"): raw = raw[4:]
                        raw = raw.strip()
                    start = raw.find("[")
                    end = raw.rfind("]") + 1
                    if start >= 0 and end > start:
                        raw = raw[start:end]
                    try:
                        items = json.loads(raw)
                    except Exception:
                        from json_repair import repair_json
                        items = json.loads(repair_json(raw))
                    base_idx = len(all_items)
                    all_items.extend(items)
                    for item in items:
                        item["_page_start"] = page_num + 1
                        item["_page_end"] = page_num + 1
                        if scan_id:
                            item["_page_img"] = f"/api/auction/page-image/{scan_id}/{base_idx + items.index(item)}"
                    yield {
                        "data": json.dumps({
                            "chunk": i + 1,
                            "total_chunks": total_chunks,
                            "items": items,
                            "scan_id": scan_id,
                            "done": False
                        }, separators=(',', ':'))
                    }
                except Exception as e:
                    print(f"Image page {page_num+1} error: {e}")
                await asyncio.sleep(0.1)
            yield {"data": json.dumps({"done": True, "total": len(all_items), "scan_id": scan_id})}
            return

        total_chunks = len(page_chunks)

        for i, chunk_text in enumerate(page_chunks):
            try:
                raw = await loop.run_in_executor(executor, call_gemini, chunk_text, i, total_chunks)
                raw = " ".join(raw.splitlines())
                if "```" in raw:
                    raw = raw.split("```")[1]
                    if raw.startswith("json"):
                        raw = raw[4:]
                    raw = raw.strip()
                start = raw.find("[")
                end = raw.rfind("]") + 1
                if start >= 0 and end > start:
                    raw = raw[start:end]
                try:
                    items = json.loads(raw)
                except Exception as search_err:
                    print(f"   Search grounding failed: {search_err}")
                    from json_repair import repair_json
                    items = json.loads(repair_json(raw))
                base_idx = len(all_items)
                all_items.extend(items)
                page_start = i * chunk_size + 1
                page_end = min((i + 1) * chunk_size, total_pages)
                for item in items:
                    item["_page_start"] = page_start
                    item["_page_end"] = page_end
                    if scan_id:
                        item["_page_img"] = f"/api/auction/page-image/{scan_id}/{base_idx + items.index(item)}"
                yield {
                    "data": json.dumps({
                        "chunk": i + 1,
                        "total_chunks": total_chunks,
                        "items": items,
                        "scan_id": scan_id,
                        "done": False
                    }, separators=(',', ':'))
                }
            except Exception as e:
                print(f"Chunk {i+1} error: {e}")
            await asyncio.sleep(0.1)

        yield {"data": json.dumps({"done": True, "total": len(all_items), "scan_id": scan_id})}

    return EventSourceResponse(generate())


# ── API: AUCTION LOT CAPTURE (Claude-driven browse -> Supabase -> Gemini itemize) ── #

class AuctionCaptureSessionCreate(BaseModel):
    source_url: str
    name: Optional[str] = None
    capture_scope: Optional[str] = None  # free-text note for Claude, e.g. "pages 1-2" or "lots 300-1300", default = all

# ── Auction-site auto-detection ──────────────────────────────────────────── #
# New sites get their own `_fetch_<site>_lots(url, scope) -> list[dict]` function
# (dict shape must match AuctionLotBulkItem) plus one line in SITE_SCRAPERS below.
# Anything not recognized falls back to the manual "ask Claude" flow untouched.

def _detect_auction_site(url: str) -> Optional[str]:
    from urllib.parse import urlparse
    host = urlparse(url).netloc.lower()
    if "rollerauction.com" in host:
        return "roller"
    if "bidspotter.com" in host:
        return "bidspotter"
    if "dickensheet.com" in host:
        return "dickensheet"
    return None

def _fetch_bidspotter_lots(source_url: str, capture_scope: Optional[str]) -> list:
    """Fetches lots from a BidSpotter auction-catalogue page. Unlike Roller, this
    page is plain server-rendered HTML (confirmed via direct fetch) — no API/GraphQL
    guessing needed. Uses each lot's detail-page URL (a stable /lot-<uuid> pattern)
    as the anchor for finding lot blocks, rather than guessing CSS class names,
    since exact HTML structure wasn't directly inspected before writing this.
    LIMITATION: pagination is client-side JS (confirmed ?page=2 returns identical
    content to page 1), so this only captures the first batch of lots shown on
    load (60 by default) — good for smaller catalogues, not verified yet for
    multi-page ones like large 1000+ lot catalogues."""
    import requests as _requests, re as _re
    from bs4 import BeautifulSoup

    resp = _requests.get(source_url, headers={
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                      "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.9",
        "Referer": "https://www.bidspotter.com/en-us/auction-catalogues",
        "Upgrade-Insecure-Requests": "1",
    }, timeout=30)
    if resp.status_code >= 400:
        raise Exception(f"BidSpotter returned {resp.status_code}: {resp.text[:300]}")
    soup = BeautifulSoup(resp.text, "html.parser")

    lot_link_re = _re.compile(r"/lot-[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}")
    seen_hrefs = set()
    out = []

    for a in soup.find_all("a", href=lot_link_re):
        href = a["href"]
        if href in seen_hrefs:
            continue
        seen_hrefs.add(href)

        # Walk up to a parent block that also contains this lot's price/location text
        block = a
        block_text = ""
        for _ in range(8):
            block = block.parent
            if block is None:
                break
            block_text = block.get_text(" ", strip=True)
            if "Opening price" in block_text or "Location:" in block_text:
                break

        title = a.get_text(strip=True)
        if not title:
            img = a.find("img")
            title = (img.get("alt") or "").strip() if img else ""

        lot_num_m = _re.search(r"\bLot\s+([A-Za-z0-9-]+)\b", block_text)
        price_m = _re.search(r"Opening price\s*\$?([\d,]+\.?\d*)", block_text)
        bid_m = _re.search(r"Current bid\s*\$?([\d,]+\.?\d*)", block_text)
        loc_m = _re.search(r"Location:\s*([^$]+?)(?:\s{2,}|$)", block_text)

        price = None
        for m in (bid_m, price_m):
            if m:
                try:
                    price = float(m.group(1).replace(",", ""))
                    break
                except ValueError:
                    pass

        img_tag = block.find("img", src=True) if block else None
        photo = img_tag["src"] if img_tag and "ajax-loader" not in img_tag["src"] and "blank-image" not in img_tag["src"] else None

        out.append({
            "lot_number": lot_num_m.group(1) if lot_num_m else None,
            "title": title[:300] if title else None,
            "description": (loc_m.group(1).strip() if loc_m else None),
            "listing_url": href if href.startswith("http") else f"https://www.bidspotter.com{href}",
            "current_bid": price,
            "photo_urls": [photo] if photo else [],
            "is_bulk_lot": False,
        })
    return out

def _decode_graphql_crunch(data: list):
    """Decodes a 'graphql-crunch' response: every value in the tree is interned
    once into a flat array, and every integer anywhere in the tree (except
    floats, which are real numeric values) is a back-reference to that array,
    expanded recursively. The root of the whole tree is the LAST element."""
    memo = {}
    def expand(v):
        if isinstance(v, bool):
            return v
        if isinstance(v, int):
            if v in memo:
                return memo[v]
            memo[v] = None
            result = expand(data[v])
            memo[v] = result
            return result
        if isinstance(v, list):
            return [expand(x) for x in v]
        if isinstance(v, dict):
            return {k: expand(x) for k, x in v.items()}
        return v
    return expand(data[-1])

_ROLLER_LOT_FIELDS = """
  auction_lot_id
  auction_id
  quantity
  lot_number
  title
  description
  lot_location
  start_time
  end_time
  winning_bid_amount
  bid_count
  required_bid
  starting_bid
  price
  buy_it_now_active
  buy_it_now_price
  category { name }
  primary_image { large medium }
  image_count
"""

def _fetch_roller_lots(source_url: str, capture_scope: Optional[str]) -> list:
    """Fetches lots for a Roller Auction (bid.rollerauction.com) listing directly
    via their internal GraphQL API — no browser/Claude involvement needed. Page
    number comes from capture_scope (e.g. "page 18"); defaults to page 1.
    NOTE: the exact query/argument names below are reconstructed from a captured
    browser request, not from official docs — if Roller's schema rejects this
    (e.g. "Unknown argument"), the error message will name the bad field/arg
    directly, making it a one-line fix rather than another DevTools session."""
    import requests as _requests, re as _re

    m = _re.search(r"/auctions/(\d+)", source_url)
    if not m:
        raise Exception(f"Could not find an auction ID in URL: {source_url}")
    auction_id = m.group(1)

    page = 1  # parsed but not yet sent — Roller rejected our guessed page/per_page args
              # and their errors don't name the real ones, so this only fetches whatever
              # their default (unpaginated) response returns until we learn the real names
    if capture_scope:
        pm = _re.search(r"page\s*(\d+)", capture_scope, _re.IGNORECASE)
        if pm:
            page = int(pm.group(1))

    query = f"""
    query LotList($auctionId: ID!) {{
      auction(auction_id: $auctionId) {{
        auction_id
        lots {{
          total
          lots {{
{_ROLLER_LOT_FIELDS}
          }}
        }}
      }}
    }}
    """
    resp = _requests.post(
        "https://bid.rollerauction.com/api",
        json={
            "operationName": "LotList",
            "query": query,
            "variables": {"auctionId": auction_id},
        },
        headers={
            "Content-Type": "application/json",
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                          "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        },
        timeout=30,
    )
    try:
        payload = resp.json()
    except Exception:
        resp.raise_for_status()
        raise Exception(f"Roller returned non-JSON response (status {resp.status_code}): {resp.text[:500]}")
    if payload.get("errors"):
        raise Exception(f"Roller GraphQL error (status {resp.status_code}): {payload['errors']}")
    if resp.status_code >= 400:
        raise Exception(f"Roller returned {resp.status_code} with no GraphQL errors field: {resp.text[:500]}")

    raw = payload.get("data")
    decoded = _decode_graphql_crunch(raw) if isinstance(raw, list) else raw
    auction = decoded.get("auction") if isinstance(decoded, dict) else None
    lots = ((auction or {}).get("lots") or {}).get("lots") or []

    out = []
    for lot in lots:
        img = lot.get("primary_image") or {}
        price = lot.get("winning_bid_amount") or lot.get("price") or lot.get("required_bid") or lot.get("starting_bid")
        out.append({
            "lot_number": lot.get("lot_number"),
            "title": lot.get("title"),
            "description": lot.get("description"),
            "listing_url": f"{source_url.split('/auctions/')[0]}/auctions/{auction_id}/lot/{lot.get('auction_lot_id')}",
            "current_bid": float(price) if price not in (None, "") else None,
            "photo_urls": [u for u in [img.get("large") or img.get("medium")] if u],
            "is_bulk_lot": False,
        })
    return out

def _fetch_dickensheet_lots(source_url: str, capture_scope: Optional[str]) -> list:
    """Fetches lots from a Dickensheet auction using their items-search API. This
    endpoint (and field shape) was already confirmed working against live data in
    an earlier session (used to build dickens_scrape.py) — not guessed here."""
    import requests as _requests, re as _re

    m = _re.search(r"(\d{4,})", source_url)
    if not m:
        raise Exception(f"Could not find an auction ID in URL: {source_url}")
    auction_id = m.group(1)

    page = 1
    if capture_scope:
        pm = _re.search(r"page\s*(\d+)", capture_scope, _re.IGNORECASE)
        if pm:
            page = int(pm.group(1))

    resp = _requests.get(
        "https://bid.dickensheet.com/api/items/search",
        params={
            "auction_id": auction_id,
            "query": "",
            "category": "All",
            "per_page": 50,
            "exact_category_match": "true",
            "page": page,
        },
        headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                                "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"},
        timeout=30,
    )
    resp.raise_for_status()
    payload = resp.json()
    items = payload.get("items") if isinstance(payload, dict) else payload
    items = items or []

    out = []
    for item in items:
        images = item.get("images") or []
        photo_urls = []
        for img in images:
            url = img.get("xl") or img.get("lg") or img.get("sm") or img.get("xs")
            if url:
                photo_urls.append(url)
        out.append({
            "lot_number": item.get("lot_identifier"),
            "title": item.get("name"),
            "description": item.get("simple_description") or item.get("description_without_html"),
            "listing_url": f"https://bid.dickensheet.com/auctions/{auction_id}",
            "current_bid": None,
            "photo_urls": photo_urls,
            "is_bulk_lot": False,
        })
    return out

SITE_SCRAPERS = {
    "roller": _fetch_roller_lots,
    "bidspotter": _fetch_bidspotter_lots,
    "dickensheet": _fetch_dickensheet_lots,
}

@app.get("/api/auction/capture/_debug/scrape-photos")
async def debug_scrape_photos(request: Request, url: str, title: str = ""):
    """Diagnostic: runs _scrape_all_lot_photos exactly as flag-as-lot does, but
    returns the raw result (pre-sanity-check) instead of writing anything, so we can
    see from the server's own network vantage point (which may behave differently
    than a local/dev machine -- datacenter IPs sometimes get treated differently by
    a site's bot-detection) whether the scrape itself is finding photos at all."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    import requests as _requests
    try:
        probe = _requests.get(url, headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                          "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        }, timeout=20)
        probe_status = probe.status_code
        probe_len = len(probe.text)
    except Exception as e:
        probe_status, probe_len = None, None
        print(f"debug_scrape_photos probe failed: {e}")
    found = _scrape_all_lot_photos(url, title or None)
    return {"probe_status": probe_status, "probe_html_length": probe_len, "found_count": len(found), "found": found}

@app.get("/api/auction/capture/_debug/roller-schema")
async def debug_roller_schema(request: Request):
    """One-time diagnostic: asks Roller's own GraphQL API what its real 'lots' field
    arguments are (via GraphQL introspection), instead of guessing. Visit this URL
    directly in the browser while logged into Lister — no DevTools needed. Once the
    real argument names are known, _fetch_roller_lots gets fixed once, for good."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    import requests as _requests

    introspect_query = """
    query IntrospectAuctionLots {
      auctionType: __type(name: "Auction") {
        fields { name args { name type { name kind ofType { name kind ofType { name } } } } }
      }
      queryType: __type(name: "Query") {
        fields(includeDeprecated: false) {
          name
          args { name type { name kind ofType { name } } }
        }
      }
    }
    """
    resp = _requests.post(
        "https://bid.rollerauction.com/api",
        json={"operationName": "IntrospectAuctionLots", "query": introspect_query},
        headers={
            "Content-Type": "application/json",
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                          "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        },
        timeout=30,
    )
    try:
        payload = resp.json()
    except Exception:
        return {"status_code": resp.status_code, "raw_text": resp.text[:2000]}
    return {"status_code": resp.status_code, "payload": payload}

def _run_auto_capture(session_id: str, source_url: str, capture_scope: Optional[str], site: str):
    """Runs in the background right after session creation. Fetches lots via the
    matching site scraper and saves them one at a time (not bulk) so progress can
    be tracked and the capture can be stopped mid-way. Progress and stop requests
    are both stored in the same 'status' column to avoid a schema migration:
    status becomes 'capturing:<done>/<total>' while running, and the frontend
    can request a stop by PATCHing status to 'stop_requested', which this loop
    checks for between every lot save."""
    try:
        scraper = SITE_SCRAPERS[site]
        raw_lots = scraper(source_url, capture_scope)
        total = len(raw_lots)
        supabase.table("auction_capture_sessions").update({"status": f"capturing:0/{total}"}).eq("id", session_id).execute()

        done = 0
        for item in raw_lots:
            current = (supabase.table("auction_capture_sessions").select("status")
                       .eq("id", session_id).execute().data or [{}])
            if current and current[0].get("status") == "stop_requested":
                supabase.table("auction_capture_sessions").update({"status": f"stopped:{done}/{total}"}).eq("id", session_id).execute()
                return
            try:
                _create_one_lot(AuctionLotCreate(session_id=session_id, **item))
            except Exception as e:
                print(f"[auto-capture] failed to save lot {item.get('lot_number')}: {e}")
            done += 1
            supabase.table("auction_capture_sessions").update({"status": f"capturing:{done}/{total}"}).eq("id", session_id).execute()

        supabase.table("auction_capture_sessions").update({"status": "done"}).eq("id", session_id).execute()
    except Exception as e:
        print(f"[auto-capture] session {session_id} failed: {e}")
        supabase.table("auction_capture_sessions").update({"status": "auto_capture_failed", "capture_scope": f"{capture_scope or ''} (auto-capture error: {e})"}).eq("id", session_id).execute()

# Roller and BidSpotter's server-side scrapers currently always fail (Roller: GraphQL
# 400s / no auction ID on non-catalog URLs like watchlists; BidSpotter: bot-detection
# wall). Auto-attempting them just produces a confusing failed session every time, so
# these two route straight to "needs Claude" instead of trying and failing first.
CLAUDE_ONLY_SITES = {"roller", "bidspotter"}

@app.post("/api/auction/capture/sessions")
async def create_capture_session(request: Request, body: AuctionCaptureSessionCreate, background_tasks: BackgroundTasks):
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    site = _detect_auction_site(body.source_url)
    auto_capture = bool(site) and site not in CLAUDE_ONLY_SITES
    if site in CLAUDE_ONLY_SITES:
        status = "awaiting_claude_capture"
    elif auto_capture:
        status = "capturing"
    else:
        status = "in_progress"
    row = {
        "business_id": str(business_id),
        "source_url": body.source_url,
        "name": body.name or body.source_url,
        "capture_scope": body.capture_scope or "all",
        "status": status,
    }
    res = supabase.table("auction_capture_sessions").insert(row).execute()
    session = res.data[0]
    if auto_capture:
        background_tasks.add_task(_run_auto_capture, session["id"], body.source_url, body.capture_scope, site)
    return session

@app.get("/api/auction/capture/sessions")
async def list_capture_sessions(request: Request, include_archived: bool = False):
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    q = (supabase.table("auction_capture_sessions")
         .select("*")
         .eq("business_id", str(business_id)))
    if not include_archived:
        q = q.eq("archived", False)
    res = q.order("created_at", desc=True).limit(100).execute()
    return res.data

@app.patch("/api/auction/capture/sessions/{session_id}")
async def update_capture_session(session_id: str, body: dict = Body(...)):
    patch = {}
    if "status" in body:
        patch["status"] = body["status"]
    if "name" in body:
        patch["name"] = body["name"]
    if not patch:
        raise HTTPException(400, "Nothing to update")
    res = supabase.table("auction_capture_sessions").update(patch).eq("id", session_id).execute()
    if not res.data:
        raise HTTPException(404, "Session not found")
    return res.data[0]

@app.delete("/api/auction/capture/sessions/{session_id}")
async def delete_capture_session(session_id: str, request: Request):
    """Deletes a capture session and everything under it (lots + their itemized
    contents). Does not delete re-hosted photos from Storage — cheap to leave
    orphaned, not worth the extra round trips here."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    owned = (supabase.table("auction_capture_sessions").select("id")
             .eq("id", session_id).eq("business_id", str(business_id)).execute().data)
    if not owned:
        raise HTTPException(404, "Session not found")
    lot_ids = [l["id"] for l in supabase.table("auction_lots").select("id").eq("session_id", session_id).execute().data]
    if lot_ids:
        supabase.table("auction_lot_items").delete().in_("lot_id", lot_ids).execute()
        supabase.table("auction_lots").delete().eq("session_id", session_id).execute()
    supabase.table("auction_capture_sessions").delete().eq("id", session_id).execute()
    return {"ok": True, "deleted_session": session_id, "deleted_lots": len(lot_ids)}

class AuctionLotBulkDelete(BaseModel):
    lot_ids: List[str]

@app.post("/api/auction/capture/lots/bulk-delete")
async def bulk_delete_capture_lots(body: AuctionLotBulkDelete, request: Request):
    """Deletes specific lots (e.g. junk/duplicate rows from a messy capture) without
    touching the rest of the session. Scoped to sessions the caller owns, same as
    delete_capture_session, so one business can't delete another's lots by guessing IDs."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    if not body.lot_ids:
        return {"ok": True, "deleted": 0}
    owned_sessions = {s["id"] for s in supabase.table("auction_capture_sessions").select("id").eq("business_id", str(business_id)).execute().data}
    lots = supabase.table("auction_lots").select("id,session_id").in_("id", body.lot_ids).execute().data
    owned_lot_ids = [l["id"] for l in lots if l["session_id"] in owned_sessions]
    if not owned_lot_ids:
        raise HTTPException(404, "No matching lots found")
    supabase.table("auction_lot_items").delete().in_("lot_id", owned_lot_ids).execute()
    supabase.table("auction_lots").delete().in_("id", owned_lot_ids).execute()
    return {"ok": True, "deleted": len(owned_lot_ids)}

class AuctionLotFlagUpdate(BaseModel):
    is_bulk_lot: bool = True

@app.patch("/api/auction/capture/lots/{lot_id}/set-lot-flag")
async def set_lot_flag(lot_id: str, body: AuctionLotFlagUpdate):
    """Pure tag toggle for the 'Flag as LOT' bucket — flips is_bulk_lot and nothing
    else. No photo re-scrape, no itemization, so it can't fail on a missing
    listing_url and is effectively instant. The heavy scrape+itemize pass
    (flag_capture_lot_as_bulk below) is now a separate, opt-in step run once the
    user has actually decided which lots are real bulk lots."""
    res = supabase.table("auction_lots").update({"is_bulk_lot": body.is_bulk_lot}).eq("id", lot_id).execute()
    if not res.data:
        raise HTTPException(404, "Lot not found")
    return res.data[0]

class AuctionLotBulkFlag(BaseModel):
    lot_ids: List[str]
    is_bulk_lot: bool = True

@app.post("/api/auction/capture/lots/bulk-flag")
async def bulk_flag_lots(body: AuctionLotBulkFlag, request: Request):
    """Multi-select version of set-lot-flag: tags (or un-tags) a batch of lots as
    LOT in one call, instant, no scraping. Scoped to sessions the caller owns, same
    ownership check as bulk-delete."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    if not body.lot_ids:
        return {"ok": True, "updated": 0}
    owned_sessions = {s["id"] for s in supabase.table("auction_capture_sessions").select("id").eq("business_id", str(business_id)).execute().data}
    lots = supabase.table("auction_lots").select("id,session_id").in_("id", body.lot_ids).execute().data
    owned_lot_ids = [l["id"] for l in lots if l["session_id"] in owned_sessions]
    if not owned_lot_ids:
        raise HTTPException(404, "No matching lots found")
    supabase.table("auction_lots").update({"is_bulk_lot": body.is_bulk_lot}).in_("id", owned_lot_ids).execute()
    return {"ok": True, "updated": len(owned_lot_ids)}

def _download_and_store_lot_photo(url: str, session_id: str, lot_number: str, idx: int) -> Optional[str]:
    """Downloads an external lot photo and re-uploads it into Supabase Storage so it
    survives even if the auction listing is later removed. Returns the public URL,
    or None if the download/upload failed (caller should fall back to the original URL)."""
    import requests as _requests, uuid as _uuid
    try:
        resp = _requests.get(url, headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                          "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        }, timeout=20)
        resp.raise_for_status()
        content_type = resp.headers.get("Content-Type", "image/jpeg").split(";")[0].strip()
        ext = {"image/jpeg": "jpg", "image/png": "png", "image/webp": "webp"}.get(content_type, "jpg")
        safe_lot = "".join(c if c.isalnum() else "_" for c in (lot_number or "lot"))
        path = f"{session_id}/{safe_lot}_{idx}_{_uuid.uuid4().hex[:8]}.{ext}"
        supabase.storage.from_("auction-lot-photos").upload(
            path=path,
            file=resp.content,
            file_options={"content-type": content_type, "upsert": "true"}
        )
        return supabase.storage.from_("auction-lot-photos").get_public_url(path)
    except Exception as e:
        print(f"Lot photo download/store failed for {url}: {e}")
        return None

class AuctionLotCreate(BaseModel):
    session_id: str
    lot_number: Optional[str] = None
    title: Optional[str] = None
    description: Optional[str] = None
    listing_url: Optional[str] = None
    current_bid: Optional[float] = None
    photo_urls: List[str] = []  # external URLs found while browsing; server re-hosts them
    is_bulk_lot: bool = False  # False (default): single item, no AI call, item = title. True: deep multi-photo itemize.
    auction_ends_at: Optional[str] = None  # ISO datetime, when known — drives auto-archiving

def _create_one_lot(body: AuctionLotCreate) -> dict:
    stored_urls = []
    for i, url in enumerate(body.photo_urls):
        stored = _download_and_store_lot_photo(url, body.session_id, body.lot_number or "lot", i)
        stored_urls.append(stored or url)  # fall back to the original URL if re-hosting failed

    row = {
        "session_id": body.session_id,
        "lot_number": body.lot_number,
        "title": body.title,
        "description": body.description,
        "listing_url": body.listing_url,
        "current_bid": body.current_bid,
        "photo_urls": stored_urls,
        "is_bulk_lot": body.is_bulk_lot,
        "itemized": False,
        "auction_ends_at": body.auction_ends_at,
    }
    res = supabase.table("auction_lots").insert(row).execute()
    lot = res.data[0]

    if not body.is_bulk_lot:
        # Default case: single item, no Gemini call — the item IS the title.
        item_row = {
            "lot_id": lot["id"],
            "item_name": (body.title or "")[:300],
            "item_description": (body.description or "")[:1000],
            "quantity": "1",
            "confidence": "high",
        }
        ins = supabase.table("auction_lot_items").insert(item_row).execute()
        supabase.table("auction_lots").update({"itemized": True}).eq("id", lot["id"]).execute()
        lot["itemized"] = True
        lot["items"] = ins.data

    return lot

@app.post("/api/auction/capture/lots")
async def create_capture_lot(body: AuctionLotCreate):
    return _create_one_lot(body)

class AuctionLotBulkItem(BaseModel):
    lot_number: Optional[str] = None
    title: Optional[str] = None
    description: Optional[str] = None
    listing_url: Optional[str] = None
    current_bid: Optional[float] = None
    photo_urls: List[str] = []
    is_bulk_lot: bool = False
    auction_ends_at: Optional[str] = None

class AuctionLotBulkCreate(BaseModel):
    session_id: str
    lots: List[AuctionLotBulkItem]

@app.post("/api/auction/capture/lots/bulk")
async def create_capture_lots_bulk(body: AuctionLotBulkCreate):
    """Saves many lots in parallel (photo download + Storage upload is network-bound,
    so this is the actual bottleneck — worth parallelizing across a small thread pool
    rather than doing it one lot at a time over N sequential round trips)."""
    import asyncio
    from concurrent.futures import ThreadPoolExecutor

    lot_bodies = [AuctionLotCreate(session_id=body.session_id, **item.model_dump()) for item in body.lots]

    loop = asyncio.get_event_loop()
    executor = ThreadPoolExecutor(max_workers=6)

    async def run_one(lot_body):
        try:
            lot = await loop.run_in_executor(executor, _create_one_lot, lot_body)
            return {"status": "ok", "lot_number": lot_body.lot_number, "lot_id": lot["id"]}
        except Exception as e:
            return {"status": "error", "lot_number": lot_body.lot_number, "error": str(e)}

    results = await asyncio.gather(*[run_one(lb) for lb in lot_bodies])
    ok = sum(1 for r in results if r["status"] == "ok")
    return {"session_id": body.session_id, "total": len(lot_bodies), "ok": ok, "errors": len(lot_bodies) - ok, "results": results}

@app.get("/api/auction/capture/sessions/{session_id}/lots")
async def list_capture_lots(session_id: str):
    lots_res = (supabase.table("auction_lots")
                .select("*")
                .eq("session_id", session_id)
                .order("created_at")
                .execute())
    lots = lots_res.data
    if not lots:
        return []
    lot_ids = [l["id"] for l in lots]
    items_res = (supabase.table("auction_lot_items")
                 .select("*")
                 .in_("lot_id", lot_ids)
                 .execute())
    items_by_lot = {}
    for it in items_res.data:
        items_by_lot.setdefault(it["lot_id"], []).append(it)
    for lot in lots:
        lot["items"] = items_by_lot.get(lot["id"], [])
    return lots

@app.get("/api/auction/capture/public/{session_id}")
async def get_public_capture_session(session_id: str):
    """No-auth read-only view of one capture session's lots, for sharing a link to a
    single catalog externally (e.g. to a friend). session_id is an unguessable UUID —
    same trust model as the existing auction_research_sessions share_id link."""
    sess_res = (supabase.table("auction_capture_sessions")
                .select("id,name,source_url,created_at")
                .eq("id", session_id).execute())
    if not sess_res.data:
        raise HTTPException(404, "Session not found")
    session = sess_res.data[0]
    lots_res = (supabase.table("auction_lots")
                .select("*")
                .eq("session_id", session_id)
                .order("created_at")
                .execute())
    lots = lots_res.data
    lot_ids = [l["id"] for l in lots]
    items_by_lot = {}
    if lot_ids:
        items_res = supabase.table("auction_lot_items").select("*").in_("lot_id", lot_ids).execute()
        for it in items_res.data:
            items_by_lot.setdefault(it["lot_id"], []).append(it)
    for lot in lots:
        lot["items"] = items_by_lot.get(lot["id"], [])
    return {"session": session, "lots": lots}

@app.patch("/api/auction/capture/lots/{lot_id}/notes")
async def update_capture_lot_notes(lot_id: str, body: dict = Body(...)):
    res = (supabase.table("auction_lots")
           .update({"notes": body.get("notes", "")})
           .eq("id", lot_id)
           .execute())
    if not res.data:
        raise HTTPException(404, "Lot not found")
    return res.data[0]

ITEMIZE_PROMPT = """You are cataloguing the contents of a single auction lot for a resale buyer.

Lot title: {title}
Lot description: {description}

Look at the attached photo(s) of this lot AND read the title/description above. Identify EVERY
distinct item present. Rules:
- If the lot is clearly one single item, return exactly one entry for it.
- If the lot is a mixed/bulk lot (shelf, pallet, box, "assorted"), enumerate each distinct item or
  item group you can actually see or that is explicitly named in the title/description. Do not
  invent items that aren't shown or named.
- If you can read a specific brand + model/part number, put it in item_name. If you cannot
  confidently identify the exact item, describe it generically in item_name (e.g. "unlabeled steel
  bracket") and set confidence to "low" — do NOT guess a specific brand/model you can't actually
  read, and do NOT blend two different products into one invented name.
- If a photo shows only a closed box/case/container and the title/description claims specific
  contents you cannot see, still list the claimed contents as one entry but set confidence to
  "low" and note in item_description that contents are unverified from the photo.
- Multiple photos may show the SAME physical item from different angles, or the same item
  repeated across photos. Treat all attached photos as views of ONE lot: return one entry per
  distinct physical item/item-group in the whole lot, not one entry per photo. Do not duplicate
  an item just because it appears in more than one photo.
- quantity: a short string like "1", "4", "unknown".

Return ONLY a raw JSON array, no markdown, no backticks:
[{{"item_name": "...", "item_description": "...", "quantity": "...", "confidence": "high"|"low"}}]
If you truly cannot identify anything, return []."""

def _gemini_itemize_call(model, prompt: str, photo_urls) -> list:
    """One Gemini call with all of a lot's photos attached together, so the model can
    see the whole lot at once and dedupe items that appear in multiple photos instead
    of emitting one copy per photo. photo_urls may be a single url string, a list of
    urls, or None/empty for a text-only call."""
    import json, re
    from json_repair import repair_json
    import requests as _requests

    if isinstance(photo_urls, str):
        photo_urls = [photo_urls]
    photo_urls = photo_urls or []

    parts = [prompt]
    for photo_url in photo_urls:
        try:
            img = _requests.get(photo_url, timeout=20)
            img.raise_for_status()
            ctype = img.headers.get("Content-Type", "image/jpeg").split(";")[0].strip()
            if ctype not in ("image/jpeg", "image/png", "image/webp"):
                ctype = "image/jpeg"
            parts.append({"mime_type": ctype, "data": img.content})
        except Exception as e:
            print(f"itemize: could not fetch photo {photo_url}: {e}")

    response = model.generate_content(parts, generation_config={"max_output_tokens": 4000})
    raw = (response.text or "").strip()
    raw = re.sub(r"^```[a-z]*\n?", "", raw, flags=re.IGNORECASE)
    raw = re.sub(r"\n?```$", "", raw).strip()
    start, end = raw.find("["), raw.rfind("]") + 1
    if start >= 0 and end > start:
        raw = raw[start:end]
    try:
        return json.loads(raw) if raw else []
    except Exception:
        return json.loads(repair_json(raw))

def _itemize_lot_deep(lot: dict) -> list:
    """Bulk-lot itemization: ONE Gemini call with all of the lot's photos attached
    together, so the model can dedupe items that show up in multiple photos itself
    instead of us getting one copy of each item per photo. Deletes any existing
    (e.g. title-only) items first."""
    import google.generativeai as genai

    gemini_key = os.getenv("GEMINI_API_KEY", "")
    if not gemini_key:
        raise HTTPException(400, "GEMINI_API_KEY not set")

    lot_id = lot["id"]
    genai.configure(api_key=gemini_key)
    model = genai.GenerativeModel("gemini-2.5-flash")

    prompt = ITEMIZE_PROMPT.format(title=lot.get("title") or "(no title)",
                                    description=lot.get("description") or "(no description)")
    photo_urls = lot.get("photo_urls") or []

    try:
        all_items = _gemini_itemize_call(model, prompt, photo_urls)
    except Exception as e:
        raise HTTPException(500, f"Gemini itemize error: {e}")

    supabase.table("auction_lot_items").delete().eq("lot_id", lot_id).execute()
    inserted = []
    for it in all_items:
        row = {
            "lot_id": lot_id,
            "item_name": (it.get("item_name") or "")[:300],
            "item_description": (it.get("item_description") or "")[:1000],
            "quantity": str(it.get("quantity") or "")[:50],
            "confidence": it.get("confidence") if it.get("confidence") in ("high", "low") else "low",
        }
        ins = supabase.table("auction_lot_items").insert(row).execute()
        inserted.append(ins.data[0])

    supabase.table("auction_lots").update({"itemized": True, "is_bulk_lot": True}).eq("id", lot_id).execute()
    return inserted

_IMG_SKIP_PATTERNS = ("logo", "icon", "spinner", "favicon", "avatar", "sprite", ".svg")

def _scrape_all_lot_photos(listing_url: str, lot_title: Optional[str] = None) -> List[str]:
    """Server-side re-fetch of a single lot's detail page to pull every content photo.
    Best-effort: works for server-rendered pages, may return few/no results on
    heavily JS-rendered sites. If lot_title is given, tries to scope the search to the
    DOM container holding that title first (avoids sweeping up unrelated 'related
    lots'/sidebar thumbnails elsewhere on the page); falls back to whole-page search.

    Two bugs fixed here after they silently broke every Roller "Flag as LOT" re-scrape:
    1. soup.find(string=...) also matches the invisible <title> tag in <head> (which
       repeats the lot title) -- that node has no images, so the climb-up-the-DOM
       loop below it kept climbing until it hit <html>/[document], scoping to the
       ENTIRE page instead of the actual photo gallery. Searching soup.body instead
       of soup keeps it out of <head> entirely.
    2. Even body-scoped, the real content container still mixes in the site's own
       nav/UI graphics (logos, badges, a Facebook tracking pixel), pushing the count
       over the ">6 images = untrustworthy" sanity check below and silently
       reverting to the single already-known photo every time. Real listing photos
       are essentially always .jpg/.jpeg (camera/phone photos); site UI chrome is
       .png/.svg/.gif. Preferring .jpg/.jpeg when any are present filters that
       chrome out without needing a per-site selector."""
    import re
    import requests as _requests
    from bs4 import BeautifulSoup
    from urllib.parse import urljoin

    urls, seen = [], set()

    def _collect(img_tags):
        for img in img_tags:
            src = img.get("src") or img.get("data-src") or img.get("data-lazy-src") or ""
            if not src:
                continue
            if any(p in src.lower() for p in _IMG_SKIP_PATTERNS):
                continue
            full = urljoin(listing_url, src)
            if full in seen:
                continue
            seen.add(full)
            urls.append(full)

    try:
        resp = _requests.get(listing_url, headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                          "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        }, timeout=20)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "html.parser")

        scoped = None
        if lot_title and soup.body:
            title_node = soup.body.find(string=lambda t: t and lot_title.strip()[:40] in t)
            if title_node:
                container = title_node.parent
                for _ in range(6):
                    if container is None:
                        break
                    if len(container.find_all("img")) >= 1:
                        scoped = container
                    container = container.parent
        if scoped is not None:
            _collect(scoped.find_all("img"))
        if not urls:
            _collect((soup.body or soup).find_all("img"))

        jpg_urls = [u for u in urls if re.search(r"\.jpe?g(?:$|\?)", u, re.IGNORECASE)]
        if jpg_urls:
            urls = jpg_urls
    except Exception as e:
        print(f"lot photo re-scrape failed for {listing_url}: {e}")
    return urls[:10]

def _fetch_main_photo_sync(lot_id: str, listing_url_override: Optional[str] = None) -> dict:
    """Grabs just the first (main) full-res photo for a lot and re-hosts it —
    deliberately lighter than flag-as-lot: no is_bulk_lot flip, no Gemini itemize
    call. Built for backfilling photos on lots that were inserted directly (e.g.
    from a watchlist import) and so never went through the normal capture flow
    that scrapes photos automatically. If the row has no listing_url yet, pass
    one in and it gets saved."""
    lot_res = supabase.table("auction_lots").select("*").eq("id", lot_id).single().execute()
    if not lot_res.data:
        raise Exception("404: Lot not found")
    lot = lot_res.data

    listing_url = listing_url_override or lot.get("listing_url")
    if not listing_url:
        raise Exception("400: No listing_url available to scrape a photo from")

    found_urls = _scrape_all_lot_photos(listing_url, lot.get("title"))
    if not found_urls:
        raise Exception("404: No photo found on listing page")

    main_url = found_urls[0]
    stored = _download_and_store_lot_photo(main_url, lot["session_id"], lot.get("lot_number") or "lot", 0)
    stored_url = stored or main_url

    patch = {"photo_urls": [stored_url]}
    if listing_url_override and not lot.get("listing_url"):
        patch["listing_url"] = listing_url_override
    supabase.table("auction_lots").update(patch).eq("id", lot_id).execute()
    return {"lot_id": lot_id, "lot_number": lot.get("lot_number"), "photo_url": stored_url}

class MainPhotoFetchItem(BaseModel):
    lot_id: str
    listing_url: Optional[str] = None

class MainPhotoFetchBulk(BaseModel):
    items: List[MainPhotoFetchItem]

def _fetch_main_photos_bulk_sync(items: List["MainPhotoFetchItem"]) -> dict:
    results = []
    for it in items:
        try:
            r = _fetch_main_photo_sync(it.lot_id, it.listing_url)
            results.append({**r, "status": "ok"})
        except Exception as e:
            msg = str(e)
            for prefix in ("404:", "400:"):
                if msg.startswith(prefix):
                    msg = msg[len(prefix):].strip()
                    break
            results.append({"lot_id": it.lot_id, "status": "error", "error": msg})
    return {"results": results}

@app.post("/api/auction/capture/lots/fetch-main-photos-bulk")
async def fetch_main_photos_bulk(body: MainPhotoFetchBulk, request: Request):
    """Batch version of the main-photo backfill above, scoped to sessions the
    caller owns (same ownership check as bulk-delete). Runs off-thread since
    it's a chain of blocking network scrapes."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    if not body.items:
        return {"results": []}
    lot_ids = [it.lot_id for it in body.items]
    owned_sessions = {s["id"] for s in supabase.table("auction_capture_sessions").select("id").eq("business_id", str(business_id)).execute().data}
    lots = supabase.table("auction_lots").select("id,session_id").in_("id", lot_ids).execute().data
    owned_ids = {l["id"] for l in lots if l["session_id"] in owned_sessions}
    filtered = [it for it in body.items if it.lot_id in owned_ids]
    if not filtered:
        raise HTTPException(404, "No matching owned lots found")
    import asyncio
    loop = asyncio.get_event_loop()
    return await loop.run_in_executor(None, _fetch_main_photos_bulk_sync, filtered)

def _itemize_capture_lot_sync(lot_id: str) -> dict:
    lot_res = supabase.table("auction_lots").select("*").eq("id", lot_id).single().execute()
    if not lot_res.data:
        raise Exception("404: Lot not found")
    inserted = _itemize_lot_deep(lot_res.data)
    return {"lot_id": lot_id, "items": inserted}

@app.post("/api/auction/capture/lots/{lot_id}/itemize")
async def itemize_capture_lot(lot_id: str):
    """Same blocking-Gemini-call issue as flag-as-lot — offloaded to a thread so
    it doesn't freeze the server for everyone else while it runs."""
    import asyncio
    loop = asyncio.get_event_loop()
    try:
        return await loop.run_in_executor(None, _itemize_capture_lot_sync, lot_id)
    except Exception as e:
        msg = str(e)
        if msg.startswith("404:"):
            raise HTTPException(404, msg[4:].strip())
        raise HTTPException(500, msg)

def _flag_lot_as_bulk_sync(lot_id: str) -> dict:
    """Blocking body of flag-as-lot: a page scrape (network I/O), several photo
    downloads, and a synchronous Gemini call. Pulled out into a plain function so
    the route can run it in a thread pool instead of on the asyncio event loop —
    previously this ran directly inside `async def`, which on this single-worker
    Uvicorn process meant one person's Flag-as-LOT click (10-60s of blocking I/O)
    froze every other request on the server: other browsers' page loads, the
    best-offers polling, and every other lot's Flag-as-LOT button, all queued
    behind it. Raises plain Exception on failure; the route translates that back
    into an HTTPException since HTTPException raised off the event loop thread
    doesn't get FastAPI's normal handling."""
    lot_res = supabase.table("auction_lots").select("*").eq("id", lot_id).single().execute()
    if not lot_res.data:
        raise Exception("404: Lot not found")
    lot = lot_res.data

    if not lot.get("listing_url"):
        raise Exception("400: Lot has no listing_url to re-scrape photos from")

    existing_urls = lot.get("photo_urls") or []
    found_urls = _scrape_all_lot_photos(lot["listing_url"], lot.get("title"))

    # Sanity check: on JS-rendered sites (confirmed on Roller Auctions) the static HTML
    # doesn't contain the real photo gallery at all, so a "successful" scrape can still
    # just be page icons/badges near the title. A believable single lot rarely has more
    # than ~6 real photos — if we got more than that, trust the re-scrape less than the
    # photo we already know is real (captured live during the initial pass).
    if not found_urls:
        found_urls = existing_urls
    elif len(found_urls) > 6:
        print(f"flag-as-lot: scrape returned {len(found_urls)} images for {lot['listing_url']} — "
              f"looks unreliable (likely a JS-rendered gallery not present in static HTML), "
              f"falling back to the {len(existing_urls)} already-known photo(s)")
        found_urls = existing_urls or found_urls[:1]

    stored_urls = []
    for i, url in enumerate(found_urls):
        stored = _download_and_store_lot_photo(url, lot["session_id"], lot.get("lot_number") or "lot", i)
        stored_urls.append(stored or url)

    supabase.table("auction_lots").update({
        "photo_urls": stored_urls,
        "is_bulk_lot": True,
    }).eq("id", lot_id).execute()
    lot["photo_urls"] = stored_urls

    inserted = _itemize_lot_deep(lot)
    lot["items"] = inserted
    lot["itemized"] = True
    lot["is_bulk_lot"] = True
    return lot

@app.post("/api/auction/capture/lots/{lot_id}/flag-as-lot")
async def flag_capture_lot_as_bulk(lot_id: str):
    """Self-serve 'this is actually a mixed lot' action: re-scrapes the listing page
    for all photos, re-hosts them, then runs the deep per-photo itemization. Runs
    off-thread (see _flag_lot_as_bulk_sync) so it doesn't block the whole server
    while it works."""
    import asyncio
    loop = asyncio.get_event_loop()
    try:
        return await loop.run_in_executor(None, _flag_lot_as_bulk_sync, lot_id)
    except Exception as e:
        msg = str(e)
        if msg.startswith("404:"):
            raise HTTPException(404, msg[4:].strip())
        if msg.startswith("400:"):
            raise HTTPException(400, msg[4:].strip())
        raise HTTPException(500, msg)

def _itemize_all_capture_lots_sync(session_id: str) -> dict:
    lots_res = (supabase.table("auction_lots")
                .select("*")
                .eq("session_id", session_id)
                .eq("itemized", False)
                .execute())
    results = []
    for lot in lots_res.data:
        try:
            items = _itemize_lot_deep(lot)
            results.append({"lot_id": lot["id"], "lot_number": lot.get("lot_number"), "status": "ok", "item_count": len(items)})
        except Exception as e:
            results.append({"lot_id": lot["id"], "lot_number": lot.get("lot_number"), "status": "error", "error": str(e)})
    return {"session_id": session_id, "results": results}

@app.post("/api/auction/capture/sessions/{session_id}/itemize-all")
async def itemize_all_capture_lots(session_id: str):
    """Same blocking-Gemini-call issue as flag-as-lot, times however many lots are
    unitemized in the session — offloaded to a thread so it doesn't freeze the
    server for everyone else while it churns through the whole session."""
    import asyncio
    loop = asyncio.get_event_loop()
    return await loop.run_in_executor(None, _itemize_all_capture_lots_sync, session_id)

# ── API: AUCTION LOT PRICING METRICS (independent, per-column, opt-in) ── #

METRIC_PROMPTS = {
    "resale_value": {
        "instructions": """Estimate the realistic USED RESALE VALUE range (low-high, in USD) if this were
resold on eBay/marketplaces. Search the web for real comparable sold/active listings and base your
estimate on them — do not guess without searching. If items are unverified/claimed-only (low
confidence in the itemization), factor that uncertainty into a wider range.""",
        "schema": '{"value_low": number, "value_high": number, "notes": "brief reasoning citing what you found"}',
    },
    "liquidity": {
        "instructions": """Rate how easily and quickly this could be resold: "high" (common demand,
many buyers, sells within days), "medium" (resells within a few weeks), or "low" (niche/specialized,
may take months or require a specific buyer). Base this on how common the item(s) are and how many
active/sold listings you find when searching.""",
        "schema": '{"rating": "high"|"medium"|"low", "notes": "brief reasoning"}',
    },
    "weight": {
        "instructions": """Estimate the combined physical shipping weight in pounds. Use known specs
for identified items where possible (search if needed); otherwise give a reasonable estimate based on
item type and size. This matters for whether resale requires freight/local pickup vs. easy parcel
shipping.""",
        "schema": '{"weight_lbs": number, "notes": "brief reasoning, note if freight/pickup likely required"}',
    },
    "max_bid": {
        "instructions": """Determine the maximum price a reseller should pay at this auction to still
make a reasonable profit margin after resale fees, shipping, and time invested. Search for real resale
value first, then work backward with a sensible margin (bigger margin for slower/riskier items, smaller
for fast-moving common items).""",
        "schema": '{"max_bid": number, "notes": "brief reasoning showing the math"}',
    },
}

def _compute_metric_for_lot(lot: dict, items: list, metric: str) -> dict:
    import json, re, time
    import requests as _requests
    from json_repair import repair_json

    gemini_key = os.getenv("GEMINI_API_KEY", "")
    if not gemini_key:
        raise HTTPException(400, "GEMINI_API_KEY not set")

    spec = METRIC_PROMPTS[metric]
    items_desc = "\n".join(
        f"- {it.get('item_name','?')} (qty {it.get('quantity','?')}, confidence {it.get('confidence','?')}): {it.get('item_description','')}"
        for it in items
    ) or f"(not itemized in detail — use the lot title/description) {lot.get('title','')}"

    prompt = f"""You are pricing a single auction lot for a professional reseller.

Lot title: {lot.get('title') or '(no title)'}
Lot description: {lot.get('description') or '(no description)'}
Current auction bid: {lot.get('current_bid')}

Itemized contents:
{items_desc}

TASK: {spec['instructions']}

Return ONLY a raw JSON object, no markdown, no backticks:
{spec['schema']}"""

    # Gemini's free tier rate-limits (429) under sustained per-lot calls in a session with
    # many lots — without a retry here, a rate-limited lot just permanently errors out and
    # looks to the user like the job "gave up" partway through. Back off and retry those
    # (and transient 5xx) a few times before actually giving up on this lot.
    resp = None
    for attempt in range(4):
        resp = _requests.post(
            f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent?key={gemini_key}",
            json={
                "contents": [{"parts": [{"text": prompt}]}],
                "tools": [{"google_search": {}}],
                "generationConfig": {"maxOutputTokens": 1500},
            },
            timeout=30,
        )
        if resp.status_code == 429 or resp.status_code >= 500:
            if attempt < 3:
                time.sleep(3 * (2 ** attempt))  # 3s, 6s, 12s
                continue
        break
    resp.raise_for_status()
    resp_data = resp.json()
    raw = (resp_data["candidates"][0]["content"]["parts"][0].get("text") or "").strip()
    raw = re.sub(r"^```[a-z]*\n?", "", raw, flags=re.IGNORECASE)
    raw = re.sub(r"\n?```$", "", raw).strip()
    start, end = raw.find("{"), raw.rfind("}") + 1
    if start >= 0 and end > start:
        raw = raw[start:end]
    try:
        data = json.loads(raw) if raw else {}
    except Exception:
        data = json.loads(repair_json(raw))
    return data

class ComputeMetricRequest(BaseModel):
    metric: str  # "resale_value" | "liquidity" | "weight" | "max_bid"
    lot_ids: Optional[List[str]] = None  # None/omitted = all lots in the session

_metric_jobs: dict = {}  # job_id -> progress state; single-process in-memory tracker, fine for this app's deploy shape

@app.post("/api/auction/capture/sessions/{session_id}/compute-metric")
async def compute_metric_for_session(session_id: str, body: ComputeMetricRequest):
    """Starts a background job and returns immediately — the actual Gemini loop keeps
    running on the server independent of the client's connection, so closing the
    browser tab/app does NOT stop it. Poll GET .../metric-jobs/{job_id} for progress."""
    import asyncio, uuid as _uuid

    if body.metric not in METRIC_PROMPTS:
        raise HTTPException(400, f"metric must be one of {list(METRIC_PROMPTS.keys())}")

    q = supabase.table("auction_lots").select("*").eq("session_id", session_id)
    if body.lot_ids:
        q = q.in_("id", body.lot_ids)
    lots = q.execute().data

    # Default run (no explicit lot_ids): skip lots that already have this metric filled
    # in, so re-clicking "run" after a partial run (errors, quota limits, etc.) only
    # retries what's missing instead of re-burning Gemini calls redoing lots that already
    # succeeded. An explicit lot_ids selection always means "run these", already-filled
    # or not — that's the escape hatch for deliberately redoing specific lots.
    if not body.lot_ids:
        already_filled_field = {
            "resale_value": "resale_value_low",
            "liquidity": "liquidity_rating",
            "weight": "weight_lbs",
            "max_bid": "max_bid",
        }[body.metric]
        lots = [l for l in lots if l.get(already_filled_field) is None]

    job_id = _uuid.uuid4().hex
    _metric_jobs[job_id] = {
        "job_id": job_id, "session_id": session_id, "metric": body.metric,
        "total": len(lots), "processed": 0, "ok": 0, "errors": 0,
        "status": "running", "last": None,
    }

    async def run_job():
        job = _metric_jobs[job_id]
        if not lots:
            job["status"] = "done"
            return
        try:
            lot_ids = [l["id"] for l in lots]
            items_res = supabase.table("auction_lot_items").select("*").in_("lot_id", lot_ids).execute()
            items_by_lot = {}
            for it in items_res.data:
                items_by_lot.setdefault(it["lot_id"], []).append(it)

            loop = asyncio.get_event_loop()
            from concurrent.futures import ThreadPoolExecutor
            executor = ThreadPoolExecutor(max_workers=1)

            for lot in lots:
                try:
                    def call(l=lot, its=items_by_lot.get(lot["id"], [])):
                        return _compute_metric_for_lot(l, its, body.metric)
                    data = await loop.run_in_executor(executor, call)
                    if body.metric == "resale_value":
                        patch = {"resale_value_low": data.get("value_low"), "resale_value_high": data.get("value_high"), "resale_value_notes": data.get("notes")}
                    elif body.metric == "liquidity":
                        patch = {"liquidity_rating": data.get("rating"), "liquidity_notes": data.get("notes")}
                    elif body.metric == "weight":
                        patch = {"weight_lbs": data.get("weight_lbs"), "weight_notes": data.get("notes")}
                    elif body.metric == "max_bid":
                        patch = {"max_bid": data.get("max_bid"), "max_bid_notes": data.get("notes")}
                    supabase.table("auction_lots").update(patch).eq("id", lot["id"]).execute()
                    job["ok"] += 1
                    job["last"] = {"lot_id": lot["id"], "lot_number": lot.get("lot_number"), "title": lot.get("title"), "status": "ok", **patch}
                except Exception as e:
                    job["errors"] += 1
                    job["last"] = {"lot_id": lot["id"], "lot_number": lot.get("lot_number"), "title": lot.get("title"), "status": "error", "error": str(e)}
                job["processed"] += 1
        finally:
            job["status"] = "done"

    asyncio.create_task(run_job())
    return {"job_id": job_id, "total": len(lots)}

@app.get("/api/auction/capture/metric-jobs/{job_id}")
async def get_metric_job(job_id: str):
    job = _metric_jobs.get(job_id)
    if not job:
        raise HTTPException(404, "Job not found (server may have restarted)")
    return job


def _scrape_current_bid(listing_url: str) -> Optional[float]:
    """Generic (site-agnostic) current-bid scrape: works off the plain text near a
    'Current bid'/'Current Bid' label rather than a bespoke DOM selector per site,
    since a session's lots can come from BidSpotter, Roller, Dickensheet, etc. Less
    precise than a tailored selector but avoids maintaining one scraper per site
    just for a single number."""
    import re
    import requests as _requests
    from bs4 import BeautifulSoup

    resp = _requests.get(listing_url, headers={
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                      "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    }, timeout=20)
    if resp.status_code != 200:
        raise Exception(f"listing page returned {resp.status_code}")
    soup = BeautifulSoup(resp.text, "html.parser")
    text = soup.get_text(" ", strip=True)
    m = (re.search(r"Current\s+[Bb]id[:\s]*\$?\s?([\d,]+\.?\d*)", text)
         or re.search(r"\$\s?([\d,]+\.?\d*)\s*Current\s+[Bb]id", text)
         or re.search(r"Opening\s+[Bb]id[:\s]*\$?\s?([\d,]+\.?\d*)", text))
    if not m:
        return None
    return float(m.group(1).replace(",", ""))

_bid_refresh_jobs: dict = {}  # job_id -> progress state, same shape/lifetime as _metric_jobs

@app.post("/api/auction/capture/sessions/{session_id}/refresh-bids")
async def refresh_session_bids(session_id: str):
    """Starts a background job that re-scrapes each lot's own listing page for its
    live current bid, one at a time with a delay between requests -- these source
    sites (esp. BidSpotter) have bot-protection that a rapid burst of requests can
    trip, so this deliberately trades speed for not getting the session's IP
    temporarily blocked. Safe to close the tab; poll /refresh-bids-jobs/{job_id}."""
    import asyncio, uuid as _uuid, time

    lots = (supabase.table("auction_lots").select("id,lot_number,title,listing_url,current_bid")
            .eq("session_id", session_id).execute().data)
    lots = [l for l in lots if l.get("listing_url")]

    job_id = _uuid.uuid4().hex
    _bid_refresh_jobs[job_id] = {
        "job_id": job_id, "session_id": session_id,
        "total": len(lots), "processed": 0, "ok": 0, "errors": 0, "changed": 0,
        "status": "running", "last": None,
    }

    async def run_job():
        job = _bid_refresh_jobs[job_id]
        if not lots:
            job["status"] = "done"
            return
        loop = asyncio.get_event_loop()
        from concurrent.futures import ThreadPoolExecutor
        executor = ThreadPoolExecutor(max_workers=1)
        try:
            for lot in lots:
                try:
                    new_bid = await loop.run_in_executor(executor, _scrape_current_bid, lot["listing_url"])
                    if new_bid is not None:
                        changed = lot.get("current_bid") != new_bid
                        supabase.table("auction_lots").update({"current_bid": new_bid}).eq("id", lot["id"]).execute()
                        job["ok"] += 1
                        if changed:
                            job["changed"] += 1
                        job["last"] = {"lot_id": lot["id"], "lot_number": lot.get("lot_number"), "title": lot.get("title"),
                                        "status": "ok", "old_bid": lot.get("current_bid"), "new_bid": new_bid}
                    else:
                        job["errors"] += 1
                        job["last"] = {"lot_id": lot["id"], "lot_number": lot.get("lot_number"), "title": lot.get("title"),
                                        "status": "error", "error": "no bid found on page"}
                except Exception as e:
                    job["errors"] += 1
                    job["last"] = {"lot_id": lot["id"], "lot_number": lot.get("lot_number"), "title": lot.get("title"),
                                    "status": "error", "error": str(e)}
                job["processed"] += 1
                await asyncio.sleep(2)  # pace requests against the source site
        finally:
            job["status"] = "done"

    asyncio.create_task(run_job())
    return {"job_id": job_id, "total": len(lots)}

@app.get("/api/auction/capture/refresh-bids-jobs/{job_id}")
async def get_bid_refresh_job(job_id: str):
    job = _bid_refresh_jobs.get(job_id)
    if not job:
        raise HTTPException(404, "Job not found (server may have restarted)")
    return job


# ── EBAY OAUTH ────────────────────────────────────────────────── #

@app.get("/api/ebay/oauth/start")
async def ebay_oauth_start(request: Request):
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    settings = get_ebay_settings(business_id)
    app_id = settings.get("EBAY_APP_ID", "")
    runame = settings.get("EBAY_RUNAME", "")
    if not app_id or not runame:
        raise HTTPException(400, "Set EBAY_APP_ID and EBAY_RUNAME in Settings first")
    import urllib.parse as _up
    params = {
        "client_id": app_id,
        "redirect_uri": runame,
        "response_type": "code",
        "scope": EBAY_OAUTH_SCOPES,
        "state": business_id,
    }
    url = "https://auth.ebay.com/oauth2/authorize?" + _up.urlencode(params)
    from fastapi.responses import RedirectResponse
    return RedirectResponse(url)

@app.get("/api/ebay/oauth/callback")
async def ebay_oauth_callback(code: str = None, state: str = None, error: str = None):
    if error:
        return HTMLResponse(f"<h3>eBay authorization failed: {error}</h3><p>You can close this tab and try again.</p>")
    if not code or not state:
        raise HTTPException(400, "Missing code/state from eBay")
    business_id = state
    settings = get_ebay_settings(business_id)
    app_id = settings.get("EBAY_APP_ID", "")
    cert_id = settings.get("EBAY_CERT_ID", "")
    runame = settings.get("EBAY_RUNAME", "")
    if not app_id or not cert_id or not runame:
        raise HTTPException(400, "Missing EBAY_APP_ID / EBAY_CERT_ID / EBAY_RUNAME in Settings")

    import requests as _req, time, base64 as _b64
    basic = _b64.b64encode(f"{app_id}:{cert_id}".encode()).decode()
    r = _req.post(
        "https://api.ebay.com/identity/v1/oauth2/token",
        headers={"Authorization": f"Basic {basic}", "Content-Type": "application/x-www-form-urlencoded"},
        data={"grant_type": "authorization_code", "code": code, "redirect_uri": runame},
        timeout=15,
    )
    if r.status_code != 200:
        return HTMLResponse(f"<h3>Token exchange failed ({r.status_code})</h3><pre>{r.text[:1000]}</pre>")

    data = r.json()
    save_ebay_setting(business_id, "EBAY_USER_TOKEN", data["access_token"])
    save_ebay_setting(business_id, "EBAY_TOKEN_EXPIRES_AT", str(time.time() + int(data.get("expires_in", 7200))))
    save_ebay_setting(business_id, "EBAY_REFRESH_TOKEN", data["refresh_token"])
    return HTMLResponse("<h3>eBay account connected \u2705</h3><p>You can close this tab and go back to Lister AI.</p>")

# ── SHOPIFY ───────────────────────────────────────────────────── #

def get_shopify_access_token(business_id: str, force_refresh: bool = False) -> str:
    """Shopify's current auth model (post Jan 2026): Dev Dashboard apps use a
    client_credentials grant instead of a static copy-once token. Tokens last
    24h, so this caches per-business in app_settings and refreshes transparently,
    the same pattern as get_ebay_access_token. force_refresh bypasses the cache —
    needed after an uninstall/reinstall, which invalidates old tokens early
    without our cache knowing."""
    import requests as _req, time

    settings = get_ebay_settings(business_id)
    domain = (settings.get("SHOPIFY_STORE_DOMAIN", "") or "").strip()
    client_id = (settings.get("SHOPIFY_CLIENT_ID", "") or "").strip()
    client_secret = (settings.get("SHOPIFY_CLIENT_SECRET", "") or "").strip()
    if not domain or not client_id or not client_secret:
        raise Exception("Shopify not connected — set Store Domain, Client ID, and Client Secret in Settings")
    domain = domain.replace("https://", "").replace("http://", "").strip("/")

    cached_token = settings.get("SHOPIFY_ACCESS_TOKEN", "")
    expires_at = float(settings.get("SHOPIFY_TOKEN_EXPIRES_AT", "0") or 0)
    if not force_refresh and cached_token and time.time() < (expires_at - 60):
        return cached_token

    r = _req.post(
        f"https://{domain}/admin/oauth/access_token",
        headers={"Content-Type": "application/x-www-form-urlencoded"},
        data={"grant_type": "client_credentials", "client_id": client_id, "client_secret": client_secret},
        timeout=15,
    )
    if r.status_code != 200:
        raise Exception(f"Shopify token request failed ({r.status_code}): {r.text[:300]}")
    data = r.json()
    token = data["access_token"]
    new_expires_at = time.time() + int(data.get("expires_in", 86399))
    save_ebay_setting(business_id, "SHOPIFY_ACCESS_TOKEN", token)
    save_ebay_setting(business_id, "SHOPIFY_TOKEN_EXPIRES_AT", str(new_expires_at))
    return token

def push_listing_to_shopify(listing: dict, image_urls_override: list = None) -> dict:
    """Creates an active product on Shopify via the Admin REST API.
    image_urls_override, when given, is used as-is (skips the photo_id/group_photos
    lookup and the eBay-live-fallback below entirely) — for items that never had a
    Lister listings row at all (a genuinely eBay-only inventory item), so there's no
    photo_id to look up in the first place; the caller (see the eBay-only publish
    endpoint) already resolved real photo URLs from Lister's own storage."""
    import requests as _req

    biz_id = listing.get("business_id")
    if not biz_id:
        raise Exception("Listing has no business_id — cannot look up Shopify settings safely")
    settings = get_ebay_settings(biz_id)  # generic per-business key/value store, not eBay-specific
    domain = (settings.get("SHOPIFY_STORE_DOMAIN", "") or "").strip()
    if not domain:
        raise Exception("Shopify not connected — set Store Domain in Settings")
    domain = domain.replace("https://", "").replace("http://", "").strip("/")
    token = get_shopify_access_token(biz_id)
    api_base = f"https://{domain}/admin/api/2024-10"

    title = (listing.get("title") or "Untitled item")[:255]
    desc  = listing.get("description") or EBAY_DESCRIPTION
    price = float(listing.get("price") or 0)
    qty   = int(listing.get("quantity") or 1)
    brand = (listing.get("brand") or "").strip() or "Unbranded"
    sku   = listing.get("ebay_sku") or f"lister-{listing['id']}"
    if image_urls_override is not None:
        images = [{"src": u} for u in image_urls_override]
    else:
        pid = str(listing.get("photo_id") or "")
        images = [{"src": photo_url(p)} for p in get_all_photo_ids(pid) if photo_url(p)] if pid else []
        if not images and listing.get("ebay_item_id"):
            # Lister's own storage has nothing for this listing (e.g. it was published
            # under an old business account whose photos live in a different, never-
            # copied-over storage instance) -- fall back to pulling the real, currently-
            # live photos straight off the eBay listing itself, since it's already
            # published there and eBay's image CDN is always publicly reachable.
            try:
                ebay_token = get_ebay_access_token(biz_id)
                ebay_urls = fetch_ebay_listing_photo_urls(ebay_token, listing["ebay_item_id"])
                images = [{"src": u} for u in ebay_urls]
            except Exception as e:
                print(f"push_listing_to_shopify: eBay photo fallback failed: {e}")

    body = {
        "product": {
            "title": title,
            "body_html": desc,
            "vendor": brand,
            "status": "active",
            "images": images,
            "variants": [{
                "price": f"{price:.2f}",
                "sku": sku,
                "inventory_management": "shopify",
                "inventory_quantity": qty,
            }],
        }
    }
    headers = {"X-Shopify-Access-Token": token, "Content-Type": "application/json"}
    # Shopify has to fetch every image URL itself before it can respond to this
    # call (images are sent as external src URLs here, not embedded data) --
    # for a multi-photo listing that can legitimately take well past 20s, which
    # is exactly what was timing out real, otherwise-fine publishes. Not adding
    # a retry-on-timeout here on purpose: this is a non-idempotent create, and
    # blindly re-POSTing after a timeout risks creating a duplicate product if
    # Shopify actually finished server-side and the response just came back
    # slow -- a longer timeout is the safe fix, a retry here is not.
    try:
        r = _req.post(f"{api_base}/products.json", headers=headers, json=body, timeout=75)
        if r.status_code == 401:
            # cached token was invalidated (e.g. app was uninstalled/reinstalled) — force a fresh one and retry once
            token = get_shopify_access_token(biz_id, force_refresh=True)
            headers["X-Shopify-Access-Token"] = token
            r = _req.post(f"{api_base}/products.json", headers=headers, json=body, timeout=75)
    except _req.exceptions.Timeout:
        # Deliberately NOT retrying here -- see comment above. Surfacing a
        # clear, actionable message instead of the raw connection-pool string
        # that used to leak straight through to the user.
        raise Exception(
            "Shopify didn't respond within 75 seconds while creating this listing (it has to fetch "
            f"{len(images)} photo(s) itself before it can respond). Check Shopify directly before "
            "retrying -- if the product actually went through and this just timed out waiting on the "
            "response, retrying now would create a duplicate."
        )
    if r.status_code not in (200, 201):
        raise Exception(f"Shopify product create failed ({r.status_code}): {r.text[:400]}")

    data = r.json().get("product", {})
    product_id = data.get("id")
    variants = data.get("variants") or []
    variant_id = variants[0].get("id") if variants else None

    channel_result = {}
    if product_id:
        try:
            channel_result = publish_product_to_channels(domain, token, product_id)
        except Exception as e:
            # product itself was created successfully — a channel-publish failure shouldn't fail the whole call
            channel_result = {"error": str(e)}

    return {
        "product_id": product_id, "variant_id": variant_id, "status": data.get("status"),
        "handle": data.get("handle"), "channels": channel_result,
        "sku": variants[0].get("sku") if variants else sku,
        "price": float(variants[0].get("price")) if variants else price,
        "quantity": variants[0].get("inventory_quantity") if variants else qty,
        "title": data.get("title") or title,
    }

def _record_new_shopify_publish_locally(business_id: str, result: dict, ebay_id: str = None) -> None:
    """Writes the just-created Shopify product/variant into shopify_inventory and
    links it on the matching inventory_match row directly, using the data
    push_listing_to_shopify already returned -- instead of triggering a full
    eBay+Shopify catalog re-sync just to reflect ONE new item, which is what
    the frontend was doing before (confirmed: this was the actual cause of
    publishing feeling 'crazy slow' and blocking back-to-back publishes,
    since a full re-sync ran after every single one). Same column mapping
    fetch_shopify_inventory_items/api_inventory_sync already use, so this
    row looks identical to one the normal sync would have produced."""
    variant_id = result.get("variant_id")
    if not variant_id:
        return  # nothing to record locally if Shopify didn't return a variant
    try:
        supabase.table("shopify_inventory").upsert({
            "id": str(variant_id), "business_id": business_id, "product_id": str(result.get("product_id") or ""),
            "sku": result.get("sku"), "title": result.get("title"), "price": result.get("price"),
            "quantity": result.get("quantity"), "status": result.get("status"), "handle": result.get("handle"),
        }).execute()
        if ebay_id:
            # Fetch by business_id only (a safe, definitely-consistent comparison)
            # and match ebay_id as a plain STRING comparison in Python, rather
            # than a SQL-level .eq("ebay_id", ebay_id) -- a row_id/ebay_id coming
            # in as a plain string from a URL path, compared directly against
            # whatever type that column actually is in Postgres, is exactly the
            # class of silent-zero-rows-matched bug that's hit this session
            # multiple times already (text vs uuid, text vs date). Comparing as
            # strings in Python sidesteps that risk entirely, at the cost of
            # fetching more rows than strictly needed -- an acceptable tradeoff
            # for a once-per-publish operation, not something on the hot path.
            candidates = _fetch_all_for_business(business_id, "inventory_match", "id,ebay_id")
            match_row = next((r for r in candidates if str(r.get("ebay_id")) == str(ebay_id)), None)
            if match_row:
                supabase.table("inventory_match").update({"shopify_id": str(variant_id)})\
                    .eq("id", match_row["id"]).execute()
            else:
                print(f"_record_new_shopify_publish_locally: no inventory_match row found for ebay_id={ebay_id} -- "
                      f"local link not made, item may keep showing as eBay-only until the next full sync")
    except Exception as e:
        # Never let this block the publish response -- the Shopify product itself
        # already exists successfully at this point; worst case here is the local
        # tables stay stale until the next real sync, same as before this existed.
        print(f"_record_new_shopify_publish_locally failed (publish itself still succeeded): {e}")

def publish_product_to_channels(domain: str, token: str, product_id, target_channel_names=None) -> dict:
    """Publishes an already-created product to additional sales channels (e.g. Google & YouTube),
    since REST product creation only auto-publishes to the Online Store channel.
    Requires the write_publications scope on top of write_products."""
    import requests as _req

    if target_channel_names is None:
        target_channel_names = ["Google & YouTube", "Point of Sale"]

    gql_url = f"https://{domain}/admin/api/2024-10/graphql.json"
    headers = {"X-Shopify-Access-Token": token, "Content-Type": "application/json"}

    list_r = _req.post(gql_url, headers=headers, json={
        "query": "{ publications(first: 20) { edges { node { id name } } } }"
    }, timeout=15)
    if list_r.status_code != 200:
        raise Exception(f"Could not list Shopify sales channels ({list_r.status_code}): {list_r.text[:300]}")
    pubs = list_r.json().get("data", {}).get("publications", {}).get("edges", [])

    matched = [
        {"publicationId": p["node"]["id"], "name": p["node"]["name"]}
        for p in pubs
        if any(target.lower() in p["node"]["name"].lower() for target in target_channel_names)
    ]
    if not matched:
        return {"published_to": [], "note": "No matching sales channels found — is Google & YouTube installed as a channel on this store?"}

    product_gid = f"gid://shopify/Product/{product_id}"
    pub_r = _req.post(gql_url, headers=headers, json={
        "query": """
            mutation publishablePublish($id: ID!, $input: [PublicationInput!]!) {
              publishablePublish(id: $id, input: $input) {
                userErrors { field message }
              }
            }
        """,
        "variables": {"id": product_gid, "input": [{"publicationId": m["publicationId"]} for m in matched]},
    }, timeout=15)
    if pub_r.status_code != 200:
        raise Exception(f"Channel publish failed ({pub_r.status_code}): {pub_r.text[:300]}")
    errors = pub_r.json().get("data", {}).get("publishablePublish", {}).get("userErrors", [])
    if errors:
        raise Exception(f"Channel publish errors: {errors}")
    return {"published_to": [m["name"] for m in matched]}

@app.get("/api/ebay/debug-category-requirements/{item_id}")
async def ebay_debug_category_requirements(item_id: str, request: Request):
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    import requests as _req
    res = supabase.table("listings").select("*").eq("id", item_id).limit(1).execute()
    if not res.data:
        raise HTTPException(404, "Listing not found")
    listing = res.data[0]
    biz_id = listing.get("business_id")
    settings = get_ebay_settings(biz_id)
    raw_category_id = listing.get("ebay_category_id")
    category_id = raw_category_id if raw_category_id and str(raw_category_id) != "0" else settings.get("EBAY_DEFAULT_CATEGORY_ID", "")
    try:
        token = get_ebay_access_token(biz_id)
        r = _req.get(
            f"{EBAY_API_BASE}/commerce/taxonomy/v1/category_tree/0/get_item_aspects_for_category",
            headers=ebay_headers(token, content_language=False),
            params={"category_id": category_id},
            timeout=20,
        )
        if r.status_code != 200:
            return {"category_id": category_id, "status": r.status_code, "body": r.text[:1000]}
        aspects = r.json().get("aspects", [])
        relevant = [
            {
                "name": a.get("localizedAspectName"),
                "required": a.get("aspectConstraint", {}).get("aspectRequired"),
                "mode": a.get("aspectConstraint", {}).get("aspectMode"),
                "dataType": a.get("aspectConstraint", {}).get("aspectDataType"),
                "cardinality": a.get("aspectConstraint", {}).get("itemToAspectCardinality"),
            }
            for a in aspects
            if a.get("aspectConstraint", {}).get("aspectRequired")
            or "mpn" in (a.get("localizedAspectName","").lower())
            or "brand" in (a.get("localizedAspectName","").lower())
        ]
        return {"category_id": category_id, "required_or_brand_mpn_aspects": relevant}
    except Exception as e:
        raise HTTPException(400, str(e))

@app.post("/api/financials/resync-one-order/{order_id}")
async def resync_one_order(order_id: str, request: Request):
    """Re-pulls and re-upserts just ONE order — a handful of API calls, seconds not
    minutes. Useful for verifying/fixing a specific order without re-walking the
    entire multi-month sync range."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    import requests as _req, datetime as _dt

    token = get_ebay_access_token(business_id)
    r = _req.get(f"{EBAY_API_BASE}/sell/fulfillment/v1/order/{order_id}",
                 headers=ebay_headers(token, content_language=False), timeout=15)
    if r.status_code != 200:
        raise HTTPException(400, f"Could not fetch order: {r.status_code} {r.text[:300]}")
    order = r.json()

    created = order.get("creationDate", "")
    order_delivery_cost = float((order.get("pricingSummary") or {}).get("deliveryCost", {}).get("value", 0) or 0)
    tax_addr = ((order.get("buyer") or {}).get("taxAddress") or {})
    buyer_state, buyer_zip, buyer_country = tax_addr.get("stateOrProvince", ""), tax_addr.get("postalCode", ""), tax_addr.get("countryCode", "")

    # Narrow fee-fetch window around just this order's date — a few days is plenty,
    # avoids re-scanning months of transactions for one order.
    order_date = _dt.datetime.strptime(created[:10], "%Y-%m-%d") if created else _dt.datetime.utcnow()
    fee_start = (order_date - _dt.timedelta(days=1)).strftime("%Y-%m-%dT00:00:00.000Z")
    fee_end = (order_date + _dt.timedelta(days=2)).strftime("%Y-%m-%dT00:00:00.000Z")
    try:
        fee_data = fetch_ebay_fees_by_line_item(business_id, fee_start, fee_end)
        fees_by_line = fee_data["fees_by_line"]
        ebay_labels_by_order = fee_data["ebay_labels_by_order"]
    except Exception as e:
        fees_by_line, ebay_labels_by_order = {}, {}

    # Tracking number for this order
    trackings = []
    try:
        rf = _req.get(f"{EBAY_API_BASE}/sell/fulfillment/v1/order/{order_id}/shipping_fulfillment",
                      headers=ebay_headers(token, content_language=False), timeout=15)
        if rf.status_code == 200:
            for f in rf.json().get("fulfillments", []):
                tn = f.get("shipmentTrackingNumber")
                if tn:
                    trackings.append(tn)
    except Exception:
        pass

    cost_by_tracking = {}
    if trackings:
        labels_res = supabase.table("shipping_labels").select("tracking_number,cost")\
            .eq("business_id", business_id).in_("tracking_number", trackings).execute()
        for row in (labels_res.data or []):
            cost_by_tracking[row["tracking_number"]] = row.get("cost") or 0
    pirate_ship_cost = sum(cost_by_tracking.get(tn, 0) or 0 for tn in trackings)
    shipping_cost = pirate_ship_cost if pirate_ship_cost > 0 else ebay_labels_by_order.get(order_id, 0)

    order_refund_total = sum(
        float((r.get("amount") or {}).get("value", 0) or 0)
        for r in (order.get("paymentSummary") or {}).get("refunds", []) or []
    )
    all_line_items = order.get("lineItems", [])
    has_line_level_refunds = any(li.get("refunds") for li in all_line_items)
    order_subtotal_for_proration = sum(
        float((li.get("lineItemCost") or {}).get("value", 0) or 0) for li in all_line_items
    ) if order_refund_total and not has_line_level_refunds else 0

    order_total_due_seller = float((order.get("paymentSummary") or {}).get("totalDueSeller", {}).get("value", 0) or 0)
    order_gross_subtotal = sum(
        float((li.get("lineItemCost") or {}).get("value", 0) or 0)
        + float((li.get("deliveryCost") or {}).get("shippingCost", {}).get("value", 0) or 0)
        for li in all_line_items
    )

    upserted = []
    for li in all_line_items:
        item_price = float((li.get("lineItemCost") or {}).get("value", 0) or 0)
        buyer_shipping = float((li.get("deliveryCost") or {}).get("shippingCost", {}).get("value", 0) or 0)
        if has_line_level_refunds:
            refund = sum(float((r.get("amount") or {}).get("value", 0) or 0) for r in (li.get("refunds") or []))
        elif order_refund_total and order_subtotal_for_proration > 0:
            refund = order_refund_total * (item_price / order_subtotal_for_proration)
        else:
            refund = 0.0
        revenue = item_price + buyer_shipping
        line_item_id = li.get("lineItemId", "")
        fee = fees_by_line.get((order_id, line_item_id), 0.0)
        net = (order_total_due_seller * (revenue / order_gross_subtotal)) if order_gross_subtotal > 0 else order_total_due_seller
        record_id = f"ebay:{order_id}:{line_item_id}"
        # Same SKU-preservation rule as the bulk sync (sync_orders_for_business) —
        # this endpoint used to have none at all, so a single-order resync would
        # silently wipe a manual correction that the bulk sync would have protected.
        existing = supabase.table("orders").select("sku").eq("id", record_id).limit(1).execute().data
        existing_sku = existing[0].get("sku") if existing else None
        fresh_sku = li.get("sku") or "(no SKU)"
        final_sku = existing_sku if (existing_sku and not _is_blank_sku(existing_sku)) else fresh_sku
        record = {
            "id": record_id,
            "business_id": business_id, "platform": "eBay", "order_id": order_id,
            "sku": final_sku, "title": li.get("title", ""), "quantity": int(li.get("quantity", 1)),
            "order_date": created[:10] if created else "",
            "gross_revenue": round(revenue, 2), "buyer_shipping": round(buyer_shipping, 2),
            "refund": round(refund, 2),
            "order_delivery_cost": round(order_delivery_cost, 2),
            "buyer_state": buyer_state, "buyer_zip": buyer_zip, "buyer_country": buyer_country,
            "fee": round(fee, 2), "net": round(net, 2),
            "tracking_number": ",".join(trackings) if trackings else None,
            "shipping_cost": round(shipping_cost, 2),
            "final_net": round(net - shipping_cost, 2),
        }
        supabase.table("orders").upsert(record, on_conflict="id").execute()
        upserted.append(record)

    return {"ok": True, "order_id": order_id, "line_items_updated": len(upserted), "records": upserted}

@app.get("/api/shopify/debug-product-metadata")
async def debug_shopify_product_metadata(request: Request, limit: int = 5):
    """Pulls a handful of real Shopify products with EVERY piece of metadata that
    could plausibly link back to an original eBay listing: SKU, handle, tags,
    and — most importantly — metafields, which is where migration tools most
    commonly stash a source-platform ID."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    import requests as _req

    settings = get_ebay_settings(business_id)
    domain = (settings.get("SHOPIFY_STORE_DOMAIN", "") or "").strip().replace("https://", "").replace("http://", "").strip("/")
    if not domain:
        raise HTTPException(400, "Shopify not connected")
    token = get_shopify_access_token(business_id)
    headers = {"X-Shopify-Access-Token": token}

    r = _req.get(f"https://{domain}/admin/api/2024-10/products.json",
                 headers=headers, params={"limit": limit}, timeout=20)
    if r.status_code == 401:
        token = get_shopify_access_token(business_id, force_refresh=True)
        headers["X-Shopify-Access-Token"] = token
        r = _req.get(f"https://{domain}/admin/api/2024-10/products.json",
                     headers=headers, params={"limit": limit}, timeout=20)
    if r.status_code != 200:
        raise HTTPException(400, f"Shopify products fetch failed ({r.status_code}): {r.text[:300]}")

    products = r.json().get("products", [])
    results = []
    for p in products:
        pid = p["id"]
        # Metafields aren't included in the basic product payload — fetch separately
        mf_r = _req.get(f"https://{domain}/admin/api/2024-10/products/{pid}/metafields.json",
                        headers=headers, timeout=15)
        metafields = mf_r.json().get("metafields", []) if mf_r.status_code == 200 else []
        results.append({
            "id": pid,
            "title": p.get("title"),
            "handle": p.get("handle"),
            "tags": p.get("tags"),
            "vendor": p.get("vendor"),
            "product_type": p.get("product_type"),
            "variants_sku_barcode": [{"sku": v.get("sku"), "barcode": v.get("barcode")} for v in p.get("variants", [])],
            "metafields": [{"namespace": m.get("namespace"), "key": m.get("key"), "value": m.get("value")} for m in metafields],
        })

    return {"products_checked": len(results), "products": results}

def fetch_ebay_inventory_items(business_id: str) -> list:
    """Reads eBay listing data from `ebay_listing_status` — populated by the Lots
    page's "Sync Active Listings" (Trading API GetMyeBaySelling), which correctly
    covers ALL real active listings regardless of which publish path created them.
    Does NOT call /sell/inventory/v1/inventory_item directly (the original
    implementation here) — that REST endpoint only lists items created through
    that specific API, which turned out to be a small fraction (134 of ~4,500) of
    real inventory, since most listings here were published via the classic
    Trading API instead. This means eBay-side freshness for the Inventory Match
    tab depends on when Active Listings was last synced on the Lots page, not a
    fresh live pull at the moment "Sync Inventory" is clicked — a real trade-off,
    but far safer than re-implementing the Trading API pagination/XML parsing a
    second time here when a correct, checkpointed version already exists."""
    res_rows = []
    start = 0
    while True:
        page = (supabase.table("ebay_listing_status").select("item_id,sku,title,quantity_available,gallery_url,price,start_time")
                .eq("business_id", business_id).eq("listing_status", "Active")
                .range(start, start + 999).execute().data or [])
        res_rows.extend(page)
        if len(page) < 1000:
            break
        start += 1000
    return [{
        "sku": r.get("sku") or "", "title": r.get("title") or "",
        "quantity": r.get("quantity_available") or 0, "condition": "",
        "item_id": r.get("item_id"), "gallery_url": r.get("gallery_url"),
        "price": r.get("price"), "start_time": r.get("start_time"),
    } for r in res_rows if r.get("item_id")]

def fetch_shopify_inventory_items(business_id: str) -> list:
    """Lists every Shopify product/variant with price and quantity — all in the
    normal product payload, no extra calls needed."""
    import requests as _req
    settings = get_ebay_settings(business_id)
    domain = (settings.get("SHOPIFY_STORE_DOMAIN", "") or "").strip().replace("https://", "").replace("http://", "").strip("/")
    if not domain:
        return []
    token = get_shopify_access_token(business_id)
    headers = {"X-Shopify-Access-Token": token}
    items = []
    url = f"https://{domain}/admin/api/2024-10/products.json"
    params = {"limit": 250}
    while url:
        r = _req.get(url, headers=headers, params=params, timeout=20)
        if r.status_code == 401:
            token = get_shopify_access_token(business_id, force_refresh=True)
            headers["X-Shopify-Access-Token"] = token
            r = _req.get(url, headers=headers, params=params, timeout=20)
        if r.status_code != 200:
            raise Exception(f"Shopify products fetch failed ({r.status_code}): {r.text[:300]}")
        data = r.json()
        for p in data.get("products", []):
            for v in p.get("variants", []):
                items.append({
                    "variant_id": str(v.get("id", "")),
                    "product_id": str(p.get("id", "")),
                    "sku": v.get("sku", ""),
                    "title": p.get("title", ""),
                    "price": float(v.get("price", 0) or 0),
                    "quantity": v.get("inventory_quantity", 0),
                    "status": p.get("status", ""),
                    "handle": p.get("handle", ""),
                })
        link = r.headers.get("Link", "")
        next_url = None
        for part in link.split(","):
            if 'rel="next"' in part:
                next_url = part.split(";")[0].strip().strip("<>")
        url = next_url
        params = None
    return items

def _normalize_title(t: str) -> str:
    return " ".join((t or "").strip().lower().split())

def _maybe_confirm_inventory_match(business_id: str, listing_id):
    """Call this right after EITHER platform's publish writes ebay_item_id or
    shopify_product_id onto a listings row. If the row now has BOTH, this is a
    known-certain pairing — no title-matching needed at all, since Lister itself
    is the one that published both sides. Creates (or upgrades) the
    inventory_match row with hd_id = listings.id, so it's confirmed/permanent
    from the moment it exists, same guarantee as a backfilled old pairing.
    BUG FIXED: listings.id is a plain integer, but hd_id was originally a uuid-
    typed column — writing an integer into it threw a Postgres type error every
    single time, silently caught by the except block below, meaning this
    function has never actually succeeded once since it was written. hd_id is
    now a text column, and this always casts to str() to be safe regardless."""
    try:
        res = supabase.table("listings").select("id,ebay_item_id,shopify_product_id,title")\
            .eq("id", listing_id).limit(1).execute()
        if not res.data:
            return
        listing = res.data[0]
        ebay_id = listing.get("ebay_item_id")
        shopify_product_id = listing.get("shopify_product_id")
        if not (ebay_id and shopify_product_id):
            return  # only one side published so far — nothing to confirm yet

        # BUG FIXED (confirmed 100% of 139 existing dual-publish matches affected):
        # shopify_product_id is exactly what its name says -- a PRODUCT id -- but
        # every other consumer of inventory_match.shopify_id (title-matching,
        # the quantity-push feature, _fetch_shopify_variants_by_ids) expects a
        # VARIANT id, same as shopify_inventory.id. Storing the product id here
        # meant the regular title-matcher never recognized this item as already
        # matched (wrong ID type), silently creating a second, orphaned
        # "Shopify only" row for the same real listing every single time -- and
        # separately meant any feature keying off shopify_id (like the eBay+
        # Shopify quantity push) would have failed outright on any dual-
        # published item. Resolved here to the real variant id before storing.
        variant_res = (supabase.table("shopify_inventory").select("id")
                       .eq("business_id", business_id).eq("product_id", shopify_product_id)
                       .limit(1).execute().data or [])
        if not variant_res:
            print(f"_maybe_confirm_inventory_match: no shopify_inventory row for product "
                  f"{shopify_product_id} yet (not synced there yet?) — skipping for now")
            return
        shopify_id = variant_res[0]["id"]

        hd_id_value = str(listing["id"])
        existing = (supabase.table("inventory_match").select("id,hd_id")
                    .eq("business_id", business_id)
                    .or_(f"ebay_id.eq.{ebay_id},shopify_id.eq.{shopify_id}")
                    .execute().data or [])
        if existing:
            row = existing[0]
            if not row.get("hd_id"):
                supabase.table("inventory_match").update({
                    "ebay_id": ebay_id, "shopify_id": shopify_id,
                    "hd_id": hd_id_value, "matched_by": "lister_dual_publish",
                }).eq("id", row["id"]).execute()
        else:
            supabase.table("inventory_match").insert({
                "business_id": business_id, "title": listing.get("title"),
                "ebay_id": ebay_id, "shopify_id": shopify_id,
                "hd_id": hd_id_value, "matched_by": "lister_dual_publish",
            }).execute()
    except Exception as e:
        print(f"_maybe_confirm_inventory_match FAILED for listing {listing_id}: {type(e).__name__}: {e}")

_rematch_inventory_lock = {}  # business_id -> True while a rematch is in progress

def _rematch_inventory_from_cache(business_id: str) -> dict:
    """The matching/confirming logic only — reads whatever's already stored in
    ebay_inventory/shopify_inventory (from the last time a real sync ran), makes
    ZERO live API calls. Split out from sync_inventory specifically so confirming
    matches never requires a fresh eBay/Shopify pull — there was no reason those
    had to be coupled together."""
    if _rematch_inventory_lock.get(business_id):
        raise Exception("A rematch is already running for this account — wait for it to finish "
                         "before starting another (two overlapping runs were exactly what caused "
                         "the duplicate-row explosion earlier).")
    _rematch_inventory_lock[business_id] = True
    try:
        return _rematch_inventory_from_cache_inner(business_id)
    finally:
        _rematch_inventory_lock[business_id] = False

def _rematch_inventory_from_cache_inner(business_id: str) -> dict:
    errors = {}

    # SELF-HEALING dual-publish confirms — root cause of confirmed-real bug
    # ("published to both platforms but shows eBay-only, no HD_ID"): right
    # after a dual publish, _maybe_confirm_inventory_match needs the brand-new
    # Shopify product to already be in the shopify_inventory cache to resolve
    # its variant id — but a just-created product isn't there until the next
    # inventory sync, so the confirm hit "skipping for now" and NOTHING ever
    # retried. "For now" was forever. This backfill re-runs the confirm for
    # every dual-published listing that still lacks a confirmed match, every
    # time a rematch runs — by which point the variant is in the cache, so the
    # confirm that raced the sync at publish time simply completes on the next
    # cycle instead of stranding the item permanently.
    try:
        dual, start_dp = [], 0
        while True:
            page = (supabase.table("listings").select("id,ebay_item_id")
                    .eq("business_id", business_id)
                    .not_.is_("ebay_item_id", "null").not_.is_("shopify_product_id", "null")
                    .range(start_dp, start_dp + 999).execute().data or [])
            dual.extend(page)
            if len(page) < 1000:
                break
            start_dp += 1000
        if dual:
            confirmed_ebay_ids = set()
            start_c = 0
            while True:
                page = (supabase.table("inventory_match").select("ebay_id")
                        .eq("business_id", business_id).not_.is_("hd_id", "null")
                        .range(start_c, start_c + 999).execute().data or [])
                confirmed_ebay_ids.update(str(r.get("ebay_id")) for r in page if r.get("ebay_id"))
                if len(page) < 1000:
                    break
                start_c += 1000
            healed = 0
            for l in dual:
                if str(l.get("ebay_item_id")) not in confirmed_ebay_ids:
                    _maybe_confirm_inventory_match(business_id, l["id"])
                    healed += 1
            if healed:
                print(f"_rematch: retried dual-publish confirm for {healed} previously-stranded listing(s)")
    except Exception as e:
        errors["dual_publish_backfill"] = str(e)

    def _fetch_paginated(table):
        # Same 1000-row Supabase cap hit repeatedly today elsewhere — a plain
        # .execute() here was silently truncating both tables to 1000 rows each,
        # which is exactly why the summary showed 'eBay: 1000, Shopify: 1000'
        # regardless of the real totals.
        rows, start = [], 0
        while True:
            page = (supabase.table(table).select("*").eq("business_id", business_id)
                    .range(start, start + 999).execute().data or [])
            rows.extend(page)
            if len(page) < 1000:
                break
            start += 1000
        return rows

    ebay_records = _fetch_paginated("ebay_inventory")
    shopify_records = _fetch_paginated("shopify_inventory")

    confirmed_rows = []
    start_cf = 0
    while True:
        page = (supabase.table("inventory_match").select("ebay_id,shopify_id")
                .eq("business_id", business_id).not_.is_("hd_id", "null")
                .range(start_cf, start_cf + 999).execute().data or [])
        confirmed_rows.extend(page)
        if len(page) < 1000:
            break
        start_cf += 1000
    confirmed_ebay_ids = {r["ebay_id"] for r in confirmed_rows if r.get("ebay_id")}
    confirmed_shopify_ids = {r["shopify_id"] for r in confirmed_rows if r.get("shopify_id")}

    ebay_by_title = {}
    for r in ebay_records:
        if r["id"] in confirmed_ebay_ids:
            continue
        ebay_by_title.setdefault(_normalize_title(r["title"]), []).append(r["id"])
    shopify_by_title = {}
    for r in shopify_records:
        if r["id"] in confirmed_shopify_ids:
            continue
        shopify_by_title.setdefault(_normalize_title(r["title"]), []).append(r["id"])

    all_titles = set(ebay_by_title) | set(shopify_by_title)
    import uuid as _uuid
    inventory_rows = []
    for norm_title in all_titles:
        ebay_ids = ebay_by_title.get(norm_title, [])
        shopify_ids = shopify_by_title.get(norm_title, [])
        pair_count = max(len(ebay_ids), len(shopify_ids), 1)
        for i in range(pair_count):
            e_id = ebay_ids[i] if i < len(ebay_ids) else None
            s_id = shopify_ids[i] if i < len(shopify_ids) else None
            # Every row gets the SAME set of keys, including hd_id (null for
            # non-matches, a real id for matches) -- PostgREST's bulk insert
            # rejects a batch where different rows have different KEY SETS
            # (error PGRST102), but a shared key with a null value on some rows
            # is fine. This confirms matches inline, in the same insert, with no
            # separate update/upsert step needed at all -- and no risk of that
            # step ever creating a garbage NULL row, since there isn't one.
            inventory_rows.append({
                "id": str(_uuid.uuid4()),
                "business_id": business_id,
                "title": ebay_ids and norm_title or norm_title,
                "ebay_id": e_id,
                "shopify_id": s_id,
                "matched_by": "title_exact" if (e_id and s_id) else None,
                "hd_id": str(_uuid.uuid4()) if (e_id and s_id) else None,
                "hidden": False,  # baseline for every row -- see hidden_before below,
                                  # which overwrites this to True for rows that were
                                  # hidden before this rebuild. Must stay present on
                                  # every row (never conditionally added), since a
                                  # batch insert with inconsistent key sets across
                                  # rows fails (PGRST102), same constraint noted above.
            })

    newly_confirmed = sum(1 for r in inventory_rows if r["hd_id"])

    # Preserve `hidden` across this rebuild. Rows without a confirmed hd_id get
    # wiped and re-inserted fresh on every sync (by design, to re-run title
    # matching) -- but a fresh insert has no way to know a row used to be
    # hidden, so every hide would silently get undone on the next sync. Look
    # up what was hidden before the wipe, keyed by the (ebay_id, shopify_id)
    # pairing (not the row's own id, which is regenerated every rebuild), and
    # re-apply it to whichever new row represents that same pairing.
    hidden_before = {}
    try:
        hidden_res = supabase.table("inventory_match").select("ebay_id,shopify_id").eq("business_id", business_id)\
            .is_("hd_id", "null").eq("hidden", True).execute()
        hidden_before = {(r.get("ebay_id"), r.get("shopify_id")) for r in (hidden_res.data or [])}
    except Exception as e:
        print(f"inventory sync: couldn't look up previously-hidden rows, hidden flags may not survive this sync: {e}")

    for row in inventory_rows:
        if (row["ebay_id"], row["shopify_id"]) in hidden_before:
            row["hidden"] = True

    try:
        supabase.table("inventory_match").delete().eq("business_id", business_id)\
            .is_("hd_id", "null").execute()
        for i in range(0, len(inventory_rows), 500):
            supabase.table("inventory_match").insert(inventory_rows[i:i+500]).execute()
    except Exception as e:
        errors["inventory_write"] = str(e)

    return {
        "ebay_synced": len(ebay_records), "shopify_synced": len(shopify_records),
        "matched": sum(1 for r in inventory_rows if r["matched_by"]) + len(confirmed_rows),
        "ebay_only": sum(1 for r in inventory_rows if r["ebay_id"] and not r["shopify_id"]),
        "shopify_only": sum(1 for r in inventory_rows if r["shopify_id"] and not r["ebay_id"]),
        "confirmed_preserved": len(confirmed_rows),
        "newly_confirmed": newly_confirmed,
        "errors": errors,
    }

def sync_inventory(business_id: str) -> dict:
    """Pulls fresh eBay + Shopify inventory (live API calls), stores each raw, then
    calls _rematch_inventory_from_cache for the matching/confirming step. Use this
    when you actually need fresh data from the platforms; use the rematch-only
    endpoint instead when you just want to re-run matching against what's already
    stored locally."""
    errors = {}
    ebay_items, shopify_items = [], []
    try:
        ebay_items = fetch_ebay_inventory_items(business_id)
    except Exception as e:
        errors["ebay"] = str(e)
    try:
        shopify_items = fetch_shopify_inventory_items(business_id)
    except Exception as e:
        errors["shopify"] = str(e)

    ebay_records = [{
        "id": it["item_id"], "business_id": business_id, "sku": it["sku"], "title": it["title"],
        "quantity": it["quantity"], "condition": it["condition"], "price": it.get("price"),
        "category_id": None, "offer_id": None, "listing_status": None, "item_id": it["item_id"],
        "gallery_url": it.get("gallery_url"), "start_time": it.get("start_time"),
    } for it in ebay_items if it.get("item_id")]
    shopify_records = [{
        "id": it["variant_id"], "business_id": business_id, "product_id": it["product_id"],
        "sku": it["sku"], "title": it["title"], "price": it["price"], "quantity": it["quantity"],
        "status": it["status"], "handle": it["handle"],
    } for it in shopify_items if it.get("variant_id")]

    for i in range(0, len(ebay_records), 500):
        try:
            supabase.table("ebay_inventory").upsert(ebay_records[i:i+500]).execute()
        except Exception as e:
            errors["ebay_write"] = str(e)
    for i in range(0, len(shopify_records), 500):
        try:
            supabase.table("shopify_inventory").upsert(shopify_records[i:i+500]).execute()
        except Exception as e:
            errors["shopify_write"] = str(e)

    result = _rematch_inventory_from_cache(business_id)
    result["errors"] = {**errors, **result.get("errors", {})}
    return result

@app.post("/api/inventory/sync")
async def sync_inventory_now(request: Request):
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    import asyncio
    if _sync_status.get(business_id, {}).get("running"):
        return {"ok": True, "already_running": True}
    _sync_status[business_id] = {"running": True, "result": None, "started_at": __import__("datetime").datetime.utcnow().isoformat(), "finished_at": None}
    async def _run():
        import datetime as _dt
        try:
            result = await asyncio.to_thread(sync_inventory, business_id)
            _sync_status[business_id] = {"running": False, "result": result, "started_at": _sync_status.get(business_id, {}).get("started_at"), "finished_at": _dt.datetime.utcnow().isoformat()}
        except Exception as e:
            _sync_status[business_id] = {"running": False, "result": {"error": str(e)}, "started_at": _sync_status.get(business_id, {}).get("started_at"), "finished_at": _dt.datetime.utcnow().isoformat()}
    asyncio.create_task(_run())
    return {"ok": True, "started": True}

# Invisible in any renderer — but real characters to a string comparison. Confirmed
# root cause of a title that would never exact-match no matter how many times it was
# synced: the eBay-sourced title had a leading U+200B (zero-width space) that Python's
# .strip()/.split() don't treat as whitespace, so it survived every normalization
# pass and made the string permanently unequal to Shopify's clean title. Built with
# chr(), never literal characters, so the source stays reviewable — an invisible
# character typed directly into this file would defeat the whole point.
_ZERO_WIDTH_CHARS = (
    chr(0x200b),  # zero-width space
    chr(0x200c),  # zero-width non-joiner
    chr(0x200d),  # zero-width joiner
    chr(0xfeff),  # zero-width no-break space / BOM
    chr(0x00ad),  # soft hyphen
)

def _shopify_sync_norm(t):
    t = t or ""
    for ch in _ZERO_WIDTH_CHARS:
        t = t.replace(ch, "")
    return " ".join(t.strip().lower().split())

def _shopify_sync_sold_by_title(business_id: str, start_date: str, end_date: str) -> dict:
    res = supabase.table("orders").select("sku,title,quantity,order_id,legacy_item_id").eq("business_id", business_id)\
        .eq("platform", "eBay").gte("order_date", start_date).lte("order_date", end_date).execute()
    sold_by_title = {}
    for r in (res.data or []):
        title = r.get("title") or ""
        if not title:
            continue
        key = _shopify_sync_norm(title)
        entry = sold_by_title.setdefault(key, {"title": title, "sku": r.get("sku") or "", "legacy_item_id": "", "qty_sold_today": 0, "order_ids": []})
        entry["qty_sold_today"] += r.get("quantity", 0) or 0
        entry["order_ids"].append(r.get("order_id"))
        if not entry["legacy_item_id"] and r.get("legacy_item_id"):
            entry["legacy_item_id"] = r.get("legacy_item_id")
    return sold_by_title

def _fetch_all_shopify_products(domain: str, shopify_token: str, location_id: str) -> tuple:
    """Paginates the entire Shopify catalog once and returns (catalog, complete, pages)
    where catalog is norm_title -> {title, qty, inventory_item_id}. Replaces per-item
    Shopify search: their query-string search syntax silently returns zero results
    whenever a title contains '&' or '+' (confirmed — a listing with both in its title
    matched nothing, even after stripping quotes), so search was never reliable for
    real part titles. A full catalog scan sidesteps that class of bug entirely and
    costs far fewer API calls for stores with more sold titles than pages of products.

    qty is read from inventoryLevel at `location_id` specifically, NOT the variant's
    inventoryQuantity field — that field is the total across ALL locations, which
    disagreed with reality for a multi-location store (showed 1 in aggregate while
    the actual primary-location stock, the one push writes to, was 0).

    A large catalog is ~20 paginated requests; Shopify's GraphQL cost-based rate limit
    can throttle (429) partway through that run. Backs off and retries on 429, and
    paces itself using Shopify's own throttle-status extension — but rate limiting can
    still eventually exhaust the retry budget. `complete` tells the caller whether
    pagination actually reached the end or gave up partway through: the caller MUST
    NOT treat a non-complete catalog as ground truth, since previously-correct
    products.get("hasNextPage") products absent only because their page was never
    fetched would wrongly look deleted/not-found (confirmed: this silently overwrote
    a correct 'found' row back to 'not found' on a later, incomplete run)."""
    import requests as _req, time
    catalog = {}
    cursor = None
    pages = 0
    headers = {"X-Shopify-Access-Token": shopify_token, "Content-Type": "application/json"}
    while True:
        after_clause = f', after: "{cursor}"' if cursor else ""
        query = ('{ products(first: 250' + after_clause + ') { pageInfo { hasNextPage endCursor } '
                 'edges { node { title variants(first: 1) { edges { node { inventoryItem { id '
                 'inventoryLevel(locationId: "' + location_id + '") { quantities(names: ["available"]) { quantity } } '
                 '} } } } } } } }')

        body = None
        for attempt in range(8):
            r = _req.post(f"https://{domain}/admin/api/2024-10/graphql.json", headers=headers,
                           json={"query": query}, timeout=30)
            if r.status_code == 200:
                body = r.json()
                break
            if r.status_code == 429:
                wait = float(r.headers.get("Retry-After", 2 ** attempt))
                time.sleep(min(wait, 60))
                continue
            break
        if body is None:
            # Genuinely failed (not just throttled) after retries — stop here and
            # tell the caller this catalog is incomplete, don't pretend it's whole.
            return catalog, False, pages
        pages += 1

        products = (body.get("data", {}) or {}).get("products", {}) or {}
        for e in products.get("edges", []):
            node = e["node"]
            title = node.get("title")
            variant_edges = (node.get("variants") or {}).get("edges", [])
            qty = None
            inv_item_id = None
            if variant_edges:
                inv_item = variant_edges[0]["node"].get("inventoryItem") or {}
                inv_item_id = inv_item.get("id")
                quantities = (inv_item.get("inventoryLevel") or {}).get("quantities") or []
                qty = quantities[0]["quantity"] if quantities else None
            catalog[_shopify_sync_norm(title)] = {"title": title, "qty": qty, "inventory_item_id": inv_item_id}

        # Pace ourselves against Shopify's own cost budget instead of waiting to get
        # throttled — cheaper than reactive 429 handling for a long paginated run.
        throttle = ((body.get("extensions", {}) or {}).get("cost", {}) or {}).get("throttleStatus", {})
        available, cost, restore_rate = throttle.get("currentlyAvailable"), (body.get("extensions", {}) or {}).get("cost", {}).get("requestedQueryCost"), throttle.get("restoreRate")
        if available is not None and cost and restore_rate and available < cost:
            time.sleep(min((cost - available) / restore_rate, 10))

        page_info = products.get("pageInfo", {})
        if not page_info.get("hasNextPage"):
            break
        cursor = page_info.get("endCursor")
    return catalog, True, pages

def _fetch_shopify_variants_by_ids(domain: str, shopify_token: str, location_id: str, variant_ids: list) -> dict:
    """Direct Shopify GraphQL lookup for a known, exact list of variant IDs (from
    inventory_match.shopify_id, locked via hd_id) -- returns {variant_id: {title,
    qty, inventory_item_id}}. This is what makes hd_id-matched items skip the
    expensive full-catalog title scan entirely: instead of paginating every
    product and fuzzy-matching by normalized title, this asks Shopify for
    exactly the variants already known to be the right ones. Same request
    shape/auth/retry/throttle-pacing as _fetch_all_shopify_products above, just
    querying nodes(ids: ...) instead of products(first: ...) -- kept consistent
    on purpose. A batch that fails after retries is simply left out of the
    result; the caller falls back to the title-matched catalog scan for
    anything missing here, so a failure here can never break a sync, only
    make it fall back to the slower path for that batch."""
    import requests as _req, time
    result = {}
    headers = {"X-Shopify-Access-Token": shopify_token, "Content-Type": "application/json"}
    for i in range(0, len(variant_ids), 100):
        batch = variant_ids[i:i + 100]
        gids = ", ".join(f'"gid://shopify/ProductVariant/{vid}"' for vid in batch)
        query = ('{ nodes(ids: [' + gids + ']) { ... on ProductVariant { id product { title } '
                 'inventoryItem { id inventoryLevel(locationId: "' + location_id + '") '
                 '{ quantities(names: ["available"]) { quantity } } } } } }')
        body = None
        for attempt in range(8):
            r = _req.post(f"https://{domain}/admin/api/2024-10/graphql.json", headers=headers,
                           json={"query": query}, timeout=30)
            if r.status_code == 200:
                body = r.json()
                break
            if r.status_code == 429:
                wait = float(r.headers.get("Retry-After", 2 ** attempt))
                time.sleep(min(wait, 60))
                continue
            break
        if body is None:
            continue  # this batch failed after retries -- caller's catalog-scan fallback covers it
        for node in ((body.get("data") or {}).get("nodes") or []):
            if not node:
                continue
            variant_gid = node.get("id", "")
            variant_id = variant_gid.rsplit("/", 1)[-1] if variant_gid else None
            if not variant_id:
                continue
            inv_item = node.get("inventoryItem") or {}
            quantities = (inv_item.get("inventoryLevel") or {}).get("quantities") or []
            qty = quantities[0]["quantity"] if quantities else None
            product = node.get("product") or {}
            result[variant_id] = {"title": product.get("title"), "qty": qty, "inventory_item_id": inv_item.get("id")}
        throttle = ((body.get("extensions", {}) or {}).get("cost", {}) or {}).get("throttleStatus", {})
        available = throttle.get("currentlyAvailable")
        cost = (body.get("extensions", {}) or {}).get("cost", {}).get("requestedQueryCost")
        restore_rate = throttle.get("restoreRate")
        if available is not None and cost and restore_rate and available < cost:
            time.sleep(min((cost - available) / restore_rate, 10))
    return result

def _safe_int(v):
    try:
        return int(v)
    except (TypeError, ValueError):
        return None

def _safe_float(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None

def _ebay_xml_to_dict(elem):
    """Strips XML namespaces and turns an ElementTree element into nested dicts/lists
    — eBay's Trading API responses are deeply nested and namespace-qualified
    (urn:ebay:apis:eBLBaseComponents on every tag). Repeated child tags become a list."""
    children = list(elem)
    if not children:
        return elem.text
    result = {}
    for child in children:
        tag = child.tag.split("}")[-1]
        val = _ebay_xml_to_dict(child)
        if tag in result:
            if not isinstance(result[tag], list):
                result[tag] = [result[tag]]
            result[tag].append(val)
        else:
            result[tag] = val
    return result

def fetch_ebay_listing_photo_urls(token: str, item_id: str) -> list:
    """Pulls the real, currently-live photo URLs straight off an active eBay
    listing via GetItem -- eBay's own image CDN is always publicly reachable
    (that's how buyers see them), so this works as a fallback source of photos
    for anything Lister's own storage doesn't have (e.g. items published before
    a business-account migration, whose original photos live in a different,
    never-copied-over storage instance)."""
    try:
        data = _ebay_get_item_status(token, item_id)
        item = data.get("Item", data)
        pics = (item.get("PictureDetails", {}) or {}).get("PictureURL")
        if not pics:
            return []
        return pics if isinstance(pics, list) else [pics]
    except Exception as e:
        print(f"fetch_ebay_listing_photo_urls failed for {item_id}: {e}")
        return []

def _download_and_store_ebay_photos(item_id: str, ebay_token: str, max_photos: int = 2) -> list:
    """Downloads up to max_photos of an eBay listing's real photos and stores them in
    Lister's own 'part-photos' Supabase bucket -- so eBay-only inventory items get a
    permanent local copy in our own environment instead of only ever referencing
    eBay's external CDN URL. Returns the local photo_ids actually stored (empty list
    if eBay had no photos or every download/upload failed)."""
    import requests as _req
    urls = fetch_ebay_listing_photo_urls(ebay_token, item_id)[:max_photos]
    stored_ids = []
    for i, url in enumerate(urls):
        try:
            resp = _req.get(url, timeout=15)
            resp.raise_for_status()
            photo_id = f"ebay-{item_id}-{i+1}.jpg"
            supabase.storage.from_("part-photos").upload(
                photo_id, resp.content,
                {"content-type": "image/jpeg", "upsert": "true"},
            )
            stored_ids.append(photo_id)
        except Exception as e:
            print(f"_download_and_store_ebay_photos: item {item_id} photo {i+1} failed: {e}")
    return stored_ids

_ebay_photo_pull_job_status = {}  # business_id -> {"running": bool, "result": dict|None, "started_at": iso, "finished_at": iso|None}

def _ebay_photo_pull_work(business_id: str) -> dict:
    """For every genuinely eBay-ONLY inventory row (no matching Shopify product —
    same definition the Inventory tab itself uses) that doesn't already have local
    photos pulled down, downloads up to 2 real photos from the live eBay listing
    into Lister's own storage (see _download_and_store_ebay_photos) and records the
    resulting photo_ids on the ebay_inventory row itself (local_photo_ids, comma-
    separated). Matched items are skipped entirely -- they already have real photos
    from the Shopify side, no need to spend eBay API calls on them. Deliberately
    does NOT touch anything about Shopify publishing -- that's a separate step,
    this only guarantees a local, permanent copy exists for the items that need it."""
    token = get_ebay_access_token(business_id)

    # Which ebay_inventory rows are actually eBay-only, per inventory_match (the
    # same table /api/inventory itself reads to decide "eBay only" vs "Matched").
    match_rows = []
    start = 0
    while True:
        page = supabase.table("inventory_match").select("ebay_id,shopify_id")\
            .eq("business_id", business_id).range(start, start + 999).execute().data or []
        match_rows.extend(page)
        if len(page) < 1000:
            break
        start += 1000
    ebay_only_ids = list({r["ebay_id"] for r in match_rows if r.get("ebay_id") and not r.get("shopify_id")})

    if not ebay_only_ids:
        return {"checked": 0, "downloaded": 0, "already_had_photos": 0, "no_photos_found": 0, "ebay_only_count": 0}

    rows = []
    for i in range(0, len(ebay_only_ids), 200):
        chunk = ebay_only_ids[i:i+200]
        page = supabase.table("ebay_inventory").select("id,item_id,local_photo_ids")\
            .eq("business_id", business_id).in_("id", chunk).execute().data or []
        rows.extend(page)

    checked, downloaded, already_had, no_photos_found = 0, 0, 0, 0
    for row in rows:
        checked += 1
        if row.get("local_photo_ids"):
            already_had += 1
            continue
        item_id = row.get("item_id") or row.get("id")
        if not item_id:
            continue
        ids = _download_and_store_ebay_photos(str(item_id), token, max_photos=2)
        if ids:
            supabase.table("ebay_inventory").update({"local_photo_ids": ",".join(ids)}).eq("id", row["id"]).execute()
            downloaded += 1
        else:
            no_photos_found += 1
    return {"checked": checked, "downloaded": downloaded, "already_had_photos": already_had,
            "no_photos_found": no_photos_found, "ebay_only_count": len(ebay_only_ids)}

async def _run_ebay_photo_pull_background(business_id: str):
    import asyncio, datetime as _dt
    try:
        result = await asyncio.to_thread(_ebay_photo_pull_work, business_id)
        _ebay_photo_pull_job_status[business_id] = {
            "running": False, "result": result,
            "started_at": _ebay_photo_pull_job_status.get(business_id, {}).get("started_at"),
            "finished_at": _dt.datetime.utcnow().isoformat(),
        }
    except Exception as e:
        _ebay_photo_pull_job_status[business_id] = {
            "running": False, "result": {"error": str(e)},
            "started_at": _ebay_photo_pull_job_status.get(business_id, {}).get("started_at"),
            "finished_at": _dt.datetime.utcnow().isoformat(),
        }

@app.post("/api/inventory/pull-ebay-photos")
async def pull_ebay_photos(request: Request):
    """Kicks off the eBay-photo-download job as a background task (a full inventory
    can have thousands of items, well past a reverse proxy's request timeout — same
    pattern as every other bulk sync in this file)."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    import asyncio, datetime as _dt
    if _ebay_photo_pull_job_status.get(business_id, {}).get("running"):
        return {"started": False, "already_running": True}
    _ebay_photo_pull_job_status[business_id] = {
        "running": True, "result": None,
        "started_at": _dt.datetime.utcnow().isoformat(), "finished_at": None,
    }
    asyncio.create_task(_run_ebay_photo_pull_background(business_id))
    return {"started": True}

@app.get("/api/inventory/pull-ebay-photos-status")
async def pull_ebay_photos_status(request: Request, response: Response):
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    response.headers["Cache-Control"] = "no-store"
    return _ebay_photo_pull_job_status.get(business_id, {"running": False, "result": None, "started_at": None, "finished_at": None})

def _ebay_get_item_status(token: str, item_id: str) -> dict:
    """Trading API GetItem — the genuine SELLER-facing endpoint for 'what's the
    status of my own listing', works for any item regardless of how it was
    created. NOT the Buy Browse API: that's the public buyer-search/browse API,
    sharing its rate-limit quota with every other use of that API (confirmed:
    that shared quota is exactly what got exhausted before this fix). GetItem is
    a separate seller-account quota bucket entirely. OAuth token goes in the XML
    body (eBayAuthToken), not an Authorization header — that's how Trading API
    auth works, unlike the REST APIs used elsewhere in this file."""
    import requests as _req
    import xml.etree.ElementTree as ET

    xml_body = (
        '<?xml version="1.0" encoding="utf-8"?>'
        '<GetItemRequest xmlns="urn:ebay:apis:eBLBaseComponents">'
        f'<RequesterCredentials><eBayAuthToken>{token}</eBayAuthToken></RequesterCredentials>'
        f'<ItemID>{item_id}</ItemID>'
        '<DetailLevel>ReturnAll</DetailLevel>'
        '<IncludeItemSpecifics>false</IncludeItemSpecifics>'
        '</GetItemRequest>'
    )
    headers = {
        "X-EBAY-API-COMPATIBILITY-LEVEL": "1193",
        "X-EBAY-API-CALL-NAME": "GetItem",
        "X-EBAY-API-SITEID": "0",
        "Content-Type": "text/xml",
    }
    r = _req.post("https://api.ebay.com/ws/api.dll", headers=headers, data=xml_body.encode("utf-8"), timeout=20)
    root = ET.fromstring(r.content)
    return _ebay_xml_to_dict(root)

def _shopify_sync_check_one(norm_title: str, entry: dict, ebay_token: str, shopify_catalog: dict,
                             direct_variant_matches: dict = None) -> dict:
    """The eBay side of the lookup for one sold title (Shopify's side prefers a
    direct hd_id-based variant match when one exists -- see
    _fetch_shopify_variants_by_ids -- and only falls back to the pre-fetched
    title-matched catalog for items with no locked pairing). Runs off the
    request thread (see refresh endpoint) since it's pure blocking I/O —
    parallelizing this is what keeps a multi-day resync from taking forever."""
    import requests as _req

    sku = entry["sku"]
    ebay_live_qty = None
    ebay_ended = False
    if sku and sku != "(no SKU)" and not sku.lower().startswith("lister-"):
        try:
            r = _req.get(f"{EBAY_API_BASE}/sell/inventory/v1/inventory_item/{sku}",
                         headers=ebay_headers(ebay_token, content_language=False), timeout=15)
            if r.status_code == 200:
                ebay_live_qty = (r.json().get("availability", {}) or {}).get("shipToLocationAvailability", {}).get("quantity")
        except Exception:
            pass

    # Most of these listings weren't created via the Inventory API (no SKU-keyed
    # inventory item exists for them), so the lookup above returns nothing for
    # them. Fall back to Trading API's GetItem by the listing's own (legacy) item
    # ID — the real seller-facing endpoint for this, works for any listing
    # regardless of how it was created, on its own separate rate-limit quota.
    legacy_item_id = entry.get("legacy_item_id")
    if ebay_live_qty is None and legacy_item_id:
        try:
            resp = _ebay_get_item_status(ebay_token, legacy_item_id)
            ack = resp.get("Ack")
            item = resp.get("Item") or {}
            if ack in ("Success", "Warning") and item:
                selling_status = item.get("SellingStatus") or {}
                listing_status = selling_status.get("ListingStatus")
                total_qty = _safe_int(item.get("Quantity"))
                sold_qty = _safe_int(selling_status.get("QuantitySold")) or 0
                if listing_status and listing_status != "Active":
                    ebay_live_qty = 0
                    ebay_ended = True
                elif total_qty is not None:
                    ebay_live_qty = max(total_qty - sold_qty, 0)
            else:
                # GetItem fails outright (item deleted / genuinely gone) for a
                # listing that no longer exists at all — treat the same as ended.
                ebay_live_qty = 0
                ebay_ended = True
        except Exception:
            pass

    shopify_id = entry.get("shopify_id")
    match = (direct_variant_matches or {}).get(shopify_id) if shopify_id else None
    if match is None:
        match = shopify_catalog.get(norm_title)
    return {
        "norm_title": norm_title, "sku": sku, "legacy_item_id": legacy_item_id,
        "ebay_live_qty": ebay_live_qty, "ebay_ended": ebay_ended, "shopify_found": match is not None,
        "shopify_live_qty": match["qty"] if match else None,
        "shopify_title": match["title"] if match else None,
        "shopify_inventory_item_id": match["inventory_item_id"] if match else None,
    }

_shopify_sync_job_status = {}  # business_id -> {"running": bool, "result": dict|None, "started_at": iso, "finished_at": iso|None}

def _shopify_sync_refresh_work(business_id: str, items: list) -> dict:
    """The actual work, run off the event loop entirely (see the endpoint below).
    A full catalog scan with rate-limit backoff can now run well past the ~30-60s a
    reverse proxy will wait for one HTTP response (confirmed: Railway's edge
    returned a plain-text "upstream error" on a wide-range sync) — this has to be a
    background job with a polled status, the same pattern already used for
    Financials' Sync Now, not a single request/response.

    `items` is the exact list the page already has on screen (from /today) — this
    checks live quantity for exactly those items, nothing else. There is no date
    range here on purpose: /today already turned the user's filtered range into a
    concrete list of sold items (a snapshot of `orders` as of right now), so a
    second, independent date window on this side was never meaningful — it just
    silently drifted from whatever range the page was actually showing (confirmed
    bug: this used to hardcode a 30-day lookback regardless of the page's filter)."""
    import datetime as _dt, concurrent.futures

    sold_by_title = {}
    for it in items:
        title = (it.get("title") or "").strip()
        if not title:
            continue
        key = _shopify_sync_norm(title)
        sold_by_title[key] = {"title": title, "sku": it.get("sku") or "", "legacy_item_id": it.get("legacy_item_id") or ""}
    if not sold_by_title:
        return {"checked": 0}

    # Resolve each sold item's locked Shopify variant (via hd_id/inventory_match,
    # matched on legacy_item_id -- the real eBay item ID) BEFORE deciding whether
    # a full catalog scan is even needed. As of this being written, coverage is
    # 4568/4569 active listings, so on a typical day every sold item already has
    # a locked pairing and the expensive full-catalog scan below can be skipped
    # entirely -- it only runs as a fallback for genuinely unmatched items.
    legacy_ids = [v["legacy_item_id"] for v in sold_by_title.values() if v.get("legacy_item_id")]
    if legacy_ids:
        match_res = supabase.table("inventory_match").select("ebay_id,shopify_id")\
            .eq("business_id", business_id).in_("ebay_id", legacy_ids).execute()
        match_by_ebay_id = {r["ebay_id"]: r["shopify_id"] for r in (match_res.data or []) if r.get("shopify_id")}
        for v in sold_by_title.values():
            v["shopify_id"] = match_by_ebay_id.get(v.get("legacy_item_id"))
    else:
        for v in sold_by_title.values():
            v["shopify_id"] = None

    # Only ever call eBay's API for a title we've never successfully gotten a live
    # quantity for — re-checking every item in the window on every single run is
    # what exhausted eBay's daily Browse API rate limit (confirmed: hundreds of
    # redundant re-checks of already-known items across repeated syncs). No
    # .in_() filter here — real titles can contain commas/quotes that corrupt
    # PostgREST's in.() syntax (confirmed elsewhere on this page) — fetch
    # everything for the business and match locally instead.
    existing_res = supabase.table("shopify_sync_snapshot").select("norm_title,ebay_live_qty,ebay_ended,updated_at")\
        .eq("business_id", business_id).execute()
    now_check = _dt.datetime.utcnow()
    known_ebay = {}
    recent_failures = set()  # titles whose eBay lookup failed within the last 4h — skip
                              # entirely this run (don't touch/re-upsert them) so their
                              # timestamp keeps aging naturally toward the retry window,
                              # instead of being refreshed every run and never retried.
    for row in (existing_res.data or []):
        if row.get("ebay_live_qty") is not None:
            known_ebay[row["norm_title"]] = {"ebay_live_qty": row.get("ebay_live_qty"), "ebay_ended": bool(row.get("ebay_ended"))}
            continue
        updated_at = row.get("updated_at")
        if updated_at:
            try:
                age = now_check - _dt.datetime.fromisoformat(updated_at.replace("Z", "+00:00")).replace(tzinfo=None)
                if age.total_seconds() < 4 * 3600:
                    recent_failures.add(row["norm_title"])
            except Exception:
                pass

    ebay_token = get_ebay_access_token(business_id)
    settings = get_ebay_settings(business_id)
    domain = (settings.get("SHOPIFY_STORE_DOMAIN", "") or "").strip().replace("https://", "").replace("http://", "").strip("/")
    shopify_token = get_shopify_access_token(business_id) if domain else None

    shopify_catalog = {}
    direct_variant_matches = {}
    if domain and shopify_token:
        loc_headers = {"X-Shopify-Access-Token": shopify_token, "Content-Type": "application/json"}
        location_id = _get_shopify_primary_location_id(domain, loc_headers)
        if location_id:
            resolved_ids = list({v["shopify_id"] for v in sold_by_title.values() if v.get("shopify_id")})
            if resolved_ids:
                direct_variant_matches = _fetch_shopify_variants_by_ids(domain, shopify_token, location_id, resolved_ids)

            # Only pay for the expensive full-catalog title scan when something
            # is left unresolved after the direct hd_id-based lookup above --
            # either it never had a locked pairing, or the direct fetch didn't
            # return a result for it (e.g. a stale/deleted variant ID).
            still_unresolved = any(
                not v.get("shopify_id") or v["shopify_id"] not in direct_variant_matches
                for v in sold_by_title.values()
            )
            if still_unresolved:
                shopify_catalog, catalog_complete, catalog_pages = _fetch_all_shopify_products(domain, shopify_token, location_id)
                if not catalog_complete:
                    # Don't write ANYTHING — an incomplete catalog would mark real,
                    # existing products "not found" just because their page never got
                    # fetched, silently overwriting correct data from a prior run.
                    # Leaving the snapshot untouched and reporting the failure is safer
                    # than a run that "succeeds" while quietly making things wrong.
                    return {"error": f"Shopify catalog scan was rate-limited and only got through "
                                      f"{catalog_pages} page(s) ({len(shopify_catalog)} products) before giving up. "
                                      f"Nothing was changed — try Sync Now again in a minute."}

    need_ebay_check = {k: v for k, v in sold_by_title.items() if k not in known_ebay and k not in recent_failures}
    already_known = {k: v for k, v in sold_by_title.items() if k in known_ebay}
    skipped_recent_failures = {k: v for k, v in sold_by_title.items() if k in recent_failures}

    with concurrent.futures.ThreadPoolExecutor(max_workers=8) as pool:
        checked_results = list(pool.map(
            lambda kv: _shopify_sync_check_one(kv[0], kv[1], ebay_token, shopify_catalog, direct_variant_matches),
            need_ebay_check.items(),
        ))

    # Shopify data is always refreshed (cheap — it's just a dict lookup against the
    # direct matches / catalog we already fetched above), but the eBay side is
    # reused as-is rather than re-hitting an API we've already gotten a real
    # answer from.
    reused_results = []
    for norm_title, entry in already_known.items():
        known = known_ebay[norm_title]
        match = direct_variant_matches.get(entry.get("shopify_id")) if entry.get("shopify_id") else None
        if match is None:
            match = shopify_catalog.get(norm_title)
        reused_results.append({
            "norm_title": norm_title, "sku": entry["sku"], "legacy_item_id": entry.get("legacy_item_id"),
            "ebay_live_qty": known["ebay_live_qty"], "ebay_ended": known["ebay_ended"],
            "shopify_found": match is not None,
            "shopify_live_qty": match["qty"] if match else None,
            "shopify_title": match["title"] if match else None,
            "shopify_inventory_item_id": match["inventory_item_id"] if match else None,
        })

    results = checked_results + reused_results
    now_iso = _dt.datetime.utcnow().isoformat()
    rows = [{
        "business_id": business_id, "norm_title": r["norm_title"],
        "title": sold_by_title[r["norm_title"]]["title"], "sku": r["sku"],
        "legacy_item_id": r["legacy_item_id"], "ebay_live_qty": r["ebay_live_qty"], "ebay_ended": r["ebay_ended"],
        "shopify_found": r["shopify_found"], "shopify_live_qty": r["shopify_live_qty"],
        "shopify_title": r["shopify_title"], "shopify_inventory_item_id": r["shopify_inventory_item_id"],
        "updated_at": now_iso,
    } for r in results]
    supabase.table("shopify_sync_snapshot").upsert(rows, on_conflict="business_id,norm_title").execute()

    return {
        "checked": len(rows), "ebay_checked": len(checked_results), "ebay_reused": len(reused_results),
        "skipped_recent_failures": len(skipped_recent_failures),
        "synced_at": now_iso, "catalog_products": len(shopify_catalog),
    }

async def _run_shopify_sync_refresh_background(business_id: str, items: list):
    import asyncio, datetime as _dt
    try:
        result = await asyncio.to_thread(_shopify_sync_refresh_work, business_id, items)
        _shopify_sync_job_status[business_id] = {
            "running": False, "result": result,
            "started_at": _shopify_sync_job_status.get(business_id, {}).get("started_at"),
            "finished_at": _dt.datetime.utcnow().isoformat(),
        }
    except Exception as e:
        _shopify_sync_job_status[business_id] = {
            "running": False, "result": {"error": str(e)},
            "started_at": _shopify_sync_job_status.get(business_id, {}).get("started_at"),
            "finished_at": _dt.datetime.utcnow().isoformat(),
        }

@app.post("/api/shopify-sync/refresh")
async def shopify_sync_refresh(request: Request, items: List[dict] = Body(...)):
    """Kicks off the catalog/eBay live-quantity check as a background task and returns
    immediately — see _shopify_sync_refresh_work for why this can't be a synchronous
    request/response. `items` is exactly what /today returned to the page (the
    caller's currentItems) — this checks live quantity for those items only, there is
    no date range on this endpoint at all."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    import asyncio, datetime as _dt
    if _shopify_sync_job_status.get(business_id, {}).get("running"):
        return {"started": False, "already_running": True}
    _shopify_sync_job_status[business_id] = {
        "running": True, "result": None,
        "started_at": _dt.datetime.utcnow().isoformat(), "finished_at": None,
    }
    asyncio.create_task(_run_shopify_sync_refresh_background(business_id, items))
    return {"started": True}

@app.get("/api/shopify-sync/refresh-status")
async def shopify_sync_refresh_status(request: Request, response: Response):
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    response.headers["Cache-Control"] = "no-store"
    return _shopify_sync_job_status.get(business_id, {"running": False, "result": None, "started_at": None, "finished_at": None})

def _ebay_get_unsold_listings_page(token: str, page_number: int, entries_per_page: int = 200) -> dict:
    """One page of eBay's Trading API GetMyeBaySelling / UnsoldList — eBay's own
    'ended without selling, still sitting there inactive' bucket, the same thing
    Seller Hub itself calls Unsold. Same seller-facing Trading API family as
    _ebay_get_item_status (OAuth token in the XML body, not an Authorization header,
    on its own quota — NOT the Buy Browse API)."""
    import requests as _req
    import xml.etree.ElementTree as ET

    xml_body = (
        '<?xml version="1.0" encoding="utf-8"?>'
        '<GetMyeBaySellingRequest xmlns="urn:ebay:apis:eBLBaseComponents">'
        f'<RequesterCredentials><eBayAuthToken>{token}</eBayAuthToken></RequesterCredentials>'
        '<UnsoldList><Include>true</Include>'
        f'<Pagination><EntriesPerPage>{entries_per_page}</EntriesPerPage><PageNumber>{page_number}</PageNumber></Pagination>'
        '</UnsoldList>'
        '<DetailLevel>ReturnAll</DetailLevel>'
        '</GetMyeBaySellingRequest>'
    )
    headers = {
        "X-EBAY-API-COMPATIBILITY-LEVEL": "1193",
        "X-EBAY-API-CALL-NAME": "GetMyeBaySelling",
        "X-EBAY-API-SITEID": "0",
        "Content-Type": "text/xml",
    }
    r = _req.post("https://api.ebay.com/ws/api.dll", headers=headers, data=xml_body.encode("utf-8"), timeout=30)
    root = ET.fromstring(r.content)
    return _ebay_xml_to_dict(root)

def _sync_ebay_unsold_listings_work(business_id: str) -> dict:
    """Pulls every page of eBay's real UnsoldList and upserts it into
    ebay_listing_status. Never deletes -- a listing eBay's UnsoldList doesn't
    return this run (relisted, sold, or otherwise no longer unsold) just keeps
    its last-known row rather than being wiped; any sku_override on it is
    preserved regardless. (Previously did a full delete+replace scoped to
    listing_status='Unsold' -- changed after a real, confirmed case of a manual
    sku_override being permanently destroyed by an equivalent delete pattern on
    the Active-listings side, with zero audit trail to even detect it happened.)"""
    import datetime as _dt

    token = get_ebay_access_token(business_id)
    all_items = []
    page = 1
    total_pages = 1
    while page <= total_pages:
        resp = _ebay_get_unsold_listings_page(token, page)
        ack = resp.get("Ack")
        if ack not in ("Success", "Warning"):
            raise Exception(f"eBay GetMyeBaySelling failed (Ack={ack}): {resp.get('Errors')}")
        unsold = resp.get("UnsoldList") or {}
        item_array = unsold.get("ItemArray") or {}
        items = item_array.get("Item") or []
        if isinstance(items, dict):
            items = [items]
        all_items.extend(items)
        pagination = unsold.get("PaginationResult") or {}
        total_pages = _safe_int(pagination.get("TotalNumberOfPages")) or 1
        page += 1

    now_iso = _dt.datetime.utcnow().isoformat()
    rows = []
    for it in all_items:
        title = it.get("Title")
        listing_details = it.get("ListingDetails") or {}
        selling_status = it.get("SellingStatus") or {}
        rows.append({
            "business_id": business_id, "item_id": it.get("ItemID"),
            "sku": it.get("SKU"), "title": title, "norm_title": _shopify_sync_norm(title),
            # QuantityAvailable is eBay's own real remaining-count field — QuantitySold
            # isn't actually present on these items at all (confirmed against a real
            # response), so deriving "remaining" from quantity-quantity_sold was
            # silently wrong; QuantityAvailable is what's actually accurate.
            "quantity": _safe_int(it.get("Quantity")), "quantity_available": _safe_int(it.get("QuantityAvailable")),
            "price": _safe_float(selling_status.get("CurrentPrice")),
            "start_time": listing_details.get("StartTime"),
            "listing_status": "Unsold", "end_time": listing_details.get("EndTime"),
            "updated_at": now_iso,
        })

    # Upsert only -- never delete. A row this fetch doesn't include anymore (sold,
    # relisted, or just paginated differently this run) simply keeps whatever it
    # last had rather than being wiped; any sku_override on it survives
    # regardless. If it later shows up as Active, the Active-listings sync
    # updates its listing_status itself the next time that runs.
    if rows:
        supabase.table("ebay_listing_status").upsert(rows, on_conflict="business_id,item_id").execute()

    return {"checked": len(rows), "synced_at": now_iso}

def _ebay_get_active_listings_page(token: str, page_number: int, entries_per_page: int = 200) -> dict:
    """One page of eBay's Trading API GetMyeBaySelling / ActiveList — same call
    family as _ebay_get_unsold_listings_page, just the Active bucket instead of
    Unsold. Used by the Lots page to count each lot's currently-active listings."""
    import requests as _req
    import xml.etree.ElementTree as ET

    xml_body = (
        '<?xml version="1.0" encoding="utf-8"?>'
        '<GetMyeBaySellingRequest xmlns="urn:ebay:apis:eBLBaseComponents">'
        f'<RequesterCredentials><eBayAuthToken>{token}</eBayAuthToken></RequesterCredentials>'
        '<ActiveList><Include>true</Include>'
        f'<Pagination><EntriesPerPage>{entries_per_page}</EntriesPerPage><PageNumber>{page_number}</PageNumber></Pagination>'
        '</ActiveList>'
        '<IncludeWatchCount>true</IncludeWatchCount>'
        '<DetailLevel>ReturnAll</DetailLevel>'
        '</GetMyeBaySellingRequest>'
    )
    headers = {
        "X-EBAY-API-COMPATIBILITY-LEVEL": "1193",
        "X-EBAY-API-CALL-NAME": "GetMyeBaySelling",
        "X-EBAY-API-SITEID": "0",
        "Content-Type": "text/xml",
    }
    r = _req.post("https://api.ebay.com/ws/api.dll", headers=headers, data=xml_body.encode("utf-8"), timeout=30)
    root = ET.fromstring(r.content)
    return _ebay_xml_to_dict(root)

class EbayCallLimitError(Exception):
    """eBay error 518 -- this specific Trading API call name has hit its usage
    limit (per-app, resets on eBay's own schedule). Distinct from other
    failures so callers can stop burning further calls immediately instead of
    retrying into a dead quota, and so a rate-limit is never silently
    indistinguishable from a confirmed 'no offer' result."""
    def __init__(self, item_id, errors):
        self.item_id = item_id
        self.errors = errors
        super().__init__(f"eBay call limit reached (item {item_id}): {errors}")

def _ebay_get_best_offers_for_item(token: str, item_id: str) -> dict:
    """GetBestOffers for a single ItemID -- this is the documented, individually
    tested mode (eBay's own Sandbox validation steps call this out explicitly
    as step 2, after confirming step 1 -- no ItemID at all -- works). Used
    instead of the account-wide no-ItemID call, which returned an empty
    ItemBestOffersArray in practice on this account despite Ack=Success."""
    import requests as _req
    import xml.etree.ElementTree as ET

    xml_body = (
        '<?xml version="1.0" encoding="utf-8"?>'
        '<GetBestOffersRequest xmlns="urn:ebay:apis:eBLBaseComponents">'
        f'<RequesterCredentials><eBayAuthToken>{token}</eBayAuthToken></RequesterCredentials>'
        f'<ItemID>{item_id}</ItemID>'
        '<BestOfferStatus>Active</BestOfferStatus>'
        '<DetailLevel>ReturnAll</DetailLevel>'
        '</GetBestOffersRequest>'
    )
    headers = {
        "X-EBAY-API-COMPATIBILITY-LEVEL": "1193",
        "X-EBAY-API-CALL-NAME": "GetBestOffers",
        "X-EBAY-API-SITEID": "0",
        "Content-Type": "text/xml",
    }
    r = _req.post("https://api.ebay.com/ws/api.dll", headers=headers, data=xml_body.encode("utf-8"), timeout=30)
    root = ET.fromstring(r.content)
    return _ebay_xml_to_dict(root)

_CLOSED_OFFER_STATUSES = {"Declined", "Expired", "Retracted", "Withdrawn", "AdminEnded", "Accepted"}

def _sync_one_item_best_offers(business_id: str, token: str, item_id: str) -> list:
    """Shared by the notification handler, the 10-min poll worker (safety
    net), and the full-scan. Returns the list of currently-open best-offer-
    ids found for this item after upsert. IMPORTANT: raises on a failed Ack
    instead of silently returning [] -- a rate-limit or auth failure must
    never look identical to a confirmed no-offer-here result to any caller
    (real bug found 8/7: the full-scan's 0-found result was untrustworthy
    because API failures were being counted the same as true negatives)."""
    import datetime as _dt

    resp = _ebay_get_best_offers_for_item(token, item_id)
    ack = resp.get("Ack")
    if ack not in ("Success", "Warning"):
        errors = resp.get("Errors") or {}
        error_code = (errors.get("ErrorCode") if isinstance(errors, dict) else None)
        raise EbayCallLimitError(item_id, errors) if error_code == "518" else Exception(f"item {item_id} Ack={ack}: {errors}")
    best_offer_array = resp.get("BestOfferArray") or {}
    offers = best_offer_array.get("BestOffer") or []
    if isinstance(offers, dict):
        offers = [offers]
    now_iso = _dt.datetime.utcnow().isoformat()
    seen_ids = []
    rows = []
    for off in offers:
        status = off.get("Status")
        # REAL BUG FIXED 8/8: this used to keep only status == "Active",
        # which silently dropped offers eBay reports as "Pending" -- the
        # state a currently-open counter-offer negotiation is actually in.
        # That's exactly why an offer someone was actively waiting to
        # respond to never showed up here. Now excludes only genuinely
        # closed/terminal statuses and keeps everything else, since the
        # safe failure mode for "did I miss an offer" is to over-include,
        # never to silently drop something still open.
        if status in _CLOSED_OFFER_STATUSES:
            continue
        best_offer_id = off.get("BestOfferID")
        if not best_offer_id:
            continue
        buyer = (off.get("Buyer") or {}).get("UserID")
        price_raw = off.get("Price")
        price = (price_raw.get("value") if isinstance(price_raw, dict) else price_raw)
        seen_ids.append(best_offer_id)
        rows.append({
            "business_id": business_id, "best_offer_id": best_offer_id,
            "item_id": item_id, "buyer_user_id": buyer,
            "offer_price": _safe_float(price),
            "quantity": _safe_int(off.get("Quantity")),
            "status": status, "expiration_time": off.get("ExpirationTime"),
            "buyer_message": off.get("BuyerMessage"),
            "updated_at": now_iso,
        })
    if rows:
        supabase.table("ebay_best_offers").upsert(rows, on_conflict="business_id,best_offer_id").execute()
    # Clear any previously-stored offer for THIS item that's no longer open
    # (accepted/declined/expired/withdrawn) -- scoped to item_id so a partial
    # (single-item) sync never touches other items' offers.
    existing = (supabase.table("ebay_best_offers").select("best_offer_id")
                .eq("business_id", business_id).eq("item_id", item_id).execute()).data or []
    stale = [r["best_offer_id"] for r in existing if r["best_offer_id"] not in seen_ids]
    if stale:
        supabase.table("ebay_best_offers").delete().eq("business_id", business_id).in_("best_offer_id", stale).execute()
    return seen_ids

def _find_item_ids_in_notification(obj, found=None) -> list:
    """eBay Platform Notification payloads vary in exact shape by event type and
    aren't worth hard-coding a schema for -- this just recursively hunts the
    parsed XML for any ItemID value(s) present anywhere, which every Best-Offer
    notification carries somewhere. The notification is only ever used as a
    wake-up trigger; the authoritative offer data always comes from the
    already-proven GetBestOffers(ItemID=X) call, never parsed out of the
    notification body itself."""
    if found is None:
        found = []
    if isinstance(obj, dict):
        for k, v in obj.items():
            if k == "ItemID" and isinstance(v, str):
                found.append(v)
            else:
                _find_item_ids_in_notification(v, found)
    elif isinstance(obj, list):
        for v in obj:
            _find_item_ids_in_notification(v, found)
    return found

@app.post("/api/ebay/notifications")
async def ebay_platform_notification(request: Request):
    """Receives eBay Platform Notifications (BestOfferPlaced) -- this is what
    makes the Offers tab need zero manual refreshing: eBay pushes here the
    instant an offer is placed, instead of Lister ever having to poll and hope
    the timing lines up. No auth on this route (eBay is the caller, not a
    logged-in user); always returns 200 quickly since eBay retries/flags the
    URL unreliable on repeated non-200s. Business is resolved by trying the
    item against every business with an eBay token connected (single-tenant
    today, but written to not assume that)."""
    import xml.etree.ElementTree as ET
    try:
        body = await request.body()
        root = ET.fromstring(body)
        parsed = _ebay_xml_to_dict(root)
        item_ids = list(set(_find_item_ids_in_notification(parsed)))
        if item_ids:
            res = supabase.table("app_settings").select("business_id").eq("key", "EBAY_REFRESH_TOKEN").execute()
            business_ids = list(set(r["business_id"] for r in (res.data or [])))
            for biz_id in business_ids:
                token = get_ebay_access_token(biz_id)
                for item_id in item_ids:
                    try:
                        found = _sync_one_item_best_offers(biz_id, token, item_id)
                        if found:
                            print(f"ebay_platform_notification: item {item_id} -> {len(found)} active offer(s) for business {biz_id}")
                    except Exception as e:
                        print(f"ebay_platform_notification: item {item_id} failed for business {biz_id}: {e}")
    except Exception as e:
        print(f"ebay_platform_notification: failed to process payload: {e}")
    return Response(status_code=200)

def _ebay_set_notification_preferences(business_id: str, token: str, application_url: str) -> dict:
    """Subscribes this app to BestOffer push notifications, delivered to
    application_url (/api/ebay/notifications). BestOffer is the SELLER-facing
    'you received an offer' event -- BestOfferPlaced (tried first) is actually
    the BUYER's own confirmation notification and would never fire for this
    account. SetNotificationPreferences also requires the app's identity
    headers (App/Dev/Cert ID), unlike the other Trading API calls this app
    makes, which only need the user's auth token. Idempotent -- safe to call
    every startup, always just resets to this same preference set."""
    import requests as _req
    import xml.etree.ElementTree as ET

    settings = get_ebay_settings(business_id)
    app_id = settings.get("EBAY_APP_ID", "")
    dev_id = settings.get("EBAY_DEV_ID", "")
    cert_id = settings.get("EBAY_CERT_ID", "")

    xml_body = (
        '<?xml version="1.0" encoding="utf-8"?>'
        '<SetNotificationPreferencesRequest xmlns="urn:ebay:apis:eBLBaseComponents">'
        f'<RequesterCredentials><eBayAuthToken>{token}</eBayAuthToken></RequesterCredentials>'
        '<ApplicationDeliveryPreferences>'
        '<ApplicationEnable>Enable</ApplicationEnable>'
        f'<ApplicationURL>{application_url}</ApplicationURL>'
        '<DeviceType>Platform</DeviceType>'
        '</ApplicationDeliveryPreferences>'
        '<UserDeliveryPreferenceArray>'
        '<NotificationEnable><EventType>BestOffer</EventType><EventEnable>Enable</EventEnable></NotificationEnable>'
        '</UserDeliveryPreferenceArray>'
        '</SetNotificationPreferencesRequest>'
    )
    headers = {
        "X-EBAY-API-COMPATIBILITY-LEVEL": "1193",
        "X-EBAY-API-CALL-NAME": "SetNotificationPreferences",
        "X-EBAY-API-SITEID": "0",
        "X-EBAY-API-APP-NAME": app_id,
        "X-EBAY-API-DEV-NAME": dev_id,
        "X-EBAY-API-CERT-NAME": cert_id,
        "Content-Type": "text/xml",
    }
    r = _req.post("https://api.ebay.com/ws/api.dll", headers=headers, data=xml_body.encode("utf-8"), timeout=30)
    root = ET.fromstring(r.content)
    return _ebay_xml_to_dict(root)

async def ebay_notification_subscribe_worker():
    """Keeps retrying for the life of the process (not just at startup) so
    BestOffer push self-heals with zero manual action. REAL BUG FIXED 8/7:
    this used to give up permanently after 5 backoff attempts (~7.5 min) --
    fine for a cold-boot-token-not-ready hiccup, but fatal if eBay's Trading
    API is returning 518 (call usage limit) at boot, since that quota can
    stay dead for hours. A permanent give-up meant the subscription would
    never be (re)established once the quota reset unless someone manually
    redeployed. Now: fast backoff for the first few tries (cold-boot case),
    then settles into a steady 30-min retry forever until it succeeds once
    per business -- cheap (one call per business per 30 min) and safe to
    keep calling (idempotent, see _ebay_set_notification_preferences)."""
    import asyncio
    app_url = "https://lister-web-aocopy-production-halfdome.up.railway.app/api/ebay/notifications"
    subscribed = set()
    attempt = 0
    while True:
        attempt += 1
        try:
            res = supabase.table("app_settings").select("business_id").eq("key", "EBAY_REFRESH_TOKEN").execute()
            business_ids = list(set(r["business_id"] for r in (res.data or [])))
            for biz_id in business_ids:
                if biz_id in subscribed:
                    continue
                try:
                    token = await asyncio.to_thread(get_ebay_access_token, biz_id)
                    result = await asyncio.to_thread(_ebay_set_notification_preferences, biz_id, token, app_url)
                    if result.get("Ack") in ("Success", "Warning"):
                        print(f"ebay_notification_subscribe_worker: subscribed business {biz_id} to BestOffer")
                        subscribed.add(biz_id)
                    else:
                        print(f"ebay_notification_subscribe_worker: business {biz_id} failed (attempt {attempt}): {result.get('Errors')}")
                except Exception as e:
                    print(f"ebay_notification_subscribe_worker: business {biz_id} error (attempt {attempt}): {e}")
            if business_ids and all(b in subscribed for b in business_ids):
                return  # every known business subscribed -- nothing left to retry
        except Exception as e:
            print(f"ebay_notification_subscribe_worker error (attempt {attempt}): {e}")
        await asyncio.sleep(min(30 * attempt, 1800))  # fast backoff early, capped at 30 min steady-state

def _ebay_get_all_best_offers(token: str) -> list:
    """The real, documented, account-wide GetBestOffers call -- confirmed
    working 8/8 after adding the Pagination container eBay's own docs
    describe as being for exactly this case ('the GetBestOffers call will
    retrieve a large number of results'). Its absence on the first attempt
    is why that earlier try returned an empty result on this ~4,700-listing
    account and made a whole per-item-candidate funnel seem necessary.
    Loops pages defensively in case entries ever exceed one page."""
    import requests as _req
    import xml.etree.ElementTree as ET

    all_entries = []
    page = 1
    while True:
        xml_body = (
            '<?xml version="1.0" encoding="utf-8"?>'
            '<GetBestOffersRequest xmlns="urn:ebay:apis:eBLBaseComponents">'
            f'<RequesterCredentials><eBayAuthToken>{token}</eBayAuthToken></RequesterCredentials>'
            '<BestOfferStatus>Active</BestOfferStatus>'
            f'<Pagination><EntriesPerPage>200</EntriesPerPage><PageNumber>{page}</PageNumber></Pagination>'
            '<DetailLevel>ReturnAll</DetailLevel>'
            '</GetBestOffersRequest>'
        )
        headers = {
            "X-EBAY-API-COMPATIBILITY-LEVEL": "1193",
            "X-EBAY-API-CALL-NAME": "GetBestOffers",
            "X-EBAY-API-SITEID": "0",
            "Content-Type": "text/xml",
        }
        r = _req.post("https://api.ebay.com/ws/api.dll", headers=headers, data=xml_body.encode("utf-8"), timeout=30)
        root = ET.fromstring(r.content)
        resp = _ebay_xml_to_dict(root)
        ack = resp.get("Ack")
        if ack not in ("Success", "Warning"):
            errors = resp.get("Errors") or {}
            error_code = (errors.get("ErrorCode") if isinstance(errors, dict) else None)
            raise EbayCallLimitError(f"bulk-page-{page}", errors) if error_code == "518" else Exception(f"bulk best offers page {page} Ack={ack}: {errors}")
        array = (resp.get("ItemBestOffersArray") or {}).get("ItemBestOffers") or []
        if isinstance(array, dict):
            array = [array]
        for entry in array:
            item = entry.get("Item") or {}
            item_id = item.get("ItemID")
            offers = (entry.get("BestOfferArray") or {}).get("BestOffer") or []
            if isinstance(offers, dict):
                offers = [offers]
            for off in offers:
                if off.get("Status") in _CLOSED_OFFER_STATUSES:
                    continue
                all_entries.append({"item_id": item_id, "item_title": item.get("Title"), "offer": off})
        total_pages = _safe_int((resp.get("PaginationResult") or {}).get("TotalNumberOfPages")) or 1
        if page >= total_pages:
            break
        page += 1
    return all_entries

def _sync_ebay_best_offers_work(business_id: str) -> dict:
    """PRIMARY mechanism as of 8/8 -- one documented paginated call
    (_ebay_get_all_best_offers) instead of a bulk-listing-count filter
    followed by per-item verification. Still runs on its own 10-min clock
    as a safety net alongside push notifications, but no longer depends on
    ebay_listing_status.best_offer_count at all (that field turned out to
    be a lifetime/historical count, not 'currently open' -- see 8/8 notes)."""
    import datetime as _dt

    token = get_ebay_access_token(business_id)
    now_iso = _dt.datetime.utcnow().isoformat()

    entries = _ebay_get_all_best_offers(token)
    seen_ids = []
    rows = []
    for e in entries:
        off = e["offer"]
        best_offer_id = off.get("BestOfferID")
        if not best_offer_id:
            continue
        price_raw = off.get("Price")
        price = (price_raw.get("value") if isinstance(price_raw, dict) else price_raw)
        seen_ids.append(best_offer_id)
        rows.append({
            "business_id": business_id, "best_offer_id": best_offer_id,
            "item_id": e["item_id"], "buyer_user_id": (off.get("Buyer") or {}).get("UserID"),
            "offer_price": _safe_float(price),
            "quantity": _safe_int(off.get("Quantity")),
            "status": off.get("Status"), "expiration_time": off.get("ExpirationTime"),
            "buyer_message": off.get("BuyerMessage"),
            "updated_at": now_iso,
        })
    if rows:
        supabase.table("ebay_best_offers").upsert(rows, on_conflict="business_id,best_offer_id").execute()
    existing = (supabase.table("ebay_best_offers").select("best_offer_id")
                .eq("business_id", business_id).execute()).data or []
    stale = [r["best_offer_id"] for r in existing if r["best_offer_id"] not in seen_ids]
    if stale:
        supabase.table("ebay_best_offers").delete().eq("business_id", business_id).in_("best_offer_id", stale).execute()

    return {"active_offers": len(rows), "synced_at": now_iso}

async def ebay_best_offers_sync_worker():
    """Safety-net worker, own 10-minute clock -- see _sync_ebay_best_offers_work.
    Primary path is push notifications; this just guards against a missed one."""
    import asyncio
    while True:
        try:
            res = supabase.table("app_settings").select("business_id").eq("key", "EBAY_REFRESH_TOKEN").execute()
            business_ids = list(set(r["business_id"] for r in (res.data or [])))
            for biz_id in business_ids:
                try:
                    result = await asyncio.to_thread(_sync_ebay_best_offers_work, biz_id)
                    print(f"ebay_best_offers_sync_worker: business {biz_id}: {result}")
                except Exception as e:
                    print(f"ebay_best_offers_sync_worker: business {biz_id} failed: {e}")
        except Exception as e:
            print(f"ebay_best_offers_sync_worker error: {e}")
        await asyncio.sleep(600)  # 10 min

@app.get("/api/best-offers")
async def list_best_offers(request: Request):
    """Backs the top section of the Offers page: currently-open incoming
    Best Offers, joined against ebay_listing_status for title/price/photo so
    the page doesn't need a second round trip. Read-only for now -- accepting,
    declining, or countering an offer is a separate build (real eBay actions,
    wants its own confirmation step rather than being bolted on here)."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    # REAL BUG FIXED 8/8: this filtered status == "Active" only, which hid
    # any offer eBay reports as "Pending" (an open counter-offer) even
    # though the sync worker (fixed earlier the same day) was already
    # writing those rows correctly -- the write path got fixed, this read
    # path didn't, so nothing changed on the actual page. Now excludes
    # only genuinely closed/terminal statuses, matching the sync worker.
    offers = (supabase.table("ebay_best_offers").select("*")
              .eq("business_id", business_id).not_.in_("status", list(_CLOSED_OFFER_STATUSES))
              .order("expiration_time").execute()).data or []
    item_ids = list({o["item_id"] for o in offers if o.get("item_id")})
    listings_by_id = {}
    if item_ids:
        # Supabase .in_() is fine at this scale -- offer volume is small (~20/day),
        # unlike the listings table itself which needs real pagination.
        listing_rows = (supabase.table("ebay_listing_status")
                         .select("item_id,title,price,gallery_url,sku")
                         .eq("business_id", business_id).in_("item_id", item_ids).execute()).data or []
        listings_by_id = {r["item_id"]: r for r in listing_rows}
    for o in offers:
        listing = listings_by_id.get(o.get("item_id")) or {}
        o["title"] = listing.get("title")
        o["listing_price"] = listing.get("price")
        o["gallery_url"] = listing.get("gallery_url")
        o["sku"] = listing.get("sku")
    return {"offers": offers, "count": len(offers)}

@app.post("/api/best-offers/sync-now")
async def best_offers_sync_now(request: Request):
    """Manual trigger for the same job the 10-min safety-net worker runs.
    Not the primary mechanism anymore (push notifications are) but kept for
    an immediate on-demand check."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    import asyncio
    result = await asyncio.to_thread(_sync_ebay_best_offers_work, business_id)
    return result

_best_offers_full_scan_status = {}  # business_id -> {"running": bool, "checked": int, "total": int, "found": int, "started_at": iso, "finished_at": iso|None}

async def _run_best_offers_full_scan(business_id: str):
    """Checks EVERY active listing directly via GetBestOffers, concurrently
    (semaphore-limited) -- does NOT depend on best_offer_count, so it finds a
    real pending offer even though the Inventory sync that populates that
    count has been stuck. This is a one-time heavy fallback (thousands of API
    calls), not something to run routinely -- the push notification handler
    and the 10-min safety-net poller are the real ongoing mechanisms; this
    exists to find what's ALREADY sitting there right now, since nothing else
    currently can."""
    import asyncio

    token = get_ebay_access_token(business_id)
    # BUG FIXED: was a single unpaginated .execute(), which Supabase silently
    # caps at 1000 rows -- this codebase already has a documented fix for
    # exactly this (Supabase caps queries around 1000 rows, paginate
    # everything) and this call didn't use it, so the scan was silently only
    # ever checking the first 1000 of ~4,700 active listings and would report
    # "0 found" even with a real offer sitting in the other ~3,700 untouched.
    all_items = []
    start_i, page_size = 0, 1000
    while True:
        page = (supabase.table("ebay_listing_status").select("item_id")
                .eq("business_id", business_id).eq("listing_status", "Active")
                .range(start_i, start_i + page_size - 1).execute()).data or []
        all_items.extend(page)
        if len(page) < page_size:
            break
        start_i += page_size
    item_ids = [r["item_id"] for r in all_items if r.get("item_id")]

    status = _best_offers_full_scan_status[business_id]
    status["total"] = len(item_ids)
    status["errors"] = 0
    status["rate_limited"] = False

    sem = asyncio.Semaphore(15)

    async def _check_one(item_id):
        async with sem:
            if status["rate_limited"]:
                return  # quota already confirmed dead this run -- don't waste more calls
            try:
                found = await asyncio.to_thread(_sync_one_item_best_offers, business_id, token, item_id)
                status["checked"] += 1
                status["found"] += len(found)
            except EbayCallLimitError as e:
                status["errors"] += 1
                status["rate_limited"] = True
                print(f"_run_best_offers_full_scan: eBay call limit hit at item {item_id} -- stopping scan early, remaining items untested")
            except Exception as e:
                status["errors"] += 1
                print(f"_run_best_offers_full_scan: item {item_id} failed: {e}")

    await asyncio.gather(*[_check_one(i) for i in item_ids])

    import datetime as _dt
    status["running"] = False
    status["finished_at"] = _dt.datetime.utcnow().isoformat()
    print(f"_run_best_offers_full_scan: business {business_id} done -- checked {status['checked']}/{status['total']}, found {status['found']} active offer(s)")

@app.post("/api/best-offers/full-scan")
async def best_offers_full_scan(request: Request):
    """One-time direct scan of every active listing -- bypasses best_offer_count
    entirely, so it works even though the Inventory sync is currently stuck.
    Runs as a background task (thousands of listings, concurrency-limited) and
    returns immediately; poll /api/best-offers/full-scan-status for progress."""
    import asyncio, datetime as _dt
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    if _best_offers_full_scan_status.get(business_id, {}).get("running"):
        return {"started": False, "already_running": True}
    _best_offers_full_scan_status[business_id] = {
        "running": True, "checked": 0, "total": 0, "found": 0, "errors": 0, "rate_limited": False,
        "started_at": _dt.datetime.utcnow().isoformat(), "finished_at": None,
    }
    asyncio.create_task(_run_best_offers_full_scan(business_id))
    return {"started": True}

@app.get("/api/best-offers/full-scan-status")
async def best_offers_full_scan_status(request: Request):
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    return _best_offers_full_scan_status.get(business_id, {"running": False, "checked": 0, "total": 0, "found": 0})

@app.get("/api/best-offers/debug-item/{item_id}")
async def best_offers_debug_item(item_id: str, request: Request):
    """One-off diagnostic -- returns eBay's RAW GetBestOffers response for a
    single item, unparsed/unfiltered, so a specific 'I have an offer and it's
    not showing' report can be checked against ground truth directly instead
    of trusting our own parsing logic."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    import asyncio
    token = await asyncio.to_thread(get_ebay_access_token, business_id)
    raw = await asyncio.to_thread(_ebay_get_best_offers_for_item, token, item_id)
    return raw

# ── SELL ANALYTICS (real view counts) ──────────────────────────── #
# Trading API's HitCount is effectively dead (deprecated for AddItem/GetItem,
# and GetSellerList/GetMyeBaySelling only return it to specially-authorized
# apps -- confirmed via eBay's own docs and dev-support forum, 8/8). Real
# page-view data lives in the separate Sell Analytics API's getTrafficReport
# method, which needs its own OAuth scope (sell.analytics.readonly) --
# added to EBAY_OAUTH_SCOPES above, requires the account to re-consent once.
#
# NOTE: this API's own daily call limit is real and separate from the
# Trading API's -- eBay's own docs say "use this method for up to a few
# thousand listings per day; it is not intended to return daily metrics for
# all your listings," and sellers report an effective ~500 calls/day cap.
# listing_ids can be batched (pipe-separated) into one call, so a full sweep
# of ~4,700 active listings is ~35-50 calls in batches of 100-150 -- fine for
# a once-a-day job, NOT something to run on a 10-min clock like best offers.
#
# WINDOW: eBay's own docs state the hard max span for date_range on this
# call is 90 days -- there is no lifetime/since-listing-start option at all.
# watch_count (Trading API) is NOT windowed -- it's the current cumulative
# total since the listing went live, which can be months old. That mismatch
# (a 90-day view window vs. a lifetime watcher count) is real and will still
# make an old, slow-moving listing look like it has "more watchers than
# views" sometimes -- 90 days is just the closest we can get within eBay's
# own limit, not a full fix. This used to be set to 30 days as an arbitrary
# "trending" choice, not because eBay required it -- widened to the true 90-
# day max since more real signal is strictly better here.

_ANALYTICS_BATCH_SIZE = 100  # ids per call -- keeps URL length and per-call cost modest
_ANALYTICS_WINDOW_DAYS = 90  # eBay's documented max span for getTrafficReport's date_range

def _ebay_get_traffic_report(token: str, item_ids: list, days: int = _ANALYTICS_WINDOW_DAYS) -> dict:
    """One getTrafficReport call, dimension=LISTING, metric=LISTING_VIEWS_TOTAL,
    for up to _ANALYTICS_BATCH_SIZE item_ids at once, over a trailing `days`-day
    window (eBay's own docs cap date_range at 90 days max for this call --
    there's no lifetime option, so this is the most history obtainable)."""
    import requests as _req, datetime as _dt, urllib.parse as _up

    end = _dt.datetime.utcnow().date()
    start = end - _dt.timedelta(days=days)
    ids_part = "|".join(item_ids)
    filter_val = f"marketplace_ids:{{EBAY_US}},date_range:[{start:%Y%m%d}..{end:%Y%m%d}],listing_ids:{{{ids_part}}}"
    params = {
        "filter": filter_val,
        "dimension": "LISTING",
        "metric": "LISTING_VIEWS_TOTAL",
    }
    url = "https://api.ebay.com/sell/analytics/v1/traffic_report?" + _up.urlencode(params, safe="{}[]|,:")
    r = _req.get(url, headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"}, timeout=30)
    if r.status_code != 200:
        raise Exception(f"getTrafficReport {r.status_code}: {r.text[:500]}")
    return r.json()

@app.get("/api/analytics/debug-traffic-test")
async def analytics_debug_traffic_test(request: Request, item_ids: str = ""):
    """MANUAL, one-time-use diagnostic -- NOT called automatically anywhere.
    Confirms the real getTrafficReport response shape against a small,
    caller-specified batch of item_ids (comma-separated) before any
    unattended batch job is built to rely on parsing it. Requires the
    account to have already re-consented with sell.analytics.readonly,
    or this will fail with an insufficient-scope error, which is itself
    useful confirmation that re-consent is still needed."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    ids = [i.strip() for i in item_ids.split(",") if i.strip()]
    if not ids:
        raise HTTPException(400, "Pass ?item_ids=123,456 (a few real item IDs from this account)")
    import asyncio
    token = await asyncio.to_thread(get_ebay_access_token, business_id)
    raw = await asyncio.to_thread(_ebay_get_traffic_report, token, ids)
    return raw

def _sync_ebay_analytics_work(business_id: str) -> dict:
    """Full sweep: pulls LISTING_VIEWS_TOTAL for every Active listing in
    batches of _ANALYTICS_BATCH_SIZE, writes into ebay_listing_status.view_count.
    REAL BUG FIXED 8/8 (caught on the very first live run): the active-item-id
    query had no pagination, so it silently hit Supabase's default 1000-row
    cap and only ever processed the first 1000 of 4,698 active listings --
    the other 3,698 were never checked, with nothing in the return value to
    say so. Now paginates the same way _fetch_all_for_business does."""
    import datetime as _dt

    token = get_ebay_access_token(business_id)
    item_ids, start_i, page_size = [], 0, 1000
    while True:
        page = (supabase.table("ebay_listing_status").select("item_id")
                .eq("business_id", business_id).eq("listing_status", "Active")
                .range(start_i, start_i + page_size - 1).execute()).data or []
        item_ids.extend(r["item_id"] for r in page if r.get("item_id"))
        if len(page) < page_size:
            break
        start_i += page_size

    updated = 0
    now_iso = _dt.datetime.utcnow().isoformat()
    for i in range(0, len(item_ids), _ANALYTICS_BATCH_SIZE):
        batch = item_ids[i:i + _ANALYTICS_BATCH_SIZE]
        resp = _ebay_get_traffic_report(token, batch)
        records = resp.get("records") or []
        for rec in records:
            dim_vals = rec.get("dimensionValues") or []
            met_vals = rec.get("metricValues") or []
            if not dim_vals or not met_vals:
                continue
            item_id = dim_vals[0].get("value")
            views = _safe_int(met_vals[0].get("value"))
            if item_id is None:
                continue
            supabase.table("ebay_listing_status").update({"view_count": views, "updated_at": now_iso}) \
                .eq("business_id", business_id).eq("item_id", item_id).execute()
            updated += 1

    return {"total_active": len(item_ids), "updated": updated, "synced_at": now_iso}

@app.post("/api/analytics/sync-views")
async def analytics_sync_views(request: Request):
    """Manual trigger for the full view-count sweep. Kept manual (not an
    automatic worker) until confirmed working end-to-end at least once --
    same pattern as every other 8/8 fix: prove the response shape, then
    automate."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    import asyncio
    result = await asyncio.to_thread(_sync_ebay_analytics_work, business_id)
    return result

def _sync_ebay_active_listings_work(business_id: str, resume: bool = True) -> dict:
    """Pulls eBay's real ActiveList one page at a time, upserting each page
    immediately and checkpointing progress in app_settings (via
    get_ebay_settings/save_ebay_setting — the same key-value store already used by
    sync_orders_for_business's ORDERS_SYNC_CHECKPOINT) — so a large account survives
    a server restart or crash mid-run by resuming from the last completed page
    instead of starting over. The user's own internet connection has no bearing on
    this at all, since it already runs as a detached server-side background task."""
    import datetime as _dt, time as _time

    token = get_ebay_access_token(business_id)
    settings = get_ebay_settings(business_id)

    page = 1
    run_started_at = None
    if resume:
        checkpoint_page = _safe_int(settings.get("EBAY_ACTIVE_LISTINGS_SYNC_CHECKPOINT_PAGE"))
        run_started_at = settings.get("EBAY_ACTIVE_LISTINGS_SYNC_RUN_STARTED_AT") or None
        if checkpoint_page and run_started_at:
            page = checkpoint_page + 1

    if not run_started_at:
        run_started_at = _dt.datetime.utcnow().isoformat()
        save_ebay_setting(business_id, "EBAY_ACTIVE_LISTINGS_SYNC_RUN_STARTED_AT", run_started_at)
        save_ebay_setting(business_id, "EBAY_ACTIVE_LISTINGS_SYNC_CHECKPOINT_PAGE", "0")

    total_pages = page  # guarantees at least one fetch even when resuming past page 1
    total_rows = 0
    while page <= total_pages:
        resp = None
        for attempt in range(4):
            try:
                resp = _ebay_get_active_listings_page(token, page)
                break
            except Exception:
                if attempt == 3:
                    raise
                _time.sleep(2 ** attempt)
        ack = resp.get("Ack")
        if ack not in ("Success", "Warning"):
            raise Exception(f"eBay GetMyeBaySelling failed (Ack={ack}): {resp.get('Errors')}")

        active = resp.get("ActiveList") or {}
        item_array = active.get("ItemArray") or {}
        items = item_array.get("Item") or []
        if isinstance(items, dict):
            items = [items]
        pagination = active.get("PaginationResult") or {}
        total_pages = _safe_int(pagination.get("TotalNumberOfPages")) or 1

        now_iso = _dt.datetime.utcnow().isoformat()
        rows = []
        for it in items:
            title = it.get("Title")
            listing_details = it.get("ListingDetails") or {}
            selling_status = it.get("SellingStatus") or {}
            picture_details = it.get("PictureDetails") or {}
            gallery_url = picture_details.get("GalleryURL")
            # BUG FIXED (confirmed 0/4605 populated on first real sync): this
            # account publishes shipping via a Business Policy (SellerShippingProfile
            # / ShippingProfileID -- see push_listing_to_ebay_v2), NOT a manual
            # ShippingServiceOptions block. GetMyeBaySelling's ActiveList items
            # reflect that same policy-based structure back, so ShippingDetails.
            # ShippingServiceOptions.ShippingServiceCost is never populated for
            # business-policy listings -- it's the legacy manual-shipping-only
            # field. Real source: SellerProfiles.SellerShippingProfile.ShippingProfileID,
            # mapped to a dollar figure via the same EBAY_SHIPPING_POLICY_OPTIONS
            # list used when publishing. Kept the old field as a fallback in case
            # any older listing still carries manual (non-policy) shipping.
            seller_profiles = it.get("SellerProfiles") or {}
            ship_profile = seller_profiles.get("SellerShippingProfile") or {}
            profile_id = str(ship_profile.get("ShippingProfileID") or "").strip()
            shipping_cost = _EBAY_SHIP_POLICY_COST_BY_ID.get(profile_id)
            if shipping_cost is None:
                ship_details = it.get("ShippingDetails") or {}
                sso = ship_details.get("ShippingServiceOptions")
                if isinstance(sso, list):
                    sso = sso[0] if sso else {}
                shipping_cost = _safe_float((sso or {}).get("ShippingServiceCost"))
            rows.append({
                "business_id": business_id, "item_id": it.get("ItemID"),
                "sku": it.get("SKU"), "title": title, "norm_title": _shopify_sync_norm(title),
                "quantity": _safe_int(it.get("Quantity")), "quantity_available": _safe_int(it.get("QuantityAvailable")),
                "price": _safe_float(selling_status.get("CurrentPrice")),
                "start_time": listing_details.get("StartTime"),
                "listing_status": "Active", "end_time": listing_details.get("EndTime"),
                "gallery_url": gallery_url,
                "shipping_cost": shipping_cost,
                # WatchCount needs IncludeWatchCount=true on the request (added above) and
                # is only returned by eBay at all once it's >0 -- so a missing/null value
                # here means "0 or not yet known", not necessarily zero. HitCount (page
                # views) needs no special request flag, just DetailLevel=ReturnAll (already
                # set) and a hit-counter style that isn't HiddenStyle-for-non-seller.
                "watch_count": _safe_int(it.get("WatchCount")),
                "view_count": _safe_int(it.get("HitCount")),
                # BestOfferCount only appears in the response at all when >0 --
                # this is the cheap per-item signal used to decide which items
                # are worth an individual GetBestOffers call (see
                # _sync_ebay_best_offers_work): calling GetBestOffers with no
                # ItemID/BestOfferID to get everything account-wide in one shot
                # is documented as supported but returned an empty
                # ItemBestOffersArray in practice on this ~4,700-listing
                # account (Ack=Success, no error) -- so the reliable, actually-
                # tested path (per eBay's own step-by-step call validation) is
                # per-ItemID, and this count is what keeps that from meaning
                # one API call per active listing.
                "best_offer_count": _safe_int((it.get("BestOfferDetails") or {}).get("BestOfferCount")),
                "updated_at": now_iso,
            })
        if rows:
            supabase.table("ebay_listing_status").upsert(rows, on_conflict="business_id,item_id").execute()
        total_rows += len(rows)

        # Checkpoint AFTER this page lands, so an interruption mid-next-page still
        # resumes from here rather than re-doing this completed one.
        save_ebay_setting(business_id, "EBAY_ACTIVE_LISTINGS_SYNC_CHECKPOINT_PAGE", str(page))
        page += 1

    # Every page landed — drop any Active row this run never touched (sold, ended,
    # or relisted under a different item_id since the run started). EXCEPT rows
    # carrying a manual sku_override: that correction has no other home anywhere
    # in the system (see set_sku_override), so deleting the row here would
    # silently and permanently destroy it the moment a listing ends -- with zero
    # audit trail, unrecoverable. Confirmed this actually happened to a real
    # item. A listing that's no longer active still gets marked as such
    # everywhere else (it just won't show as "Active" in listing_status), but
    # the override itself is preserved for good.
    # Every page landed. A listing this run never touched again (sold, ended, or
    # relisted under a different item_id since the run started) is marked Ended
    # here -- NEVER deleted. This table is the only place sku_override lives, has
    # no audit trail, and the person building on top of this correctly pointed
    # out there's no such thing as "delete" in a real database when the data is
    # still meaningfully useful later (a since-ended listing's row is exactly
    # what a past order needs to resolve its SKU against). The active-listings
    # count elsewhere already filters on listing_status="Active" specifically,
    # so marking Ended (instead of leaving stale "Active") keeps that count
    # correct without erasing anything.
    supabase.table("ebay_listing_status").update({"listing_status": "Ended"})\
        .eq("business_id", business_id).eq("listing_status", "Active")\
        .lt("updated_at", run_started_at).execute()

    save_ebay_setting(business_id, "EBAY_ACTIVE_LISTINGS_SYNC_CHECKPOINT_PAGE", "")
    save_ebay_setting(business_id, "EBAY_ACTIVE_LISTINGS_SYNC_RUN_STARTED_AT", "")

    # Snapshotting runs AFTER the live sync is fully committed and checkpoints
    # are cleared, and is wrapped so it can never affect this function's
    # result or leave the checkpoint in a bad state. The live listings data
    # (what drives Shopify matching) has zero tolerance for this step's
    # failure; the history table has real tolerance for it -- if a snapshot
    # write fails, today's history point is simply missing, logged, and
    # nothing else about the sync is affected.
    try:
        _snapshot_active_listings(business_id)
    except Exception as e:
        print(f"_snapshot_active_listings failed for business {business_id} (live sync unaffected): {e}")

    return {"checked": total_rows, "synced_at": _dt.datetime.utcnow().isoformat()}

def _shopify_gql(business_id: str, query: str, variables: dict = None):
    import requests as _req
    token = get_shopify_access_token(business_id)
    settings = get_ebay_settings(business_id)
    domain = (settings.get("SHOPIFY_STORE_DOMAIN", "") or "").replace("https://", "").replace("http://", "").strip("/")
    r = _req.post(f"https://{domain}/admin/api/2024-10/graphql.json",
                  headers={"X-Shopify-Access-Token": token, "Content-Type": "application/json"},
                  json={"query": query, "variables": variables or {}}, timeout=30)
    out = r.json()
    if out.get("errors"):
        raise Exception(f"Shopify GraphQL error: {str(out['errors'])[:300]}")
    return out.get("data") or {}

def _ensure_shipping_profile(business_id: str, cost: int) -> str:
    """Returns the Shopify delivery-profile GID for the given flat cost,
    creating 'eBay $<cost> Shipping' (US flat rate, all locations) if it
    doesn't exist yet. Profiles are looked up by that exact name."""
    name = f"eBay ${cost} Shipping"
    data = _shopify_gql(business_id, """
      query { deliveryProfiles(first: 25) { edges { node { id name } } } }""")
    for edge in (data.get("deliveryProfiles") or {}).get("edges", []):
        if edge["node"]["name"] == name:
            return edge["node"]["id"]
    locs = _shopify_gql(business_id, """
      query { locations(first: 10, includeLegacy: true) { edges { node { id } } } }""")
    loc_ids = [e["node"]["id"] for e in (locs.get("locations") or {}).get("edges", [])]
    if not loc_ids:
        raise Exception("No Shopify locations found — cannot create shipping profile")
    created = _shopify_gql(business_id, """
      mutation createProfile($profile: DeliveryProfileInput!) {
        deliveryProfileCreate(profile: $profile) {
          profile { id }
          userErrors { field message }
        }
      }""", {"profile": {
        "name": name,
        "locationGroupsToCreate": [{
            "locations": loc_ids,
            "zonesToCreate": [{
                "name": "United States",
                "countries": [{"code": "US", "includeAllProvinces": True}],
                "methodDefinitionsToCreate": [{
                    "name": f"Flat Rate (${cost})", "active": True,
                    "rateDefinition": {"price": {"amount": float(cost), "currencyCode": "USD"}},
                }],
            }],
        }],
      }})
    res = (created.get("deliveryProfileCreate") or {})
    if res.get("userErrors"):
        raise Exception(f"deliveryProfileCreate: {res['userErrors']}")
    return res["profile"]["id"]

def apply_shopify_shipping_rule(business_id: str) -> dict:
    """Standing rule: eBay listing on the $69/$195 flat-shipping tier =>
    its matched Shopify product's variant goes into the matching
    'eBay $69/$195 Shipping' delivery profile so Shopify charges the same
    shipping. Runs automatically after every active-listings sync; idempotent
    via ebay_listing_status.shopify_ship_rule_applied (skips items already
    pushed at the same cost)."""
    els = _fetch_all_for_business(business_id, "ebay_listing_status",
                                  "item_id,listing_status,shipping_cost,shopify_ship_rule_applied")
    targets = {r["item_id"]: int(float(r["shipping_cost"]))
               for r in els
               if r.get("item_id") and r.get("listing_status") == "Active"
               and r.get("shipping_cost") is not None and float(r["shipping_cost"]) in (69.0, 195.0)
               and (r.get("shopify_ship_rule_applied") is None
                    or float(r["shopify_ship_rule_applied"]) != float(r["shipping_cost"]))}
    if not targets:
        return {"pushed": 0}
    matches = _fetch_all_for_business(business_id, "inventory_match", "ebay_id,shopify_id")
    shopify_inv = {r["id"]: r for r in _fetch_all_for_business(business_id, "shopify_inventory", "id,product_id")}
    variants_by_cost = {69: [], 195: []}
    items_by_cost = {69: [], 195: []}
    for m in matches:
        eid, sid = m.get("ebay_id"), m.get("shopify_id")
        if not eid or not sid or eid not in targets:
            continue
        cost = targets[eid]
        # shopify_id may be a variant id (title-match rows) or product id
        # (dual-publish rows) — profile association needs variant GIDs.
        if sid in shopify_inv:
            variants_by_cost[cost].append(f"gid://shopify/ProductVariant/{sid}")
            items_by_cost[cost].append(eid)
    pushed = 0
    for cost, variant_gids in variants_by_cost.items():
        if not variant_gids:
            continue
        profile_id = _ensure_shipping_profile(business_id, cost)
        for i in range(0, len(variant_gids), 100):
            chunk = variant_gids[i:i+100]
            chunk_items = items_by_cost[cost][i:i+100]
            data = _shopify_gql(business_id, """
              mutation assign($id: ID!, $profile: DeliveryProfileInput!) {
                deliveryProfileUpdate(id: $id, profile: $profile) {
                  userErrors { field message }
                }
              }""", {"id": profile_id, "profile": {"variantsToAssociate": chunk}})
            errs = ((data.get("deliveryProfileUpdate") or {}).get("userErrors")) or []
            if errs:
                print(f"apply_shopify_shipping_rule: userErrors on ${cost} chunk: {errs}")
                continue
            supabase.table("ebay_listing_status").update({"shopify_ship_rule_applied": cost})\
                .eq("business_id", business_id).in_("item_id", chunk_items).execute()
            pushed += len(chunk)
    return {"pushed": pushed}

def _snapshot_active_listings(business_id: str):
    """Appends today's active-listings state into active_listings_daily_snapshot
    (append-only, one row per item per calendar day), so inventory value can be
    reconstructed for any past date -- something ebay_listing_status alone can't
    do, since it only ever holds current state. Called once per successful sync
    (both the daily automatic worker and the manual button), tagged with today's
    date; upserts on (business_id, item_id, snapshot_date) so re-running the sync
    twice in one day never creates duplicate rows for that day."""
    rows = _fetch_all_for_business(business_id, "ebay_listing_status",
                                    "item_id,sku,title,quantity,quantity_available,price,listing_status,shipping_cost")
    active_rows = [r for r in rows if r.get("listing_status") == "Active"]
    if not active_rows:
        return
    import datetime as _dt2
    snapshot_date = _dt2.date.today().isoformat()
    snapshot_rows = [{
        "business_id": business_id, "item_id": r["item_id"], "sku": r.get("sku"),
        "title": r.get("title"), "quantity": r.get("quantity"),
        "quantity_available": r.get("quantity_available"), "price": r.get("price"),
        "shipping_cost": r.get("shipping_cost"),
        "snapshot_date": snapshot_date,
    } for r in active_rows]
    # Chunked -- a single request with thousands of rows risks hitting a payload
    # size limit; this table can hold years of daily history at full item detail.
    for i in range(0, len(snapshot_rows), 500):
        chunk = snapshot_rows[i:i + 500]
        supabase.table("active_listings_daily_snapshot").upsert(
            chunk, on_conflict="business_id,item_id,snapshot_date").execute()

async def inventory_value_sync_worker():
    """Keeps ebay_inventory/shopify_inventory fresh -- these are what
    _compute_inventory_snapshot_value actually reads from, which is what
    analytics_snapshot_worker's daily Inventory Snapshot Value history is built
    on. Found via direct DB check: both tables were stale for days (ebay_inventory
    last synced 7/25, shopify_inventory 7/27) because sync_inventory() had only
    ever been called from the manual '/api/inventory/sync-now' button -- the same
    bug shape as Active Listings before active_listings_sync_worker, just in a
    different pair of tables. This is the actual reason the daily inventory-value
    history had gone flat for 5 straight days despite the snapshot worker itself
    running correctly every day.
    Uses its own app_settings timestamp key (INVENTORY_VALUE_SYNC_LAST_RUN_AT)
    rather than the existing _sync_status dict -- that dict is already shared
    (ambiguously) between this feature's manual button and the unrelated
    order-resync endpoint, and didn't want to add a third consumer on top of
    that pre-existing collision."""
    import asyncio, datetime as _dt
    while True:
        try:
            res = supabase.table("app_settings").select("business_id").eq("key", "EBAY_REFRESH_TOKEN").execute()
            business_ids = list(set(r["business_id"] for r in (res.data or [])))
            for biz_id in business_ids:
                try:
                    settings = get_ebay_settings(biz_id)
                    last_run = settings.get("INVENTORY_VALUE_SYNC_LAST_RUN_AT")
                    needs_sync = True
                    if last_run:
                        age = _dt.datetime.utcnow() - _dt.datetime.fromisoformat(
                            last_run.replace("Z", "+00:00")).replace(tzinfo=None)
                        needs_sync = age >= _dt.timedelta(days=1)
                    if not needs_sync:
                        continue
                    result = await asyncio.to_thread(sync_inventory, biz_id)
                    save_ebay_setting(biz_id, "INVENTORY_VALUE_SYNC_LAST_RUN_AT", _dt.datetime.utcnow().isoformat())
                    print(f"inventory_value_sync_worker: business {biz_id} synced -> {result}")
                except Exception as e:
                    print(f"inventory_value_sync_worker: business {biz_id} failed: {e}")
        except Exception as e:
            print(f"inventory_value_sync_worker error: {e}")
        await asyncio.sleep(3600)  # check hourly; actual sync per business is gated to ~once/day above

_post_publish_sync_pending = {}  # business_id -> {"first": iso, "last": iso}

def _mark_needs_post_publish_sync(business_id: str):
    """Called right after a successful eBay publish. Doesn't sync anything itself —
    just timestamps that this business has fresh listings waiting, for
    post_publish_sync_worker to pick up once publishing activity settles."""
    import datetime as _dt
    now_iso = _dt.datetime.utcnow().isoformat()
    entry = _post_publish_sync_pending.get(business_id)
    if entry:
        entry["last"] = now_iso
    else:
        _post_publish_sync_pending[business_id] = {"first": now_iso, "last": now_iso}

async def post_publish_sync_worker():
    """Fires an active-listings sync shortly after a publish, instead of making
    new items wait for the once-daily cycle to pick up their shipping cost /
    inventory match. Debounced rather than immediate: a real Lister session
    publishes many items over a stretch of time, so firing once per publish
    would be wasteful and would just repeatedly interrupt itself. Fires once
    publishing goes quiet for QUIET_MINUTES, with a MAX_WAIT_MINUTES ceiling so
    a long continuous session still gets a sync at least that often instead of
    the timer resetting forever."""
    import asyncio, datetime as _dt
    QUIET_MINUTES = 20
    MAX_WAIT_MINUTES = 60
    while True:
        try:
            now = _dt.datetime.utcnow()
            for biz_id, entry in list(_post_publish_sync_pending.items()):
                if _ebay_active_listings_job_status.get(biz_id, {}).get("running"):
                    continue
                last = _dt.datetime.fromisoformat(entry["last"])
                first = _dt.datetime.fromisoformat(entry["first"])
                quiet_long_enough = (now - last) >= _dt.timedelta(minutes=QUIET_MINUTES)
                been_waiting_too_long = (now - first) >= _dt.timedelta(minutes=MAX_WAIT_MINUTES)
                if not (quiet_long_enough or been_waiting_too_long):
                    continue
                del _post_publish_sync_pending[biz_id]
                _ebay_active_listings_job_status[biz_id] = {
                    "running": True, "result": None,
                    "started_at": _dt.datetime.utcnow().isoformat(), "finished_at": None,
                }
                try:
                    result = await asyncio.to_thread(_sync_ebay_active_listings_work, biz_id)
                    _ebay_active_listings_job_status[biz_id] = {
                        "running": False, "result": result,
                        "started_at": _ebay_active_listings_job_status[biz_id]["started_at"],
                        "finished_at": _dt.datetime.utcnow().isoformat(),
                    }
                    print(f"post_publish_sync_worker: business {biz_id} synced after publish activity")
                except Exception as e:
                    _ebay_active_listings_job_status[biz_id] = {
                        "running": False, "result": {"error": str(e)},
                        "started_at": _ebay_active_listings_job_status[biz_id]["started_at"],
                        "finished_at": _dt.datetime.utcnow().isoformat(),
                    }
                    print(f"post_publish_sync_worker: business {biz_id} failed: {e}")
        except Exception as e:
            print(f"post_publish_sync_worker error: {e}")
        await asyncio.sleep(120)  # check every 2 min -- cheap, and keeps the 20-min debounce reasonably tight

async def active_listings_sync_worker():
    """Keeps the Lots page's Active Listings column fresh automatically. Previously
    this table (ebay_listing_status) only ever got refreshed by manually clicking
    'Sync Active Listings' -- which is exactly why a report showed it last synced
    days earlier despite being looked at daily. active_listings_value/count on the
    Lots page are computed LIVE from this table on every page load (see
    GET /api/acquisitions), so keeping this table itself fresh is the whole fix --
    no separate 'active listings' value needs updating anywhere else.
    Runs a check every hour, but only actually syncs a given business once its
    last sync is >= 1 day old, so it settles into roughly a daily cadence per
    business without needing a separate 24-hour-only scheduler. Skips a business
    if a sync (manual or automatic) is already running for it, so this can never
    race the manual 'Sync Active Listings' button."""
    import asyncio, datetime as _dt
    while True:
        try:
            res = supabase.table("app_settings").select("business_id").eq("key", "EBAY_REFRESH_TOKEN").execute()
            business_ids = list(set(r["business_id"] for r in (res.data or [])))
            for biz_id in business_ids:
                try:
                    if _ebay_active_listings_job_status.get(biz_id, {}).get("running"):
                        continue
                    last_res = supabase.table("ebay_listing_status").select("updated_at")\
                        .eq("business_id", biz_id).eq("listing_status", "Active")\
                        .order("updated_at", desc=True).limit(1).execute()
                    last_synced_at = (last_res.data or [{}])[0].get("updated_at")
                    needs_sync = True
                    if last_synced_at:
                        age = _dt.datetime.utcnow() - _dt.datetime.fromisoformat(
                            last_synced_at.replace("Z", "+00:00")).replace(tzinfo=None)
                        needs_sync = age >= _dt.timedelta(days=1)
                    if not needs_sync:
                        continue
                    _ebay_active_listings_job_status[biz_id] = {
                        "running": True, "result": None,
                        "started_at": _dt.datetime.utcnow().isoformat(), "finished_at": None,
                    }
                    result = await asyncio.to_thread(_sync_ebay_active_listings_work, biz_id)
                    _ebay_active_listings_job_status[biz_id] = {
                        "running": False, "result": result,
                        "started_at": _ebay_active_listings_job_status[biz_id]["started_at"],
                        "finished_at": _dt.datetime.utcnow().isoformat(),
                    }
                    print(f"active_listings_sync_worker: business {biz_id} synced automatically")
                except Exception as e:
                    print(f"active_listings_sync_worker: business {biz_id} failed: {e}")
        except Exception as e:
            print(f"active_listings_sync_worker error: {e}")
        await asyncio.sleep(3600)  # check hourly; actual sync per business is gated to ~once/day above

_ebay_active_listings_job_status = {}  # business_id -> {"running": bool, "result": dict|None, "started_at": iso, "finished_at": iso|None}

async def _run_ebay_active_listings_sync_background(business_id: str):
    import asyncio, datetime as _dt
    try:
        result = await asyncio.to_thread(_sync_ebay_active_listings_work, business_id)
        _ebay_active_listings_job_status[business_id] = {
            "running": False, "result": result,
            "started_at": _ebay_active_listings_job_status.get(business_id, {}).get("started_at"),
            "finished_at": _dt.datetime.utcnow().isoformat(),
        }
    except Exception as e:
        _ebay_active_listings_job_status[business_id] = {
            "running": False, "result": {"error": str(e)},
            "started_at": _ebay_active_listings_job_status.get(business_id, {}).get("started_at"),
            "finished_at": _dt.datetime.utcnow().isoformat(),
        }

ACTIVE_LISTINGS_SYNC_COOLDOWN_DAYS = 3  # ~twice a week — this is a full-account ActiveList pull, not a per-item lookup, so it has no business running more often than that regardless of how many times the button gets clicked

@app.post("/api/acquisitions/sync-active-listings")
async def acquisitions_sync_active_listings(request: Request, force: bool = False):
    """Kicks off the real eBay ActiveList pull as a background job — this is what
    populates the Lots page's Active Listings column. Counts get grouped by SKU
    prefix (before the first '-') in GET /api/acquisitions, the same lot-matching
    convention already used by /api/acquisitions/debug-sku. Hard-capped to once
    every ACTIVE_LISTINGS_SYNC_COOLDOWN_DAYS regardless of how often this is called —
    a full-account listings pull has no reason to run more than a couple times a
    week, and this makes that true structurally instead of just by convention."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    import asyncio, datetime as _dt
    if _ebay_active_listings_job_status.get(business_id, {}).get("running"):
        return {"started": False, "already_running": True}

    last_res = supabase.table("ebay_listing_status").select("updated_at")\
        .eq("business_id", business_id).eq("listing_status", "Active")\
        .order("updated_at", desc=True).limit(1).execute()
    last_synced_at = (last_res.data or [{}])[0].get("updated_at")
    if last_synced_at and not force:
        age = _dt.datetime.utcnow() - _dt.datetime.fromisoformat(last_synced_at.replace("Z", "+00:00")).replace(tzinfo=None)
        if age < _dt.timedelta(days=ACTIVE_LISTINGS_SYNC_COOLDOWN_DAYS):
            return {"started": False, "reason": "cooldown", "last_synced_at": last_synced_at,
                    "next_allowed_at": (_dt.datetime.fromisoformat(last_synced_at.replace("Z", "+00:00")).replace(tzinfo=None)
                                         + _dt.timedelta(days=ACTIVE_LISTINGS_SYNC_COOLDOWN_DAYS)).isoformat()}

    _ebay_active_listings_job_status[business_id] = {
        "running": True, "result": None,
        "started_at": _dt.datetime.utcnow().isoformat(), "finished_at": None,
    }
    asyncio.create_task(_run_ebay_active_listings_sync_background(business_id))
    return {"started": True}

@app.get("/api/acquisitions/sync-active-listings-status")
async def acquisitions_sync_active_listings_status(request: Request, response: Response):
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    response.headers["Cache-Control"] = "no-store"
    return _ebay_active_listings_job_status.get(business_id, {"running": False, "result": None, "started_at": None, "finished_at": None})

@app.get("/api/acquisitions/debug-active-listing-raw")
async def debug_active_listing_raw(request: Request):
    """Temporary read-only diagnostic: one raw page (3 items) of eBay's real
    ActiveList response, unmodified — so the actual price field name/shape can be
    confirmed before building a dollar-total column on top of a guess."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    token = get_ebay_access_token(business_id)
    resp = _ebay_get_active_listings_page(token, 1, entries_per_page=3)
    return resp

_ebay_listings_job_status = {}  # business_id -> {"running": bool, "result": dict|None, "started_at": iso, "finished_at": iso|None}

async def _run_ebay_listings_sync_background(business_id: str):
    import asyncio, datetime as _dt
    try:
        result = await asyncio.to_thread(_sync_ebay_unsold_listings_work, business_id)
        _ebay_listings_job_status[business_id] = {
            "running": False, "result": result,
            "started_at": _ebay_listings_job_status.get(business_id, {}).get("started_at"),
            "finished_at": _dt.datetime.utcnow().isoformat(),
        }
    except Exception as e:
        _ebay_listings_job_status[business_id] = {
            "running": False, "result": {"error": str(e)},
            "started_at": _ebay_listings_job_status.get(business_id, {}).get("started_at"),
            "finished_at": _dt.datetime.utcnow().isoformat(),
        }

@app.post("/api/shopify-sync/sync-ebay-listings")
async def shopify_sync_sync_ebay_listings(request: Request):
    """Kicks off the real eBay Unsold-listings pull as a background job — this is
    what actually populates the Inactive tab. No date range, no Shopify calls here at
    all; just eBay's own current inactive-listings bucket."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    import asyncio, datetime as _dt
    if _ebay_listings_job_status.get(business_id, {}).get("running"):
        return {"started": False, "already_running": True}
    _ebay_listings_job_status[business_id] = {
        "running": True, "result": None,
        "started_at": _dt.datetime.utcnow().isoformat(), "finished_at": None,
    }
    asyncio.create_task(_run_ebay_listings_sync_background(business_id))
    return {"started": True}

@app.get("/api/shopify-sync/sync-ebay-listings-status")
async def shopify_sync_sync_ebay_listings_status(request: Request, response: Response):
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    response.headers["Cache-Control"] = "no-store"
    return _ebay_listings_job_status.get(business_id, {"running": False, "result": None, "started_at": None, "finished_at": None})

@app.get("/api/shopify-sync/inactive-items")
async def shopify_sync_inactive_items(request: Request, response: Response, limit: int = 15):
    """Pure local read of the last eBay Unsold-listings sync — zero external calls,
    so this is instant no matter how many total inactive listings there are. Ordered
    by end_time desc (most recently ended first) and capped at `limit` so the page
    never tries to render/act on all ~260 at once; Shopify matching is a separate,
    explicit step (see /inactive-check) scoped to exactly what this returns."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    response.headers["Cache-Control"] = "no-store"
    limit = max(1, min(limit, 500))

    res = supabase.table("ebay_listing_status").select("*")\
        .eq("business_id", business_id).eq("listing_status", "Unsold").execute()
    rows = res.data or []
    total = len(rows)
    rows.sort(key=lambda r: r.get("end_time") or "", reverse=True)
    rows = rows[:limit]

    items = [{
        "item_id": r.get("item_id"), "sku": r.get("sku") or "", "title": r.get("title") or "",
        "end_time": r.get("end_time"),
        "ebay_live_qty": max((r.get("quantity") or 0) - (r.get("quantity_sold") or 0), 0),
        "shopify_found": False, "shopify_live_qty": None, "shopify_title": None, "shopify_inventory_item_id": None,
    } for r in rows]

    last_synced = max((r.get("updated_at") for r in (res.data or []) if r.get("updated_at")), default=None)
    return {"items": items, "total": total, "last_synced_at": last_synced}

_inactive_check_job_status = {}  # business_id -> {"running": bool, "result": dict|None, "started_at": iso, "finished_at": iso|None}

def _inactive_check_work(business_id: str, items: list) -> dict:
    """Matches exactly the N inactive items the page currently has on screen against
    a fresh Shopify catalog scan — the same _fetch_all_shopify_products already used
    by the Active tab. Deliberately not persisted anywhere: this is a live check on
    demand, same explicit-action contract as Sync Now, just without a snapshot table
    on this side."""
    settings = get_ebay_settings(business_id)
    domain = (settings.get("SHOPIFY_STORE_DOMAIN", "") or "").strip().replace("https://", "").replace("http://", "").strip("/")
    if not domain:
        return {"error": "Shopify not connected"}
    shopify_token = get_shopify_access_token(business_id)
    if not shopify_token:
        return {"error": "Shopify not connected"}
    loc_headers = {"X-Shopify-Access-Token": shopify_token, "Content-Type": "application/json"}
    location_id = _get_shopify_primary_location_id(domain, loc_headers)
    if not location_id:
        return {"error": "Could not determine Shopify location"}

    shopify_catalog, catalog_complete, catalog_pages = _fetch_all_shopify_products(domain, shopify_token, location_id)
    if not catalog_complete:
        return {"error": f"Shopify catalog scan was rate-limited and only got through "
                          f"{catalog_pages} page(s) ({len(shopify_catalog)} products) before giving up. "
                          f"Try again in a minute."}

    results = []
    for it in items:
        norm_title = _shopify_sync_norm(it.get("title") or "")
        match = shopify_catalog.get(norm_title)
        results.append({
            "item_id": it.get("item_id"), "shopify_found": match is not None,
            "shopify_live_qty": match["qty"] if match else None,
            "shopify_title": match["title"] if match else None,
            "shopify_inventory_item_id": match["inventory_item_id"] if match else None,
        })
    return {"checked": len(results), "results": results}

async def _run_inactive_check_background(business_id: str, items: list):
    import asyncio, datetime as _dt
    try:
        result = await asyncio.to_thread(_inactive_check_work, business_id, items)
        _inactive_check_job_status[business_id] = {
            "running": False, "result": result,
            "started_at": _inactive_check_job_status.get(business_id, {}).get("started_at"),
            "finished_at": _dt.datetime.utcnow().isoformat(),
        }
    except Exception as e:
        _inactive_check_job_status[business_id] = {
            "running": False, "result": {"error": str(e)},
            "started_at": _inactive_check_job_status.get(business_id, {}).get("started_at"),
            "finished_at": _dt.datetime.utcnow().isoformat(),
        }

@app.post("/api/shopify-sync/inactive-check")
async def shopify_sync_inactive_check(request: Request, items: List[dict] = Body(...)):
    """Kicks off a Shopify catalog match for exactly the inactive items currently on
    screen (title/sku/item_id) as a background job — the catalog scan itself is the
    same operation that needs background treatment on the Active tab, regardless of
    how few items are being matched against it."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    import asyncio, datetime as _dt
    if _inactive_check_job_status.get(business_id, {}).get("running"):
        return {"started": False, "already_running": True}
    _inactive_check_job_status[business_id] = {
        "running": True, "result": None,
        "started_at": _dt.datetime.utcnow().isoformat(), "finished_at": None,
    }
    asyncio.create_task(_run_inactive_check_background(business_id, items))
    return {"started": True}

@app.get("/api/shopify-sync/inactive-check-status")
async def shopify_sync_inactive_check_status(request: Request, response: Response):
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    response.headers["Cache-Control"] = "no-store"
    return _inactive_check_job_status.get(business_id, {"running": False, "result": None, "started_at": None, "finished_at": None})

@app.get("/api/shopify-sync/debug-catalog-search")
async def shopify_sync_debug_catalog_search(request: Request, q: str):
    """Temporary read-only diagnostic: pulls the full live Shopify catalog (same
    paginated scan Sync Now uses) and returns every product whose title contains q
    (case-insensitive substring), with both the raw and normalized title, so a
    near-miss title mismatch can be seen and compared directly against eBay's title."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    settings = get_ebay_settings(business_id)
    domain = (settings.get("SHOPIFY_STORE_DOMAIN", "") or "").strip().replace("https://", "").replace("http://", "").strip("/")
    shopify_token = get_shopify_access_token(business_id) if domain else None
    if not domain or not shopify_token:
        raise HTTPException(400, "Shopify not connected")

    loc_headers = {"X-Shopify-Access-Token": shopify_token, "Content-Type": "application/json"}
    location_id = _get_shopify_primary_location_id(domain, loc_headers)
    if not location_id:
        raise HTTPException(400, "Could not determine Shopify location")
    catalog, complete, pages = _fetch_all_shopify_products(domain, shopify_token, location_id)
    q_lower = q.lower()
    matches = [{"title": v["title"], "norm_title": k, "qty": v["qty"]}
               for k, v in catalog.items() if q_lower in k]
    return {"catalog_size": len(catalog), "complete": complete, "pages_fetched": pages, "matches": matches}

@app.get("/api/shopify-sync/debug-exact-match-by-order")
async def shopify_sync_debug_exact_match_by_order(request: Request, order_id: str):
    """Same as debug-exact-match, but pulls the eBay title straight from the orders
    table by order_id (the actual unique key) instead of accepting it as a URL param
    — a hand-retyped title would silently 'fix' whatever invisible character is
    actually the bug, testing nothing. NOT sku: sku is a lot/location code here, not
    a unique product ID (confirmed elsewhere in this file) — using it to look up a
    specific order pulled a completely unrelated listing that happened to share it."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    order_res = supabase.table("orders").select("title").eq("business_id", business_id)\
        .eq("platform", "eBay").eq("order_id", order_id).limit(1).execute()
    if not order_res.data:
        raise HTTPException(404, f"No eBay order found with order_id={order_id}")
    return await shopify_sync_debug_exact_match(request, ebay_title=order_res.data[0]["title"])

@app.get("/api/shopify-sync/debug-exact-match")
async def shopify_sync_debug_exact_match(request: Request, ebay_title: str):
    """Temporary read-only diagnostic: runs the EXACT dict-key lookup
    _shopify_sync_check_one uses (not the substring search debug-catalog-search
    does — those are two different tests, and substring matching was giving false
    confidence that titles were matching when the exact lookup was still failing).
    If it doesn't match, finds the closest substring candidate and returns both
    normalized strings as codepoint lists so an invisible character (NBSP, zero-
    width space, curly vs straight punctuation) can't hide in a text diff."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    settings = get_ebay_settings(business_id)
    domain = (settings.get("SHOPIFY_STORE_DOMAIN", "") or "").strip().replace("https://", "").replace("http://", "").strip("/")
    shopify_token = get_shopify_access_token(business_id) if domain else None
    if not domain or not shopify_token:
        raise HTTPException(400, "Shopify not connected")

    loc_headers = {"X-Shopify-Access-Token": shopify_token, "Content-Type": "application/json"}
    location_id = _get_shopify_primary_location_id(domain, loc_headers)
    if not location_id:
        raise HTTPException(400, "Could not determine Shopify location")
    catalog, complete, pages = _fetch_all_shopify_products(domain, shopify_token, location_id)

    ebay_norm = _shopify_sync_norm(ebay_title)
    exact_match = catalog.get(ebay_norm)

    def _codepoints(s):
        return [f"U+{ord(c):04X}({c!r})" for c in s]

    result = {
        "catalog_complete": complete, "catalog_pages": pages, "catalog_size": len(catalog),
        "ebay_title": ebay_title, "ebay_norm_title": ebay_norm,
        "exact_match_found": exact_match is not None,
    }
    if exact_match:
        result["matched_shopify_title"] = exact_match["title"]
        result["matched_qty"] = exact_match["qty"]
    else:
        # Substring search among catalog keys for anything that looks related, to
        # find the near-miss candidate even though the exact key lookup failed.
        words = [w for w in ebay_norm.split() if len(w) > 3]
        candidates = [k for k in catalog if any(w in k for w in words)]
        result["near_miss_candidates"] = [{
            "shopify_norm_title": k,
            "same_length": len(k) == len(ebay_norm),
            "ebay_codepoints": _codepoints(ebay_norm),
            "shopify_codepoints": _codepoints(k),
        } for k in candidates[:3]]
    return result

@app.get("/api/shopify-sync/debug-ebay-lookup")
async def shopify_sync_debug_ebay_lookup(request: Request, order_id: str):
    """Temporary read-only diagnostic: runs the exact SKU + Trading API GetItem
    lookups _shopify_sync_check_one does for one order, and returns the raw
    response instead of silently swallowing failures."""
    import requests as _req
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")

    order_res = supabase.table("orders").select("sku,title,legacy_item_id").eq("business_id", business_id)\
        .eq("platform", "eBay").eq("order_id", order_id).limit(1).execute()
    if not order_res.data:
        raise HTTPException(404, f"No eBay order found with order_id={order_id}")
    row = order_res.data[0]
    sku = row.get("sku") or ""
    legacy_item_id = row.get("legacy_item_id") or ""

    try:
        token = get_ebay_access_token(business_id)
        token_ok = True
        token_error = None
    except Exception as e:
        token = None
        token_ok = False
        token_error = str(e)

    result = {"sku": sku, "legacy_item_id": legacy_item_id, "token_acquired": token_ok, "token_error": token_error}

    if token_ok and sku and sku != "(no SKU)" and not sku.lower().startswith("lister-"):
        try:
            r = _req.get(f"{EBAY_API_BASE}/sell/inventory/v1/inventory_item/{sku}",
                         headers=ebay_headers(token, content_language=False), timeout=15)
            result["sku_lookup"] = {"status": r.status_code, "body": r.text[:1000]}
        except Exception as e:
            result["sku_lookup"] = {"exception": str(e)}
    else:
        result["sku_lookup"] = {"skipped": "no usable sku"}

    if token_ok and legacy_item_id:
        try:
            result["get_item_lookup"] = _ebay_get_item_status(token, legacy_item_id)
        except Exception as e:
            result["get_item_lookup"] = {"exception": str(e)}
    else:
        result["get_item_lookup"] = {"skipped": "no legacy_item_id"}

    return result

@app.get("/api/shopify-sync/debug-summary")
async def shopify_sync_debug_summary(request: Request):
    """Temporary read-only diagnostic: how much of the snapshot table is actually
    populated right now, broken down by whether eBay's side was ever successfully
    checked (vs still null from before the incremental-sync fix, or from items that
    hit the rate limit before it landed)."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")

    snap_res = supabase.table("shopify_sync_snapshot").select(
        "norm_title,title,sku,ebay_live_qty,ebay_ended,shopify_found,shopify_live_qty,updated_at"
    ).eq("business_id", business_id).execute()
    rows = snap_res.data or []

    ebay_known = [r for r in rows if r.get("ebay_live_qty") is not None]
    ebay_unknown = [r for r in rows if r.get("ebay_live_qty") is None]
    shopify_found = [r for r in rows if r.get("shopify_found")]

    return {
        "total_snapshot_rows": len(rows),
        "ebay_qty_known": len(ebay_known),
        "ebay_qty_still_unknown": len(ebay_unknown),
        "shopify_found_count": len(shopify_found),
        "sample_still_unknown": [
            {"title": r.get("title"), "sku": r.get("sku"), "updated_at": r.get("updated_at")}
            for r in ebay_unknown[:15]
        ],
    }

@app.get("/api/shopify-sync/debug-item")
async def shopify_sync_debug_item(request: Request, q: str):
    """Temporary read-only diagnostic: dump every push_log row, the current snapshot
    row, and the last 90 days of eBay order rows matching a title OR sku substring, so
    we can trace where a specific quantity came from even when q is a SKU, not a title
    fragment."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    import datetime as _dt
    cutoff = (_dt.datetime.utcnow() - _dt.timedelta(days=90)).strftime("%Y-%m-%d")
    push_rows = supabase.table("shopify_push_log").select("*").eq("business_id", business_id)\
        .or_(f"title.ilike.%{q}%,sku.ilike.%{q}%").execute().data or []
    snap_rows = supabase.table("shopify_sync_snapshot").select("*").eq("business_id", business_id)\
        .or_(f"title.ilike.%{q}%,sku.ilike.%{q}%").execute().data or []
    order_rows = supabase.table("orders").select("order_id,order_date,sku,title,quantity,platform").eq("business_id", business_id)\
        .eq("platform", "eBay").or_(f"title.ilike.%{q}%,sku.ilike.%{q}%").gte("order_date", cutoff).execute().data or []
    return {"push_log": push_rows, "snapshot": snap_rows, "recent_orders": order_rows}

@app.post("/api/shopify-sync/ignore")
async def shopify_sync_ignore(request: Request, body: dict = Body(...)):
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    title = (body.get("title") or "").strip()
    if not title:
        raise HTTPException(400, "title required")
    supabase.table("shopify_sync_ignored").upsert(
        {"business_id": business_id, "norm_title": _shopify_sync_norm(title), "title": title},
        on_conflict="business_id,norm_title",
    ).execute()
    return {"ok": True}

@app.post("/api/shopify-sync/unignore")
async def shopify_sync_unignore(request: Request, body: dict = Body(...)):
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    title = (body.get("title") or "").strip()
    if not title:
        raise HTTPException(400, "title required")
    supabase.table("shopify_sync_ignored").delete()\
        .eq("business_id", business_id).eq("norm_title", _shopify_sync_norm(title)).execute()
    return {"ok": True}

@app.get("/api/shopify-sync/sync-status")
async def shopify_sync_status(request: Request, response: Response):
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    response.headers["Cache-Control"] = "no-store"
    res = supabase.table("shopify_sync_snapshot").select("updated_at").eq("business_id", business_id)\
        .order("updated_at", desc=True).limit(1).execute()
    last = res.data[0]["updated_at"] if res.data else None
    return {"last_synced_at": last}

@app.get("/api/shopify-sync/today")
async def shopify_sync_today(request: Request, response: Response, date: str = None, start: str = None, end: str = None):
    """Step 1: every eBay sale in the selected date range, matched against the last
    synced snapshot of live eBay/Shopify quantities (see /refresh above). Reads only
    from local tables — zero external API calls, so this is fast no matter how wide
    the date range is. Click 'Sync Now' to refresh the snapshot itself. Shopify is
    matched by TITLE — SKU here is just a lot/location code, not a unique product ID,
    so it's useless for cross-platform matching (confirmed: it was matching completely
    unrelated items that happened to share a storage location).

    start/end are the caller's local calendar dates (YYYY-MM-DD) — server UTC time was
    used before, which drifts a day off from the user's "today" depending on time of
    day and timezone. 'date' is accepted as a single-day alias for start=end=date, for
    older callers. Falls back to server UTC "today" for both if nothing is given."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    response.headers["Cache-Control"] = "no-store"
    import datetime as _dt, re as _re

    _date_re = r"^\d{4}-\d{2}-\d{2}$"
    server_today = _dt.datetime.utcnow().strftime("%Y-%m-%d")
    if date and _re.match(_date_re, date):
        start = end = date
    start_date = start if start and _re.match(_date_re, start) else server_today
    end_date = end if end and _re.match(_date_re, end) else server_today

    sold_by_title = _shopify_sync_sold_by_title(business_id, start_date, end_date)
    if not sold_by_title:
        return {"start_date": start_date, "end_date": end_date, "items": []}

    # Fetch every snapshot/ignored row for this business rather than filtering with
    # .in_("norm_title", [...]) — a title list built from real part titles can contain
    # commas, parentheses, or other characters that break PostgREST's in.() filter
    # syntax, silently dropping rows from the result with no error raised (confirmed:
    # a verified-correct snapshot row disappeared from a same-batch .in_() query while
    # being fetchable individually). Matching locally in Python sidesteps that
    # entirely, the same fix already applied to Shopify's own search.
    snap_res = supabase.table("shopify_sync_snapshot").select("*").eq("business_id", business_id).execute()
    snapshots = {row["norm_title"]: row for row in (snap_res.data or [])}

    ignored_res = supabase.table("shopify_sync_ignored").select("norm_title").eq("business_id", business_id).execute()
    ignored_titles = {row["norm_title"] for row in (ignored_res.data or [])}

    items = []
    for norm_title, entry in sold_by_title.items():
        snap = snapshots.get(norm_title) or {}
        items.append({
            "sku": entry["sku"], "title": entry["title"], "legacy_item_id": entry["legacy_item_id"],
            "qty_sold_today": entry["qty_sold_today"], "order_ids": entry["order_ids"],
            "ebay_live_qty": snap.get("ebay_live_qty"),
            "shopify_found": bool(snap.get("shopify_found")), "shopify_live_qty": snap.get("shopify_live_qty"),
            "shopify_title": snap.get("shopify_title"), "shopify_inventory_item_id": snap.get("shopify_inventory_item_id"),
            "snapshot_updated_at": snap.get("updated_at"),
            "ignored": norm_title in ignored_titles,
        })

    # Real server-side "already pushed" check against the push audit table —
    # most recent successful push per order_id wins, since a row can be re-pushed.
    try:
        log_res = supabase.table("shopify_push_log").select("order_id,new_quantity,created_at")\
            .eq("business_id", business_id).eq("status", "success").order("created_at").execute()
        pushed_by_order_id = {}
        for row in (log_res.data or []):
            for oid in (row.get("order_id") or "").split(","):
                oid = oid.strip()
                if oid:
                    pushed_by_order_id[oid] = row.get("new_quantity")
        for it in items:
            match = next((pushed_by_order_id[oid] for oid in it["order_ids"] if oid in pushed_by_order_id), None)
            it["already_pushed"] = match is not None
            it["qty_matches"] = match is not None and match == it.get("shopify_live_qty")
    except Exception as e:
        print(f"shopify_push_log read failed: {e}")
        for it in items:
            it["already_pushed"] = False
            it["qty_matches"] = False

    return {"start_date": start_date, "end_date": end_date, "items": items}

def _get_shopify_primary_location_id(domain: str, headers: dict) -> str:
    import requests as _req
    r = _req.post(f"https://{domain}/admin/api/2024-10/graphql.json", headers=headers,
                   json={"query": "{ locations(first: 1) { edges { node { id } } } }"}, timeout=15)
    edges = (r.json().get("data", {}) or {}).get("locations", {}).get("edges", []) if r.status_code == 200 else []
    return edges[0]["node"]["id"] if edges else None

def _push_shopify_qty_updates(business_id: str, items: list) -> dict:
    """Takes rows carrying a Shopify inventoryItem id (from a live check) and SETS
    Shopify's quantity to match eBay's live quantity exactly — not a relative
    decrement. A decrement compounds drift whenever the two platforms are already
    out of sync (confirmed: found a listing at -1 in Shopify that no push ever
    touched — a relative delta would have made that worse, not fixed it). Setting
    to eBay's live number is self-correcting instead. Logs every attempt for audit/undo."""
    import requests as _req, datetime as _dt

    settings = get_ebay_settings(business_id)
    domain = (settings.get("SHOPIFY_STORE_DOMAIN", "") or "").strip().replace("https://", "").replace("http://", "").strip("/")
    if not domain:
        raise Exception("Shopify not connected")
    token = get_shopify_access_token(business_id)
    headers = {"X-Shopify-Access-Token": token, "Content-Type": "application/json"}
    location_id = _get_shopify_primary_location_id(domain, headers)
    if not location_id:
        raise Exception("Could not determine Shopify location")

    results = []
    for it in items:
        inv_item_id = it.get("shopify_inventory_item_id")
        target_qty = it.get("ebay_live_qty")
        sku = it.get("sku", "")
        order_ids = it.get("order_ids", [])
        log_row = {
            "business_id": business_id, "order_id": ",".join(order_ids) if order_ids else "",
            "sku": sku, "title": it.get("title"), "quantity_deducted": it.get("qty_sold_today", 0),
        }
        if not inv_item_id:
            log_row["status"] = "no_shopify_match"
            results.append({"title": it.get("title"), "status": "no_shopify_match"})
        elif target_qty is None:
            log_row["status"] = "no_ebay_qty"
            results.append({"title": it.get("title"), "status": "no_ebay_qty", "error": "eBay quantity unknown — click Sync Now first"})
        else:
            try:
                mutation = {
                    "query": """mutation setQty($input: InventorySetQuantitiesInput!) {
                        inventorySetQuantities(input: $input) {
                            inventoryAdjustmentGroup { changes { name delta quantityAfterChange } }
                            userErrors { field message }
                        }
                    }""",
                    "variables": {"input": {
                        "reason": "correction",
                        "name": "available",
                        "ignoreCompareQuantity": True,
                        "quantities": [{"inventoryItemId": inv_item_id, "locationId": location_id, "quantity": int(target_qty)}],
                    }},
                }
                r = _req.post(f"https://{domain}/admin/api/2024-10/graphql.json", headers=headers, json=mutation, timeout=20)
                resp = r.json() if r.status_code == 200 else {}
                gql_errors = resp.get("errors") or []
                errors = (resp.get("data", {}) or {}).get("inventorySetQuantities", {}).get("userErrors", []) if not gql_errors else gql_errors
                if r.status_code == 200 and not gql_errors and not errors:
                    new_qty = target_qty
                    log_row["status"] = "success"
                    log_row["new_quantity"] = new_qty
                    results.append({"title": it.get("title"), "status": "success", "new_quantity": new_qty})
                    try:
                        supabase.table("shopify_sync_snapshot")\
                            .update({"shopify_live_qty": new_qty, "updated_at": _dt.datetime.utcnow().isoformat()})\
                            .eq("business_id", business_id).eq("norm_title", _shopify_sync_norm(it.get("title"))).execute()
                    except Exception as e:
                        print(f"shopify_sync_snapshot update after push failed: {e}")
                else:
                    err_msg = "; ".join(e.get("message", "") for e in errors) or f"HTTP {r.status_code}"
                    log_row["status"] = "error"
                    log_row["error_message"] = err_msg
                    results.append({"title": it.get("title"), "status": "error", "error": err_msg})
            except Exception as e:
                log_row["status"] = "error"
                log_row["error_message"] = str(e)
                results.append({"title": it.get("title"), "status": "error", "error": str(e)})

        try:
            supabase.table("shopify_push_log").insert(log_row).execute()
        except Exception as e:
            print(f"shopify_push_log insert failed: {e}")

    return {"ok": True, "results": results}

@app.post("/api/shopify-sync/push")
async def shopify_sync_push(request: Request, items: List[dict] = Body(...)):
    """Step 2: takes the exact rows the user selected from Step 1 — thin wrapper
    around _push_shopify_qty_updates, which is also used by the hourly automated
    sync (see shopify_sync_worker)."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    try:
        return _push_shopify_qty_updates(business_id, items)
    except Exception as e:
        raise HTTPException(400, str(e))

class ManualMatch(BaseModel):
    ebay_row_id: str    # inventory_match.id of the eBay-only row
    shopify_row_id: str  # inventory_match.id of the Shopify-only row

@app.post("/api/inventory/manual-match")
async def manual_match_inventory(body: ManualMatch, request: Request):
    """Manually links an eBay-only row to a Shopify-only row when they're really
    the same item but title-matching missed it (e.g. eBay truncated the title).
    Merges both into a single permanently confirmed row — same guarantee as an
    automatic match, just human-confirmed instead of title-derived."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    import uuid as _uuid

    ebay_row = (supabase.table("inventory_match").select("*").eq("id", body.ebay_row_id)
                .eq("business_id", business_id).limit(1).execute().data or [None])[0]
    shopify_row = (supabase.table("inventory_match").select("*").eq("id", body.shopify_row_id)
                   .eq("business_id", business_id).limit(1).execute().data or [None])[0]
    if not ebay_row or not shopify_row:
        raise HTTPException(404, "One or both rows not found")
    if not ebay_row.get("ebay_id") or not shopify_row.get("shopify_id"):
        raise HTTPException(400, "First row must be eBay-only, second must be Shopify-only")

    # shopify_row is allowed to already carry an old ebay_id -- that's the whole
    # point of the end-and-relist scenario this also serves (the Shopify side's
    # real pairing lives on an eBay listing that's since ended, and the new
    # relisted item needs to inherit it). What's NOT allowed is stealing a
    # pairing whose eBay side is still genuinely active -- that would silently
    # orphan a live listing. Checked against ebay_listing_status directly
    # (the real current-state source), not assumed from the row's own data.
    old_ebay_id = shopify_row.get("ebay_id")
    if old_ebay_id and old_ebay_id != ebay_row["ebay_id"]:
        still_active = (supabase.table("ebay_listing_status").select("item_id")
                         .eq("business_id", business_id).eq("item_id", old_ebay_id)
                         .eq("listing_status", "Active").limit(1).execute().data)
        if still_active:
            raise HTTPException(409, f"This Shopify item is still actively matched to eBay listing "
                                      f"{old_ebay_id}, which is currently Active -- can't reassign it "
                                      f"without un-matching that listing first.")

    new_hd_id = str(_uuid.uuid4())
    supabase.table("inventory_match").insert({
        "business_id": business_id, "title": ebay_row.get("title") or shopify_row.get("title"),
        "ebay_id": ebay_row["ebay_id"], "shopify_id": shopify_row["shopify_id"],
        "hd_id": new_hd_id, "matched_by": "manual",
    }).execute()
    # Remove the two now-superseded one-sided rows.
    supabase.table("inventory_match").delete().eq("id", body.ebay_row_id).execute()
    supabase.table("inventory_match").delete().eq("id", body.shopify_row_id).execute()
    return {"ok": True, "hd_id": new_hd_id}

class InventoryQuantityPush(BaseModel):
    quantity: int

@app.post("/api/inventory/{row_id}/push-quantity")
async def push_inventory_quantity(row_id: str, body: InventoryQuantityPush, request: Request):
    """Pushes ONE quantity value to BOTH eBay and Shopify for a confirmed matched
    pair, live, right now. This is a genuinely sensitive write path -- the only
    place in this app that writes a value onto a live eBay listing outside of
    initial publish -- so it's deliberately narrow by construction, not just by
    convention:
      - Requires a confirmed match on BOTH platforms (real ebay_id + shopify_id
        via inventory_match). Refuses an eBay-only or Shopify-only row outright.
      - Verifies the eBay listing is live-Active via a fresh GetItem call right
        before writing -- not cached/stale data. Refuses if not Active.
      - The eBay write itself (_ebay_revise_quantity_only) sends ONLY ItemID +
        Quantity -- structurally incapable of touching price, title, category,
        or anything else, the same proven pattern as the SKU-only revise above.
      - eBay's Quantity is a LIFETIME total, not "available now" -- it computes
        available as Quantity - QuantitySold itself. This reads the live
        QuantitySold right before writing and adds it to the requested
        available quantity, so the number that ends up showing as available
        actually matches what was asked for.
      - Fails closed: if the eBay write fails, Shopify is never touched at all.
        If eBay succeeds but Shopify then fails, that's reported clearly as a
        partial failure -- never silently swallowed.
      - Every attempt (success, failure, or partial) gets its own row in
        ebay_quantity_push_log, independent of the ordinary sync logs.
    """
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    if body.quantity < 0:
        raise HTTPException(400, "Quantity cannot be negative")

    row_res = supabase.table("inventory_match").select("*").eq("id", row_id).eq("business_id", business_id).execute()
    if not row_res.data:
        raise HTTPException(404, "Row not found")
    row = row_res.data[0]
    ebay_item_id = row.get("ebay_id")
    shopify_variant_id = row.get("shopify_id")
    title = row.get("title") or ""
    if not ebay_item_id or not shopify_variant_id:
        raise HTTPException(400, "This row must be a confirmed match on both eBay and Shopify — "
                                  "one-sided rows can't use this endpoint.")
    if not str(ebay_item_id).isdigit():
        # Guards against a known data anomaly found earlier this session: a
        # duplicate inventory_match row where ebay_id was literally the
        # string "lister-XXX" instead of a real eBay item ID.
        raise HTTPException(400, f"ebay_id on this row ('{ebay_item_id}') doesn't look like a real "
                                  f"eBay item ID — refusing to write to it.")

    log_row = {
        "business_id": business_id, "ebay_item_id": str(ebay_item_id), "shopify_variant_id": str(shopify_variant_id),
        "title": title, "requested_available_qty": body.quantity,
    }

    # Fresh live check, not cached -- must be genuinely Active right now.
    ebay_token = get_ebay_access_token(business_id)
    try:
        status_resp = _ebay_get_item_status(ebay_token, ebay_item_id)
    except Exception as e:
        log_row.update({"ebay_status": "error", "ebay_error": f"GetItem failed: {e}"})
        supabase.table("ebay_quantity_push_log").insert(log_row).execute()
        raise HTTPException(502, f"Could not check eBay's live listing status: {e}")
    item = status_resp.get("Item") or {}
    selling_status = item.get("SellingStatus") or {}
    listing_status = selling_status.get("ListingStatus")
    if status_resp.get("Ack") not in ("Success", "Warning") or listing_status != "Active":
        log_row.update({"ebay_status": "refused", "ebay_error": f"Listing not Active (status={listing_status})"})
        supabase.table("ebay_quantity_push_log").insert(log_row).execute()
        raise HTTPException(409, f"eBay listing {ebay_item_id} is not currently Active "
                                  f"(status: {listing_status or 'unknown'}) — refusing to push a quantity update.")

    quantity_sold = _safe_int(selling_status.get("QuantitySold")) or 0
    new_total_qty = body.quantity + quantity_sold
    log_row["live_quantity_sold"] = quantity_sold
    log_row["new_total_qty_sent"] = new_total_qty

    # eBay first. If this fails, stop entirely -- Shopify is never touched.
    try:
        if body.quantity <= 0:
            # Same fix as ebay_sync_push -- ReviseItem refuses a Quantity that
            # computes to 0 available (error 515), so 0 has to end the listing
            # via EndItem instead. See _ebay_end_item for why.
            _ebay_end_item(business_id, ebay_item_id, "NotAvailable")
            final_status = _live_check_ebay_listing_status(business_id, ebay_item_id)
            log_row["ebay_status"] = f"success_ended ({final_status})"
        else:
            _ebay_revise_quantity_only(business_id, ebay_item_id, new_total_qty)
            log_row["ebay_status"] = "success"
    except Exception as e:
        log_row.update({"ebay_status": "error", "ebay_error": str(e)})
        supabase.table("ebay_quantity_push_log").insert(log_row).execute()
        raise HTTPException(502, f"eBay quantity update failed — nothing was changed on Shopify either. Error: {e}")

    # eBay succeeded — reflect it locally right away so the Inventory view
    # doesn't show a stale number until the next scheduled sync. (The ended-
    # item branch above already wrote this via _live_check_ebay_listing_status.)
    if body.quantity > 0:
        try:
            supabase.table("ebay_listing_status").update({
                "quantity": new_total_qty, "quantity_available": body.quantity,
            }).eq("business_id", business_id).eq("item_id", ebay_item_id).execute()
        except Exception as e:
            print(f"push_inventory_quantity: local ebay_listing_status update failed (eBay itself succeeded): {e}")

    # Now Shopify — same proven absolute-set mutation shape already used by
    # the automated sync (_push_shopify_qty_updates), not a relative delta.
    shopify_error = None
    try:
        import requests as _req
        settings = get_ebay_settings(business_id)
        domain = (settings.get("SHOPIFY_STORE_DOMAIN", "") or "").strip().replace("https://", "").replace("http://", "").strip("/")
        if not domain:
            raise Exception("Shopify not connected")
        shopify_token = get_shopify_access_token(business_id)
        sp_headers = {"X-Shopify-Access-Token": shopify_token, "Content-Type": "application/json"}
        location_id = _get_shopify_primary_location_id(domain, sp_headers)
        if not location_id:
            raise Exception("Could not determine Shopify location")
        variant_data = _fetch_shopify_variants_by_ids(domain, shopify_token, location_id, [str(shopify_variant_id)])
        inv_item_id = (variant_data.get(str(shopify_variant_id)) or {}).get("inventory_item_id")
        if not inv_item_id:
            raise Exception(f"Could not resolve Shopify variant {shopify_variant_id} to a live inventory item")
        mutation = {
            "query": """mutation setQty($input: InventorySetQuantitiesInput!) {
                inventorySetQuantities(input: $input) {
                    inventoryAdjustmentGroup { changes { name delta quantityAfterChange } }
                    userErrors { field message }
                }
            }""",
            "variables": {"input": {
                "reason": "correction", "name": "available", "ignoreCompareQuantity": True,
                "quantities": [{"inventoryItemId": inv_item_id, "locationId": location_id, "quantity": body.quantity}],
            }},
        }
        r = _req.post(f"https://{domain}/admin/api/2024-10/graphql.json", headers=sp_headers, json=mutation, timeout=20)
        resp = r.json() if r.status_code == 200 else {}
        gql_errors = resp.get("errors") or []
        user_errors = (resp.get("data", {}) or {}).get("inventorySetQuantities", {}).get("userErrors", []) if not gql_errors else gql_errors
        if r.status_code != 200 or gql_errors or user_errors:
            raise Exception("; ".join(e.get("message", "") for e in user_errors) or f"HTTP {r.status_code}")
        log_row["shopify_status"] = "success"
        try:
            supabase.table("shopify_inventory").update({"quantity": body.quantity}).eq("business_id", business_id).eq("id", str(shopify_variant_id)).execute()
        except Exception as e:
            print(f"push_inventory_quantity: local shopify_inventory update failed (Shopify itself succeeded): {e}")
    except Exception as e:
        shopify_error = str(e)
        log_row.update({"shopify_status": "error", "shopify_error": shopify_error})

    supabase.table("ebay_quantity_push_log").insert(log_row).execute()

    if shopify_error:
        raise HTTPException(207, f"eBay updated successfully to {body.quantity} available, but Shopify failed: "
                                  f"{shopify_error} — you'll need to fix Shopify manually.")
    return {"ok": True, "ebay_available_qty": body.quantity, "shopify_qty": body.quantity}

@app.get("/api/inventory")
async def list_inventory(request: Request):
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")

    def _fetch_all(table, select_cols):
        all_rows, start, page_size = [], 0, 1000
        while True:
            res = supabase.table(table).select(select_cols).eq("business_id", business_id)\
                .range(start, start + page_size - 1).execute()
            page = res.data or []
            all_rows.extend(page)
            if len(page) < page_size:
                break
            start += page_size
        return all_rows

    inv_rows = _fetch_all("inventory_match", "*")
    ebay_by_id = {r["id"]: r for r in _fetch_all("ebay_inventory", "*")}
    shopify_records_all = _fetch_all("shopify_inventory", "*")
    # ebay_inventory.listing_status is never actually populated (sync_inventory
    # hardcodes it to None) -- ebay_listing_status is the real, reliably-kept-
    # current source for "is this eBay item still active", used here to tell a
    # genuinely-unmatched Shopify item apart from one whose eBay side ended (the
    # end-and-relist scenario) for the Manual Match candidate list below.
    _els_rows = _fetch_all("ebay_listing_status", "item_id,listing_status,shipping_cost")
    active_ebay_ids = {r["item_id"] for r in _els_rows
                        if r.get("item_id") and r.get("listing_status") == "Active"}
    shipping_by_item_id = {r["item_id"]: r.get("shipping_cost") for r in _els_rows if r.get("item_id")}
    # shopify_inventory.id is a VARIANT id (that's what fetch_shopify_inventory_items
    # keys it by), but inventory_match.shopify_id can hold either that same variant
    # id (rows from ordinary title-matching) OR a plain PRODUCT id (rows from the
    # Lister dual-publish confirm/backfill, which only ever has listings.shopify_
    # product_id available, not a variant id) -- two different ID spaces that were
    # incorrectly treated as interchangeable. Confirmed root cause: backfilled rows
    # showed correct data in inventory_match itself but blank on this page, because
    # the lookup only ever tried the variant-id dict. Try both now.
    shopify_by_id = {r["id"]: r for r in shopify_records_all}
    shopify_by_product_id = {r["product_id"]: r for r in shopify_records_all if r.get("product_id")}

    # For the "Publish to Shopify from here" button: only actually possible for
    # eBay-only items that have a real listings row behind them (title, price,
    # photos, description all live there — a plain eBay-native item Lister never
    # touched has none of that saved locally, so there's nothing to publish from).
    all_ebay_ids = [row.get("ebay_id") for row in inv_rows if row.get("ebay_id")]
    listing_by_ebay_id = {}
    for i in range(0, len(all_ebay_ids), 200):
        chunk = all_ebay_ids[i:i+200]
        lres = supabase.table("listings").select("id,ebay_item_id,shopify_product_id").in_("ebay_item_id", chunk).execute()
        for l in (lres.data or []):
            listing_by_ebay_id[l["ebay_item_id"]] = l

    results = []
    for row in inv_rows:
        if row.get("hidden"):
            continue
        e = ebay_by_id.get(row.get("ebay_id")) or {}
        s = shopify_by_id.get(row.get("shopify_id")) or shopify_by_product_id.get(row.get("shopify_id")) or {}
        matching_listing = listing_by_ebay_id.get(row.get("ebay_id"))
        results.append({
            "id": row["id"], "title": row["title"], "matched_by": row.get("matched_by"),
            "hd_id": row.get("hd_id"),
            "ebay_sku": e.get("sku"), "ebay_qty": e.get("quantity"), "ebay_condition": e.get("condition"),
            "ebay_item_id": e.get("item_id") or (row.get("ebay_id") if row.get("ebay_id") else None),
            "ebay_gallery_url": e.get("gallery_url"), "ebay_price": e.get("price"),
            "ebay_local_photo_ids": e.get("local_photo_ids"), "ebay_created_at": e.get("start_time"),
            "shopify_sku": s.get("sku"), "shopify_qty": s.get("quantity"), "shopify_price": s.get("price"), "shopify_status": s.get("status"),
            "shopify_product_id": s.get("product_id") or (row.get("shopify_id") if row.get("shopify_id") else None),
            "ebay_still_active": (row.get("ebay_id") in active_ebay_ids) if row.get("ebay_id") else None,
            "ebay_shipping_cost": shipping_by_item_id.get(row.get("ebay_id")),
            "listing_id": (matching_listing.get("id") if matching_listing and not matching_listing.get("shopify_product_id") else None),
            "qty_variance": (e.get("quantity") is not None and s.get("quantity") is not None and e.get("quantity") != s.get("quantity")),
        })
    return {"inventory": results}

@app.get("/api/inventory/ebay-only")
async def api_inventory_ebay_only(request: Request):
    """Lean, standalone version of the 'eBay Only' subset of /api/inventory --
    filters at the query level (ebay_id set, shopify_id null) instead of fetching
    every inventory_match/ebay_inventory/shopify_inventory row (the full endpoint's
    approach) just to filter down to this smaller slice client-side. Same
    fields/actions as the Inventory tab shows for these rows."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")

    rows = []
    start = 0
    while True:
        page = (supabase.table("inventory_match").select("id,title,ebay_id")
                .eq("business_id", business_id).eq("hidden", False)
                .not_.is_("ebay_id", "null").is_("shopify_id", "null")
                .range(start, start + 999).execute().data or [])
        rows.extend(page)
        if len(page) < 1000:
            break
        start += 1000

    ebay_ids = [r["ebay_id"] for r in rows if r.get("ebay_id")]
    ebay_by_id, listing_by_ebay_id = {}, {}
    for i in range(0, len(ebay_ids), 200):
        chunk = ebay_ids[i:i + 200]
        for e in (supabase.table("ebay_inventory").select("id,sku,quantity,condition,price,gallery_url,local_photo_ids,start_time")
                  .in_("id", chunk).execute().data or []):
            ebay_by_id[e["id"]] = e
        for l in (supabase.table("listings").select("id,ebay_item_id,shopify_product_id")
                  .in_("ebay_item_id", chunk).execute().data or []):
            listing_by_ebay_id[l["ebay_item_id"]] = l

    results = []
    for row in rows:
        e = ebay_by_id.get(row.get("ebay_id")) or {}
        matching_listing = listing_by_ebay_id.get(row.get("ebay_id"))
        results.append({
            "id": row["id"], "title": row["title"],
            "ebay_item_id": e.get("id") or row.get("ebay_id"),
            "ebay_sku": e.get("sku"), "ebay_qty": e.get("quantity"), "ebay_price": e.get("price"),
            "ebay_gallery_url": e.get("gallery_url"), "ebay_local_photo_ids": e.get("local_photo_ids"),
            "ebay_created_at": e.get("start_time"),
            "listing_id": (matching_listing.get("id") if matching_listing and not matching_listing.get("shopify_product_id") else None),
        })
    results.sort(key=lambda r: r.get("ebay_created_at") or "", reverse=True)
    return {"items": results}

@app.post("/api/inventory/{row_id}/hide")
async def hide_inventory_row(row_id: str, request: Request):
    """Permanently removes a row from the main Inventory list, without deleting
    any underlying data -- just a hidden flag on inventory_match. Reversible
    via the Unhide list at the top of the page."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    res = supabase.table("inventory_match").update({"hidden": True})\
        .eq("id", row_id).eq("business_id", business_id).execute()
    if not res.data:
        raise HTTPException(404, "Row not found")
    return {"ok": True, "id": row_id, "hidden": True}

@app.post("/api/inventory/{row_id}/unhide")
async def unhide_inventory_row(row_id: str, request: Request):
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    res = supabase.table("inventory_match").update({"hidden": False})\
        .eq("id", row_id).eq("business_id", business_id).execute()
    if not res.data:
        raise HTTPException(404, "Row not found")
    return {"ok": True, "id": row_id, "hidden": False}

@app.get("/api/inventory/hidden")
async def list_hidden_inventory(request: Request):
    """Every currently-hidden row, for the Unhide panel -- title is enough to
    identify what you're restoring, doesn't need the full eBay/Shopify detail
    the main list shows."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    res = supabase.table("inventory_match").select("id,title,hd_id").eq("business_id", business_id)\
        .eq("hidden", True).execute()
    return {"hidden": res.data or []}

@app.post("/api/inventory/hide-no-created-date")
async def hide_no_created_date(request: Request):
    """Bulk-hides every eBay-only row (has an ebay_id, no shopify_id -- the
    'eBay only' status) whose ebay_inventory.start_time is empty. Per explicit
    request: these are treated as safe to clear out in bulk rather than
    reviewing one at a time. Skips already-hidden rows and anything with a
    real hd_id (a confirmed/permanent match), same protection every other
    part of this feature already respects."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")

    match_rows = _fetch_all_for_business(business_id, "inventory_match", "id,ebay_id,shopify_id,hidden,hd_id")
    ebay_rows = _fetch_all_for_business(business_id, "ebay_inventory", "id,start_time")
    start_time_by_id = {str(r.get("id")): r.get("start_time") for r in ebay_rows}

    to_hide = []
    for row in match_rows:
        if row.get("hidden") or row.get("hd_id"):
            continue
        if not row.get("ebay_id") or row.get("shopify_id"):
            continue  # not "eBay only"
        if not start_time_by_id.get(str(row.get("ebay_id"))):
            to_hide.append(row["id"])

    for row_id in to_hide:
        supabase.table("inventory_match").update({"hidden": True}).eq("id", row_id).execute()

    return {"hidden_count": len(to_hide)}


@app.get("/api/ebay/debug-raw-order/{order_id}")
async def ebay_debug_raw_order(order_id: str, request: Request):
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    import requests as _req
    token = get_ebay_access_token(business_id)
    r = _req.get(
        f"{EBAY_API_BASE}/sell/fulfillment/v1/order/{order_id}",
        headers=ebay_headers(token, content_language=False),
        timeout=15,
    )
    return {"status": r.status_code, "body": r.json() if r.headers.get("content-type","").startswith("application/json") else r.text[:2000]}

@app.get("/api/ebay/debug-inventory-item/{item_id}")
async def ebay_debug_inventory_item(item_id: str, request: Request):
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    import requests as _req
    res = supabase.table("listings").select("*").eq("id", item_id).limit(1).execute()
    if not res.data:
        raise HTTPException(404, "Listing not found")
    listing = res.data[0]
    sku = listing.get("ebay_sku") or f"lister-{listing['id']}"
    biz_id = listing.get("business_id")
    try:
        token = get_ebay_access_token(biz_id)
        r = _req.get(f"{EBAY_API_BASE}/sell/inventory/v1/inventory_item/{sku}",
                      headers=ebay_headers(token, content_language=False), timeout=15)
        return {"sku": sku, "status": r.status_code, "body": r.json() if r.headers.get("content-type","").startswith("application/json") else r.text[:1000]}
    except Exception as e:
        raise HTTPException(400, str(e))

@app.get("/api/shopify/debug-scopes")
async def shopify_debug_scopes(request: Request):
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    import requests as _req
    try:
        settings = get_ebay_settings(business_id)
        domain = (settings.get("SHOPIFY_STORE_DOMAIN", "") or "").strip().replace("https://", "").replace("http://", "").strip("/")
        token = get_shopify_access_token(business_id, force_refresh=True)
        r = _req.post(
            f"https://{domain}/admin/api/2024-10/graphql.json",
            headers={"X-Shopify-Access-Token": token, "Content-Type": "application/json"},
            json={"query": "{ currentAppInstallation { accessScopes { handle } } }"},
            timeout=15,
        )
        return {"status": r.status_code, "body": r.json() if r.headers.get("content-type","").startswith("application/json") else r.text[:500]}
    except Exception as e:
        raise HTTPException(400, str(e))

@app.post("/api/listings/{item_id}/shopify-publish")
async def api_shopify_publish(item_id: str, request: Request):
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    try:
        res = supabase.table("listings").select("*").eq("id", item_id).limit(1).execute()
        if not res.data:
            raise HTTPException(404, "Listing not found")
        listing = res.data[0]
        _require_valid_lot_sku_for_publish(business_id, listing)
        result = push_listing_to_shopify(listing)
        _record_new_shopify_publish_locally(business_id, result)  # writes shopify_inventory only here (no ebay_id arg) -- inventory_match linking is _maybe_confirm_inventory_match's job below, already existing
        try:
            supabase.table("listings").update({
                "shopify_product_id": str(result.get("product_id") or ""),
                "shopify_status": result.get("status") or "active",
                "shopify_error": None,
            }).eq("id", item_id).execute()
        except Exception as col_err:
            # shopify_* columns may not exist yet on this Supabase project
            print(f"shopify-publish: product created ({result}) but failed to save status columns: {col_err}")
        try:
            _maybe_confirm_inventory_match(business_id, item_id)
        except Exception:
            pass
        return {"ok": True, **result}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))

def _publish_one_to_shopify(business_id: str, item_id: str) -> dict:
    """Shared single-item publish logic, used by both the single endpoint above and
    the batch job below, so the two paths can never drift out of sync."""
    res = supabase.table("listings").select("*").eq("id", item_id).limit(1).execute()
    if not res.data:
        raise Exception("Listing not found")
    listing = res.data[0]
    _require_valid_lot_sku_for_publish(business_id, listing)
    result = push_listing_to_shopify(listing)
    try:
        supabase.table("listings").update({
            "shopify_product_id": str(result.get("product_id") or ""),
            "shopify_status": result.get("status") or "active",
            "shopify_error": None,
        }).eq("id", item_id).execute()
    except Exception as col_err:
        print(f"shopify-publish: product created ({result}) but failed to save status columns: {col_err}")
    try:
        _maybe_confirm_inventory_match(business_id, item_id)
    except Exception:
        pass
    return result

_shopify_batch_publish_job_status = {}  # business_id -> {"running": bool, "result": dict|None, "started_at": iso, "finished_at": iso|None}

def _shopify_batch_publish_work(business_id: str, item_ids: list) -> dict:
    published, failed_ids, errors = 0, [], {}
    for item_id in item_ids:
        try:
            _publish_one_to_shopify(business_id, item_id)
            published += 1
        except Exception as e:
            failed_ids.append(item_id)
            errors[item_id] = str(e)
            try:
                supabase.table("listings").update({"shopify_status": "failed", "shopify_error": str(e)}).eq("id", item_id).execute()
            except Exception:
                pass
        _shopify_batch_publish_job_status[business_id]["progress"] = {"done": published + len(failed_ids), "total": len(item_ids)}
    return {"published": published, "failed": len(failed_ids), "failed_ids": failed_ids, "errors": errors}

async def _run_shopify_batch_publish_background(business_id: str, item_ids: list):
    import asyncio, datetime as _dt
    try:
        result = await asyncio.to_thread(_shopify_batch_publish_work, business_id, item_ids)
        _shopify_batch_publish_job_status[business_id] = {
            "running": False, "result": result,
            "started_at": _shopify_batch_publish_job_status.get(business_id, {}).get("started_at"),
            "finished_at": _dt.datetime.utcnow().isoformat(),
        }
    except Exception as e:
        _shopify_batch_publish_job_status[business_id] = {
            "running": False, "result": {"error": str(e)},
            "started_at": _shopify_batch_publish_job_status.get(business_id, {}).get("started_at"),
            "finished_at": _dt.datetime.utcnow().isoformat(),
        }

@app.post("/api/listings/shopify-publish-batch")
async def shopify_publish_batch(request: Request, item_ids: List[str] = Body(..., embed=True)):
    """Publishes multiple listings to Shopify as a background job -- runs entirely
    server-side, so closing the tab or navigating away no longer stops it partway
    through (the previous bulk button was a plain client-side JS loop that died the
    instant the tab closed, same class of problem every other bulk action in this
    file already solved with this exact background-job + polling pattern)."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    if not item_ids:
        raise HTTPException(400, "No items selected")
    if _shopify_batch_publish_job_status.get(business_id, {}).get("running"):
        return {"started": False, "already_running": True}
    import asyncio, datetime as _dt
    _shopify_batch_publish_job_status[business_id] = {
        "running": True, "result": None, "progress": {"done": 0, "total": len(item_ids)},
        "started_at": _dt.datetime.utcnow().isoformat(), "finished_at": None,
    }
    asyncio.create_task(_run_shopify_batch_publish_background(business_id, item_ids))
    return {"started": True}

@app.get("/api/listings/shopify-publish-batch-status")
async def shopify_publish_batch_status(request: Request, response: Response):
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    response.headers["Cache-Control"] = "no-store"
    return _shopify_batch_publish_job_status.get(business_id, {"running": False, "result": None, "progress": None, "started_at": None, "finished_at": None})


@app.post("/api/inventory/{row_id}/publish-ebay-only-to-shopify")
async def publish_ebay_only_to_shopify(row_id: str, request: Request):
    """Publishes a genuinely eBay-only inventory item (no Lister listings row behind
    it at all — a plain eBay-native item Lister never touched) to Shopify, using the
    real photos already downloaded into Lister's own storage by
    /api/inventory/pull-ebay-photos. Refuses to run until that step has actually
    happened for this specific item (checks for a non-empty local_photo_ids), rather
    than silently falling back to eBay's external CDN or publishing with no photos —
    those photos are meant to live in our own environment first, by design."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    res = supabase.table("ebay_inventory").select("*").eq("id", row_id).eq("business_id", business_id).limit(1).execute()
    if not res.data:
        raise HTTPException(404, "eBay inventory item not found")
    row = res.data[0]
    local_photo_ids = (row.get("local_photo_ids") or "").strip()
    if not local_photo_ids:
        raise HTTPException(400, "No local photos yet for this item — click 'Pull eBay-Only Photos' first.")
    photo_ids = [p.strip() for p in local_photo_ids.split(",") if p.strip()]
    image_urls = [photo_url(p) for p in photo_ids if photo_url(p)]
    if not image_urls:
        raise HTTPException(400, "local_photo_ids was set but none resolved to a real photo URL — re-run the photo pull for this item.")

    synthetic_listing = {
        "business_id": business_id,
        "title": row.get("title"),
        "price": row.get("price"),
        "quantity": row.get("quantity") or 1,
        "ebay_sku": row.get("sku"),
        "id": row_id,
    }
    # Same lock every other publish path enforces — a SKU must match a real lot
    # before anything goes live, no exceptions for this path either.
    _require_valid_lot_sku_for_publish(business_id, synthetic_listing)
    try:
        result = push_listing_to_shopify(synthetic_listing, image_urls_override=image_urls)
        _record_new_shopify_publish_locally(business_id, result, ebay_id=row_id)
        return {"ok": True, **result}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))

# ── SETTINGS ──────────────────────────────────────────────────── #

@app.get("/api/settings")
async def get_settings(request: Request):
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    try:
        return get_ebay_settings(business_id)
    except Exception:
        return {}

class SaveSetting(BaseModel):
    key:   str
    value: str

@app.post("/api/settings")
async def save_setting(body: SaveSetting, request: Request):
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Unauthorized")
    try:
        existing = supabase.table("app_settings").select("key").eq("business_id", business_id).eq("key", body.key).limit(1).execute()
        if existing.data:
            supabase.table("app_settings").update({"value": body.value}).eq("business_id", business_id).eq("key", body.key).execute()
        else:
            supabase.table("app_settings").insert({"business_id": business_id, "key": body.key, "value": body.value}).execute()
        return {"ok": True}
    except Exception as e:
        raise HTTPException(500, str(e))
import hashlib, secrets

def hash_password(password: str) -> str:
    salt = secrets.token_hex(16)
    hashed = hashlib.sha256((salt + password).encode()).hexdigest()
    return f"{salt}:{hashed}"

def verify_password(password: str, stored: str) -> bool:
    try:
        salt, hashed = stored.split(":")
        return hashlib.sha256((salt + password).encode()).hexdigest() == hashed
    except Exception:
        return False

def get_business_id(request: Request):
    """Get business_id from session cookie."""
    session = request.cookies.get("session_id")
    if not session:
        return None
    try:
        res = supabase.table("sessions").select("business_id").eq("token", session).execute()
        if res.data:
            return res.data[0]["business_id"]
    except Exception:
        pass
    return None

@app.get("/login")
async def login_page(request: Request, error: str = ""):
    return templates.TemplateResponse("login.html", {"request": request, "error": error})

@app.post("/login")
async def login_submit(request: Request):
    form = await request.form()
    email = str(form.get("email", "")).strip().lower()
    password = str(form.get("password", ""))
    try:
        res = supabase.table("businesses").select("id,password_hash").eq("email", email).execute()
        if not res.data:
            return templates.TemplateResponse("login.html", {"request": request, "error": "Invalid email or password"})
        biz = res.data[0]
        if not verify_password(password, biz["password_hash"]):
            return templates.TemplateResponse("login.html", {"request": request, "error": "Invalid email or password"})
        token = secrets.token_hex(32)
        supabase.table("sessions").insert({"token": token, "business_id": biz["id"]}).execute()
        from fastapi.responses import RedirectResponse
        resp = RedirectResponse("/", status_code=302)
        resp.set_cookie("session_id", token, httponly=True, max_age=60*60*24*30)
        return resp
    except Exception as e:
        return templates.TemplateResponse("login.html", {"request": request, "error": f"Login failed: {e}"})

@app.get("/register")
async def register_page(request: Request, error: str = ""):
    return templates.TemplateResponse("register.html", {"request": request, "error": error})

@app.post("/register")
async def register_submit(request: Request):
    form = await request.form()
    business_name = str(form.get("business_name", "")).strip()
    email = str(form.get("email", "")).strip().lower()
    password = str(form.get("password", ""))
    if not business_name or not email or not password:
        return templates.TemplateResponse("register.html", {"request": request, "error": "All fields required"})
    if len(password) < 8:
        return templates.TemplateResponse("register.html", {"request": request, "error": "Password must be at least 8 characters"})
    try:
        existing = supabase.table("businesses").select("id").eq("email", email).execute()
        if existing.data:
            return templates.TemplateResponse("register.html", {"request": request, "error": "Email already registered"})
        password_hash = hash_password(password)
        res = supabase.table("businesses").insert({
            "name": business_name,
            "email": email,
            "password_hash": password_hash
        }).execute()
        business_id = res.data[0]["id"]
        token = secrets.token_hex(32)
        supabase.table("sessions").insert({"token": token, "business_id": business_id}).execute()
        from fastapi.responses import RedirectResponse
        resp = RedirectResponse("/", status_code=302)
        resp.set_cookie("session_id", token, httponly=True, max_age=60*60*24*30)
        return resp
    except Exception as e:
        return templates.TemplateResponse("register.html", {"request": request, "error": f"Registration failed: {e}"})

@app.get("/logout")
async def logout(request: Request):
    from fastapi.responses import RedirectResponse
    token = request.cookies.get("session_id")
    if token:
        try:
            supabase.table("sessions").delete().eq("token", token).execute()
        except Exception:
            pass
    resp = RedirectResponse("/login", status_code=302)
    resp.delete_cookie("session_id")
    return resp
